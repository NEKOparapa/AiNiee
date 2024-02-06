import re
import os
import json
import pandas as pd
import traceback
import openpyxl
from chardet import detect

# v1.6
class Jr_Tpp():
    def __init__(self,config:dict,path:str=False):
        self.config=config
        self.ProgramData= {} # 翻译工程数据,键为文件名，值为DataFrame，列为['原文','译文','地址','标签']，同时设置原文为索引
        self.BlackDir=config['BlackDir']   # 地址黑名单，自动打标签“BlackDir”
        self.BlackFiles=config['BlackFiles']    # 黑名单文件，不读取这些文件，需要是文件全名（也不会有人把Map加黑名单吧）
        self.BlackCode=config['BlackCode']      # 效果同blackdir，只不过这个是code
        self.NameWithout=config['NameWithout']  #   对这些字段搜索反选后，打Name标签
        self.codewithnames=config['codewithnames']  # dnb用，包裹文件名的标识符
        self.ReadCode=config['ReadCode']    # 只读取这些code的文本
        self.ja=config['ja']    # 是否为日文游戏，日文游戏的情况下，只会提取含中日字符的文本，dnb也只会处理含中日字符的文件名
        if path:
            self.load(path) # 从工程文件加载

####################################读取和注入游戏文本，保存与加载翻译工程，导入翻译结果等基本功能###################################
    # 用openpyxl读xlsx，因为用pandas会把'=xxx'的字符串读成NaN，而且解决不了
    def __Readxlsx(self, name):
        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(name)
            # 获取所有工作表的名称
            sheet_names = workbook.sheetnames
            # 选择第一个工作表
            worksheet = workbook[sheet_names[0]]
            # 读取列名
            column_names = [cell.value for cell in worksheet[1] if cell.value is not None]
            # 检查是否只有列名
            if len(column_names) == 0:
                # 创建空DataFrame
                df = pd.DataFrame(columns=column_names)
            else:
                # 读取数据
                data = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    data.append(row)

                # 创建DataFrame
                df = pd.DataFrame(data, columns=column_names)
            # 关闭Excel文件
            workbook.close()
            return df
        except Exception as e:
            print(traceback.format_exc())
            print(e)
            print('请关闭所有xlsx文件再试')
    # 用openpyxl写xlsx，默认只导出原文和译文列
    def __Writexlsx(self,df,name,full=False):
        # 创建一个Excel工作簿
        workbook = openpyxl.Workbook()
        # 获取默认的工作表
        sheet = workbook.active
        # 定义需要导出的列
        if not full:
            columns_to_export = ['原文', '译文']
        else:
            columns_to_export = ['原文', '译文','地址','标签','code']
        # 写入表头
        header_row = [column for column in columns_to_export]
        sheet.append(header_row)
        # 将DataFrame的数据写入工作表
        for index, row in df.iterrows():
            data_row = [row[column] for column in columns_to_export]
            sheet.append(data_row)
        # 将单元格格式设置为文本格式
        for column_cells in sheet.columns:
            for cell in column_cells:
                cell.number_format = '@'
        # 编码所有字符串为Unicode
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    encoded_string = cell.value.encode('utf-8')
                    cell.value = encoded_string.decode('utf-8')
        # 保存工作簿为xlsx文件
        workbook.save(name)
    # 读取json文件中含中日字符的字符串，并记录其地址。
    # 输入json文件的内容，返回其中所有文本组成的list
    def __ReadFile(self,data,FileName:str,code:int =False) -> list:
        res=[]
        tp = type(data)
        if tp == dict:
            for key in data.keys():
                # FileName用来记录地址
                code = data.get('code', False)
                res+=self.__ReadFile(data[key],FileName+'\\'+key,code)

        elif tp == list:
            for i in range(0,len(data)):
                res+=self.__ReadFile(data[i],FileName+'\\'+str(i),code)
        # 是字符串，而且含中日字符(System.json\gameTitle不论是否含中日字符，都进
        elif tp==str and (not self.ja or re.search(r'[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5]',data)
                          or r'System.json\gameTitle' in FileName) :
            if r'System.json\gameTitle' in FileName and data=='':data=' '# 游戏名为空时，变为空格，否则会报错
            if not code:code='-1'
            # 只有特定code才读取,readcode为空时，全读
            if str(code) in self.ReadCode or len(self.ReadCode)==0:
                res.append([data,'',FileName,'',str(code)])
        return res
    # 读取文件夹路径，返回包括其子文件夹内的所有文件名
    def __ReadFolder(self,dir:str) -> list:
        res=[]
        if os.path.isdir(dir):
            FileList = os.listdir(dir)
            for name in FileList:
                temp=dir+'\\'+name
                if os.path.isfile(temp):
                    res+=[temp]
                elif os.path.isdir(temp):
                    res+=self.__ReadFolder(temp)
            return res
        else:return [dir]
    # 判断地址是否为黑名单地址
    def __IfBlackDir(self,Dir:str) ->bool:
        for blackdir in self.BlackDir:
            # 如果有任意一个黑名单遍历后仍为True，则说明地址为黑名单地址，break出来
            dirsig = True
            for i in blackdir.split('*'):
                if i not in Dir:
                    dirsig = False
                    break
            if dirsig:
                return True
        return False
    # 去除DataFrame中重复的行，将重复行的地址和code添加到被保留的行中,在地址为黑名单地址或code为黑名单code的情况下，不保留其数据
    def __RemoveDuplicated(self,data:pd.DataFrame) -> pd.DataFrame:
        if '地址' and 'code' in data.columns:
            a = data[~data.index.duplicated()].copy()  # 去除重复行的
            b = data[data.index.duplicated()].copy()  # 仅含重复行的
            a = {'a': a}
            b = {'b': b}
            for index in a['a'].index:
                if index in b['b'].index:
                    Dir = list(b['b'][b['b'].index == index]['地址'])
                    code = list(b['b'][b['b'].index == index]['code'])
                    black=False
                    if self.__IfBlackDir(Dir):black=True
                    if code in self.BlackCode: black = True
                    if not black:
                        for i in range(0,len(Dir)):
                            a['a'].loc[index,'地址']+='☆↑↓'+Dir[i]
                            a['a'].loc[index, 'code'] += ','+code[i]
            return a['a']
        else:
            return data[~data.index.duplicated()]
    # 将ReadFile得到的数据转化为DataFrame
    def __toDataFrame(self,data:list) -> pd.DataFrame:
        DataFrame = pd.DataFrame(data, columns=['原文', '译文', '地址', '标签','code'])
        DataFrame.index = list(DataFrame['原文'])
        DataFrame=self.__RemoveDuplicated(DataFrame)# 去除原文重复的行
        return DataFrame
    # 将后缀json和xlsx互相转化
    def __nameswitch(self,name,csv:bool=False):
        name = name.split('\\')[-1]
        resname = ''
        temp = name.split('.')
        for i in temp[:-1]:
            resname += i + '.'
        if not csv:
            if 'xlsx' in name:
                resname += 'json'
            else:
                resname += 'xlsx'
        else:
            if 'csv' in name:
                resname += 'json'
            else:
                resname += 'csv'
        return resname
    # 按照Dir逐级读取文件内容，直到读到untrs，将其替换为trsed，然后逐级返回
    def __WriteFile(self,data,untrs:str,trsed:str,Dir:list):
        # 获取文本在文件内的地址
        if type(data)==list:
            i=int(Dir[0])
            data[i]=self.__WriteFile(data[i],untrs,trsed,Dir[1:])
        elif type(data)==dict:
            data[Dir[0]]=self.__WriteFile(data[Dir[0]],untrs,trsed,Dir[1:])
        elif type(data)==str and len(Dir)==0:
            if data==untrs or (data=='' and untrs==' '):
                data=trsed
            else:
                print(f'原文\"{data}\"不匹配')
        return data
    # 检查译文中是否存在空数据，若存在，用原文填充，并输出提示
    def __CheckNAN(self):
        nanlist=[]
        for name in self.ProgramData.keys():
            DataFrame=self.ProgramData[name]
            NANFrame=DataFrame[DataFrame['译文'].isnull()]
            if len(NANFrame):
                indexlist=list(NANFrame.index)
                nanlist+=indexlist
                for index in indexlist:
                    DataFrame.loc[index,'译文']=index
            self.ProgramData[name]=DataFrame
        if len(nanlist):
            print('以下原文没有对应译文，恢复为原文')
            for i in nanlist:
                print(i)
    # 从游戏读取文本,参数为游戏目录，自动标签黑名单地址,并对其应用原文
    def ReadGame(self,GameDir:str):
        Files=self.__ReadFolder(GameDir)
        for File in Files:
            name = File.split('\\')[-1]
            # 只读取data内的json文件
            if '\\data\\' in File and '.json' in name:
                # 黑名单文件不读取
                if name not in self.BlackFiles:
                    print(f'正在读取{name}')
                    try:
                        try:
                            with open(File, 'r', encoding='utf8') as f:
                                data = json.load(f)
                        except:
                            with open(File,'rb') as f:
                                encoding = detect(f.read())['encoding']
                                if encoding==None:
                                    encoding='ansi'
                            with open(File, 'r', encoding=encoding) as f:
                                data = json.load(f)
                    except Exception as e:
                        print(traceback.format_exc())
                        print(e)
                        print(f'无法确定{name}文件编码,且无法用ANSI编码打开，读取失败')
                    TextDatas=self.__ReadFile(data,name)
                    self.ProgramData.update({name:self.__toDataFrame(TextDatas)})
        print('########################读取游戏完成########################')
        ## 标签黑名单地址,并对其应用原文
        # self.LabelBlackDir()
        # 有可能有些文本有多个地址和code，但只有其中一个是黑的，所以不再事先标记黑名单，而在注入时，对每个地址/code单独判断是否是黑的
    # 注入翻译到游戏,BlackLabel为不注入的标签list，默认为'BlackDir',BlackCode默认self.BlackCode
    def InjectGame(self,GameDir:str,path:str,BlackLabel:list=False,BlackCode:list=False):
        self.__CheckNAN()
        if not BlackLabel:
            BlackLabel=['BlackDir']
        if not BlackCode:
            BlackCode=self.BlackCode
        Files = self.__ReadFolder(GameDir)
        for File in Files:
            # 只读取data内的json文件
            name = File.split('\\')[-1]
            if '\\data\\' in File and '.json' in name:
                print(f'正在写入{name}')
                with open(File, 'r', encoding='utf8') as f:
                    data = json.load(f)
                # 黑名单文件不做修改，直接输出
                if name not in self.BlackFiles:
                    # 写入翻译
                    if name in self.ProgramData.keys():
                        DataFrame=self.ProgramData[name]
                        for untrs in DataFrame.index:
                            trsed=DataFrame.loc[untrs,'译文']
                            Dirlist=DataFrame.loc[untrs,'地址'].split('☆↑↓')
                            codelist=DataFrame.loc[untrs,'code'].split(',')
                            labellist=DataFrame.loc[untrs,'标签'].split(',')
                            black=False
                            for label in labellist:
                                if label in BlackLabel:
                                    black=True
                            for i in range(0,len(Dirlist)):
                                Dir=Dirlist[i]
                                code=codelist[i]
                                # 标签，地址和code都不是黑的才写入
                                if not black and code not in BlackCode and not self.__IfBlackDir(Dir):
                                    Dir=Dir.split('\\')
                                    # 写入翻译
                                    data=self.__WriteFile(data,untrs,trsed,Dir[1:])
                    else:
                        print(f'{name}不在工程文件中，工程文件与游戏是否匹配')
                # 获取文件输出路径
                outputpath=(path+'\\data\\'+File.split('\\data\\')[-1]).lstrip('\\')
                # 获取并创建从path到outputpath的路径
                datadir=(outputpath.replace(name,'').replace(path,'').strip('\\')).split('\\')
                temp=path.rstrip('\\')
                for i in datadir:
                    temp+='\\'+i
                    if not os.path.exists(temp.strip('\\')): os.mkdir(temp.strip('\\'))
                # 输出文件
                out = json.dumps(data, ensure_ascii=False)
                with open(outputpath, 'w', encoding='utf8') as f1:
                    print(out, file=f1)

        print('########################写入游戏完成########################')
    # 获取翻译工程内的文件名,Mapxxx.json合并为Mapxxx~XXX.json
    def GetFileNames(self) -> list:
        namelist=list(self.ProgramData.keys())
        mapname=[]
        for name in namelist.copy():
            if 'map' in name.lower() and 'info' not in name.lower():
                namelist.remove(name)
                mapname.append(int(name.replace('Map','').replace('.json','')))
        mapname=sorted(mapname)
        if int(mapname[0])<10:
            mapname[0]=f'00{mapname[0]}'
        elif int(mapname[0])<100:
            mapname[0]=f'0{mapname[0]}'
        if int(mapname[-1])<10:
            mapname[-1]=f'00{mapname[-1]}'
        elif int(mapname[-1])<100:
            mapname[-1]=f'0{mapname[-1]}'
        namelist.append(f'Map{mapname[0]}~{mapname[-1]}.json')
        return namelist
    # 导出单个文件，默认只导出原文和译文列
    def ToXlsx(self,name:str,path:str):
        # 文件名后缀改为xlsx
        outputname = self.__nameswitch(name)
        print(f'正在导出{outputname}')
        data=self.ProgramData[name]
        try:
            self.__Writexlsx(data,path + '\\' + outputname)
        except Exception as e:
            print(traceback.format_exc())
            print(e)
            input('导出失败，请关闭所有xlsx文件后再次尝试')
    # 导出单个文件到csv
    def ToCsv(self,name:str,path:str):
        # 文件名后缀改为xlsx
        outputname = self.__nameswitch(name, True)
        print(f'正在保存{outputname}')
        data = self.ProgramData[name]
        try:
            data.to_csv(path+'\\'+outputname, sep='\uFFFC', encoding='utf8',index=False)
        except Exception as e:
            print(traceback.format_exc())
            print(e)
            input('保存失败，请关闭所有csv文件后再次尝试')
    # 导出工程,数据保存为xlsx
    def Output(self,path:str):
        if not os.path.exists(path+'\\data'): os.mkdir(path+'\\data')
        for name in self.ProgramData.keys():
            self.ToXlsx(name,path+'\\data')
        print('########################导出完成########################')
    # 保存工程文件，数据保存为csv，设置保存为json
    def Save(self,path:str):
        if not os.path.exists(path+'\\csv'): os.mkdir(path+'\\csv')
        for name in self.ProgramData.keys():
            self.ToCsv(name,path+'\\csv')
        out = json.dumps(self.config, indent=4, ensure_ascii=False)
        with open(path+'\\'+'config.json', 'w', encoding='utf8') as f1:
            print(out, file=f1)
        print('########################保存工程完成########################')
    # 导入翻译，从json导入,路径需指定到json文件。可指定某几个文件(list格式)，namelist为False则全选
    # trsdata和path二选一
    def InputFromJson(self,trsdata:dict=False,path:str=False,namelist:list=False):
        if not trsdata:
            try:
                try:
                    with open(path, 'r', encoding='utf8') as f:
                        trsdata = json.load(f)
                except:
                    with open(path, 'rb') as f:
                        encoding = detect(f.read())['encoding']
                        if encoding==None:
                            encoding='ansi'
                    with open(path, 'r', encoding=encoding) as f:
                        trsdata = json.load(f)
            except Exception as e:
                print(traceback.format_exc())
                print(e)
                input(f'读取{path}失败,请确保json文件头格式正确\n'
                      f'若提示UnicodeDecodeError，请确保该文件编码可读')
        # 全选
        if not namelist:
            namelist=self.ProgramData.keys()
        # 对每一个文件都遍历一次json文件（有可能有重复原文）
        for name in namelist:
            DataFrame=self.ProgramData[name]
            for untrs in trsdata.keys():
                if untrs in DataFrame.index:
                    DataFrame.loc[untrs,'译文']=trsdata[untrs]
            self.ProgramData[name]=DataFrame
    # 从DataFrame导入翻译，指定namelist.默认有列名，第一列为原文，第二列为译文。
    # samefile为true时，不搜索原文，直接用译文列覆盖原数据
    def InputFromDataFrame(self, data:pd.DataFrame, namelist:list=False, samefile=False):
        # 格式化data
        col=list(data.columns)
        if len(col)>=2:
            col[0]='原文'
            col[1]='译文'
            data.columns=col
            data.index=list(data['原文'])
            data['译文']=data['译文'].fillna('')
            data=self.__RemoveDuplicated(data)# 去除原文重复的行
            if not namelist:
                namelist=self.ProgramData.keys()
            for name in namelist:
                DataFrame=self.ProgramData[name]
                if not samefile:
                    for untrs in data.index:
                        if untrs in DataFrame.index:
                            DataFrame.loc[untrs,'译文']=data.loc[untrs,'译文']
                else:
                    DataFrame['译文']=data['译文']
                self.ProgramData[name]=DataFrame
    # 导入翻译，从xlsx导入，可指定到文件夹也可指定到xlsx文件。
    def InputFromeXlsx(self, path:str, namelist:list=False, samefile=False):
        FileNames=self.__ReadFolder(path) # 读取path内文件路径
        for file in FileNames:
            name=file.split('\\')[-1]
            if '.xlsx' in name:
                print('正在导入{}'.format(name))
                try:
                    data = self.__Readxlsx(file)
                except Exception as e:
                    print(traceback.format_exc())
                    print(e)
                    input(f'读取{path}失败')
                if not samefile:
                    self.InputFromDataFrame(data,namelist)
                else:
                    # 获取文件的json文件名
                    jsonname =self.__nameswitch(file)
                    # 检查是否存在该json文件
                    if jsonname in self.ProgramData.keys():
                        self.InputFromDataFrame(data,[jsonname],True)
                    else:
                        print(f'{name}没有对应的json文件，已跳过导入该文件')
    # 读取工程,只从xlsx读取，path内需包含一个config
    # xlsx必须含['原文','译文','地址','code']五列，读取时自动将xlsx名转化为json名，并作为键
    def load(self,path:str):
        self.ProgramData={} # 清空工程数据
        FileNames = self.__ReadFolder(path) # 读取path内文件路径
        for file in FileNames:
            name = file.split('\\')[-1]
            if '.csv' in name:
                print('正在加载{}'.format(name))
                data = pd.read_csv(file, sep='\uFFFC',encoding='utf8',engine='python',dtype='str')
                # 检查csv格式是否正确
                if not list(data.columns)==['原文','译文','地址','标签','code']:
                    print(f'{file}文件列名不为[\'原文\',\'译文\',\'地址\',\'标签\',\'code\']，读取失败')
                    continue
                # 设置索引
                data.index = list(data['原文'])
                # 填充译文和标签空数据
                data['译文'] = data['译文'].fillna('')
                data['标签'] = data['标签'].fillna('')
                # 去除原文重复的行
                data=self.__RemoveDuplicated(data)
                # 获取文件的json文件名
                jsonname =self.__nameswitch(file,True)
                self.ProgramData.update({jsonname:data})
        print('########################加载工程完成########################')
############################################输出，增减标签，按条件搜索及其配套操作函数###########################################
    # 添加标签,target格式为{文件名:[索引（原文)]}
    def addlabel(self,target:dict,label:str):
        for name in target.keys():
            for index in target[name]:
                if index in self.ProgramData[name].index:
                    # 如果标签已经存在，不重复添加
                    if label not in self.ProgramData[name].loc[index,'标签']:
                        self.ProgramData[name].loc[index,'标签']=(self.ProgramData[name].loc[index,'标签']+','+label).lstrip(',')
                else:
                    print(f'在{name}中没有找到{name}:\"{index}\"行')
        print(f'添加标签\"{label}\"完成')
    # 去除标签
    def removelabel(self,target:dict,label:str):
        for name in target.keys():
            for index in target[name]:
                if index in self.ProgramData[name].index:
                    # 确保标签存在
                    if label in self.ProgramData[name].loc[index,'标签']:
                        # 先在最左边加上一个逗号，然后删掉逗号+label，然后把最左边逗号去掉
                        self.ProgramData[name].loc[index,'标签']=(','+self.ProgramData[name].loc[index,'标签']).replace\
                            ((','+label),'').lstrip(',')
                    else:
                        print(f'{name}:\"{index}\"行没有{label}标签')
                else:
                    print(f'在{name}中没有找到{name}:\"{index}\"行')
        print(f'去除标签\"{label}\"完成')
    # 默认输出ProgramData，可用来输出搜索结果
    def Display(self,target:dict=False,namelist:list=False):
        if not target:
            target=self.ProgramData.copy()
        if not namelist:
            namelist=target.keys()
        for key in namelist:
            if key in target.keys():
                print(f'{key}:')
                # 完整输出DataFrame
                with pd.option_context('display.max_rows', None,    # 行数
                                       'display.max_columns', None, # 列数
                                       'display.width', None,       # 单元格长度
                                       'display.max_colwidth', None,# 不折叠单元格
                                        # 对齐表格(对不齐）
                                       # 'display.unicode.ambiguous_as_wide', True,
                                       # 'display.unicode.east_asian_width', True,
                                       # 'display.width',200
                                        ):
                    print(target[key].reset_index(drop=True,inplace=False))
            else:
                print(f'{key}不再目标范围内')
    # 按条件搜索，col是按搜索目标，0原文，1译文，2地址，3标签，4code。搜索条件为按*分割,target和返回值格式同ProgramData,notin为True，则搜索不含搜索目标的
    # BigSmall为true则不区分大小写
    def search(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False) ->dict:
        if col == 0:col = '原文'
        elif col == 1:col = '译文'
        elif col == 2:col = '地址'
        elif col == 3: col = '标签'
        elif col==4:col='code'
        string=string.split('*')
        res={}
        if not target:
            target=self.ProgramData.copy()
        if not namelist:
            namelist=target.keys()
        for name in namelist:
            if name in target.keys():
                DataFrame = target[name]
                if BigSmall:
                    temp=DataFrame.apply(lambda x: x.astype(str).str.lower())
                else:
                    temp=DataFrame.copy()
                for chara in string:
                    if BigSmall:
                        chara=chara.lower()
                    temp=temp[temp[col].str.contains(chara)]
                # 根据是否反选，返回未被改变大小写的dataframe
                if notin:
                    temp=DataFrame[~DataFrame.index.isin(temp.index)].dropna()
                else:
                    temp = DataFrame[DataFrame.index.isin(temp.index)].dropna()
                if len(list(temp.index)):
                    res.update({name:temp})
        return res
    # 搜索含A但不含B的，默认colB=colA
    def DoubleSearch(self,A:str,B:str,colA:int,colB:int=False,target:dict=False,namelist:list=False,BigSmall=False):
        res=self.search(A,colA,target,namelist,notin=False,BigSmall=BigSmall)
        if not colB:
            colB=colA
        return self.search(B,colB,res,namelist,True,BigSmall)
    # 替换,只能替换译文列
    def Replace(self,before:str,after:str,target:dict=False,namelist:list=False):
        if not target:
            target=self.ProgramData
        if not namelist:
            namelist=target.keys()
        for name in namelist:
            # DataFrame的replace不知道为什么就是换不掉
            for index in target[name].index:
                target[name].loc[index,'译文']=target[name].loc[index,'译文'].replace(before,after)
        return target
    # 根据搜索结果增减标签,add为True，添加标签,返回搜索结果
    def LabelBySearch(self,string:str,col:int,label:str,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False,add=True):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        target={}
        for name in res.keys():
            target.update({name:list(res[name].index)})
        if add:
            self.addlabel(target,label)
        else:
            self.removelabel(target,label)
        return res
    # 输出搜索结果,返回搜索结果
    def DisplayBySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        self.Display(res)
        return res
    # 将搜索结果导出到当前目录的单个xlsx中,返回搜索结果,默认只导出原文和译文
    def OutputBySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False,OutputName:str='SearchRes.xlsx',full=False):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        if len (res):
            output=pd.concat(list(res.values()),axis=0)
            self.__Writexlsx(output,OutputName,full)
            print(f'已将搜索结果保存为{OutputName}')
        else:
            print('搜索结果为空')
        return res
    # 将搜索结果导出到当前目录的单个json文件中,返回json数据,搜索为空则返回空df
    def JsonBySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False,OutputName:str='SearchRes.json'):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        if len(res):
            res = pd.concat(list(res.values()), axis=0)
            output=dict(zip(res['原文'],res['译文'].fillna('')))
            out = json.dumps(output, indent=4, ensure_ascii=False)
            with open(OutputName,'w',encoding='utf8') as f:
                print(out, file=f)
            print(f'已将结果导出为{OutputName}')
            return output
        else:
            print('搜索结果为空')
            return res
#######################################################预处理和后处理######################################################
    # 标签黑名单地址,标签为'BlackDir'。同时对其应用原文
    def LabelBlackDir(self):
        for i in self.BlackDir:
            res=self.LabelBySearch(i,2,'BlackDir')
            self.ApplyUntrs(res)
        for i in self.BlackCode:
            res = self.LabelBySearch(i, 4, 'BlackDir')
            self.ApplyUntrs(res)
        print('全部黑名单标记完成')
    # 标签名称为'Name',withoutx形如['Actors.json','Items.json','Skills.json'],可除外这些文件中的name，一般对应文件含对应对象的名字
    def LabelName(self,without:list=False):
        # target=self.search('BlackDir',3,notin=True) # 目标为不含'BlackDir'标签的行
        target=self.ProgramData
        if without:
            # 从target中依次除外without
            for i in without:
                target=self.search(i,2,target=target,notin=True)
        # 对剩下的标签'Name',不区分大小写搜索
        self.LabelBySearch('name',2,'Name',target=target,BigSmall=True)
    # 对名字标签并导出json文件
    def GetName(self,without:list=False):
        self.LabelName(without)
        if not os.path.exists('name'): os.mkdir('name')
        namedict=self.JsonBySearch('Name',3,OutputName=r'name\Name.json')
        splited_name={}
        for name in namedict.keys():
            namelist=re.sub('[^\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5]','↓☆←',name).split('↓☆←')
            trsednamelist=re.sub('[^\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5]','↓☆←',name).split('↓☆←')
            if len(namelist)==len(trsednamelist):
                splited_name.update(dict(zip(namelist, trsednamelist)))
            else:
                splited_name.update(dict(zip(namelist,namelist)))
        if '' in splited_name.keys():del splited_name['']
        out = json.dumps(splited_name, indent=4, ensure_ascii=False)
        with open(r'name\Name.json', 'w', encoding='utf8') as f:
            print(out, file=f)
    # 对target翻译应用原文
    def ApplyUntrs(self,target):
        for name in target.keys():
            if name in self.ProgramData.keys():
                for index in target[name].index:
                    if index in self.ProgramData[name].index:
                        self.ProgramData[name].loc[index,'译文']=index
        print('应用原文完成')
    # 对搜索结果应用原文
    def ApplyUntrs_BySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False):
        res = self.search(string, col, target=target, namelist=namelist, notin=notin, BigSmall=BigSmall)
        self.ApplyUntrs(res)
    # 自动对游戏标题添加水印（地址为'System.json\gameTitle'的译文末尾添加mark）
    def AddMark(self,mark:str):
        if 'System.json' in self.ProgramData.keys():
            try:
                data=self.ProgramData['System.json']
                index=list(data[data['地址']==r'System.json\gameTitle'].index)[0]
                self.ProgramData['System.json'].loc[index,'译文']+=mark
                print('########################已添加水印########################')
            except Exception as e:
                print(traceback.format_exc())
                print(e)
                input(f'没有找到游戏标题，添加水印失败')
    # dnb用，分割原文译文的函数
    def __splitbychar(self,q, l):
        b = []
        if l[0] in q and l[1] in q:
            a = q.split(l[0])
            b.append(a[0])
            for i in range(1, len(a)):
                a[i] = l[0] + a[i]
                c = a[i].split(l[1])
                for j in range(0, len(c) - 1): b.append(c[j] + l[1])
                b.append(c[-1])
        return b
    # dnb用，处理原文包含文件名，但不等于文件名时，返回处理后译文
    def __dealin(self,untrs, trsed, filename):
        check = False
        for l in self.codewithnames:
            untrs_list = self.__splitbychar(untrs, l)
            trsed_list = self.__splitbychar(trsed, l)
            length = len(trsed_list)
            # 如果按某分隔符拆分结果长度不相等，不处理并单独导出
            if len(untrs_list) != length:
                self.need2check.update({untrs: trsed})
                return trsed
            elif length:
                # 如果拆分结果长度大于3，说明有不止一个分隔符，需导出确认
                if length > 3: check = True
                # 分隔符不相等
                if l[0] != l[1]:
                    # 被分隔符包裹的字符串只会出现在奇数位
                    for i in range(0, int(length / 2)):
                        # 如果文件名这一段在原文中，把对应位置的译文替换回原文
                        if filename in untrs_list[2 * i + 1]:
                            trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                else:
                    # 分隔符包裹的字符只会出现在4*i+2位
                    for i in range(0, int(length / 4)):
                        if filename in untrs_list[4 * i + 2]:
                            trsed_list[4 * i + 2] = untrs_list[4 * i + 2]
                # 将处理后的文本拼接好，等待下一循环
                trsed = ''
                for i in range(0, length): trsed += trsed_list[i]
        if check:
            self.need2check.update({untrs: trsed})
        return trsed
    # 处理文件名被翻译的问题
    def dnb(self,GameDir):
        print('开始修正文件名')
        temp=self.__ReadFolder(GameDir)
        files=[]
        self.need2check={}
        # 得到含有中日字符的文件名和不带后缀的形式
        for filename in temp:
            filename = filename.split('\\')[-1]
            if re.search(r'[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5]',filename) or not self.ja:
                files.append(filename)
                files+=filename.split('.')[:-1]
        # 遍历
        for name in self.ProgramData.keys():
            for index in self.ProgramData[name].index:
                for filename in files:
                    # 原文等于文件名，译文替换回原文
                    if index== filename:
                        self.ProgramData[name].loc[index,'译文']=index
                    # 原文包含文件名时，在文件名被特定符号包裹的情况下，将被特定符号包裹的文本，按顺序替换回原文
                    elif filename in index:
                        self.ProgramData[name].loc[index,'译文']=self.__dealin(index,self.ProgramData[name].loc[index,'译文'],
                                                                             filename)
        print('########################修正文件名完成########################')
        # 如有需确认的文本，导出json文件
        if len(self.need2check):
            out = json.dumps(self.need2check, indent=4, ensure_ascii=False)
            with open('need2check.json', 'w', encoding='utf8') as f1:
                print(out, file=f1)
            print('已将可能需要人工修正的文本行导出到need2check.json')
    # 自动换行
    def AutoLineFeed(self,linelength:int):
        for name in self.ProgramData.keys():
            DataFrame=self.ProgramData[name]
            for index in DataFrame.index:
                data=DataFrame.loc[index,'译文']
                # 按换行符拆分译文
                lines = data.split('\n')
                res = ''
                for line in lines:
                    q = line
                    # 如果长度大于设定单行最大值，则进行拆分
                    while len(q) > linelength:
                        # 从字符串的第linelength-1个字符开始匹配中文汉字和英文字母，并返回其位置
                        n = re.search(r'[0-9a-zA-Z\u4e00-\u9fa5]', q[linelength:])
                        if n != None:
                            n = n.span()[0] + linelength
                        else:
                            break
                        res += q[:n] + '\n'
                        q = q[n:]
                    res += q + '\n'
                DataFrame.loc[index,'译文']= res.rstrip('\n')
            self.ProgramData[name]=DataFrame
    # 核对原文译文中，文本出现次数，若不同，单独导出。只导出原文至少出现一次的。需要一个名为checkdict.json的检查字典，格式为
    #{"要检查的原文":"对应的译文"}
    def checknum(self):
        count = 0
        # 加载检查字典，从长倒短排序
        try:
            with open('checkdict.json', 'r', encoding='utf8') as f:
                tempdict = json.load(f)
            # 按从长到短排序
            sortedkey = sorted(tempdict.keys(), key=lambda x: len(x), reverse=True)
            checkdict = {}
            for i in sortedkey:
                checkdict[i] = tempdict[i]
        except Exception as e:
            print(e)
            input('没有找到格式正确的checkdict文件，请确认路径设置是否正确以及文件是否符合json格式')
        res = {}
        fixdict = {}
        print('处理中，请稍候,根据checkdict的长度，可能会花费较长时间')
        for CheckUntrs in checkdict.keys():
            tempdict={}
            CheckTrsed=checkdict[CheckUntrs]
            for name in self.ProgramData.keys():
                DataFrame=self.ProgramData[name]
                for untrs in DataFrame.index:
                    trsed=DataFrame.loc[untrs,'译文']
                    if untrs.count(CheckUntrs)!=trsed.count(CheckTrsed) and untrs.count(CheckUntrs)>0:
                        tempdict.update({untrs:trsed})
            if len(tempdict):
                res.update({CheckUntrs: tempdict})
                fixdict.update({CheckUntrs: [CheckTrsed]})
                count += len(tempdict)
        out = json.dumps(res, indent=4, ensure_ascii=False)
        with open('checkres.json', 'w', encoding='utf8') as f1:
            print(out, file=f1)
        print('已将检查结果保存到checkres.json')
        # 询问是否覆盖fixdict.json
        overw = True
        if os.path.exists('fixdict.json'):
            inp = 0
            overw = False
            while inp not in ['y', 'n']:
                inp = input('fixdict.json已存在，请问要覆盖吗（y,n)\n').lower()
            if inp == 'y': overw = True
        if overw:
            self.tojson(fixdict, 'fixdict.json')
            print('对应的修正词典正确翻译导出为fixdict.json\n'
                  '修正词典的格式为：\n'
                  '{\n'
                  '"原文1":["正确译文","错误译文1","错误译文2"],\n'
                  '"原文2":["正确译文","错误译文1","错误译文2"]\n'
                  '}')
        print("########################核对完毕########################")
    # 根据核对结果，对翻译结果进行校正,只对核对结果做替换
    def fixnum(self):
        try:
            with open(r'checkres.json', 'r', encoding='utf8') as f:
                checkres = json.load(f)
        except Exception as e:
            print(e)
            input('没有找到格式正确的checkdict.json文件，请确认文件是否存在且符合json格式')
        try:
            with open(r'fixdict.json', 'r', encoding='utf8') as f:
                tempdict = json.load(f)
            sortedkey = sorted(tempdict.keys(), key=lambda x: len(x), reverse=True)
            fixdict = {}
            for i in sortedkey:
                fixdict[i] = tempdict[i]
        except Exception as e:
            print(e)
            input('没有找到格式正确的fixdict文件，请确认路径设置是否正确以及文件是否符合json格式')
        print('处理中，请稍候')
        TrsData={}  # 保存处理完毕的键值对
        for fixkey in list(fixdict.keys()):
            if fixkey in checkres.keys():
                righttrs = fixdict[fixkey][0]
                for untrs in checkres[fixkey].keys():
                    if fixkey in untrs:
                        for i in fixdict[fixkey][1:]:
                            # 如果相应文本之前已经替换过，则对TrsData内的操作
                            if untrs in TrsData.keys():
                                TrsData[untrs]=TrsData[untrs].replace(i, righttrs)
                            else:
                                checkres[fixkey][untrs] = checkres[fixkey][untrs].replace(i, righttrs)
                    if untrs not in TrsData.keys():
                        TrsData.update({untrs:checkres[fixkey][untrs]})
        self.InputFromJson(trsdata=TrsData)
        print('已应用修正结果')
    # 处理note中<SG説明:这类不能被翻译的文本，将其还原。特征是地址为含note，且包含数量相等的<和:
    # 这类文本被翻译，通常会导致图鉴等不显示
    def DNoteB(self):
        print('正在处理可能存在的note问题')
        res=self.search('note',2)
        for name in res.keys():
            DataFrame=res[name]
            for untrs in DataFrame.index:
                trsed=DataFrame.loc[untrs,'译文']
                l=['<',':'] # 分隔符
                if untrs.count(l[0])==untrs.count(l[1]) and untrs.count(l[0])==trsed.count(l[0]) and\
                    untrs.count(l[1])==trsed.count(l[1]):
                    # 将文本按照分隔符拆分
                    untrs_list = self.__splitbychar(untrs, l)
                    trsed_list = self.__splitbychar(trsed, l)
                    length = len(trsed_list)
                    # 如果按某分隔符拆分结果长度不相等，不处理
                    if len(untrs_list) != length:
                        continue
                    elif length:
                        # 被分隔符包裹的字符串只会出现在奇数位
                        for i in range(0, int(length / 2)):
                            trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                        # 将处理后的文本拼接好，等待下一循环
                        trsed = ''
                        for i in range(0, length): trsed += trsed_list[i]
                DataFrame.loc[untrs,'译文']=trsed
            # 将处理后文本导入到工程数据
            self.InputFromDataFrame(DataFrame,[name])
        print('########################note处理完毕########################')
#########################################################一键处理#########################################################
    # 读取游戏并打好预设标签，获取人名，保存数据，参数为游戏目录和保存路径，保存路径需是已存在的文件夹
    def FromGame(self,GameDir,path):
        self.ReadGame(GameDir)
        self.GetName(self.NameWithout)
        self.Output(path)
        self.Save(path)
    # 注入游戏，自动处理文件名问题，如有水印(如有），打水印，参数为游戏根目录,翻译数据路径和注入翻译后的json文件保存目录，mark为水印
    def ToGmae(self,GameDir,path,OutputPath,mark:str=False):
        self.InputFromeXlsx(path)
        self.dnb(GameDir)
        self.DNoteB()
        if mark:
            self.AddMark(mark)
        self.InjectGame(GameDir,OutputPath)
    # 游戏版本更新，path是旧版翻译文件的路径
    def Updata(self,GameDir,path,savepath):
        self.ReadGame(GameDir)
        self.InputFromeXlsx(path)
        self.GetName(self.NameWithout)
        self.Output(savepath)
        self.Save(savepath)