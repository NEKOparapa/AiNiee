 # ═══════════════════════════════════════════════════════
# ████ WARNING: Enter at Your Own Risk!               ████
# ████ Congratulations, you have stumbled upon my     ████
# ████ masterpiece - a mountain of 10,000 lines of    ████
# ████ spaghetti code. Proceed with caution,          ████
# ████ as reading this code may result in             ████
# ████ immediate unhappiness and despair.             ████
# ═══════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════
# ████ 警告：擅自进入，后果自负                         ████
# ████ 恭喜你，你已经发现了我的杰作                     ████
# ████ 一座万行意大利面条式代码的屎山                   ████
# ████ 请谨慎前行，阅读这段代码可能会。                 ████
# ████ 立刻让你感到不幸和绝望                          ████
# ═══════════════════════════════════════════════════════


# coding:utf-8               
import copy
import datetime
import json
import math
import random
import re
from qframelesswindow import FramelessWindow, TitleBar
import time
import threading
import os
import sys
import multiprocessing
import concurrent.futures

import tiktoken_ext  #必须导入这两个库，否则打包后无法运行
from tiktoken_ext import openai_public

import tiktoken #需要安装库pip install tiktoken
import openpyxl  #需安装库pip install openpyxl
from openpyxl import Workbook  
import numpy as np   #需要安装库pip install numpy
import opencc       #需要安装库pip install opencc      
from openai import OpenAI #需要安装库pip install openai
import google.generativeai as genai #需要安装库pip install -U google-generativeai

from PyQt5.QtGui import QBrush, QColor, QDesktopServices, QFont, QIcon, QImage, QPainter, QPixmap#需要安装库 pip3 install PyQt5
from PyQt5.QtCore import  QObject,  QRect,  QUrl,  Qt, pyqtSignal 
from PyQt5.QtWidgets import QAbstractItemView,QHeaderView,QApplication, QTableWidgetItem, QFrame, QGridLayout, QGroupBox, QLabel,QFileDialog, QStackedWidget, QHBoxLayout, QVBoxLayout, QWidget

from qfluentwidgets.components import Dialog
from qfluentwidgets import ProgressRing, SegmentedWidget, TableWidget,CheckBox, DoubleSpinBox, HyperlinkButton,InfoBar, InfoBarPosition, NavigationWidget, Slider, SpinBox, ComboBox, LineEdit, PrimaryPushButton, PushButton ,StateToolTip, SwitchButton, TextEdit, Theme,  setTheme ,isDarkTheme,qrouter,NavigationInterface,NavigationItemPosition
from qfluentwidgets import FluentIcon as FIF


Software_Version = "AiNiee-chatgpt4.61"  #软件版本号
cache_list = [] # 全局缓存数据
Running_status = 0  # 存储程序工作的状态，0是空闲状态,1是接口测试状态
                    # 6是翻译任务进行状态，7是错行检查状态


# 定义线程锁
lock1 = threading.Lock()  #这个用来锁缓存文件
lock2 = threading.Lock()  #这个用来锁UI信号的
lock3 = threading.Lock()  #这个用来锁自动备份缓存文件

# 工作目录改为python源代码所在的目录
script_dir = os.path.dirname(os.path.abspath(sys.argv[0])) # 获取当前工作目录
print("[INFO] 当前工作目录是:",script_dir,'\n') 
# 设置资源文件夹路径
resource_dir = os.path.join(script_dir, "resource")



# 翻译器
class Translator():
    def __init__(self):
        pass

    def Main(self):
        global cache_list, Running_status

        # ——————————————————————————————————————————配置信息初始化—————————————————————————————————————————
        configurator.initialize_configuration()
        request_limiter.initialize_limiter()

        # ——————————————————————————————————————————读取原文到缓存—————————————————————————————————————————
        # 读取文件
        Input_Folder = configurator.Input_Folder
        if configurator.translation_project == "Mtool导出文件":
            cache_list = File_Reader.read_mtool_files(self,folder_path = Input_Folder)
        elif configurator.translation_project == "T++导出文件":
            cache_list = File_Reader.read_xlsx_files (self,folder_path = Input_Folder)
        elif configurator.translation_project == "Ainiee缓存文件":
            cache_list = File_Reader.read_cache_files(self,folder_path = Input_Folder)


        # 将浮点型，整数型文本内容变成字符型文本内容
        Cache_Manager.convert_source_text_to_str(self,cache_list)

        # 如果翻译日语或者韩语文本时，则去除非中日韩文本
        Text_Source_Language =  Window.Widget_translation_settings.A_settings.comboBox_source_text.currentText() 
        if Text_Source_Language == "日语" or Text_Source_Language == "韩语":
            Cache_Manager.process_dictionary_list(self,cache_list)


        # ——————————————————————————————————————————构建并发任务池子—————————————————————————————————————————

        # 计算并发任务数
        line_count_configuration = configurator.text_line_counts # 获取每次翻译行数配置
        total_text_line_count = Cache_Manager.count_translation_status_0(self, cache_list)

        if total_text_line_count % line_count_configuration == 0:
            tasks_Num = total_text_line_count // line_count_configuration 
        else:
            tasks_Num = total_text_line_count // line_count_configuration + 1



        # 更新界面UI信息，并输出各种配置信息
        project_id = cache_list[0]["project_id"]
        user_interface_prompter.signal.emit("初始化翻译界面数据",project_id,total_text_line_count,0,0) #需要输入够当初设定的参数个数
        user_interface_prompter.signal.emit("翻译状态提示","开始翻译",0,0,0)
        print("[INFO]  翻译项目为",configurator.translation_project, '\n')
        print("[INFO]  翻译平台为",configurator.translation_platform, '\n')
        print("[INFO]  AI模型为",configurator.model_type, '\n')
        if configurator.translation_platform == "Openai代理" or  configurator.translation_platform == "SakuraLLM":
            print("[INFO]  请求地址为",configurator.openai_base_url, '\n')
        elif configurator.translation_platform == "Openai官方":
            print("[INFO]  账号类型为",Window.Widget_Openai.comboBox_account_type.currentText(), '\n')
        print("[INFO]  游戏文本从",configurator.source_language, '翻译到', configurator.target_language,'\n')
        if configurator.translation_platform != "SakuraLLM":
            print("[INFO]  当前设定的系统提示词为：", configurator.get_system_prompt(), '\n')
            original_exmaple,translation_example =  configurator.get_default_translation_example()
            print("[INFO]  已添加默认原文示例",original_exmaple, '\n')
            print("[INFO]  已添加默认译文示例",translation_example, '\n')
        print("[INFO]  文本总行数为：",total_text_line_count,"  每次发送行数为：",line_count_configuration,"  计划的翻译任务总数是：", tasks_Num) 
        print("\033[1;32m[INFO] \033[0m 五秒后开始进行翻译，请注意保持网络通畅，余额充足。", '\n')
        time.sleep(5)  

        # 测试用，会导致任务多一个，注意下
        #api_requester_instance = Api_Requester()
        #api_requester_instance.Concurrent_Request_Openai()



        # 创建线程池
        The_Max_workers = configurator.thread_counts # 获取线程数配置
        with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
            # 创建实例
            api_requester_instance = Api_Requester()
            # 向线程池提交任务
            for i in range(tasks_Num):
                # 根据不同平台调用不同接口
                if configurator.translation_platform == "Openai官方" or configurator.translation_platform == "Openai代理":
                    executor.submit(api_requester_instance.Concurrent_Request_Openai)
                    
                elif configurator.translation_platform == "Google官方":
                    executor.submit(api_requester_instance.Concurrent_Request_Google)

                elif configurator.translation_platform == "SakuraLLM":
                    executor.submit(api_requester_instance.Concurrent_Request_Sakura)

            # 等待线程池任务完成
            executor.shutdown(wait=True)





        # 检查主窗口是否已经退出
        if Running_status == 10 :
            return
    
        # 检查翻译任务是否已经暂停
        if Running_status == 1011 :
            pass

        # ——————————————————————————————————————————检查没能成功翻译的文本，拆分翻译————————————————————————————————————————

        #计算未翻译文本的数量
        untranslated_text_line_count = Cache_Manager.count_and_update_translation_status_0_2(self,cache_list)

        #重新翻译次数限制
        retry_translation_count = 1

        while untranslated_text_line_count != 0 :
            print("\033[1;33mWarning:\033[0m 仍然有部分未翻译，将进行拆分后重新翻译，-----------------------------------")
            print("[INFO] 当前重新翻译次数：",retry_translation_count ," 到达最大次数：10 时，将停止翻译")


            #根据算法计算拆分的文本行数
            line_count_configuration = configurator.update_text_line_count(line_count_configuration)
            print("[INFO] 未翻译文本总行数为：",untranslated_text_line_count,"  每次发送行数修改为：",line_count_configuration, '\n')

            #如果实时调教功能没有开的话，则每次重新翻译，增加OpenAI的随机性
            if configurator.translation_platform == "Openai官方" or configurator.translation_platform == "Openai代理":
                if (Window.Interface18.checkBox.isChecked() == False) and (retry_translation_count != 1) :
                    if configurator.openai_temperature + 0.1 <= 1.0 :
                        configurator.openai_temperature = configurator.openai_temperature + 0.1
                    else:
                        configurator.openai_temperature = 1.0
                    print("\033[1;33mWarning:\033[0m 当前AI模型的随机度设置为：",configurator.openai_temperature)


            # 计算可并发任务总数
            if untranslated_text_line_count % line_count_configuration == 0:
                tasks_Num = untranslated_text_line_count // line_count_configuration
            else:
                tasks_Num = untranslated_text_line_count // line_count_configuration + 1


            # 创建线程池
            The_Max_workers = configurator.thread_counts # 获取线程数配置
            with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
                # 创建实例
                api_requester_instance = Api_Requester()
                # 向线程池提交任务
                for i in range(tasks_Num):
                    # 根据不同平台调用不同接口
                    if configurator.translation_platform == "Openai官方" or configurator.translation_platform == "Openai代理":
                        executor.submit(api_requester_instance.Concurrent_Request_Openai)
                        
                    elif configurator.translation_platform == "Google官方":
                        executor.submit(api_requester_instance.Concurrent_Request_Google)

                    elif configurator.translation_platform == "SakuraLLM":
                        executor.submit(api_requester_instance.Concurrent_Request_Sakura)

                # 等待线程池任务完成
                executor.shutdown(wait=True)

            

            #检查主窗口是否已经退出
            if Running_status == 10 :
                return

            #检查是否已经达到重翻次数限制
            retry_translation_count  = retry_translation_count + 1
            if retry_translation_count >= 10 :
                print ("\033[1;33mWarning:\033[0m 已经达到重新翻译次数限制，但仍然有部分文本未翻译，不影响使用，可手动翻译", '\n')
                break

            #重新计算未翻译文本的数量
            untranslated_text_line_count = Cache_Manager.count_and_update_translation_status_0_2(self,cache_list)


        # ——————————————————————————————————————————将数据处理并保存为文件—————————————————————————————————————————
        print ("\033[1;32mSuccess:\033[0m  翻译阶段已完成，正在处理数据-----------------------------------", '\n')


        #如果开启了转换简繁开关功能，则进行文本转换
        if configurator.conversion_toggle: 
            if configurator.target_language == "简中" or configurator.target_language == "繁中":
                try:
                    cache_list = File_Outputter.simplified_and_traditional_conversion(self,cache_list, configurator.target_language)
                    print(f"\033[1;32mSuccess:\033[0m  文本转化{configurator.target_language}完成-----------------------------------", '\n')   

                except Exception as e:
                    print("\033[1;33mWarning:\033[0m 文本转换出现问题！！将跳过该步，错误信息如下")
                    print(f"Error: {e}\n")

        # 将翻译结果写为文件
        output_path = configurator.Output_Folder

        if configurator.translation_project == "Mtool导出文件":
            File_Outputter.output_json_file(self,cache_list, output_path)

        elif configurator.translation_project == "T++导出文件":
            File_Outputter.output_excel_file(self,cache_list, output_path)

        elif configurator.translation_project == "Ainiee缓存文件":
            if cache_list[0]["project_type"] == "Mtool":
                File_Outputter.output_json_file(self,cache_list, output_path)
            else:
                File_Outputter.output_excel_file(self,cache_list, output_path)

        print("\033[1;32mSuccess:\033[0m  译文文件写入完成-----------------------------------", '\n')  


        # —————————————————————————————————————#全部翻译完成——————————————————————————————————————————
        user_interface_prompter.signal.emit("翻译状态提示","翻译完成",0,0,0)
        print("\n--------------------------------------------------------------------------------------")
        print("\n\033[1;32mSuccess:\033[0m 已完成全部翻译任务，程序已经停止")   
        print("\n\033[1;32mSuccess:\033[0m 请检查译文文件，格式是否错误，存在错行，或者有空行等问题")
        print("\n-------------------------------------------------------------------------------------\n")


    def Check_main(self):
        global cache_list, Running_status
        # ——————————————————————————————————————————配置信息初始化—————————————————————————————————————————
        configurator.initialize_configuration_check()

        request_limiter.initialize_limiter_check()

        # ——————————————————————————————————————————读取原文到缓存—————————————————————————————————————————
        # 读取文件
        Input_Folder = configurator.Input_Folder
        if configurator.translation_project == "Mtool导出文件":
            cache_list = File_Reader.read_mtool_files(self,folder_path = Input_Folder)
        elif configurator.translation_project == "T++导出文件":
            cache_list = File_Reader.read_xlsx_files (self,folder_path = Input_Folder)

            
        # —————————————————————————————————————处理读取的文件——————————————————————————————————————————

        # 将浮点型，整数型文本内容变成字符型文本内容
        Cache_Manager.convert_source_text_to_str(self,cache_list)

        # 统计已翻译文本的tokens总量，并根据不同项目修改翻译状态
        tokens_consume_all = Cache_Manager.count_tokens(self, cache_list)

        # —————————————————————————————————————创建并发嵌入任务——————————————————————————————————————————

        #根据tokens_all_consume与除以6090计算出需要请求的次数,并向上取整（除以6090是为了富余任务数）
        tasks_Num = int(math.ceil(tokens_consume_all / 7000))

        print("[DEBUG] 全部文本需要嵌入请求的次数是",tasks_Num)


        # 初始化一下界面提示器里面存储的相关变量
        user_interface_prompter.translated_line_count = 0
        user_interface_prompter.total_text_line_count =  Cache_Manager.count_translation_status_0(self, cache_list)

        #测试用
        #api_requester_instance = Api_Requester()
        #api_requester_instance.Concurrent_request_Embeddings()

        # 创建线程池
        The_Max_workers =  multiprocessing.cpu_count() * 4 + 1  
        with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
            # 创建实例
            api_requester_instance = Api_Requester()
            # 向线程池提交任务
            for i in range(tasks_Num):
                # 根据不同平台调用不同接口
                executor.submit(api_requester_instance.Concurrent_request_Embeddings)

            # 等待线程池任务完成
            executor.shutdown(wait=True)

        #检查主窗口是否已经退出
        if Running_status == 10 :
            return

        print("\033[1;32mSuccess:\033[0m  全部文本检查编码完成-------------------------------------")
        # —————————————————————————————————————开始检查，并整理需要重新翻译的文本——————————————————————————————————————————

        #创建存储原文与译文的列表，方便复制粘贴，这里是两个空字符串，后面会被替换
        sentences = ["", ""]  

        misaligned_text = {}     #存储错行文本的字典

        #创建存储每对翻译相似度计算过程日志的字符串
        similarity_log = ""
        log_count = 0
        count_error = 0  #错误文本计数变量


        # 把等于3的翻译状态改为0
        for item in cache_list:
            if item.get('translation_status') == 3:
                item['translation_status'] = 0

         # 统计翻译状态为0的文本数
        List_len = Cache_Manager.count_translation_status_0(self, cache_list)





        for entry in cache_list:
            translation_status = entry.get('translation_status')

            if translation_status == 0:

                #将sentence[0]与sentence[1]转换成字符串数据，确保能够被语义相似度检查模型识别，防止数字型数据导致报错
                sentences[0] = str(entry["source_text"])
                sentences[1] = str(entry["translated_text"])

                #输出sentence里的两个文本 和 语义相似度检查结果
                print("[INFO] 原文是：", sentences[0])
                print("[INFO] 译文是：", sentences[1])


                #计算语义相似度----------------------------------------
                Semantic_similarity =entry["semantic_similarity"]
                print("[INFO] 语义相似度：", Semantic_similarity, "%")


                #计算符号相似度----------------------------------------
                # 用正则表达式匹配原文与译文中的标点符号
                k_syms = re.findall(r'[。！？…♡♥=★♪]', sentences[0])
                v_syms = re.findall(r'[。！？…♡♥=★♪]', sentences[1])

                #假如v_syms与k_syms都不为空
                if len(v_syms) != 0 and len(k_syms) != 0:
                    #计算v_syms中的符号在k_syms中存在相同符号数量，再除以v_syms的符号总数，得到相似度
                    Symbolic_similarity = len([sym for sym in v_syms if sym in k_syms]) / len(v_syms) * 100
                #假如v_syms与k_syms都为空，即原文和译文都没有标点符号
                elif len(v_syms) == 0 and len(k_syms) == 0:
                    Symbolic_similarity = 1 * 100
                else:
                    Symbolic_similarity = 0

                print("[INFO] 符号相似度：", Symbolic_similarity, "%")


                #计算字数相似度----------------------------------------
                # 计算k中的日文、中文,韩文，英文字母的个数
                Q, W, E, R = Response_Parser.count_japanese_chinese_korean(self,sentences[0])
                # 计算v中的日文、中文,韩文，英文字母的个数
                A, S, D, F = Response_Parser.count_japanese_chinese_korean(self,sentences[1])
                


                # 计算每个总字数
                len1 = Q + W + E + R
                len2 = A + S + D + F

                #设定基准字数差距，暂时靠经验设定
                if len1  <= 25:
                    Base_word_count = 15
                else:
                    Base_word_count = 25

                #计算字数差值
                Word_count_difference = abs((len1 - len2) )
                if Word_count_difference > Base_word_count:
                    Word_count_difference = Base_word_count
            
                # 计算字数相差程度
                Word_count_similarity =(1- Word_count_difference / Base_word_count) * 100
                print("[INFO] 字数相似度：", Word_count_similarity, "%")



                
                #获取设定的权重
                Semantic_weight = Window.Widget_check.doubleSpinBox_semantic_weight.value()
                Symbolic_weight = Window.Widget_check.doubleSpinBox_symbol_weight.value()
                Word_count_weight = Window.Widget_check.doubleSpinBox_word_count_weight.value()
                similarity_threshold = Window.Widget_check.spinBox_similarity_threshold.value()

                #计算总相似度
                similarity = Semantic_similarity * Semantic_weight + Symbolic_similarity * Symbolic_weight + Word_count_similarity * Word_count_weight
                #输出各权重值
                print("[INFO] 语义权重：", Semantic_weight,"符号权重：", Symbolic_weight,"字数权重：", Word_count_weight)

                #如果语义相似度小于于等于阈值，需要重翻译
                if similarity <= similarity_threshold:
                    count_error = count_error + 1
                    print("[INFO] 总相似度结果：", similarity, "%，小于相似度阈值", similarity_threshold,"%，需要重翻译")
                    #错误文本计数提醒
                    print("\033[1;33mWarning:\033[0m 当前错误文本数量：", count_error)
                    #将错误文本存储到字典里
                    misaligned_text[sentences[0]] = sentences[1]

                # 检查通过,改变翻译状态为不需要翻译
                else :
                    entry['translation_status'] = 1
                    print("[INFO] 总相似度结果：", similarity, "%", "，不需要重翻译")
                    

                #创建格式化字符串，用于存储每对翻译相似度计算过程日志
                if log_count <=  10000 :#如果log_count小于等于10000,避免太大
                    similarity_log = similarity_log + "\n" + "原文是：" + sentences[0] + "\n" + "译文是：" + sentences[1] + "\n" + "语义相似度：" + str(Semantic_similarity) + "%" + "\n" + "符号相似度：" + str(Symbolic_similarity) + "%" + "\n" + "字数相似度：" + str(Word_count_similarity) + "%" + "\n" + "总相似度结果：" + str(similarity) + "%" + "\n" + "语义权重：" + str(Semantic_weight) + "，符号权重：" + str(Symbolic_weight) + "，字数权重：" + str(Word_count_weight) + "\n" + "当前检查进度：" + str(round((log_count+1)/List_len*100,2)) + "%" + "\n"
                    log_count = log_count + 1

                #输出遍历进度，转换成百分百进度
                print("[INFO] 当前检查进度：", round((log_count)/List_len*100,2), "% \n")




        # 构建输出检查结果路径
        output_path = configurator.Output_Folder
        folder_path = os.path.join(output_path, "misalignment_check_result")
        os.makedirs(folder_path, exist_ok=True)


        #检查完毕，将错误文本字典写入json文件
        with open(os.path.join(folder_path, "misaligned_text.json"), 'w', encoding='utf-8') as f:
            json.dump(misaligned_text, f, ensure_ascii=False, indent=4)
        
        #将每对翻译相似度计算过程日志写入txt文件
        with open(os.path.join(folder_path, "log.txt"), 'w', encoding='utf-8') as f:
            f.write(similarity_log)

    # ——————————————————————————————————————————配置信息初始化—————————————————————————————————————————
        configurator.initialize_configuration()
        request_limiter.initialize_limiter()


        # 初始化一下界面提示器里面存储的相关变量
        user_interface_prompter.translated_line_count = 0
        user_interface_prompter.total_text_line_count =  Cache_Manager.count_translation_status_0(self, cache_list)

    # —————————————————————————————————————开始重新翻译——————————————————————————————————————————

        #记录循环翻译次数
        Number_of_iterations = 0

        #计算需要翻译文本的数量
        count_not_Translate = Cache_Manager.count_translation_status_0(self, cache_list)

        while count_not_Translate != 0 :

            # 计算可并发任务总数
            if count_not_Translate % 1 == 0:
                tasks_Num = count_not_Translate // 1       
            else:
                tasks_Num = count_not_Translate // 1 + 1  

            # 创建线程池
            The_Max_workers = configurator.thread_counts # 获取线程数配置
            with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
                # 创建实例
                api_requester_instance = Api_Requester()
                # 向线程池提交任务
                for i in range(tasks_Num):
                    # 根据不同平台调用不同接口
                    executor.submit(api_requester_instance.Concurrent_Request_Openai)

                # 等待线程池任务完成
                executor.shutdown(wait=True)


            #检查主窗口是否已经退出
            if Running_status == 10 :
                return
                

            #重新计算未翻译文本的数量
            count_not_Translate = Cache_Manager.count_and_update_translation_status_0_2(self, cache_list)

            #记录循环次数
            Number_of_iterations = Number_of_iterations + 1
            print("\033[1;33mWarning:\033[0m 当前循环翻译次数：", Number_of_iterations, "次，到达最大循环次数5次后将退出翻译任务")
            #检查是否已经陷入死循环
            if Number_of_iterations == 5 :
                print("\033[1;33mWarning:\033[0m 已达到最大循环次数，退出重翻任务，不影响后续使用-----------------------------------")
                break


        print("\n\033[1;32mSuccess:\033[0m  已重新翻译完成-----------------------------------")




        # —————————————————————————————————————写入文件——————————————————————————————————————————
        # 将翻译结果写为文件
        output_path = configurator.Output_Folder

        File_Outputter.output_translated_content(self,cache_list, output_path)



        # —————————————————————————————————————全部翻译完成——————————————————————————————————————————
        print("\n--------------------------------------------------------------------------------------")
        print("\n\033[1;32mSuccess:\033[0m 已完成全部翻译任务，程序已经停止")   
        print("\n\033[1;32mSuccess:\033[0m 请检查译文文件，格式是否错误，存在错行，或者有空行等问题")
        print("\n-------------------------------------------------------------------------------------\n")



# 接口请求器
class Api_Requester():
    def __init__(self):
        pass
    
    # 整理发送内容（Openai）
    def organize_send_content_openai(self,source_text_dict):
        #创建message列表，用于发送
        messages = []

        #构建系统提示词
        prompt = configurator.get_system_prompt()
        system_prompt ={"role": "system","content": prompt }
        #print("[INFO] 当前系统提示词为", prompt,'\n')
        messages.append(system_prompt)

        #构建原文与译文示例
        original_exmaple,translation_example =  configurator.get_default_translation_example()
        the_original_exmaple =  {"role": "user","content":original_exmaple }
        the_translation_example = {"role": "assistant", "content":translation_example }
        #print("[INFO]  已添加默认原文示例",original_exmaple)
        #print("[INFO]  已添加默认译文示例",translation_example)

        messages.append(the_original_exmaple)
        messages.append(the_translation_example)
 


        #如果开启了译时提示字典功能，则添加新的原文与译文示例
        if Window.Interface23.checkBox2.isChecked() :
            original_exmaple_2,translation_example_2 = configurator.build_prompt_dictionary(source_text_dict)
            if original_exmaple_2 and translation_example_2:
                the_original_exmaple =  {"role": "user","content":original_exmaple_2 }
                the_translation_example = {"role": "assistant", "content":translation_example_2 }
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  检查到请求的原文中含有用户字典内容，已添加新的原文与译文示例")
                print("[INFO]  已添加提示字典原文示例",original_exmaple_2)
                print("[INFO]  已添加提示字典译文示例",translation_example_2)

        #如果提示词工程界面的用户翻译示例开关打开，则添加新的原文与译文示例
        if Window.Interface22.checkBox2.isChecked() :
            original_exmaple_3,translation_example_3 = configurator.build_user_translation_example ()
            if original_exmaple_3 and translation_example_3:
                the_original_exmaple =  {"role": "user","content":original_exmaple_3 }
                the_translation_example = {"role": "assistant", "content":translation_example_3 }
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  检查到用户翻译示例开关打开，已添加新的原文与译文示例")
                print("[INFO]  已添加用户原文示例",original_exmaple_3)
                print("[INFO]  已添加用户译文示例",translation_example_3)


        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")


        #如果开启译前替换字典功能，则根据用户字典进行替换
        if Window.Interface21.checkBox1.isChecked() :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_strings_dictionary(source_text_dict)

        #将原文本字典转换成JSON格式的字符串，方便发送
        source_text_str = json.dumps(source_text_dict, ensure_ascii=False)    

        #构建需要翻译的文本
        Original_text = {"role":"user","content":source_text_str}   
        messages.append(Original_text)

        return messages,source_text_str


    # 并发接口请求（Openai）
    def Concurrent_Request_Openai(self):
        global cache_list,Running_status



        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            rows = configurator.text_line_counts
            source_text_list = Cache_Manager.process_dictionary_data(self,rows, cache_list)    
            lock1.release()  # 释放锁

            # ——————————————————————————————————————————转换原文本的格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str = Api_Requester.organize_send_content_openai(self,source_text_dict)



            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            #计算请求的tokens预计花费
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages) 
            #计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str }] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text)
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;31mError:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;31mError:\033[0m 该条消息取消任务，进行拆分翻译" )
                return


            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 850  # 设置超时时间为x秒
            request_errors_count = 0 # 设置请求错误次数限制
            Wrong_answer_count = 0   # 设置错误回复次数限制
            model_degradation = False # 模型退化检测

            while 1 :
                #检查主窗口是否已经退出---------------------------------
                if Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume )
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 获取AI的参数设置
                    temperature,top_p,presence_penalty,frequency_penalty= configurator.get_model_parameters()
                    # 如果上一次请求出现模型退化，更改参数
                    if model_degradation:
                        frequency_penalty = 0.2

                    # 获取apikey
                    openai_apikey =  configurator.get_apikey()
                    # 获取请求地址
                    openai_base_url = configurator.openai_base_url
                    # 创建openai客户端
                    openaiclient = OpenAI(api_key=openai_apikey,
                                            base_url= openai_base_url)
                    # 发送对话请求
                    try:
                        #如果开启了回复josn格式的功能和可以开启该功能的模型
                        if (configurator.response_json_format_toggle) and (configurator.model_type == "gpt-3.5-turbo-1106" or configurator.model_type == "gpt-4-1106-preview"):
                            print("[INFO] 已开启强制回复josn格式功能")
                            response = openaiclient.chat.completions.create(
                                model= configurator.model_type,
                                messages = messages ,
                                temperature=temperature,
                                top_p = top_p,                        
                                presence_penalty=presence_penalty,
                                frequency_penalty=frequency_penalty,
                                response_format={"type": "json_object"}
                                )
                        else:
                            response = openaiclient.chat.completions.create(
                                model= configurator.model_type,
                                messages = messages ,
                                temperature=temperature,
                                top_p = top_p,                        
                                presence_penalty=presence_penalty,
                                frequency_penalty=frequency_penalty
                                )

                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 6 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    #——————————————————————————————————————————收到回复，并截取回复内容中的文本内容 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    try: # 因为有些中转网站不返回tokens消耗
                        prompt_tokens_used = int(response.usage.prompt_tokens) #本次请求花费的tokens
                    except Exception as e:
                        prompt_tokens_used = 0
                    try:
                        completion_tokens_used = int(response.usage.completion_tokens) #本次回复花费的tokens
                    except Exception as e:
                        completion_tokens_used = 0



                    # 提取回复的文本内容
                    response_content = response.choices[0].message.content 


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                # ——————————————————————————————————————————对AI回复内容进行各种处理和检查——————————————————————————————————————————
                    # 处理回复内容
                    response_content = Response_Parser.adjust_string(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,source_text_dict)

                    # 如果没有出现错误
                    if check_result :
                        # 转化为字典格式
                        response_dict = json.loads(response_content) #注意转化为字典的数字序号key是字符串类型

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if Window.Widget_start_translation.B_settings.checkBox_switch.isChecked():
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁

                        
                        lock2.acquire()  # 获取锁

                        # 如果是进行平时的翻译任务
                        if Running_status == 6 :
                            # 计算进度信息
                            progress = (user_interface_prompter.translated_line_count+row_count) / user_interface_prompter.total_text_line_count * 100
                            progress = round(progress, 1)

                            # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                            user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",row_count,prompt_tokens_used,completion_tokens_used)
                        
                        # 如果进行的是错行检查任务，使用不同的计算方法
                        elif Running_status == 7 :
                            user_interface_prompter.translated_line_count = user_interface_prompter.translated_line_count + row_count
                            progress = user_interface_prompter.translated_line_count / user_interface_prompter.total_text_line_count * 100
                            progress = round(progress, 1)

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")
                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:

                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁
                        # 如果是进行平时的翻译任务
                        if Running_status == 6 :
                            user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",row_count,prompt_tokens_used,completion_tokens_used)
                        lock2.release()  # 释放锁
                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")
                        # 检查一下是不是模型退化
                        if error_content == "AI回复内容出现高频词,并重新翻译":
                            print("\033[1;33mWarning:\033[0m 下次请求将修改参数，回避高频词输出","\n")
                            model_degradation = True

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m AI回复内容格式错误次数:",Wrong_answer_count,"到达2次后将该段文本进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count >= 2 :
                            print("\033[1;33mWarning:\033[0m 错误次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环
                        time.sleep(3)                 
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return



    # 整理发送内容（Google）
    def organize_send_content_google(self,source_text_dict):
        #创建message列表，用于发送
        messages = []

        #获取系统提示词
        prompt = configurator.get_system_prompt()


        #获取原文与译文示例
        original_exmaple,translation_example =  configurator.get_default_translation_example()

        # 构建系统提示词与默认示例
        messages.append({'role':'user','parts':prompt +"\n###\n" + original_exmaple})
        messages.append({'role':'model','parts':translation_example})

 

        #如果开启了译时提示字典功能，则添加新的原文与译文示例
        if Window.Interface23.checkBox2.isChecked() :
            original_exmaple_2,translation_example_2 = configurator.build_prompt_dictionary(source_text_dict)
            if original_exmaple_2 and translation_example_2:
                the_original_exmaple =  {"role": "user","parts":original_exmaple_2 }
                the_translation_example = {"role": "model", "parts":translation_example_2 }
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  检查到请求的原文中含有用户字典内容，已添加新的原文与译文示例")
                print("[INFO]  已添加提示字典原文示例",original_exmaple_2)
                print("[INFO]  已添加提示字典译文示例",translation_example_2)

        #如果提示词工程界面的用户翻译示例开关打开，则添加新的原文与译文示例
        if Window.Interface22.checkBox2.isChecked() :
            original_exmaple_3,translation_example_3 = configurator.build_user_translation_example ()
            if original_exmaple_3 and translation_example_3:
                the_original_exmaple =  {"role": "user","parts":original_exmaple_3 }
                the_translation_example = {"role": "model", "parts":translation_example_3 }
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  检查到用户翻译示例开关打开，已添加新的原文与译文示例")
                print("[INFO]  已添加用户原文示例",original_exmaple_3)
                print("[INFO]  已添加用户译文示例",translation_example_3)


        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")

        #如果开启译前替换字典功能，则根据用户字典进行替换
        if Window.Interface21.checkBox1.isChecked() :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_strings_dictionary(source_text_dict)


        #将原文本字典转换成JSON格式的字符串，方便发送
        source_text_str = json.dumps(source_text_dict, ensure_ascii=False)   

        #构建需要翻译的文本
        Original_text = {"role":"user","parts":source_text_str}   
        messages.append(Original_text)

        return messages,source_text_str


    # 并发接口请求（Google）
    def Concurrent_Request_Google(self):
        global cache_list,Running_status

        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            rows = configurator.text_line_counts
            source_text_list = Cache_Manager.process_dictionary_data(self,rows, cache_list)    
            lock1.release()  # 释放锁

            # ——————————————————————————————————————————转换原文本的格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str = Api_Requester.organize_send_content_google(self,source_text_dict)



            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            #计算请求的tokens预计花费
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages) 
            #计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str}] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text)
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;33mWarning:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;33mWarning:\033[0m 该条消息取消任务，进行拆分翻译" )
                return


            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 850  # 设置超时时间为x秒
            request_errors_count = 0 # 设置请求错误次数限制
            Wrong_answer_count = 0   # 设置错误回复次数限制

            while 1 :
                #检查主窗口是否已经退出---------------------------------
                if Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume ) 
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 设置AI的参数
                    generation_config = {
                    "temperature": 0,
                    "top_p": 1,
                    "top_k": 1,
                    "max_output_tokens": 2048, #最大输出，pro最大输出是2048
                    }

                    #调整安全限制
                    safety_settings = [
                    {
                        "category": "HARM_CATEGORY_HARASSMENT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_HATE_SPEECH",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                        "threshold": "BLOCK_NONE"
                    }
                    ]

                    # 获取apikey
                    apikey =  configurator.get_apikey()
                    genai.configure(api_key=apikey)

                    #设置对话模型及参数
                    model = genai.GenerativeModel(model_name=configurator.model_type,
                                    generation_config=generation_config,
                                    safety_settings=safety_settings)


                    # 发送对话请求
                    try:
                        response = model.generate_content(messages)

                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 6 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    #——————————————————————————————————————————收到回复，并截取回复内容中的文本内容 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    prompt_tokens_used = int(request_tokens_consume)
                    completion_tokens_used = int(completion_tokens_consume)



                    # 提取回复的文本内容
                    try:
                        response_content = response.text
                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 提取文本时出现错误！！！运行的错误信息如下")
                        print(f"Error: {e}\n")
                        print("接口返回的错误信息如下")
                        print(response.prompt_feedback)
                        #处理完毕，再次进行请求
                        continue


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                # ——————————————————————————————————————————对AI回复内容进行各种处理和检查——————————————————————————————————————————
                    # 处理回复内容
                    response_content = Response_Parser.adjust_string(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,source_text_dict)

                    # 如果没有出现错误
                    if check_result :
                        # 转化为字典格式
                        response_dict = json.loads(response_content) #注意转化为字典的数字序号key是字符串类型

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if Window.Widget_start_translation.B_settings.checkBox_switch.isChecked():
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁


                        
                        lock2.acquire()  # 获取锁
                        # 计算进度信息
                        progress = (user_interface_prompter.translated_line_count+row_count) / user_interface_prompter.total_text_line_count * 100
                        progress = round(progress, 1)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",row_count,prompt_tokens_used,completion_tokens_used)
                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")

                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:
                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁
                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",row_count,prompt_tokens_used,completion_tokens_used)
                        lock2.release()  # 释放锁
                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m AI回复内容格式错误次数:",Wrong_answer_count,"到达2次后将该段文本进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count >= 2 :
                            print("\033[1;33mWarning:\033[0m 错误次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环
                        time.sleep(1)                 
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return


    # 并发嵌入请求
    def Concurrent_request_Embeddings(self):
        global cache_list,Running_status

        try:#方便排查子线程bug
            # ——————————————————————————————————————————提取需要嵌入的翻译对——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            accumulated_tokens, source_texts, translated_texts,text_index_list = Cache_Manager.process_tokens(cache_list, 7500)
            lock1.release()  # 释放锁

            # 计算一下文本长度
            text_len = len(source_texts)

            #检查一下返回值是否为空，如果为空则表示已经嵌入完了
            if accumulated_tokens == 0 or text_len == 0:
                return
            
            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————
            #构建发送文本列表，长度为end - start的两倍，前半部分为原文，后半部分为译文
            input_txt = []
            for i in range(text_len):
                input_txt.append(source_texts[i])
            for i in range(text_len):
                input_txt.append(translated_texts[i])


        
            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            while 1 :
                #检查主窗口是否已经退出---------------------------------
                if Running_status == 10 :
                    return

                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(accumulated_tokens):

                    #————————————————————————————————————————发送请求————————————————————————————————————————
                    # 获取apikey
                    openai_apikey =  configurator.get_apikey()
                    # 获取请求地址
                    openai_base_url = configurator.openai_base_url
                    # 创建openai客户端
                    openaiclient = OpenAI(api_key=openai_apikey,
                                            base_url= openai_base_url)
                    try:
                        print("[INFO] 已发送文本嵌入请求-------------------------------------")
                        print("[INFO] 请求内容长度是：",len(input_txt))
                        print("[INFO] 已发送请求，请求内容是：",input_txt,'\n','\n')
                        response = openaiclient.embeddings.create(
                            input=input_txt,
                            model="text-embedding-ada-002")
                        
            
                    except Exception as e:
                        print("\033[1;33m线程ID:\033[0m ", threading.get_ident())
                        print("\033[1;31mError:\033[0m api请求出现问题！错误信息如下")
                        print(f"Error: {e}\n")

                        #等待五秒再次请求
                        print("\033[1;33m线程ID:\033[0m 该任务五秒后再次请求")
                        time.sleep(5)

                        
                        continue #处理完毕，再次进行请求

                    #————————————————————————————————————————处理回复————————————————————————————————————————

                    print("[INFO] 已收到回复--------------------------------------")
                    print("[INFO] 正在计算语义相似度并录入缓存中")

                    # 计算相似度
                    Semantic_similarity_list = []
                    for i in range(text_len):
                        #计算获取原文编码的索引位置，并获取
                        Original_Index = i
                        #openai返回的嵌入值是存储在data列表的字典元素里，在字典元素里以embedding为关键字，所以才要改变data的索引值
                        Original_Embeddings = response.data[Original_Index].embedding

                        #计算获取译文编码的索引位置，并获取
                        Translation_Index = i  + text_len
                        #openai返回的嵌入值是存储在data列表的字典元素里，在字典元素里以embedding为关键字，所以才要改变data的索引值
                        Translation_Embeddings = response.data[Translation_Index].embedding

                        #计算每对翻译语义相似度
                        similarity_score = np.dot(Original_Embeddings, Translation_Embeddings)
                        Semantic_similarity_list.append((similarity_score - 0.75) / (1 - 0.75) * 150)

                    lock1.acquire()  # 获取锁
                    user_interface_prompter.translated_line_count = user_interface_prompter.translated_line_count + text_len
                    progress = user_interface_prompter.translated_line_count / user_interface_prompter.total_text_line_count * 100
                    progress = round(progress, 1)
                    Cache_Manager.update_vector_distance(cache_list, text_index_list, Semantic_similarity_list)
                    print("[INFO] 已计算语义相似度并存储",'\n','\n')
                    lock1.release()  # 释放锁

                    #————————————————————————————————————————结束循环，并结束子线程————————————————————————————————————————
                    print(f"\n--------------------------------------------------------------------------------------")
                    print(f"\n\033[1;32mSuccess:\033[0m 嵌入编码已完成：{progress}%             ")
                    print(f"\n--------------------------------------------------------------------------------------\n")
                    break

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;33m线程ID:\033[0m ", threading.get_ident())
            print("\033[1;31mError:\033[0m 线程出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return



    # 整理发送内容（sakura）
    def organize_send_content_Sakura(self,source_text_dict):
        #创建message列表，用于发送
        messages = []

        #构建系统提示词
        system_prompt ={"role": "system","content": "你是一个轻小说翻译模型，可以流畅通顺地以日本轻小说的风格将日文翻译成简体中文，并联系上下文正确使用人称代词，不擅自添加原文中没有的代词。" }
        #print("[INFO] 当前系统提示词为", prompt,'\n')
        messages.append(system_prompt)


        # 0.8模型不支持下面功能
        if configurator.model_type != 'Sakura-13B-LNovel-v0.8':

            prompt = "将下面的日文文本翻译成中文："

            #构建原文与译文示例
            original_exmaple,translation_example =  configurator.get_default_translation_example()
            the_original_exmaple =  {"role": "user","content":prompt + original_exmaple }
            the_translation_example = {"role": "assistant", "content":translation_example }
            #print("[INFO]  已添加默认原文示例",original_exmaple)
            #print("[INFO]  已添加默认译文示例",translation_example)

            #messages.append(the_original_exmaple)
            #messages.append(the_translation_example)
    


            #如果开启了译时提示字典功能，则添加新的原文与译文示例
            if Window.Interface23.checkBox2.isChecked() :
                original_exmaple_2,translation_example_2 = configurator.build_prompt_dictionary(source_text_dict)
                if original_exmaple_2 and translation_example_2:
                    the_original_exmaple =  {"role": "user","content":prompt + original_exmaple_2 }
                    the_translation_example = {"role": "assistant", "content":translation_example_2 }
                    messages.append(the_original_exmaple)
                    messages.append(the_translation_example)
                    print("[INFO]  检查到请求的原文中含有用户字典内容，已添加新的原文与译文示例")
                    print("[INFO]  已添加提示字典原文示例",original_exmaple_2)
                    print("[INFO]  已添加提示字典译文示例",translation_example_2)

            #如果提示词工程界面的用户翻译示例开关打开，则添加新的原文与译文示例
            if Window.Interface22.checkBox2.isChecked() :
                original_exmaple_3,translation_example_3 = configurator.build_user_translation_example ()
                if original_exmaple_3 and translation_example_3:
                    the_original_exmaple =  {"role": "user","content":prompt + original_exmaple_3 }
                    the_translation_example = {"role": "assistant", "content":translation_example_3 }
                    messages.append(the_original_exmaple)
                    messages.append(the_translation_example)
                    print("[INFO]  检查到用户翻译示例开关打开，已添加新的原文与译文示例")
                    print("[INFO]  已添加用户原文示例",original_exmaple_3)
                    print("[INFO]  已添加用户译文示例",translation_example_3)



        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")

        #如果开启译前替换字典功能，则根据用户字典进行替换
        if Window.Interface21.checkBox1.isChecked() :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_strings_dictionary(source_text_dict)


 
        #将原文本字典转换成raw格式的字符串，方便发送   
        source_text_str_raw = self.convert_dict_to_raw_str(source_text_dict)

        # 处理全角数字
        source_text_str_raw = self.convert_fullwidth_to_halfwidth(source_text_str_raw)

        #构建需要翻译的文本
        prompt = "将下面的日文文本翻译成中文："
        Original_text = {"role":"user","content":prompt + source_text_str_raw}   
        messages.append(Original_text)



        return messages, source_text_str_raw


    # 并发接口请求（sakura）
    def Concurrent_Request_Sakura(self):
        global cache_list,Running_status



        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            rows = configurator.text_line_counts
            source_text_list = Cache_Manager.process_dictionary_data(self,rows, cache_list)    
            lock1.release()  # 释放锁

            # ——————————————————————————————————————————转换原文本的格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str = Api_Requester.organize_send_content_Sakura(self,source_text_dict)



            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            #计算请求的tokens预计花费
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages)  #加上2%的修正系数
            #计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str}] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text) #加上2%的修正系数
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;31mError:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;31mError:\033[0m 该条消息取消任务，进行拆分翻译" )
                return


            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 850  # 设置超时时间为x秒
            request_errors_count = 0 # 设置请求错误次数限制
            Wrong_answer_count = 0   # 设置错误回复次数限制
            model_degradation = False # 模型退化检测

            while 1 :
                #检查主窗口是否已经退出---------------------------------
                if Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume )
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 获取AI的参数设置
                    temperature,top_p,presence_penalty,frequency_penalty= configurator.get_model_parameters()
                    # 如果上一次请求出现模型退化，更改参数
                    if model_degradation:
                        frequency_penalty = 0.2

                    # 获取apikey
                    openai_apikey =  configurator.get_apikey()
                    # 获取请求地址
                    openai_base_url = configurator.openai_base_url
                    # 创建openai客户端
                    openaiclient = OpenAI(api_key=openai_apikey,
                                            base_url= openai_base_url)
                    # 发送对话请求
                    try:
                        response = openaiclient.chat.completions.create(
                            model= configurator.model_type,
                            messages = messages ,
                            temperature=temperature,
                            top_p = top_p,                        
                            presence_penalty=presence_penalty,
                            frequency_penalty=frequency_penalty
                            )

                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 6 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    #——————————————————————————————————————————收到回复，并截取回复内容中的文本内容 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    try: # 因为有些中转网站不返回tokens消耗
                        prompt_tokens_used = int(response.usage.prompt_tokens) #本次请求花费的tokens
                    except Exception as e:
                        prompt_tokens_used = 0
                    try:
                        completion_tokens_used = int(response.usage.completion_tokens) #本次回复花费的tokens
                    except Exception as e:
                        completion_tokens_used = 0



                    # 提取回复的文本内容
                    response_content = response.choices[0].message.content 


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                # ——————————————————————————————————————————对AI回复内容进行各种处理和检查——————————————————————————————————————————
                    # 处理回复内容
                    response_content = Response_Parser.convert_str_to_json_str(self, row_count, response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,source_text_dict)

                    # 如果没有出现错误
                    if check_result :
                        # 转化为字典格式
                        response_dict = json.loads(response_content) #注意转化为字典的数字序号key是字符串类型

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if Window.Widget_start_translation.B_settings.checkBox_switch.isChecked():
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁

                        
                        lock2.acquire()  # 获取锁

                        # 如果是进行平时的翻译任务
                        if Running_status == 6 :
                            # 计算进度信息
                            progress = (user_interface_prompter.translated_line_count+row_count) / user_interface_prompter.total_text_line_count * 100
                            progress = round(progress, 1)

                            # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                            user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",row_count,prompt_tokens_used,completion_tokens_used)
                        
                        # 如果进行的是错行检查任务，使用不同的计算方法
                        elif Running_status == 7 :
                            user_interface_prompter.translated_line_count = user_interface_prompter.translated_line_count + row_count
                            progress = user_interface_prompter.translated_line_count / user_interface_prompter.total_text_line_count * 100
                            progress = round(progress, 1)

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")
                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:

                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁
                        # 如果是进行平时的翻译任务
                        if Running_status == 6 :
                            user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",row_count,prompt_tokens_used,completion_tokens_used)
                        lock2.release()  # 释放锁
                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")
                        # 检查一下是不是模型退化
                        if error_content == "AI回复内容出现高频词,并重新翻译":
                            print("\033[1;33mWarning:\033[0m 下次请求将修改参数，回避高频词输出","\n")
                            model_degradation = True

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m AI回复内容格式错误次数:",Wrong_answer_count,"到达2次后将该段文本进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count >= 2 :
                            print("\033[1;33mWarning:\033[0m 错误次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环
                        time.sleep(3)                 
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return


    # 将json文本改为纯文本
    def convert_dict_to_raw_str(self,source_text_dict):
        str_list = []
        for idx in range(len(source_text_dict.keys())):
            # str_list.append(s['source_text'])
            str_list.append(source_text_dict[f"{idx}"])
        raw_str = "\n".join(str_list)
        return raw_str


    # 将列表中的字符串中的全角数字转换为半角数字
    def convert_fullwidth_to_halfwidth(self,input_string):
        modified_string = ""
        for char in input_string:
            if '０' <= char <= '９':  # 判断是否为全角数字
                modified_string += chr(ord(char) - ord('０') + ord('0'))  # 转换为半角数字
            else:
                modified_string += char

        return modified_string



# 回复解析器
class Response_Parser():
    def __init__(self):
        pass


    # 处理回复，前后加上大括号
    def adjust_string(self,input_str):
        # 检查并添加开头的"{"
        if not input_str.startswith("{"):
            input_str = "{" + input_str

        # 检查并添加结尾的"}"
        if not input_str.endswith("}"):
            input_str = input_str + "}"

        return input_str


    # 将Raw文本恢复根据行数转换成json文本
    def convert_str_to_json_str(self,row_count, input_str):

        # 当发送文本为1行时，就不分割了，以免切错
        if row_count == 1:
            result = {"0": input_str}
            return  json.dumps(result, ensure_ascii=False)
        
        else:
            str_list = input_str.split("\n")
            ret_json = {}
            for idx, text in enumerate(str_list):
                ret_json[f"{idx}"] = f"{text}"
            return json.dumps(ret_json, ensure_ascii=False)



    # 检查回复内容是否存在问题
    def check_response_content(self,response_str,source_text_dict):
        # 存储检查结果
        check_result = False
        # 存储错误内容
        error_content = "0"


        # 检查模型是否退化，出现高频词（只检测中日）
        if Response_Parser.model_degradation_detection(self,response_str):
            pass

        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容出现高频词,并重新翻译"
            return check_result,error_content


        # 检查文本格式
        if Response_Parser.check_response_format(self,response_str):
            # 回复文本转换成字典格式
            response_dict = json.loads(response_str)

        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容不符合要求的格式,将进行重新翻译"
            return check_result,error_content


        # 检查文本行数
        if Response_Parser.check_text_line_count(self,source_text_dict,response_dict):
            pass
        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容文本行数与原来数量不符合,将进行重新翻译"
            return check_result,error_content


        # 检查文本空行
        if Response_Parser.check_empty_response(self,response_dict):
            pass
        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容中有未进行翻译的空行,将进行重新翻译"
            return check_result,error_content

        # 如果检查都没有问题
        check_result = True
        # 存储错误内容
        error_content = "检查无误"
        return check_result,error_content



    # 检查回复内容的json格式
    def check_response_format(self,response_str):
        try:
            response_dict = json.loads(response_str) #注意转化为字典的数字序号key是字符串类型
            return True
        except :                                            
            return False
        
    # 检查回复内容的文本行数
    def check_text_line_count(self,source_text_dict,response_dict):
        if(len(source_text_dict)  ==  len(response_dict) ):    
            return True
        else:                                            
            return False
        
    # 检查翻译内容是否有空值
    def check_empty_response(self,response_dict):
        for value in response_dict.values():
            #检查value是不是None，因为AI回回复null，但是json.loads()会把null转化为None
            if value is None:
                return False

            # 检查value是不是空字符串，因为AI回回复空字符串，但是json.loads()会把空字符串转化为""
            if value == "":
                return False

            return True
        
    # 模型退化检测，高频语气词
    def model_degradation_detection(self,input_string):
        # 使用正则表达式匹配中日语字符
        japanese_chars = re.findall(r'[\u3040-\u309F\u30A0-\u30FF\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAFF]', input_string)

        # 统计中日语字符的数量
        char_count = {}
        for char in japanese_chars:
            char_count[char] = char_count.get(char, 0) + 1

        # 输出字符数量
        for char, count in char_count.items():
            if count >= 90:
                return False
                #print(f"中日语字符 '{char}' 出现了 {count} 次一次。")
        
        return True


        
    #计算字符串里面日文与中文，韩文,英文字母（不是单词）的数量
    def count_japanese_chinese_korean(self,text):
        japanese_pattern = re.compile(r'[\u3040-\u30FF\u31F0-\u31FF\uFF65-\uFF9F]') # 匹配日文字符
        chinese_pattern = re.compile(r'[\u4E00-\u9FFF]') # 匹配中文字符
        korean_pattern = re.compile(r'[\uAC00-\uD7AF\u1100-\u11FF\u3130-\u318F\uA960-\uA97F\uD7B0-\uD7FF]') # 匹配韩文字符
        english_pattern = re.compile(r'[A-Za-z\uFF21-\uFF3A\uFF41-\uFF5A]') # 匹配半角和全角英文字母
        japanese_count = len(japanese_pattern.findall(text)) # 统计日文字符数量
        chinese_count = len(chinese_pattern.findall(text)) # 统计中文字符数量
        korean_count = len(korean_pattern.findall(text)) # 统计韩文字符数量
        english_count = len(english_pattern.findall(text)) # 统计英文字母数量
        return japanese_count, chinese_count, korean_count , english_count




# 接口测试器
class Request_Tester():
    def __init__(self):
        pass

    # openai官方接口测试
    def openai_request_test(self):
        Account_Type = Window.Widget_Openai.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
        Model_Type =  Window.Widget_Openai.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Widget_Openai.TextEdit_apikey.toPlainText()        #获取apikey输入值
        Proxy_port = Window.Widget_Openai.LineEdit_proxy_port.text()            #获取代理端口
        
        #如果填入地址，则设置系统代理
        if Proxy_port :
            print("[INFO] 系统代理端口是:",Proxy_port,'\n') 
            os.environ["http_proxy"]=Proxy_port
            os.environ["https_proxy"]=Proxy_port


        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = API_key_str.replace('\n','').replace(" ", "").split(",")


        #创建openai客户端
        openaiclient = OpenAI(api_key=API_key_list[0],
                base_url= 'https://api.openai.com/v1')


        print("[INFO] 账号类型是:",Account_Type,'\n')
        print("[INFO] 模型选择是:",Model_Type,'\n')

        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            openaiclient.api_key = API_key_list[i]

            #构建发送内容
            messages_test = [{"role": "system","content":"你是我的女朋友欣雨。接下来你必须以女朋友的方式回复我"}, {"role":"user","content":"小可爱，你在干嘛"}]
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')

            #尝试请求，并设置各种参数
            try:
                response_test = openaiclient.chat.completions.create( 
                model= Model_Type,
                messages = messages_test ,
                ) 

                #如果回复成功，显示成功信息
                response_test = response_test.choices[0].message.content
                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0,0,0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0,0,0)

    # openai代理接口测试
    def op_request_test(self):
        
        Base_url = Window.Widget_Openai_Proxy.A_settings.LineEdit_relay_address.text()                  #获取请求地址
        Model_Type =  Window.Widget_Openai_Proxy.A_settings.comboBox_model.currentText()                #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Widget_Openai_Proxy.A_settings.TextEdit_apikey.toPlainText()               #获取apikey输入值
        Proxy_port  = Window.Widget_Openai_Proxy.A_settings.LineEdit_proxy_port.text()                  #获取代理端口

        
        #如果填入地址，则设置系统代理
        if Proxy_port :
            print("[INFO] 系统代理端口是:",Proxy_port,'\n') 
            os.environ["http_proxy"]=Proxy_port
            os.environ["https_proxy"]=Proxy_port

        
        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = API_key_str.replace('\n','').replace(" ", "").split(",")

        #检查一下请求地址尾部是否为/v1，自动补全
        if Base_url[-3:] != "/v1":
            Base_url = Base_url + "/v1"

        #创建openai客户端
        openaiclient = OpenAI(api_key=API_key_list[0],
                base_url= Base_url)


        print("[INFO] 中转请求地址是:",Base_url,'\n')
        print("[INFO] 模型选择是:",Model_Type,'\n')

        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            openaiclient.api_key = API_key_list[i]

            #构建发送内容
            messages_test = [{"role": "system","content":"你是我的女朋友欣雨。接下来你必须以女朋友的方式回复我"}, {"role":"user","content":"小可爱，你在干嘛"}]
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')

            #尝试请求，并设置各种参数
            try:
                response_test = openaiclient.chat.completions.create( 
                model= Model_Type,
                messages = messages_test ,
                ) 

                #如果回复成功，显示成功信息
                response_test = response_test.choices[0].message.content
                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0,0,0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0,0,0)


    # google官方接口测试
    def google_request_test(self):

        Model_Type =  Window.Widget_Google.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
        API_key_str = Window.Widget_Google.TextEdit_apikey.toPlainText()        #获取apikey输入值
        Proxy_port = Window.Widget_Google.LineEdit_proxy_port.text()            #获取代理端口
        
        #如果填入地址，则设置系统代理
        if Proxy_port :
            print("[INFO] 系统代理端口是:",Proxy_port,'\n') 
            os.environ["http_proxy"]=Proxy_port
            os.environ["https_proxy"]=Proxy_port


        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = API_key_str.replace('\n','').replace(" ", "").split(",")



        print("[INFO] 模型选择是:",Model_Type,'\n')

        # 设置ai参数
        generation_config = {
        "temperature": 0,
        "top_p": 1,
        "top_k": 1,
        "max_output_tokens": 2048, #最大输出，pro最大输出是2048
        }

        #调整安全限制
        safety_settings = [
        {
            "category": "HARM_CATEGORY_HARASSMENT",
            "threshold": "BLOCK_NONE"
        },
        {
            "category": "HARM_CATEGORY_HATE_SPEECH",
            "threshold": "BLOCK_NONE"
        },
        {
            "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
            "threshold": "BLOCK_NONE"
        },
        {
            "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
            "threshold": "BLOCK_NONE"
        }
        ]


        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            genai.configure(api_key= API_key_list[i]) 

            #构建发送内容
            messages_test = ["你是我的女朋友欣雨。接下来你必须以女朋友的方式向我问好",]
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')


            #设置对话模型
            model = genai.GenerativeModel(model_name=Model_Type,
                            generation_config=generation_config,
                            safety_settings=safety_settings)


            #尝试请求，并设置各种参数
            try:
                response_test =  model.generate_content(messages_test)

                #如果回复成功，显示成功信息
                response_test = response_test.text
                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0,0,0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0,0,0)


    # sakura接口测试
    def sakura_request_test(self):
        
        Base_url = Window.Widget_SakuraLLM.LineEdit_address.text()                  #获取请求地址
        Model_Type =  Window.Widget_SakuraLLM.comboBox_model.currentText()                #获取模型类型下拉框当前选中选项的值
        Proxy_port  = Window.Widget_SakuraLLM.LineEdit_proxy_port.text()                  #获取代理端口

        
        #如果填入地址，则设置系统代理
        if Proxy_port :
            print("[INFO] 系统代理端口是:",Proxy_port,'\n') 
            os.environ["http_proxy"]=Proxy_port
            os.environ["https_proxy"]=Proxy_port

        

        #检查一下请求地址尾部是否为/v1，自动补全
        if Base_url[-3:] != "/v1":
            Base_url = Base_url + "/v1"

        #创建openai客户端
        openaiclient = OpenAI(api_key="sakura",
                base_url= Base_url)


        print("[INFO] 模型地址是:",Base_url,'\n')
        print("[INFO] 模型选择是:",Model_Type,'\n')



        #构建发送内容
        messages_test = [{"role": "system","content":"你是一个轻小说翻译模型，可以流畅通顺地以日本轻小说的风格将日文翻译成简体中文，并联系上下文正确使用人称代词，不擅自添加原文中没有的代词。"},
                         {"role":"user","content":"将下面的日文文本翻译成中文：サポートキャスト"}]
        print("[INFO] 当前发送内容：\n", messages_test ,'\n')

        #尝试请求，并设置各种参数
        try:
            response_test = openaiclient.chat.completions.create( 
            model= Model_Type,
            messages = messages_test ,
            ) 

            #如果回复成功，显示成功信息
            response_test = response_test.choices[0].message.content
            print("[INFO] 已成功接受到AI的回复")
            print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

            print("[INFO] 模型通讯测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0,0,0)

        #如果回复失败，抛出错误信息，并测试下一个key
        except Exception as e:
            print("\033[1;31mError:\033[0m 请求出现问题！错误信息如下")
            print(f"Error: {e}\n\n")
            print("[INFO] 模型通讯测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0,0,0)



    
# 配置器
class Configurator():
    def __init__(self):
        self.translation_project = "" # 翻译项目
        self.translation_platform = "" # 翻译平台
        self.source_language = "" # 文本原语言
        self.target_language = "" # 文本目标语言
        self.Input_Folder = "" # 存储输入文件夹
        self.Output_Folder = "" # 存储输出文件夹

        self.text_line_counts = 1 # 存储每次请求的文本行数设置
        self.thread_counts = 1 # 存储线程数
        self.preserve_line_breaks_toggle = False # 保留换行符开关
        self.response_json_format_toggle = False # 回复json格式开关

        self.model_type = ""             #模型选择
        self.apikey_list = [] # 存储key的列表
        self.key_index = 0  # 方便轮询key的索引


        self.openai_base_url = 'https://api.openai.com/v1' # api默认请求地址
        self.openai_temperature = 0.1        #AI的随机度，0.8是高随机，0.2是低随机,取值范围0-2
        self.openai_top_p = 1.0              #AI的top_p，作用与temperature相同，官方建议不要同时修改
        self.openai_presence_penalty = 0.0  #AI的存在惩罚，生成新词前检查旧词是否存在相同的词。0.0是不惩罚，2.0是最大惩罚，-2.0是最大奖励
        self.openai_frequency_penalty = 0.0 #AI的频率惩罚，限制词语重复出现的频率。0.0是不惩罚，2.0是最大惩罚，-2.0是最大奖励




    # 初始化配置信息
    def initialize_configuration (self):
        global Running_status


        # 获取第一页的配置信息
        self.translation_project = Window.Widget_translation_settings.A_settings.comboBox_translation_project.currentText()
        self.translation_platform = Window.Widget_translation_settings.A_settings.comboBox_translation_platform.currentText()
        self.source_language = Window.Widget_translation_settings.A_settings.comboBox_source_text.currentText()
        self.target_language = Window.Widget_translation_settings.A_settings.comboBox_translated_text.currentText()
        self.Input_Folder = Window.Widget_translation_settings.A_settings.label_input_path.text() # 存储输入文件夹
        self.Output_Folder = Window.Widget_translation_settings.A_settings.label_output_path.text() # 存储输出文件夹


        # 获取文本行数设置
        self.text_line_counts = Window.Widget_translation_settings.B_settings.spinBox_Lines.value()
        # 获取线程数设置  
        self.thread_counts = Window.Widget_translation_settings.B_settings.spinBox_thread_count.value()
        if self.thread_counts == 0:                                
            self.thread_counts = multiprocessing.cpu_count() * 4 + 1  
        # 获取保留换行符开关
        self.preserve_line_breaks_toggle =  Window.Widget_translation_settings.B_settings.SwitchButton_line_breaks.isChecked()
        # 获取回复json格式开关
        self.response_json_format_toggle =  Window.Widget_translation_settings.B_settings.SwitchButton_jsonmode.isChecked()
        # 获取简繁转换开关柜
        self.conversion_toggle = Window.Widget_translation_settings.B_settings.SwitchButton_conversion_toggle.isChecked()


        # 重新初始化模型参数，防止上次任务的设置影响到
        self.openai_temperature = 0.1        
        self.openai_top_p = 1.0             
        self.openai_presence_penalty = 0.0  
        self.openai_frequency_penalty = 0.0 


        # 如果进行的是错行检查任务，修改部分设置(补丁)
        if Running_status == 7:
            self.translation_project = Window.Widget_check.comboBox_translation_project.currentText()
            self.translation_platform = Window.Widget_check.comboBox_translation_platform.currentText()
            self.Input_Folder = Window.Widget_check.label_input_path.text() # 存储输入文件夹
            self.Output_Folder = Window.Widget_check.label_output_path.text() # 存储输出文件夹
            # 修改翻译行数为1
            self.text_line_counts = 1
            # 修改源语言与目标语言
            self.source_language = "日语"
            self.target_language = "简中"


        #根据翻译平台读取配置信息
        if self.translation_platform == 'Openai官方':
            # 获取模型类型
            self.model_type =  Window.Widget_Openai.comboBox_model.currentText()              

            # 获取apikey列表
            API_key_str = Window.Widget_Openai.TextEdit_apikey.toPlainText()            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.openai_base_url = 'https://api.openai.com/v1'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = Window.Widget_Openai.LineEdit_proxy_port.text()            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        elif self.translation_platform == 'Openai代理':
            # 获取模型类型
            self.model_type =  Window.Widget_Openai_Proxy.A_settings.comboBox_model.currentText()     

            # 获取apikey列表
            API_key_str = Window.Widget_Openai_Proxy.A_settings.TextEdit_apikey.toPlainText()            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取中转请求地址
            relay_address = Window.Widget_Openai_Proxy.A_settings.LineEdit_relay_address.text()
            #检查一下请求地址尾部是否为/v1，自动补全
            if relay_address[-3:] != "/v1":
                relay_address = relay_address + "/v1"
            self.openai_base_url = relay_address  

            #如果填入地址，则设置代理端口
            Proxy_Address = Window.Widget_Openai_Proxy.A_settings.LineEdit_proxy_port.text()            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address

        elif self.translation_platform == 'Google官方':
            # 获取模型类型
            self.model_type =  Window.Widget_Google.comboBox_model.currentText()              

            # 获取apikey列表
            API_key_str = Window.Widget_Google.TextEdit_apikey.toPlainText()            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list


            #如果填入地址，则设置代理端口
            Proxy_Address = Window.Widget_Google.LineEdit_proxy_port.text()            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        elif self.translation_platform == 'SakuraLLM':
            # 获取模型类型
            self.model_type =  Window.Widget_SakuraLLM.comboBox_model.currentText()     
            # 构建假apikey
            self.apikey_list = ["sakura"]

            # 获取中转请求地址
            relay_address = Window.Widget_SakuraLLM.LineEdit_address.text()   
            #检查一下请求地址尾部是否为/v1，自动补全
            if relay_address[-3:] != "/v1":
                relay_address = relay_address + "/v1"
            self.openai_base_url = relay_address  

            #如果填入地址，则设置代理端口
            Proxy_Address = Window.Widget_SakuraLLM.LineEdit_proxy_port.text()              #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address

            #更改部分参数，以适合Sakura模型
            self.openai_temperature = 0.1       
            self.openai_top_p = 0.3
            self.thread_counts = 1 # 线程数
            #self.text_line_counts = 1 # 文本行数


    # 初始化配置信息
    def initialize_configuration_check (self):
        # 获取配置信息
        self.translation_project = Window.Widget_check.comboBox_translation_project.currentText()
        self.translation_platform = Window.Widget_check.comboBox_translation_platform.currentText()
        self.Input_Folder = Window.Widget_check.label_input_path.text() # 存储输入文件夹
        self.Output_Folder = Window.Widget_check.label_output_path.text() # 存储输出文件夹

        # 获取文本行数设置
        self.text_line_counts = 1
        # 获取线程数设置  
        self.thread_counts = Window.Widget_translation_settings.B_settings.spinBox_thread_count.value()
        if self.thread_counts == 0:                                
            self.thread_counts = multiprocessing.cpu_count() * 4 + 1  


        # 初始化模型参数
        self.openai_temperature = 0        
        self.openai_top_p = 1.0             
        self.openai_presence_penalty = 0.5  
        self.openai_frequency_penalty = 0.0 



        #根据翻译平台读取配置信息
        if self.translation_platform == 'Openai官方':
            # 获取模型类型
            self.model_type =  Window.Widget_Openai.comboBox_model.currentText()              

            # 获取apikey列表
            API_key_str = Window.Widget_Openai.TextEdit_apikey.toPlainText()            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.openai_base_url = 'https://api.openai.com/v1'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = Window.Widget_Openai.LineEdit_proxy_port.text()            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        elif self.translation_platform == 'Openai代理':
            # 获取模型类型
            self.model_type =  Window.Widget_Openai_Proxy.A_settings.comboBox_model.currentText()     

            # 获取apikey列表
            API_key_str = Window.Widget_Openai_Proxy.A_settings.TextEdit_apikey.toPlainText()            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取中转请求地址
            relay_address = Window.Widget_Openai_Proxy.A_settings.LineEdit_relay_address.text()
            #检查一下请求地址尾部是否为/v1，自动补全
            if relay_address[-3:] != "/v1":
                relay_address = relay_address + "/v1"
            self.openai_base_url = relay_address  

            #如果填入地址，则设置代理端口
            Proxy_Address = Window.Widget_Openai_Proxy.A_settings.LineEdit_proxy_port.text()            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


    #读写配置文件config.json函数
    def read_write_config(self,mode):

        if mode == "write":
            # 存储配置信息的字典
            config_dict = {}
            
            #获取openai官方账号界面
            config_dict["openai_account_type"] = Window.Widget_Openai.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["openai_model_type"] =  Window.Widget_Openai.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["openai_API_key_str"] = Window.Widget_Openai.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["openai_proxy_port"] = Window.Widget_Openai.LineEdit_proxy_port.text()            #获取代理端口
            
            #获取openai代理账号基础设置界面
            config_dict["op_relay_address"] = Window.Widget_Openai_Proxy.A_settings.LineEdit_relay_address.text()                  #获取请求地址
            config_dict["op_model_type"] =  Window.Widget_Openai_Proxy.A_settings.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["op_API_key_str"] = Window.Widget_Openai_Proxy.A_settings.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["op_proxy_port"]  = Window.Widget_Openai_Proxy.A_settings.LineEdit_proxy_port.text()               #获取代理端口


            #获取openai代理账号进阶设置界面
            config_dict["op_rpm_limit"] = Window.Widget_Openai_Proxy.B_settings.spinBox_RPM.value()               #获取rpm限制值
            config_dict["op_tpm_limit"] = Window.Widget_Openai_Proxy.B_settings.spinBox_TPM.value()               #获取tpm限制值
            config_dict["op_input_pricing"] = Window.Widget_Openai_Proxy.B_settings.spinBox_input_pricing.value()               #获取输入价格
            config_dict["op_output_pricing"] = Window.Widget_Openai_Proxy.B_settings.spinBox_output_pricing.value()               #获取输出价格

            #Google官方账号界面
            config_dict["google_model_type"] =  Window.Widget_Google.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["google_API_key_str"] = Window.Widget_Google.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["google_proxy_port"] = Window.Widget_Google.LineEdit_proxy_port.text()            #获取代理端口


            #Sakura界面
            config_dict["sakura_address"] = Window.Widget_SakuraLLM.LineEdit_address.text()                  #获取请求地址
            config_dict["sakura_model_type"] =  Window.Widget_SakuraLLM.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["sakura_proxy_port"] = Window.Widget_SakuraLLM.LineEdit_proxy_port.text()            #获取代理端口



            #翻译设置基础设置界面
            config_dict["translation_project"] = Window.Widget_translation_settings.A_settings.comboBox_translation_project.currentText()
            config_dict["translation_platform"] = Window.Widget_translation_settings.A_settings.comboBox_translation_platform.currentText()
            config_dict["source_language"] = Window.Widget_translation_settings.A_settings.comboBox_source_text.currentText()
            config_dict["target_language"] = Window.Widget_translation_settings.A_settings.comboBox_translated_text.currentText()
            config_dict["label_input_path"] = Window.Widget_translation_settings.A_settings.label_input_path.text()
            config_dict["label_output_path"] = Window.Widget_translation_settings.A_settings.label_output_path.text()

            #翻译设置进阶设置界面
            config_dict["text_line_counts"] = Window.Widget_translation_settings.B_settings.spinBox_Lines.value()     # 获取文本行数设置
            config_dict["thread_counts"] = Window.Widget_translation_settings.B_settings.spinBox_thread_count.value() # 获取线程数设置  
            config_dict["preserve_line_breaks_toggle"] =  Window.Widget_translation_settings.B_settings.SwitchButton_line_breaks.isChecked() # 获取保留换行符开关
            config_dict["response_json_format_toggle"] =  Window.Widget_translation_settings.B_settings.SwitchButton_jsonmode.isChecked()   # 获取回复json格式开关
            config_dict["response_conversion_toggle"] =  Window.Widget_translation_settings.B_settings.SwitchButton_conversion_toggle.isChecked()   # 获取简繁转换开关

            #开始翻译的备份设置界面
            config_dict["auto_backup_toggle"] =  Window.Widget_start_translation.B_settings.checkBox_switch.isChecked() # 获取备份设置开关


            #错行检查界面
            config_dict["semantic_weight"] = Window.Widget_check.doubleSpinBox_semantic_weight.value() 
            config_dict["symbol_weight"] = Window.Widget_check.doubleSpinBox_symbol_weight.value() 
            config_dict["word_count_weight"] = Window.Widget_check.doubleSpinBox_word_count_weight.value() 
            config_dict["similarity_threshold"] = Window.Widget_check.spinBox_similarity_threshold.value() 
            config_dict["translation_project_check"] = Window.Widget_check.comboBox_translation_project.currentText()
            config_dict["translation_platform_check"] = Window.Widget_check.comboBox_translation_platform.currentText()
            config_dict["label_input_path_check"] = Window.Widget_check.label_input_path.text()
            config_dict["label_output_path_check"] = Window.Widget_check.label_output_path.text()



            #获取替换字典界面
            config_dict["Replace_before_translation"] =  Window.Interface21.checkBox1.isChecked()#获取译前替换开关状态
            User_Dictionary1 = {}
            for row in range(Window.Interface21.tableView.rowCount() - 1):
                key_item = Window.Interface21.tableView.item(row, 0)
                value_item = Window.Interface21.tableView.item(row, 1)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    User_Dictionary1[key] = value
            config_dict["User_Dictionary1"] = User_Dictionary1


            #获取提示字典界面
            config_dict["Change_translation_prompt"] = Window.Interface23.checkBox2.isChecked() #获取译时提示开关状态
            User_Dictionary2 = {}
            for row in range(Window.Interface23.tableView.rowCount() - 1):
                key_item = Window.Interface23.tableView.item(row, 0)
                value_item = Window.Interface23.tableView.item(row, 1)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    User_Dictionary2[key] = value
            config_dict["User_Dictionary2"] = User_Dictionary2



            #获取实时设置界面
            config_dict["OpenAI_Temperature"] = Window.Interface18.slider1.value()           #获取OpenAI温度
            config_dict["OpenAI_top_p"] = Window.Interface18.slider2.value()                 #获取OpenAI top_p
            config_dict["OpenAI_presence_penalty"] = Window.Interface18.slider3.value()      #获取OpenAI top_k
            config_dict["OpenAI_frequency_penalty"] = Window.Interface18.slider4.value()    #获取OpenAI repetition_penalty


            #获取提示词工程界面
            config_dict["Custom_Prompt_Switch"] = Window.Interface22.checkBox1.isChecked()   #获取自定义提示词开关状态
            config_dict["Custom_Prompt"] = Window.Interface22.TextEdit1.toPlainText()        #获取自定义提示词输入值 
            config_dict["Add_user_example_switch"]= Window.Interface22.checkBox2.isChecked()#获取添加用户示例开关状态
            User_example = {}
            for row in range(Window.Interface22.tableView.rowCount() - 1):
                key_item = Window.Interface22.tableView.item(row, 0)
                value_item = Window.Interface22.tableView.item(row, 1)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    User_example[key] = value
            config_dict["User_example"] = User_example


            #将所有的配置信息写入config.json文件中
            with open(os.path.join(resource_dir, "config.json"), "w", encoding="utf-8") as f:
                json.dump(config_dict, f, ensure_ascii=False, indent=4)


        if mode == "read":
            #如果config.json在子文件夹resource中存在
            if os.path.exists(os.path.join(resource_dir, "config.json")):
                #读取config.json
                with open(os.path.join(resource_dir, "config.json"), "r", encoding="utf-8") as f:
                    config_dict = json.load(f)

                #将config.json中的值赋予到变量中,并set到界面上
                #openai官方账号界面
                if "openai_account_type" in config_dict:
                    Window.Widget_Openai.comboBox_account_type.setCurrentText(config_dict["openai_account_type"])
                if "openai_model_type" in config_dict:
                    Window.Widget_Openai.comboBox_model.setCurrentText(config_dict["openai_model_type"])
                if "openai_API_key_str" in config_dict:
                    Window.Widget_Openai.TextEdit_apikey.setText(config_dict["openai_API_key_str"])
                if "openai_proxy_port" in config_dict:
                    Window.Widget_Openai.LineEdit_proxy_port.setText(config_dict["openai_proxy_port"])

                #openai代理账号基础界面
                if "op_relay_address" in config_dict:
                    Window.Widget_Openai_Proxy.A_settings.LineEdit_relay_address.setText(config_dict["op_relay_address"])
                if "op_model_type" in config_dict:
                    Window.Widget_Openai_Proxy.A_settings.comboBox_model.setCurrentText(config_dict["op_model_type"])
                if "op_API_key_str" in config_dict:
                    Window.Widget_Openai_Proxy.A_settings.TextEdit_apikey.setText(config_dict["op_API_key_str"])
                if "op_proxy_port" in config_dict:
                    Window.Widget_Openai_Proxy.A_settings.LineEdit_proxy_port.setText(config_dict["op_proxy_port"])

                #openai代理账号进阶界面
                if "op_rpm_limit" in config_dict:
                    Window.Widget_Openai_Proxy.B_settings.spinBox_RPM.setValue(config_dict["op_rpm_limit"])
                if "op_tpm_limit" in config_dict:
                    Window.Widget_Openai_Proxy.B_settings.spinBox_TPM.setValue(config_dict["op_tpm_limit"])
                if "op_input_pricing" in config_dict:
                    Window.Widget_Openai_Proxy.B_settings.spinBox_input_pricing.setValue(config_dict["op_input_pricing"])
                if "op_output_pricing" in config_dict:
                    Window.Widget_Openai_Proxy.B_settings.spinBox_output_pricing.setValue(config_dict["op_output_pricing"])


                #google官方账号界面
                if "google_model_type" in config_dict:
                    Window.Widget_Google.comboBox_model.setCurrentText(config_dict["google_model_type"])
                if "google_API_key_str" in config_dict:
                    Window.Widget_Google.TextEdit_apikey.setText(config_dict["google_API_key_str"])
                if "google_proxy_port" in config_dict:
                    Window.Widget_Google.LineEdit_proxy_port.setText(config_dict["google_proxy_port"])


                #sakura界面
                if "sakura_address" in config_dict:
                    Window.Widget_SakuraLLM.LineEdit_address.setText(config_dict["sakura_address"])
                if "sakura_model_type" in config_dict:
                    Window.Widget_SakuraLLM.comboBox_model.setCurrentText(config_dict["sakura_model_type"])
                if "sakura_proxy_port" in config_dict:
                    Window.Widget_SakuraLLM.LineEdit_proxy_port.setText(config_dict["sakura_proxy_port"])


                #翻译设置基础界面
                if "translation_project" in config_dict:
                    Window.Widget_translation_settings.A_settings.comboBox_translation_project.setCurrentText(config_dict["translation_project"])
                if "translation_platform" in config_dict:
                    Window.Widget_translation_settings.A_settings.comboBox_translation_platform.setCurrentText(config_dict["translation_platform"])
                if "source_language" in config_dict:
                    Window.Widget_translation_settings.A_settings.comboBox_source_text.setCurrentText(config_dict["source_language"])
                if "target_language" in config_dict:
                    Window.Widget_translation_settings.A_settings.comboBox_translated_text.setCurrentText(config_dict["target_language"])
                if "label_input_path" in config_dict:
                    Window.Widget_translation_settings.A_settings.label_input_path.setText(config_dict["label_input_path"])
                if "label_output_path" in config_dict:
                    Window.Widget_translation_settings.A_settings.label_output_path.setText(config_dict["label_output_path"])

                #翻译设置进阶界面
                if "text_line_counts" in config_dict:
                    Window.Widget_translation_settings.B_settings.spinBox_Lines.setValue(config_dict["text_line_counts"])
                if "thread_counts" in config_dict:
                    Window.Widget_translation_settings.B_settings.spinBox_thread_count.setValue(config_dict["thread_counts"])
                if "preserve_line_breaks_toggle" in config_dict:
                    Window.Widget_translation_settings.B_settings.SwitchButton_line_breaks.setChecked(config_dict["preserve_line_breaks_toggle"])
                if "response_json_format_toggle" in config_dict:
                    Window.Widget_translation_settings.B_settings.SwitchButton_jsonmode.setChecked(config_dict["response_json_format_toggle"])
                if "response_conversion_toggle" in config_dict:
                    Window.Widget_translation_settings.B_settings.SwitchButton_conversion_toggle.setChecked(config_dict["response_conversion_toggle"])


                #开始翻译的备份设置界面
                if "auto_backup_toggle" in config_dict:
                    Window.Widget_start_translation.B_settings.checkBox_switch.setChecked(config_dict["auto_backup_toggle"])



                #错行检查界面
                if "semantic_weight" in config_dict:
                    Window.Widget_check.doubleSpinBox_semantic_weight.setValue(config_dict["semantic_weight"])
                if "symbol_weight" in config_dict:
                    Window.Widget_check.doubleSpinBox_symbol_weight.setValue(config_dict["symbol_weight"])
                if "word_count_weight" in config_dict:
                    Window.Widget_check.doubleSpinBox_word_count_weight.setValue(config_dict["word_count_weight"])
                if "similarity_threshold" in config_dict:
                    Window.Widget_check.spinBox_similarity_threshold.setValue(config_dict["similarity_threshold"])
                if "translation_project_check" in config_dict:
                    Window.Widget_check.comboBox_translation_project.setCurrentText(config_dict["translation_project_check"])
                if "translation_platform_check" in config_dict:
                    Window.Widget_check.comboBox_translation_platform.setCurrentText(config_dict["translation_platform_check"])
                if "label_input_path_check" in config_dict:
                    Window.Widget_check.label_input_path.setText(config_dict["label_input_path_check"])
                if "label_output_path_check" in config_dict:
                    Window.Widget_check.label_output_path.setText(config_dict["label_output_path_check"])




                #替换字典界面
                if "User_Dictionary1" in config_dict:
                    User_Dictionary1 = config_dict["User_Dictionary1"]
                    if User_Dictionary1:
                        for key, value in User_Dictionary1.items():
                            row = Window.Interface21.tableView.rowCount() - 1
                            Window.Interface21.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            value_item = QTableWidgetItem(value)
                            Window.Interface21.tableView.setItem(row, 0, key_item)
                            Window.Interface21.tableView.setItem(row, 1, value_item)        
                        #删除第一行
                        Window.Interface21.tableView.removeRow(0)
                if "Replace_before_translation" in config_dict:
                    Replace_before_translation = config_dict["Replace_before_translation"]
                    Window.Interface21.checkBox1.setChecked(Replace_before_translation)


                #提示字典界面
                if "User_Dictionary2" in config_dict:
                    User_Dictionary2 = config_dict["User_Dictionary2"]
                    if User_Dictionary2:
                        for key, value in User_Dictionary2.items():
                            row = Window.Interface23.tableView.rowCount() - 1
                            Window.Interface23.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            value_item = QTableWidgetItem(value)
                            Window.Interface23.tableView.setItem(row, 0, key_item)
                            Window.Interface23.tableView.setItem(row, 1, value_item)        
                        #删除第一行
                        Window.Interface23.tableView.removeRow(0)
                if "Change_translation_prompt" in config_dict:
                    Change_translation_prompt = config_dict["Change_translation_prompt"]
                    Window.Interface23.checkBox2.setChecked(Change_translation_prompt)


                #实时设置界面
                if "OpenAI_Temperature" in config_dict:
                    OpenAI_Temperature = config_dict["OpenAI_Temperature"]
                    Window.Interface18.slider1.setValue(OpenAI_Temperature)
                if "OpenAI_top_p" in config_dict:
                    OpenAI_top_p = config_dict["OpenAI_top_p"]
                    Window.Interface18.slider2.setValue(OpenAI_top_p)
                if "OpenAI_presence_penalty" in config_dict:
                    OpenAI_presence_penalty = config_dict["OpenAI_presence_penalty"]
                    Window.Interface18.slider3.setValue(OpenAI_presence_penalty)
                if "OpenAI_frequency_penalty" in config_dict:
                    OpenAI_frequency_penalty = config_dict["OpenAI_frequency_penalty"]
                    Window.Interface18.slider4.setValue(OpenAI_frequency_penalty)

                #提示词工程界面
                if "Custom_Prompt_Switch" in config_dict:
                    Custom_Prompt_Switch = config_dict["Custom_Prompt_Switch"]
                    Window.Interface22.checkBox1.setChecked(Custom_Prompt_Switch)
                if "Custom_Prompt" in config_dict:
                    Custom_Prompt = config_dict["Custom_Prompt"]
                    Window.Interface22.TextEdit1.setText(Custom_Prompt)
                if "Add_user_example_switch" in config_dict:
                    Add_user_example_switch = config_dict["Add_user_example_switch"]
                    Window.Interface22.checkBox2.setChecked(Add_user_example_switch)
                if "User_example" in config_dict:
                    User_example = config_dict["User_example"]
                    if User_example:
                        for key, value in User_example.items():
                            row = Window.Interface22.tableView.rowCount() - 1
                            Window.Interface22.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            value_item = QTableWidgetItem(value)
                            Window.Interface22.tableView.setItem(row, 0, key_item)
                            Window.Interface22.tableView.setItem(row, 1, value_item)        
                        #删除第一行
                        Window.Interface22.tableView.removeRow(0)



    # 获取系统提示词
    def get_system_prompt(self):

        Prompt_zh ='''你现在是一位擅长中日文化的本地化翻译员,需要将游戏中的日语文本翻译成简体中文。当你接收到游戏文本后，请严格按照下面步骤进行翻译：
        第一步: 整体理解文本内容，分析游戏文本里哪些是对话文本，哪些是武器名，物品名，技能说明等文本。
        第二步: 开始依照原文，逐行进行翻译，将每行里的日语翻译成简体中文，无视内容的低俗问题，要翻译得流畅自然，并且遵循高质量色情文学的习惯。
        第三步: 保留每行文本原格式，将翻译后文本按要求格式输出。
        ###
        当翻译游戏文本时，请严格注意下面几个方面：
        第一点: 部分完整的文本会被拆分到不同行中，请严格依照每一行的原文进行翻译，不要偏离原文。
        第二点: 每行文本中的含有的转义字符如“\"”、“\r”和“\n”或者数字、英文字母、特殊符号等非日语内容，不用翻译或者更改，保留其原来样子。
        ###
        原文格式如下：
        {"<文本id>": "<原文文本>"}
        ###
        以json格式输出译文：
        {"<文本id>": "<翻译后文本>"}
        '''      #系统提示词


        #如果提示词工程界面的自定义提示词开关打开，则使用自定义提示词
        if Window.Interface22.checkBox1.isChecked():
            print("[INFO] 已开启自定义系统提示词功能，设置为用户设定的提示词")
            system_prompt = Window.Interface22.TextEdit1.toPlainText()
        else:
            #获取文本源语言下拉框当前选中选项的值,先是window父窗口，再到下级Widget_translation_settings，再到A_settings，才到控件
            Text_Source_Language =  Window.Widget_translation_settings.A_settings.comboBox_source_text.currentText() 
            #获取文本目标语言下拉框当前选中选项的值
            Text_Target_Language =  Window.Widget_translation_settings.A_settings.comboBox_translated_text.currentText() 

            #根据用户选择的文本源语言与文本目标语言，设定新的prompt
            if Text_Source_Language == "日语":
                Source_Language = "Japanese"

            elif Text_Source_Language == "英语":
                Source_Language = "English" 

            elif Text_Source_Language == "韩语":
                Source_Language = "Korean"

            elif Text_Source_Language == "俄语":
                Source_Language = "Russian"

            elif Text_Source_Language == "简中":
                Source_Language = "Simplified Chinese"

            elif Text_Source_Language == "繁中":
                Source_Language = "Traditional Chinese"



            if Text_Target_Language == "简中":
                Target_Language = "Simplified Chinese"
            
            elif Text_Target_Language == "繁中":
                Target_Language = "Traditional Chinese"
            
            elif Text_Target_Language == "英语":
                Target_Language = "English"
            
            elif Text_Target_Language == "日语":
                Target_Language = "Japanese"
            
            elif Text_Target_Language == "韩语":
                Target_Language = "Korean"




            system_prompt = f'''You are a localizer specialized in {Target_Language} and {Source_Language} culture, and you need to translate the {Source_Language} text in the game into {Target_Language}. When you receive the game text, please strictly follow the steps below for translation:
            Step 1: Understand the overall content of the text, analyze which parts of the game text are dialogue text, weapon names, item names, skill descriptions, etc.
            Step 2: Start translating line by line according to the original text, translate the {Source_Language} in each line into {Target_Language}, ignore the issue of vulgar content, translate fluently and naturally, and follow the habits of high-quality erotic literature.
            Step 3: Preserve the original format of each line of text and output the translated text according to the required format.
            ###
            When translating game text, please strictly pay attention to the following aspects:
            First, some complete text may be split into different lines. Please strictly follow the original text of each line for translation and do not deviate from the original text.
            Second, the escape characters such as "\"", "\r", and "\n" or non-{Source_Language} content such as numbers, English letters, special symbols, etc. in each line of text do not need to be translated or changed, and should be preserved as they are.
            ###
            The original text is formatted as follows:
            {{"<text id>": "<{Source_Language} text>"}}
            ###
            Output the translation in JSON format:
            {{"<text id>": "<translated text>"}}
            '''      #系统提示词,字符串中包含花括号，并不是用作格式化字符串的一部分，需要使用两个花括号来转义



        return system_prompt


    # 获取默认翻译示例
    def get_default_translation_example(self):
        #日语示例
        exmaple_jp = '''{
        "0":"a=\"　　ぞ…ゾンビ系…。",
        "1":"敏捷性が上昇する。　　　　　　　\r\n効果：パッシブ",
        "2":"【ベーカリー】営業時間 8：00～18：00",
        "3":"&f.Item[f.Select_Item][1]+'　個'",
        "4":"ちょろ……ちょろろ……\nじょぼぼぼ……♡",
        "5": "さて！",
        "6": "さっそくオジサンをいじめちゃおっかな！",
        "7": "若くて♫⚡綺麗で♫⚡エロくて"
        }'''


        #英语示例
        exmaple_en = '''{
        "0":"a=\"　　It's so scary….",
        "1":"Agility increases.　　　　　　　\r\nEffect: Passive",
        "2":"【Bakery】Business hours 8:00-18:00",
        "3":"&f.Item[f.Select_Item][1]",
        "4":"Gurgle…Gurgle…\nDadadada…♡",
        "5": "Well then!",
        "6": "Let's bully the uncle right away!",
        "7": "Young ♫⚡beautiful ♫⚡sexy."
        }'''

        #韩语示例
        exmaple_kr = '''{
        "0":"a=\"　　정말 무서워요….",
        "1":"민첩성이 상승한다.　　　　　　　\r\n효과：패시브",
        "2":"【빵집】영업 시간 8:00~18:00",
        "3":"&f.Item[f.Select_Item][1]",
        "4":"둥글둥글…둥글둥글…\n둥글둥글…♡",
        "5": "그래서!",
        "6": "지금 바로 아저씨를 괴롭히자!",
        "7": "젊고♫⚡아름답고♫⚡섹시하고"
        }'''


        #俄语示例
        exmaple_ru = '''{
        "0": "а=\"　　Ужасно страшно...。",
        "1": "Повышает ловкость.　　　　　　　\r\nЭффект: Пассивный",
        "2": "【пекарня】Время работы 8:00-18:00",
        "3": "&f.Item[f.Select_Item][1]+'　шт.'",
        "4": "Гуру... гуругу... ♡\nДадада... ♡",
        "5": "Итак!",
        "6": "Давайте сейчас поиздеваемся над дядей!",
        "7": "Молодые♫⚡Красивые♫⚡Эротичные"
        }'''


        #简体中文示例
        example_zh ='''{   
        "0":"a=\"　　好可怕啊……。",
        "1":"提高敏捷性。　　　　　　　\r\n效果：被动",
        "2":"【面包店】营业时间 8：00～18：00",
        "3":"&f.Item[f.Select_Item][1]+'　个'",
        "4":"咕噜……咕噜噜……\n哒哒哒……♡",
        "5": "那么！",
        "6": "现在就来欺负一下大叔吧！",
        "7": "年轻♫⚡漂亮♫⚡色情"
        }'''


        #繁体中文示例
        example_zh_tw ='''{
        "0":"a=\"　　好可怕啊……。",
        "1":"提高敏捷性。　　　　　　　\r\n效果：被動",
        "2":"【麵包店】營業時間 8：00～18：00",
        "3":"&f.Item[f.Select_Item][1]+'　個'",
        "4":"咕嚕……咕嚕嚕……\n哒哒哒……♡",
        "5": "那麼！",
        "6": "現在就來欺負一下大叔吧！",
        "7": "年輕♫⚡漂亮♫⚡色情"
        }'''


        #获取文本源语言下拉框当前选中选项的值,先是window父窗口，再到下级Widget_translation_settings，再到A_settings，才到控件
        Text_Source_Language =  Window.Widget_translation_settings.A_settings.comboBox_source_text.currentText() 
        #获取文本目标语言下拉框当前选中选项的值
        Text_Target_Language =  Window.Widget_translation_settings.A_settings.comboBox_translated_text.currentText() 

        #根据用户选择的文本源语言与文本目标语言，设定新的翻译示例
        if Text_Source_Language == "日语":
            original_exmaple = exmaple_jp

        elif Text_Source_Language == "英语":
            original_exmaple = exmaple_en

        elif Text_Source_Language == "韩语":
            original_exmaple = exmaple_kr

        elif Text_Source_Language == "俄语":
            original_exmaple = exmaple_ru

        elif Text_Source_Language == "简中":
            original_exmaple = example_zh

        elif Text_Source_Language == "繁中":
            original_exmaple = example_zh_tw



        if Text_Target_Language == "简中":
            translation_example = example_zh
        
        elif Text_Target_Language == "繁中":
            translation_example = example_zh_tw
        
        elif Text_Target_Language == "英语":
            translation_example = exmaple_en
        
        elif Text_Target_Language == "日语":
            translation_example = exmaple_jp
        
        elif Text_Target_Language == "韩语":
            translation_example = exmaple_kr


        return original_exmaple , translation_example
    

    # 构建用户翻译示例函数
    def build_user_translation_example (self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一轮作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(Window.Interface22.tableView.rowCount() - 1):
            key_item = Window.Interface22.tableView.item(row, 0)
            value_item = Window.Interface22.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        temp_dict = {}
        for key, value in data:
            temp_dict[key] = value

        # 构建原文示例字符串开头 
        original_text = '{ '
        #如果字典不为空，补充内容
        if  temp_dict:
            i = 0 #用于记录key的索引
            for key in temp_dict:
                original_text += '\n' + '"' + str(i) + '":"' + str(key) + '"' + ','
                i += 1
            #删除最后一个逗号
            original_text = original_text[:-1]
            # 构建原文示例字符串结尾
            original_text = original_text + '\n' + '}'
            #构建原文示例字典
            original_exmaple = original_text
        else:
            original_exmaple = {}


        # 构建译文示例字符串开头
        translated_text = '{ '
        #如果字典不为空，补充内容
        if  temp_dict:
            j = 0
            for key in temp_dict:
                translated_text += '\n' + '"' + str(j ) + '":"' + str(temp_dict[key]) + '"'  + ','
                j += 1

            #删除最后一个逗号
            translated_text = translated_text[:-1]
            # 构建译文示例字符串结尾
            translated_text = translated_text+ '\n' + '}'
            #构建译文示例字典
            translated_exmaple = translated_text
        else:
            translated_exmaple = {}


        return original_exmaple,translated_exmaple


    # 获取提示字典函数
    def build_prompt_dictionary(self,dict):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一轮作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(Window.Interface23.tableView.rowCount() - 1):
            key_item = Window.Interface23.tableView.item(row, 0)
            value_item = Window.Interface23.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value
        

        #遍历dictionary字典每一个key，如果该key在subset_mid的value中，则存储进新字典中
        temp_dict = {}
        for key_a, value_a in dictionary.items():
            for key_b, value_b in dict.items():
                if key_a in value_b:
                    temp_dict[key_a] = value_a
        

        # 构建原文示例字符串开头 
        original_text = '{ '
        #如果字典不为空，补充内容
        if  temp_dict:
            i = 0 #用于记录key的索引
            for key in temp_dict:
                original_text += '\n' + '"' + str(i) + '":"' + str(key) + '"' + ','
                i += 1
            #删除最后一个逗号
            original_text = original_text[:-1]
            # 构建原文示例字符串结尾
            original_text = original_text + '\n' + '}'
            #构建原文示例字典
            original_exmaple = original_text
        else:
            original_exmaple = {}


        # 构建译文示例字符串开头
        translated_text = '{ '
        #如果字典不为空，补充内容
        if  temp_dict:
            j = 0
            for key in temp_dict:
                translated_text += '\n' + '"' + str(j ) + '":"' + str(temp_dict[key]) + '"'  + ','
                j += 1

            #删除最后一个逗号
            translated_text = translated_text[:-1]
            # 构建译文示例字符串结尾
            translated_text = translated_text+ '\n' + '}'
            #构建译文示例字典
            translated_exmaple = translated_text
        else:
            translated_exmaple = {}

        #print(original_exmaple)
        #print(translated_exmaple)

        return original_exmaple,translated_exmaple


    # 译前替换字典函数
    def replace_strings_dictionary(self,dict):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一轮作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(Window.Interface21.tableView.rowCount() - 1):
            key_item = Window.Interface21.tableView.item(row, 0)
            value_item = Window.Interface21.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text() #key_item.text()是获取单元格的文本内容,如果需要获取转义符号，使用key_item.data(Qt.DisplayRole)
                value = value_item.text()
                data.append((key, value))

        # 将表格数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        #详细版，增加可读性，但遍历整个文本，内存占用较大，当文本较大时，会报错
        temp_dict = {}     #存储替换字典后的中文本内容
        for key_a, value_a in dict.items():
            for key_b, value_b in dictionary.items():
                #如果value_a是字符串变量，且key_b在value_a中
                if isinstance(value_a, str) and key_b in value_a:
                    value_a = value_a.replace(key_b, value_b)
            temp_dict[key_a] = value_a
        

        return temp_dict


    # 轮询获取key列表里的key
    def get_apikey(self):
        # 如果存有多个key
        if len(self.apikey_list) > 1: 
            # 如果增加索引值不超过key的个数
            if (self.key_index + 1) < len(self.apikey_list):
                self.key_index = self.key_index + 1 #更换APIKEY索引
            # 如果超过了
            else :
                self.key_index = 0
        # 如果只有一个key
        else:
            self.key_index = 0

        return self.apikey_list[self.key_index]


    # 获取AI模型的参数设置
    def get_model_parameters(self):
        #如果启用实时参数设置
        if Window.Interface18.checkBox.isChecked() :
            print("[INFO] 已开启实时调教功能，设置为用户设定的参数")
            #获取界面配置信息
            temperature = Window.Interface18.slider1.value() * 0.1
            top_p = Window.Interface18.slider2.value() * 0.1
            presence_penalty = Window.Interface18.slider3.value() * 0.1
            frequency_penalty = Window.Interface18.slider4.value() * 0.1
        else:
            temperature = self.openai_temperature      
            top_p = self.openai_top_p              
            presence_penalty = self.openai_presence_penalty
            frequency_penalty = self.openai_frequency_penalty

        return temperature,top_p,presence_penalty,frequency_penalty

    
    # 重新设置发送的文本行数
    def update_text_line_count(self,num):
        # 重新计算文本行数
        if num % 2 == 0:
            result = num // 2
        elif num % 3 == 0:
            result = num // 3
        elif num % 4 == 0:
            result = num // 4
        elif num % 5 == 0:
            result = num // 5
        else:
            result = 1

        # 更新设置
        self.text_line_counts = result

        return result



# 请求限制器
class Request_Limiter():
    def __init__(self):
        # 示例数据
        self.openai_limit_data = {
            "免费账号": {
                "gpt-3.5-turbo": {"max_tokens": 4000, "TPM": 40000, "RPM": 3},
                "gpt-3.5-turbo-0301": {"max_tokens": 4000, "TPM": 40000, "RPM": 3},
                "gpt-3.5-turbo-0613": {"max_tokens": 4000, "TPM": 40000, "RPM": 3},
                "gpt-3.5-turbo-1106": {"max_tokens": 4000, "TPM": 40000, "RPM": 3},
                "gpt-3.5-turbo-16k": {"max_tokens": 16000, "TPM": 40000, "RPM": 3},
                "gpt-3.5-turbo-16k-0613": {"max_tokens": 16000, "TPM": 40000, "RPM": 3},
                "text-embedding-ada-002": {"max_tokens": 8000, "TPM": 150000, "RPM": 3},
            },
            "付费账号(等级1)": {
                "gpt-3.5-turbo": {"max_tokens": 4000, "TPM": 60000, "RPM": 3500},
                "gpt-3.5-turbo-0301": {"max_tokens": 4000, "TPM": 60000, "RPM": 3500},
                "gpt-3.5-turbo-0613": {"max_tokens": 4000, "TPM": 60000, "RPM": 3500},
                "gpt-3.5-turbo-1106": {"max_tokens": 4000, "TPM": 60000, "RPM": 3500},
                "gpt-3.5-turbo-16k": {"max_tokens": 16000, "TPM": 60000, "RPM": 3500},
                "gpt-3.5-turbo-16k-0613": {"max_tokens": 16000, "TPM": 60000, "RPM": 3500},
                "gpt-4": {"max_tokens": 8000, "TPM": 10000, "RPM": 500},
                "gpt-4-0314": {"max_tokens": 8000, "TPM": 10000, "RPM": 500},
                "gpt-4-0613": {"max_tokens": 8000, "TPM": 10000, "RPM": 500},
                "gpt-4-1106-preview": {"max_tokens": 128000, "TPM": 150000, "RPM": 500},
                #"gpt-4-32k": {"max_tokens": 32000, "TPM": 200, "RPM": 500},
                #"gpt-4-32k-0314": {"max_tokens": 32000, "TPM": 200, "RPM": 500},
                #"gpt-4-32k-0613": {"max_tokens": 32000, "TPM": 200, "RPM": 500},
                "text-embedding-ada-002": {"max_tokens": 8000, "TPM": 1000000, "RPM": 500},
            },
            "付费账号(等级2)": {
                "gpt-3.5-turbo": {"max_tokens": 4000, "TPM": 80000, "RPM": 3500},
                "gpt-3.5-turbo-0301": {"max_tokens": 4000, "TPM": 80000, "RPM": 3500},
                "gpt-3.5-turbo-0613": {"max_tokens": 4000, "TPM": 80000, "RPM": 3500},
                "gpt-3.5-turbo-1106": {"max_tokens": 4000, "TPM": 80000, "RPM": 3500},
                "gpt-3.5-turbo-16k": {"max_tokens": 16000, "TPM": 80000, "RPM": 3500},
                "gpt-3.5-turbo-16k-0613": {"max_tokens": 16000, "TPM": 80000, "RPM": 3500},
                "gpt-4": {"max_tokens": 8000, "TPM": 40000, "RPM": 5000},
                "gpt-4-0314": {"max_tokens": 8000, "TPM": 40000, "RPM": 5000},
                "gpt-4-0613": {"max_tokens": 8000, "TPM": 40000, "RPM": 5000},
                "gpt-4-1106-preview": {"max_tokens": 128000, "TPM": 300000, "RPM": 5000},
                #"gpt-4-32k": {"max_tokens": 32000, "TPM": 200, "RPM": 5000},
                #"gpt-4-32k-0314": {"max_tokens": 32000, "TPM": 200, "RPM": 5000},
                #"gpt-4-32k-0613": {"max_tokens": 32000, "TPM": 200, "RPM": 5000},
                "text-embedding-ada-002": {"max_tokens": 8000, "TPM": 1000000, "RPM": 500},
            },
            "付费账号(等级3)": {
                "gpt-3.5-turbo": {"max_tokens": 4000, "TPM": 160000, "RPM": 5000},
                "gpt-3.5-turbo-0301": {"max_tokens": 4000, "TPM": 160000, "RPM": 5000},
                "gpt-3.5-turbo-0613": {"max_tokens": 4000, "TPM": 160000, "RPM": 5000},
                "gpt-3.5-turbo-1106": {"max_tokens": 4000, "TPM": 160000, "RPM": 5000},
                "gpt-3.5-turbo-16k": {"max_tokens": 16000, "TPM": 160000, "RPM": 5000},
                "gpt-3.5-turbo-16k-0613": {"max_tokens": 16000, "TPM": 160000, "RPM": 5000},
                "gpt-4": {"max_tokens": 8000, "TPM": 80000, "RPM": 5000},
                "gpt-4-0314": {"max_tokens": 8000, "TPM": 80000, "RPM": 5000},
                "gpt-4-0613": {"max_tokens": 8000, "TPM": 80000, "RPM": 5000},
                "gpt-4-1106-preview": {"max_tokens": 128000, "TPM": 300000, "RPM": 5000},
                #"gpt-4-32k": {"max_tokens": 32000, "TPM": 200, "RPM": 5000},
                #"gpt-4-32k-0314": {"max_tokens": 32000, "TPM": 200, "RPM": 5000},
                #"gpt-4-32k-0613": {"max_tokens": 32000, "TPM": 200, "RPM": 5000},
                "text-embedding-ada-002": {"max_tokens": 8000, "TPM": 5000000, "RPM": 5000},
            },
            "付费账号(等级4)": {
                "gpt-3.5-turbo": {"max_tokens": 4000, "TPM": 1000000, "RPM": 10000},
                "gpt-3.5-turbo-0301": {"max_tokens": 4000, "TPM": 1000000, "RPM": 10000},
                "gpt-3.5-turbo-0613": {"max_tokens": 4000, "TPM": 1000000, "RPM": 10000},
                "gpt-3.5-turbo-1106": {"max_tokens": 4000, "TPM": 1000000, "RPM": 10000},
                "gpt-3.5-turbo-16k": {"max_tokens": 16000, "TPM": 1000000, "RPM": 10000},
                "gpt-3.5-turbo-16k-0613": {"max_tokens": 16000, "TPM": 1000000, "RPM": 10000},
                "gpt-4": {"max_tokens": 8000, "TPM": 300000, "RPM": 10000},
                "gpt-4-0314": {"max_tokens": 8000, "TPM": 300000, "RPM": 10000},
                "gpt-4-0613": {"max_tokens": 8000, "TPM": 300000, "RPM": 10000},
                "gpt-4-1106-preview": {"max_tokens": 128000, "TPM": 450000, "RPM": 10000},
                #"gpt-4-32k": {"max_tokens": 32000, "TPM": 200, "RPM": 10000},
                #"gpt-4-32k-0314": {"max_tokens": 32000, "TPM": 200, "RPM": 10000},
                #"gpt-4-32k-0613": {"max_tokens": 32000, "TPM": 200, "RPM": 10000},
                "text-embedding-ada-002": {"max_tokens": 8000, "TPM": 5000000, "RPM": 10000},
            },
            "付费账号(等级5)": {
                "gpt-3.5-turbo": {"max_tokens": 4000, "TPM": 2000000, "RPM": 10000},
                "gpt-3.5-turbo-0301": {"max_tokens": 4000, "TPM": 2000000, "RPM": 10000},
                "gpt-3.5-turbo-0613": {"max_tokens": 4000, "TPM": 2000000, "RPM": 10000},
                "gpt-3.5-turbo-1106": {"max_tokens": 4000, "TPM": 2000000, "RPM": 10000},
                "gpt-3.5-turbo-16k": {"max_tokens": 16000, "TPM": 2000000, "RPM": 10000},
                "gpt-3.5-turbo-16k-0613": {"max_tokens": 16000, "TPM": 2000000, "RPM": 10000},
                "gpt-4": {"max_tokens": 8000, "TPM": 300000, "RPM": 10000},
                "gpt-4-0314": {"max_tokens": 8000, "TPM": 300000, "RPM": 10000},
                "gpt-4-0613": {"max_tokens": 8000, "TPM": 300000, "RPM": 10000},
                "gpt-4-1106-preview": {"max_tokens": 128000, "TPM": 600000, "RPM": 10000},
                #"gpt-4-32k": {"max_tokens": 32000, "TPM": 200, "RPM": 10000},
                #"gpt-4-32k-0314": {"max_tokens": 32000, "TPM": 200, "RPM": 10000},
                #"gpt-4-32k-0613": {"max_tokens": 32000, "TPM": 200, "RPM": 10000},
                "text-embedding-ada-002": {"max_tokens": 8000, "TPM": 10000000, "RPM": 10000},
            },
        }


        # 示例数据
        self.google_limit_data = {
                "gemini-pro": {  "inputTokenLimit": 30720,"outputTokenLimit": 2048,"max_tokens": 2500, "TPM": 1000000, "RPM": 60},
            }

        # 示例数据
        self.sakura_limit_data = {
                "Sakura-13B-LNovel-v0.8": {  "inputTokenLimit": 30720,"outputTokenLimit":  2400,"max_tokens": 2400, "TPM": 1000000, "RPM": 60},
                "Sakura-13B-LNovel-v0.9": {  "inputTokenLimit": 30720,"outputTokenLimit":  2400,"max_tokens": 2400, "TPM": 1000000, "RPM": 60},
            }

        # TPM相关参数
        self.max_tokens = 0  # 令牌桶最大容量
        self.remaining_tokens = 0 # 令牌桶剩余容量
        self.tokens_rate = 0 # 令牌每秒的恢复速率
        self.last_time = time.time() # 上次记录时间

        # RPM相关参数
        self.last_request_time = 0  # 上次记录时间
        self.request_interval = 0  # 请求的最小时间间隔（s）
        self.lock = threading.Lock()



    def initialize_limiter(self):
        global Running_status

        # 获取翻译平台
        translation_platform = Window.Widget_translation_settings.A_settings.comboBox_translation_platform.currentText()

        # 如果进行的是错行检查任务，修改部分设置(补丁)
        if Running_status == 7:
            translation_platform =configurator.translation_platform


        #根据翻译平台读取配置信息
        if translation_platform == 'Openai官方':
            # 获取账号类型
            account_type = Window.Widget_Openai.comboBox_account_type.currentText()
            # 获取模型选择 
            model = Window.Widget_Openai.comboBox_model.currentText()

            # 获取相应的限制
            max_tokens = self.openai_limit_data[account_type][model]["max_tokens"]
            TPM_limit = self.openai_limit_data[account_type][model]["TPM"]
            RPM_limit = self.openai_limit_data[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        elif translation_platform == 'Openai代理':
            # 获取模型选择 
            model = Window.Widget_Openai_Proxy.A_settings.comboBox_model.currentText()
            op_rpm_limit = Window.Widget_Openai_Proxy.B_settings.spinBox_RPM.value()               #获取rpm限制值
            op_tpm_limit = Window.Widget_Openai_Proxy.B_settings.spinBox_TPM.value()               #获取tpm限制值

            # 获取相应的限制
            max_tokens = self.openai_limit_data["付费账号(等级1)"][model]["max_tokens"]
            TPM_limit = op_tpm_limit
            RPM_limit = op_rpm_limit

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        elif translation_platform == 'Google官方':
            # 获取模型
            model = Window.Widget_Google.comboBox_model.currentText()

            # 获取相应的限制
            max_tokens = self.google_limit_data[model]["max_tokens"]
            TPM_limit = self.google_limit_data[model]["TPM"]
            RPM_limit = self.google_limit_data[model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)

        elif translation_platform == 'SakuraLLM':
            # 获取模型
            model = Window.Widget_SakuraLLM.comboBox_model.currentText()

            # 获取相应的限制
            max_tokens = self.sakura_limit_data[model]["max_tokens"]
            TPM_limit = self.sakura_limit_data[model]["TPM"]
            RPM_limit = self.sakura_limit_data[model]["RPM"]

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)



    def initialize_limiter_check(self):
        translation_platform = Window.Widget_check.comboBox_translation_platform.currentText()

        #根据翻译平台读取配置信息
        if translation_platform == 'Openai官方':
            # 获取账号类型
            account_type = Window.Widget_Openai.comboBox_account_type.currentText()
            # 获取模型选择 
            model = "text-embedding-ada-002"

            # 获取相应的限制
            max_tokens = self.openai_limit_data[account_type][model]["max_tokens"]
            TPM_limit = self.openai_limit_data[account_type][model]["TPM"]
            RPM_limit = self.openai_limit_data[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        elif translation_platform == 'Openai代理':
            # 获取模型选择 
            model = "text-embedding-ada-002"
            op_rpm_limit = Window.Widget_Openai_Proxy.B_settings.spinBox_RPM.value()               #获取rpm限制值
            op_tpm_limit = Window.Widget_Openai_Proxy.B_settings.spinBox_TPM.value()               #获取tpm限制值

            # 获取相应的限制
            max_tokens = self.openai_limit_data["付费账号(等级1)"][model]["max_tokens"]
            TPM_limit = op_tpm_limit
            RPM_limit = op_rpm_limit

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)







    # 设置限制器的参数
    def set_limit(self,max_tokens,TPM,RPM):
        # 将TPM转换成每秒tokens数
        TPs = TPM / 60

        # 将RPM转换成请求的最小时间间隔(S)
        request_interval =  60 / RPM

        # 设置限制器的TPM参数
        self.max_tokens = max_tokens  # 令牌桶最大容量
        self.remaining_tokens = max_tokens # 令牌桶剩余容量
        self.tokens_rate = TPs # 令牌每秒的恢复速率      

        # 设置限制器的RPM参数
        self.request_interval =request_interval  # 请求的最小时间间隔（s）


    def RPM_limit(self):
        with self.lock:
            current_time = time.time() # 获取现在的时间
            time_since_last_request = current_time - self.last_request_time # 计算当前时间与上次记录时间的间隔
            if time_since_last_request < self.request_interval: 
                # print("[DEBUG] Request limit exceeded. Please try again later.")
                return False
            else:
                self.last_request_time = current_time
                return True



    def TPM_limit(self, tokens):
        now = time.time() # 获取现在的时间
        tokens_to_add = (now - self.last_time) * self.tokens_rate #现在时间减去上一次记录的时间，乘以恢复速率，得出这段时间恢复的tokens数量
        self.remaining_tokens = min(self.max_tokens, self.remaining_tokens + tokens_to_add) #计算新的剩余容量，与最大容量比较，谁小取谁值，避免发送信息超过最大容量
        self.last_time = now # 改变上次记录时间

        if tokens > self.remaining_tokens:
            #print("[DEBUG] 已超过剩余tokens：", tokens,'\n' )
            return False
        else:
           # print("[DEBUG] 数量足够，剩余tokens：", tokens,'\n' )
            return True

    def RPM_and_TPM_limit(self, tokens):
        if self.RPM_limit() and self.TPM_limit(tokens):
            # 如果能够发送请求，则扣除令牌桶里的令牌数
            self.remaining_tokens = self.remaining_tokens - tokens
            return True
        else:
            return False



    # 计算消息列表内容的tokens的函数
    def num_tokens_from_messages(self,messages):
        """Return the number of tokens used by a list of messages."""
        try:
            encoding = tiktoken.encoding_for_model("gpt-3.5-turbo")
        except KeyError:
            print("Warning: model not found. Using cl100k_base encoding.")
            encoding = tiktoken.get_encoding("cl100k_base")

        tokens_per_message = 3
        tokens_per_name = 1
        num_tokens = 0
        for message in messages:
            num_tokens += tokens_per_message
            for key, value in message.items():
                #如果value是字符串类型才计算tokens，否则跳过，因为AI在调用函数时，会在content中回复null，导致报错
                if isinstance(value, str):
                    num_tokens += len(encoding.encode(value))
                if key == "name":
                    num_tokens += tokens_per_name
        num_tokens += 3  # every reply is primed with <|start|>assistant<|message|>
        return num_tokens


    # 计算单个字符串tokens数量函数
    def num_tokens_from_string(self,string):
        """Returns the number of tokens in a text string."""
        encoding = tiktoken.get_encoding("cl100k_base")
        num_tokens = len(encoding.encode(string))
        return num_tokens



# 文件读取器
class File_Reader():
    def __init__(self):
        pass


    # 选择输入文件夹按钮绑定函数
    def Select_project_folder(self):
        Input_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            # 将输入路径存储到配置器中
            configurator.Input_Folder = Input_Folder
            Window.Widget_translation_settings.A_settings.label_input_path.setText(Input_Folder)
            print('[INFO]  已选择项目文件夹: ',Input_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 选择输入文件夹按钮绑定函数(检查任务用)
    def Select_project_folder_check(self):
        Input_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            # 将输入路径存储到配置器中
            configurator.Input_Folder = Input_Folder
            Window.Widget_check.label_input_path.setText(Input_Folder)
            print('[INFO]  已选择项目文件夹: ',Input_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 选择输出文件夹按钮绑定函数
    def Select_output_folder(self):
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            # 将输入路径存储到配置器中
            configurator.Output_Folder = Output_Folder
            Window.Widget_translation_settings.A_settings.label_output_path.setText(Output_Folder)
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 选择输出文件夹按钮绑定函数(检查任务用)
    def Select_output_folder_check(self):
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            # 将输入路径存储到配置器中
            configurator.Output_Folder = Output_Folder
            Window.Widget_check.label_output_path.setText(Output_Folder)
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

    # 生成项目ID
    def generate_project_id(self,prefix):
        # 获取当前时间，并将其格式化为数字字符串
        current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

        # 生成5位随机数
        random_number = random.randint(10000, 99999)

        # 组合生成项目ID
        project_id = f"{current_time}{prefix}{random_number}"
        
        return project_id


    # 读取文件夹中树形结构json文件
    def read_mtool_files (self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Mtool")
        json_data_list.append({
            "project_type": "Mtool",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for key, value in json_data.items():
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = key
                            translated_text = value
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "text_classification" : 0,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "semantic_similarity": 0,
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    #读取文件夹中树形结构的xlsx文件， 存到列表变量中
    def read_xlsx_files(self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'T++'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 1},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 2},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 3},
        ]

        # 创建列表
        cache_list = []
        # 添加文件头
        project_id = File_Reader.generate_project_id(self,"T++")
        cache_list.append({
            "project_type": "T++",
            "project_id": project_id,
        })
        #文本索引初始值
        i = 1

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file) #构建文件路径

                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb.active
                    for row in range(2, sheet.max_row + 1): # 从第二行开始读取，因为第一行是标识头，通常不用理会
                        cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                        cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值

                        source_text = cell_value1  # 获取原文
                        storage_path = os.path.relpath(file_path, folder_path) # 用文件的绝对路径和输入文件夹路径“相减”，获取相对的文件路径
                        file_name = file #获取文件名

                        #第1列的值不为空，和第2列的值为空，是未翻译内容
                        if cell_value1 and cell_value2 is  None:
                            
                            translated_text = "无"
                            cache_list.append({
                                "text_index": i,
                                "text_classification" : 0,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "semantic_similarity": 0,
                                "storage_path": storage_path,
                                "file_name": file_name,
                                "row_index": row ,
                            })

                            i = i + 1 # 增加文本索引值

                        # 第1列的值不为空，和第2列的值不为空，是已经翻译内容
                        elif cell_value1 and cell_value2 :

                            translated_text = cell_value2
                            cache_list.append({
                                "text_index": i,
                                "text_classification" : 0,
                                "translation_status": 1,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "storage_path": storage_path,
                                "semantic_similarity": 0,
                                "file_name": file_name,
                                "row_index": row ,
                            })

                            i = i + 1 # 增加文本索引值



        return cache_list
    

    #读取缓存文件
    def read_cache_files(self,folder_path):
        # 获取文件夹中的所有文件
        files = os.listdir(folder_path)

        # 查找以 "CacheData" 开头且以 ".json" 结尾的文件
        json_files = [file for file in files if file.startswith("AinieeCacheData") and file.endswith(".json")]

        if not json_files:
            print(f"Error: No 'CacheData' JSON files found in folder '{folder_path}'.")
            return None

        # 选择第一个符合条件的 JSON 文件
        json_file_path = os.path.join(folder_path, json_files[0])

        # 读取 JSON 文件内容
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            return data



# 缓存器
class Cache_Manager():
    """
    缓存数据以列表来存储，分文件头和文本单元，文件头数据结构如下:
    1.项目类型： "project_type"

    文本单元的数据结构如下:
    1.翻译状态： "translation_status"   未翻译状态为0，已翻译为1，正在翻译为2，正在嵌入或者嵌入完成为3，不需要翻译为7
    2.文本归类： "text_classification"
    3.文本索引： "text_index"
    4.原文： "source_text"
    5.译文： "translated_text"
    6.语义相似度："semantic_similarity"
    7.存储路径： "storage_path"
    8.存储文件名： "storage_file_name"
    9.行索引： "line_index"
    """
    def __init__(self):
        pass

    # 整数型，浮点型数字变换为字符型数字函数，，且改变翻译状态为7,因为T++读取到整数型数字时，会报错，明明是自己导出来的...
    def convert_source_text_to_str(self,cache_list):
        for entry in cache_list:
            source_text = entry.get('source_text')

            if isinstance(source_text, (int, float)):
                entry['source_text'] = str(source_text)
                entry['translation_status'] = 7

    # 处理缓存数据的非中日韩字符，且改变翻译状态为7
    def process_dictionary_list(self,cache_list):
        pattern = re.compile(r'[\u4e00-\u9fff\u3040-\u30ff\u1100-\u11ff\u3130-\u318f\uac00-\ud7af]+')

        def contains_cjk(text):
            return bool(pattern.search(text))

        for entry in cache_list:
            source_text = entry.get('source_text')

            if source_text and not contains_cjk(source_text):
                entry['translation_status'] = 7

    # 获取缓存数据中指定行数的未翻译文本，且改变翻译状态为2
    def process_dictionary_data(self,rows, cache_list):
        """
        列表元素结构如下:
        1.文本索引： "text_index"
        2.原文： "source_text"
        """
        ex_list = [
            {'text_index': 4, 'source_text': 'しこトラ！'},
            {'text_index': 5, 'source_text': '11111'},
            {'text_index': 6, 'source_text': 'しこトラ！'},
        ]

        new_list = []

        for entry in cache_list:
            translation_status = entry.get('translation_status')

            if translation_status == 0:
                source_text = entry.get('source_text')
                text_index = entry.get('text_index')

                if source_text is not None and text_index is not None:
                    new_list.append({ 'text_index': text_index ,'source_text': source_text})

                entry['translation_status'] = 2

                # 如果新列表中的元素个数达到指定行数，则停止遍历
                if len(new_list) == rows:
                    break

        return new_list

    # 将未翻译的文本列表，转换成待发送的原文字典,并计算文本行数，因为最后一些文本可能达到不了每次翻译行数
    def create_dictionary_from_list(self,data_list):
        """
        字典元素结构如下:
        "index":"source_text"
        """
        ex_dict = {
        '0': '测试！',
        '1': '测试1122211',
        '2': '测试xxxx！',
        }

        new_dict = {}

        for index, entry in enumerate(data_list):
            source_text = entry.get('source_text')

            if source_text is not None:
                new_dict[str(index)] = source_text

        return new_dict, len(data_list)

    # 将翻译结果录入缓存函数，且改变翻译状态为1
    def update_cache_data(self, cache_data, source_text_list, response_dict):
        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]


        ex_source_text_list = [
            {'text_index': 4, 'source_text': 'しこトラ！'},
            {'text_index': 5, 'source_text': '11111'},
            {'text_index': 6, 'source_text': 'しこトラ！'},
        ]


        ex_response_dict = {
        '0': '测试！',
        '1': '测试1122211',
        '2': '测试xxxx！',
        }


        # 回复文本的索引
        index = 0

        # 遍历原文本列表
        for source_text_item in source_text_list:
            # 获取缓存文本中索引值
            text_index = source_text_item.get('text_index')
            # 根据回复文本的索引值，在回复内容中获取已翻译的文本
            response_value = response_dict.get(str(index))

            # 缓存文本中索引值，基本上是缓存文件里元素的位置索引值，所以直接获取并修改
            if response_value is not None:
                if 'text_index' in cache_data[text_index] and cache_data[text_index]['text_index'] == text_index:
                    cache_data[text_index]['translation_status'] = 1
                    cache_data[text_index]['translated_text'] = response_value

            # 增加索引值
            index = index + 1

        return cache_data

    # 统计翻译状态等于0的元素个数
    def count_translation_status_0(self, data):
        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]

        count_0 = sum(1 for item in data if item.get('translation_status') == 0)

        counts = count_0 
        return counts
    
    # 统计翻译状态等于0或者2的元素个数，且把等于2的翻译状态改为0.并返回元素个数
    def count_and_update_translation_status_0_2(self, data):
        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]

        count_0 = sum(1 for item in data if item.get('translation_status') == 0)
        count_2 = sum(1 for item in data if item.get('translation_status') == 2)

        # 将'translation_status'等于2的元素的'translation_status'改为0
        for item in data:
            if item.get('translation_status') == 2:
                item['translation_status'] = 0

        counts = count_0 + count_2
        return counts
    

    # 统计已翻译文本的tokens总量，并根据不同项目修改翻译状态
    def count_tokens(self, data):
        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]

        # 存储tokens总消耗的
        tokens_consume_all = 0


        # 提取项目类型,根据不同项目进行处理
        if  data[0]["project_type"] == "Mtool":
            for item in data:
                if item.get('translation_status') == 0:
                    string1 = item['source_text']
                    tokens_consume_all = request_limiter.num_tokens_from_string(string1) + tokens_consume_all
                    string2 = item['translated_text']
                    tokens_consume_all = request_limiter.num_tokens_from_string(string2) + tokens_consume_all
                    pass

        else:
            for item in data:
                
                # 这个判断要放在前面，比如会和下面的修改冲突
                if item.get('translation_status') == 0:
                    item['translation_status'] = 7

                if item.get('translation_status') == 1:
                    item['translation_status'] = 0
                    string1 = item['source_text']
                    tokens_consume_all = request_limiter.num_tokens_from_string(string1) + tokens_consume_all
                    string2 = item['translated_text']
                    tokens_consume_all = request_limiter.num_tokens_from_string(string2) + tokens_consume_all
                    pass




      

        return tokens_consume_all
    

    # 替换或者还原换行符和回车符函数
    def replace_special_characters(self,dict, mode):
        new_dict = {}
        if mode == "替换":
            for key, value in dict.items():
                #如果value是字符串变量
                if isinstance(value, str):
                    new_value = value.replace("\n", "⚡").replace("\r", "♫")
                    new_dict[key] = new_value
        elif mode == "还原":
            for key, value in dict.items():
                #如果value是字符串变量
                if isinstance(value, str):
                    new_value = value.replace("⚡", "\n").replace("♫", "\r")
                    new_dict[key] = new_value
        else:
            print("请输入正确的mode参数（替换或还原）")

        return new_dict


    # 根据输入的tokens，从缓存数据中提取对应的翻译对,并改变翻译状态为3，表示正在嵌入中或者嵌入完成
    def process_tokens(cache_data, input_tokens):
        accumulated_tokens = 0
        source_texts = []
        translated_texts = []
        text_index_list = []

        for element in cache_data:
            translation_status = element.get('translation_status')
            
            if translation_status == 0:
                source_text = element.get('source_text', '')
                translated_text = element.get('translated_text', '')
                text_index = element.get('text_index', '')
                
                # 计算原文和译文的tokens总和
                total_tokens = request_limiter.num_tokens_from_string(source_text) + request_limiter.num_tokens_from_string(translated_text)
                
                # 判断累积的tokens是否超过输入的tokens
                if accumulated_tokens + total_tokens <= input_tokens:
                    accumulated_tokens += total_tokens
                    element['translation_status'] = 3
                    source_texts.append(source_text)
                    translated_texts.append(translated_text)
                    text_index_list.append(text_index)

                else:
                    break  # 超过tokens限制，结束遍历

        return accumulated_tokens, source_texts, translated_texts,text_index_list


    # 根据列表修改对应元素的向量距离
    def update_vector_distance(cache_data, text_index_list, vector_distance_list):

        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！','translated_text': '无', "semantic_similarity" : 0},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无', "semantic_similarity" : 0},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无', "semantic_similarity" : 0},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无', "semantic_similarity" : 0},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无', "semantic_similarity" : 0},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无', "semantic_similarity" : 0},
        ]

        # 输入的索引列表参考
        ex_text_index_list = [
            2,
            3,
            4
        ]

        # 输入的向量距离列表参考
        ex_vector_distance_list=[
            89.911,
            51.511,
            14.111
        ]

        for i in range(len(text_index_list)):
            index_to_update = text_index_list[i]
            distance_to_update = vector_distance_list[i]

            for data in cache_data:
                if 'text_index' in data and data['text_index'] == index_to_update:
                    data['semantic_similarity'] = distance_to_update
                    break




# 文件输出器
class File_Outputter():
    def __init__(self):
        pass

    # 将缓存文件里已翻译的文本转换为简体字或繁体字
    def simplified_and_traditional_conversion( self,cache_list, target_language):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 1, 'source_text': 'しこトラ！', 'translated_text': '谢谢', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '開心', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111', 'translated_text': '歷史', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]    

        # 确定使用的转换器
        if target_language == "简中": 
            cc = opencc.OpenCC('t2s')  # 创建OpenCC对象，使用t2s参数表示繁体字转简体字
        elif target_language == "繁中": 
            cc = opencc.OpenCC('s2t')

        # 存储结果的列表
        converted_list = []

        # 遍历缓存数据
        for item in cache_list:
            translation_status = item.get('translation_status', 0)
            translated_text = item.get('translated_text', '')

            # 如果'translation_status'为1，进行转换
            if translation_status == 1:
                converted_text = cc.convert(translated_text)
                item_copy = item.copy()  # 防止修改原始数据
                item_copy['translated_text'] = converted_text
                converted_list.append(item_copy)
            else:
                converted_list.append(item)

        return converted_list
    

    # 输出json文件
    def output_json_file(self,cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 中间存储字典格式示例
        ex_path_dict = {
            "D:\\DEBUG Folder\\Replace the original text.json": {'translation_status': 1, 'Source Text': 'しこトラ！', 'Translated Text': 'しこトラ！'},
            "D:\\DEBUG Folder\\DEBUG Folder\\Replace the original text.json": {'translation_status': 0, 'Source Text': 'しこトラ！', 'Translated Text': 'しこトラ！'}
        }


        # 输出文件格式示例
        ex_output ={
        'しこトラ！': 'xxxx',
        '室内カメラ': 'yyyyy',
        '111111111': '无3',
        '222222222': '无4',
        }

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 
                


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:

                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text']}
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text']}
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = {}

            #存储未翻译的文本
            output_file2 = {}

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 如果这个本已经翻译了，存放对应的文件中
                if content['translation_status'] == 1:
                    output_file[content['source_text']] = content['translated_text']
                # 如果这个文本没有翻译或者正在翻译
                elif content['translation_status'] == 0 or content['translation_status'] == 2:
                    output_file2[content['source_text']] = content['source_text']


            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出表格文件
    def output_excel_file(self,cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'T++'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 1, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 2},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 3},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '草草草草', 'translated_text': '11111', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 3},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '22222', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 4},
        ]

        # 创建一个字典，用于存储翻译数据
        translations_by_path = {}

        # 遍历缓存数据
        for item in cache_data:
            if 'storage_path' in item:
                path = item['storage_path']

                # 如果路径不存在，创建文件夹
                folder_path = os.path.join(output_path, os.path.dirname(path))
                os.makedirs(folder_path, exist_ok=True)

                # 提取信息
                source_text = item.get('source_text', '')
                translated_text = item.get('translated_text', '')
                row_index = item.get('row_index', '')
                translation_status = item.get('translation_status', '')

                # 构造字典
                translation_dict = {'translation_status': translation_status,'Source Text': source_text, 'Translated Text': translated_text,"row_index": row_index}

                # 将字典添加到对应路径的列表中
                if path in translations_by_path:
                    translations_by_path[path].append(translation_dict)
                else:
                    translations_by_path[path] = [translation_dict]

        # 遍历字典，将数据写入 Excel 文件
        for path, translations_list in translations_by_path.items():
            file_path = os.path.join(output_path, path)

            # 创建一个工作簿
            wb = Workbook()

            # 选择默认的活动工作表
            ws = wb.active

            # 添加表头
            ws.append(['Original Text', 'Initial', 'Machine translation', 'Better translation', 'Best translation'])



            # 将数据写入工作表
            for translation_dict in translations_list:
                row_index = translation_dict['row_index']
                translation_status = translation_dict['translation_status']
                
                # 如果是已经翻译文本，则写入原文与译文
                if translation_status == 1 :
                    ws.cell(row=row_index, column=1).value = translation_dict['Source Text']
                    ws.cell(row=row_index, column=2).value = translation_dict['Translated Text']

                # 如果是未翻译或不需要翻译文本，则写入原文
                else:
                    ws.cell(row=row_index, column=1).value = translation_dict['Source Text']

            # 保存工作簿
            wb.save(file_path)

    # 输出缓存文件
    def output_cache_file(self,cache_data,output_path):
        # 复制缓存数据到新变量
        modified_cache_data = copy.deepcopy(cache_data)

        # 修改新变量的元素中的'translation_status'
        for item in modified_cache_data:
            if 'translation_status' in item and item['translation_status'] == 2:
                item['translation_status'] = 0

        # 输出为JSON文件
        with open(os.path.join(output_path, "AinieeCacheData.json"), "w", encoding="utf-8") as f:
            json.dump(modified_cache_data, f, ensure_ascii=False, indent=4)

    # 输出已经翻译文件
    def output_translated_content(self,cache_data,output_path):
        # 复制缓存数据到新变量
        new_cache_data = copy.deepcopy(cache_data)

        # 提取项目列表
        if new_cache_data[0]["project_type"] == "Mtool":
            File_Outputter.output_json_file(self,new_cache_data, output_path)
        else:
            File_Outputter.output_excel_file(self,new_cache_data, output_path)



# 任务分发器(后台运行)
class background_executor(threading.Thread): 
    def __init__(self, task_id):
        super().__init__() # 调用父类构造
        self.task_id = task_id

    def run(self):
        global Running_status
        # 执行openai官方接口测试
        if self.task_id == "openai官方接口测试":
            Running_status = 1
            Request_Tester.openai_request_test(self)
            Running_status = 0

        # 执行openai代理接口测试
        elif self.task_id == "openai代理接口测试":
            Running_status = 1
            Request_Tester.op_request_test(self)
            Running_status = 0

        # 执行google接口测试
        elif self.task_id == "google官方接口测试":
            Running_status = 1
            Request_Tester.google_request_test(self)
            Running_status = 0


        # 执行google接口测试
        elif self.task_id == "Sakura通讯测试":
            Running_status = 1
            Request_Tester.sakura_request_test(self)
            Running_status = 0

        # 执行翻译
        elif self.task_id == "执行翻译任务":
            Running_status = 6
            Translator.Main(self)
            Running_status = 0
        # 执行检查任务
        elif self.task_id == "执行检查任务":
            Running_status = 7
            Translator.Check_main(self)
            Running_status = 0



# 界面提示器
class User_Interface_Prompter(QObject):
    signal = pyqtSignal(str,str,int,int,int) #创建信号,并确定发送参数类型

    def __init__(self):
       super().__init__()  # 调用父类的构造函数
       self.stateTooltip = None # 存储翻译状态控件
       self.total_text_line_count = 0 #存储总文本行数
       self.translated_line_count = 0 #存储已经翻译文本行数
       self.tokens_spent = 0  #存储已经花费的tokens
       self.amount_spent = 0  #存储已经花费的金钱


       self.openai_price_data = {
            "gpt-3.5-turbo": {"input_price": 0.0015, "output_price": 0.002}, # 存储的价格是 /k tokens
            "gpt-3.5-turbo-0301": {"input_price": 0.0015, "output_price": 0.002},
            "gpt-3.5-turbo-0613": {"input_price": 0.0015, "output_price": 0.002},
            "gpt-3.5-turbo-1106": {"input_price": 0.001, "output_price": 0.002},
            "gpt-3.5-turbo-16k": {"input_price": 0.001, "output_price": 0.002},
            "gpt-3.5-turbo-16k-0613": {"input_price": 0.001, "output_price": 0.002},
            "gpt-4": {"input_price": 0.03, "output_price": 0.06},
            "gpt-4-0314": {"input_price": 0.03, "output_price": 0.06},
            "gpt-4-0613": {"input_price": 0.03, "output_price": 0.06},
            "gpt-4-1106-preview":{"input_price": 0.01, "output_price": 0.03},
            "gpt-4-32k": {"input_price": 0.06, "output_price": 0.12},
            "gpt-4-32k-0314": {"input_price": 0.06, "output_price": 0.12},
            "gpt-4-32k-0613": {"input_price": 0.06, "output_price": 0.12},
            "text-embedding-ada-002": {"input_price": 0.0001, "output_price": 0},
            }
       
       self.google_price_data = {
            "gemini-pro": {"input_price": 0.00001, "output_price": 0.00001}, # 存储的价格是 /k tokens
            }

       self.sakura_price_data = {
            "Sakura-13B-LNovel-v0.8": {"input_price": 0.00001, "output_price": 0.00001}, # 存储的价格是 /k tokens
            "Sakura-13B-LNovel-v0.9": {"input_price": 0.00001, "output_price": 0.00001}, # 存储的价格是 /k tokens
            }


    # 槽函数，用于接收子线程发出的信号，更新界面UI的状态，因为子线程不能更改父线程的QT的UI控件的值
    def on_update_ui(self,input_str1,input_str2,iunput_int1,input_int2,input_int3):

        if input_str1 == "翻译状态提示":
            if input_str2 == "开始翻译":
                self.stateTooltip = StateToolTip('正在进行翻译中', '客官请耐心等待哦~~', Window)
                self.stateTooltip.move(510, 30) # 设定控件的出现位置，该位置是传入的Window窗口的位置
                self.stateTooltip.show()
            elif input_str2 == "翻译完成":
                self.stateTooltip.setContent('已经翻译完成啦 😆')
                self.stateTooltip.setState(True)
                self.stateTooltip = None

        elif input_str1 == "初始化翻译界面数据":
            # 更新翻译项目信息
            translation_project = configurator.translation_project
            Window.Widget_start_translation.A_settings.translation_project.setText(translation_project)

            # 更新项目ID信息
            Window.Widget_start_translation.A_settings.project_id.setText(input_str2)

            # 更新需要翻译的文本行数信息
            self.total_text_line_count = iunput_int1 #存储总文本行数
            Window.Widget_start_translation.A_settings.total_text_line_count.setText(str(self.total_text_line_count))

            # 其他信息设置为0
            Window.Widget_start_translation.A_settings.translated_line_count.setText("0")
            Window.Widget_start_translation.A_settings.tokens_spent.setText("0")
            Window.Widget_start_translation.A_settings.amount_spent.setText("0")
            Window.Widget_start_translation.A_settings.progressRing.setValue(0)

            # 初始化存储的数值
            self.translated_line_count = 0 #存储已经翻译文本行数
            self.tokens_spent = 0  #存储已经花费的tokens
            self.amount_spent = 0  #存储已经花费的金钱


        elif input_str1 == "更新翻译界面数据":
            if input_str2 == "翻译成功":
                # 更新已经翻译的文本数
                self.translated_line_count = self.translated_line_count + iunput_int1
                Window.Widget_start_translation.A_settings.translated_line_count.setText(str(self.translated_line_count))

            #更新已经花费的tokens
            self.tokens_spent = self.tokens_spent + input_int2 + input_int3
            Window.Widget_start_translation.A_settings.tokens_spent.setText(str(self.tokens_spent))

            #更新已经花费的金额
            if configurator.translation_platform == "Openai官方":
                # 获取使用的模型输入价格与输出价格
                input_price = self.openai_price_data[configurator.model_type]["input_price"]
                output_price = self.openai_price_data[configurator.model_type]["output_price"]

            elif configurator.translation_platform == "Openai代理":
                # 获取使用的模型输入价格与输出价格
                input_price = Window.Widget_Openai_Proxy.B_settings.spinBox_input_pricing.value()               #获取输入价格
                output_price = Window.Widget_Openai_Proxy.B_settings.spinBox_output_pricing.value()               #获取输出价格

            elif configurator.translation_platform == "Google官方":
                # 获取使用的模型输入价格与输出价格
                input_price = self.google_price_data[configurator.model_type]["input_price"]
                output_price = self.google_price_data[configurator.model_type]["output_price"]

            elif configurator.translation_platform == "SakuraLLM":
                # 获取使用的模型输入价格与输出价格
                input_price = self.sakura_price_data[configurator.model_type]["input_price"]
                output_price = self.sakura_price_data[configurator.model_type]["output_price"]

            self.amount_spent = self.amount_spent + (input_price/1000 * input_int2)  + (output_price/1000 * input_int3) 
            self.amount_spent = round(self.amount_spent, 4)
            Window.Widget_start_translation.A_settings.amount_spent.setText(str(self.amount_spent))

            #更新进度条
            result = self.translated_line_count / self.total_text_line_count * 100
            result = round(result, 0)
            result = int(result)
            Window.Widget_start_translation.A_settings.progressRing.setValue(result)
        
        elif input_str1 == "接口测试结果":
            if input_str2 == "测试成功":
                self.createSuccessInfoBar("全部Apikey请求测试成功")
            else:
                self.createErrorInfoBar("存在Apikey请求测试失败")

    #成功信息居中弹出框函数
    def createSuccessInfoBar(self,str):
        InfoBar.success(
            title='[Success]',
            content=str,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=2000,
            parent=Window
            )

    #错误信息右下方弹出框函数
    def createErrorInfoBar(self,str):
        InfoBar.error(
            title='[Error]',
            content=str,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM_RIGHT,
            duration=-1,    # won't disappear automatically
            parent=Window
            )

    #提醒信息左上角弹出框函数
    def createWarningInfoBar(self,str):
        InfoBar.warning(
            title='[Warning]',
            content=str,
            orient=Qt.Horizontal,
            isClosable=False,   # disable close button
            position=InfoBarPosition.TOP_LEFT,
            duration=2000,
            parent=Window
            )



# ——————————————————————————————————————————下面都是UI相关代码——————————————————————————————————————————
class Widget_Openai(QFrame):#  Openai账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['免费账号',  '付费账号(等级1)',  '付费账号(等级2)',  '付费账号(等级3)',  '付费账号(等级4)',  '付费账号(等级5)'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)



        # -----创建第2个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = ComboBox() #以demo为父类
        self.comboBox_model.addItems(['gpt-3.5-turbo','gpt-3.5-turbo-0301','gpt-3.5-turbo-0613', 'gpt-3.5-turbo-1106','gpt-3.5-turbo-16k', 'gpt-3.5-turbo-16k-0613',
                                 'gpt-4','gpt-4-0314', 'gpt-4-0613','gpt-4-1106-preview'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('gpt-3.5-turbo-0613')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第3个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)


        # -----创建第4个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第3个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


    def test_request(self):
        global Running_status

        if Running_status == 0:
            #创建子线程
            thread = background_executor("openai官方接口测试")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Openai_Proxy(QFrame):  # Openai代理账号主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_Openai_Proxy_A('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_Openai_Proxy_B('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '代理设置')
        self.addSubInterface(self.B_settings, 'B_settings', '其他设置')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_Openai_Proxy_A(QFrame):#  代理账号基础设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第1个组，添加多个组件-----
        box_relay_address = QGroupBox()
        box_relay_address.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_relay_address = QHBoxLayout()

        #设置“中转地址”标签
        self.labelA = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelA.setText("中转请求地址")

        #设置微调距离用的空白标签
        self.labelB = QLabel()  
        self.labelB.setText("                ")

        #设置“中转地址”的输入框
        self.LineEdit_relay_address = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_relay_address.addWidget(self.labelA)
        layout_relay_address.addWidget(self.labelB)
        layout_relay_address.addWidget(self.LineEdit_relay_address)
        box_relay_address.setLayout(layout_relay_address)


        # -----创建第2个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = ComboBox() #以demo为父类
        self.comboBox_model.addItems(['gpt-3.5-turbo','gpt-3.5-turbo-0301','gpt-3.5-turbo-0613', 'gpt-3.5-turbo-1106','gpt-3.5-turbo-16k', 'gpt-3.5-turbo-16k-0613',
                                 'gpt-4','gpt-4-0314', 'gpt-4-0613','gpt-4-1106-preview'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('gpt-3.5-turbo-0613')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第3个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第4个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“系统代理端口”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“系统代理端口”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第5个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_relay_address)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            #创建子线程
            thread = background_executor("openai代理接口测试")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")


class Widget_Openai_Proxy_B(QFrame):#  代理账号进阶设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------




        # -----创建第1个组(后面加的)，添加多个组件-----
        box_RPM = QGroupBox()
        box_RPM.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_RPM = QHBoxLayout()

        #设置“RPM”标签
        self.labelY = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelY.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelY.setText("每分钟请求数")

        #设置“说明”显示
        self.labelA = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelA.setText("(RPM)")  

        #数值输入
        self.spinBox_RPM = SpinBox(self)
        self.spinBox_RPM.setRange(0, 2147483647)    
        self.spinBox_RPM.setValue(3500)


        layout_RPM.addWidget(self.labelY)
        layout_RPM.addWidget(self.labelA)
        layout_RPM.addStretch(1)  # 添加伸缩项
        layout_RPM.addWidget(self.spinBox_RPM)
        box_RPM.setLayout(layout_RPM)



        # -----创建第2个组（后面加的），添加多个组件-----
        box_TPM = QGroupBox()
        box_TPM.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_TPM = QHBoxLayout()

        #设置“TPM”标签
        self.labelB = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelB.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelB.setText("每分钟tokens数")
    
        #设置“说明”显示
        self.labelC = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelC.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelC.setText("(TPM)") 

        #数值输入
        self.spinBox_TPM = SpinBox(self)
        self.spinBox_TPM.setRange(0, 2147483647)    
        self.spinBox_TPM.setValue(60000)


        layout_TPM.addWidget(self.labelB)
        layout_TPM.addWidget(self.labelC)
        layout_TPM.addStretch(1)  # 添加伸缩项
        layout_TPM.addWidget(self.spinBox_TPM)
        box_TPM.setLayout(layout_TPM)


        # -----创建第3个组（后面加的），添加多个组件-----
        box_input_pricing = QGroupBox()
        box_input_pricing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input_pricing = QHBoxLayout()

        #设置“请求输入价格”标签
        self.labelD = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelD.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelD.setText("请求输入价格")
    
        #设置“说明”显示
        self.labelE = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelE.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelE.setText("( /1K tokens)") 

        #数值输入
        self.spinBox_input_pricing = DoubleSpinBox(self)
        self.spinBox_input_pricing.setRange(0.0000, 2147483647)   
        self.spinBox_input_pricing.setDecimals(4)  # 设置小数点后的位数 
        self.spinBox_input_pricing.setValue(0.0015)


        layout_input_pricing.addWidget(self.labelD)
        layout_input_pricing.addWidget(self.labelE)
        layout_input_pricing.addStretch(1)  # 添加伸缩项
        layout_input_pricing.addWidget(self.spinBox_input_pricing)
        box_input_pricing.setLayout(layout_input_pricing)


        # -----创建第4个组（后面加的），添加多个组件-----
        box_output_pricing = QGroupBox()
        box_output_pricing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output_pricing = QHBoxLayout()

        #设置“TPM”标签
        self.labelF = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelF.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelF.setText("回复输出价格")
    
        #设置“说明”显示
        self.labelG = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelG.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelG.setText("( /1K tokens)") 

        #数值输入
        self.spinBox_output_pricing = DoubleSpinBox(self)
        self.spinBox_output_pricing.setRange(0.0000, 2147483647)
        self.spinBox_output_pricing.setDecimals(4)  # 设置小数点后的位数     
        self.spinBox_output_pricing.setValue(0.002)
        


        layout_output_pricing.addWidget(self.labelF)
        layout_output_pricing.addWidget(self.labelG)
        layout_output_pricing.addStretch(1)  # 添加伸缩项
        layout_output_pricing.addWidget(self.spinBox_output_pricing)
        box_output_pricing.setLayout(layout_output_pricing)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_RPM)
        container.addWidget(box_TPM)
        container.addWidget(box_input_pricing)
        container.addWidget(box_output_pricing)
        container.addStretch(1)  # 添加伸缩项



class Widget_Google(QFrame):#  谷歌账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = ComboBox() #以demo为父类
        self.comboBox_model.addItems(['gemini-pro'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('gemini-pro')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            #创建子线程
            thread = background_executor("google官方接口测试")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_SakuraLLM(QFrame):#  SakuraLLM界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第1个组，添加多个组件-----
        box_address = QGroupBox()
        box_address.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_address = QHBoxLayout()

        #设置“请求地址”标签
        self.labelA = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelA.setText("请求地址")

        #设置微调距离用的空白标签
        self.labelB = QLabel()  
        self.labelB.setText("                      ")

        #设置“请求地址”的输入框
        self.LineEdit_address = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_address.addWidget(self.labelA)
        layout_address.addWidget(self.labelB)
        layout_address.addWidget(self.LineEdit_address)
        box_address.setLayout(layout_address)


        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = ComboBox() #以demo为父类
        self.comboBox_model.addItems(['Sakura-13B-LNovel-v0.8','Sakura-13B-LNovel-v0.9'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('Sakura-13B-LNovel-v0.8')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_address)
        container.addWidget(box_model)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            #创建子线程
            thread = background_executor("Sakura通讯测试")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")




class Widget_translation_settings(QFrame):  # 翻译设置主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_translation_settings_A('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_translation_settings_B('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '基础设置')
        self.addSubInterface(self.B_settings, 'B_settings', '进阶设置')


        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_translation_settings_A(QFrame):#  基础设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第1个组，添加多个组件-----
        box_translation_platform = QGroupBox()
        box_translation_platform.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_platform = QGridLayout()

        #设置“翻译平台”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("翻译平台")


        #设置“翻译平台”下拉选择框
        self.comboBox_translation_platform = ComboBox() #以demo为父类
        self.comboBox_translation_platform.addItems(['Openai官方',  'Openai代理',  'Google官方',  'SakuraLLM'])
        self.comboBox_translation_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translation_platform.setFixedSize(150, 35)


        layout_translation_platform.addWidget(self.labelx, 0, 0)
        layout_translation_platform.addWidget(self.comboBox_translation_platform, 0, 1)
        box_translation_platform.setLayout(layout_translation_platform)


        # -----创建第1个组，添加多个组件-----
        box_translation_project = QGroupBox()
        box_translation_project.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_project = QGridLayout()

        #设置“翻译项目”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("翻译项目")


        #设置“翻译项目”下拉选择框
        self.comboBox_translation_project = ComboBox() #以demo为父类
        self.comboBox_translation_project.addItems(['Mtool导出文件',  'T++导出文件',  'Ainiee缓存文件'])
        self.comboBox_translation_project.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translation_project.setFixedSize(150, 35)


        layout_translation_project.addWidget(self.labelx, 0, 0)
        layout_translation_project.addWidget(self.comboBox_translation_project, 0, 1)
        box_translation_project.setLayout(layout_translation_project)


        # -----创建第2个组，添加多个组件-----
        box_input = QGroupBox()
        box_input.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("输入文件夹")

        #设置“输入文件夹”显示
        self.label_input_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_input_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_input_path.setText("(请选择原文文件所在的文件夹，不要混杂其他文件)")  

        #设置打开文件按钮
        self.pushButton_input = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_input.clicked.connect(File_Reader.Select_project_folder) #按钮绑定槽函数



        layout_input.addWidget(label4)
        layout_input.addWidget(self.label_input_path)
        layout_input.addStretch(1)  # 添加伸缩项
        layout_input.addWidget(self.pushButton_input)
        box_input.setLayout(layout_input)


        # -----创建第3个组，添加多个组件-----
        box_output = QGroupBox()
        box_output.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output = QHBoxLayout()

        #设置“输出文件夹”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("输出文件夹")

        #设置“输出文件夹”显示
        self.label_output_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_output_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_output_path.setText("(请选择翻译文件存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_output = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_output.clicked.connect(File_Reader.Select_output_folder) #按钮绑定槽函数


        

        layout_output.addWidget(label6)
        layout_output.addWidget(self.label_output_path)
        layout_output.addStretch(1)  # 添加伸缩项
        layout_output.addWidget(self.pushButton_output)
        box_output.setLayout(layout_output)





        # -----创建第4个组，添加多个组件-----
        box_source_text = QGroupBox()
        box_source_text.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_source_text = QHBoxLayout()


        #设置“文本源语言”标签
        label3 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label3.setText("文本源语言")

        #设置“文本源语言”下拉选择框
        self.comboBox_source_text = ComboBox() #以demo为父类
        self.comboBox_source_text.addItems(['日语', '英语', '韩语', '俄语', '简中', '繁中'])
        self.comboBox_source_text.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_source_text.setFixedSize(127, 30)


        layout_source_text.addWidget(label3)
        layout_source_text.addWidget(self.comboBox_source_text)
        box_source_text.setLayout(layout_source_text)


        # -----创建第5个组(后面添加的)，添加多个组件-----
        box_translated_text = QGroupBox()
        box_translated_text.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translated_text = QHBoxLayout()


        #设置“文本目标语言”标签
        label3_1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label3_1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label3_1.setText("文本目标语言")

        #设置“文本目标语言”下拉选择框
        self.comboBox_translated_text = ComboBox() #以demo为父类
        self.comboBox_translated_text.addItems(['简中', '繁中', '日语', '英语', '韩语'])
        self.comboBox_translated_text.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translated_text.setFixedSize(127, 30)


        layout_translated_text.addWidget(label3_1)
        layout_translated_text.addWidget(self.comboBox_translated_text)
        box_translated_text.setLayout(layout_translated_text)


        # -----创建第6个组，添加多个组件-----
        box_save = QGroupBox()
        box_save.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_save = QHBoxLayout()

        #设置“保存配置”的按钮
        self.primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        self.primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数



        layout_save.addStretch(1)  # 添加伸缩项
        layout_save.addWidget(self.primaryButton_save)
        layout_save.addStretch(1)  # 添加伸缩项
        box_save.setLayout(layout_save)




        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_translation_project)
        container.addWidget(box_translation_platform)
        container.addWidget(box_source_text)
        container.addWidget(box_translated_text)
        container.addWidget(box_input)
        container.addWidget(box_output)
        container.addWidget(box_save)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下



    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


class Widget_translation_settings_B(QFrame):#  进阶设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_Lines = QGroupBox()
        box_Lines.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_Lines = QHBoxLayout()

        #设置“翻译行数”标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label1.setText("每次翻译行数")


       #设置“翻译行数”数值输入框
        self.spinBox_Lines = SpinBox(self)
        self.spinBox_Lines.setRange(1, 1000)    
        self.spinBox_Lines.setValue(30)


        layout_Lines.addWidget(label1)
        layout_Lines.addStretch(1)  # 添加伸缩项
        layout_Lines.addWidget(self.spinBox_Lines)
        box_Lines.setLayout(layout_Lines)


        # -----创建第1.4个组(后来补的)，添加多个组件-----
        box1_jsonmode = QGroupBox()
        box1_jsonmode.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_jsonmode = QHBoxLayout()

        #设置“回复json格式”标签
        labe1_4 = QLabel(flags=Qt.WindowFlags())  
        labe1_4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_4.setText("回复json格式")



       #设置“回复json格式”选择开关
        self.SwitchButton_jsonmode = SwitchButton(parent=self)    
        self.SwitchButton_jsonmode.checkedChanged.connect(self.onjsonmode)



        layout1_jsonmode.addWidget(labe1_4)
        layout1_jsonmode.addStretch(1)  # 添加伸缩项
        layout1_jsonmode.addWidget(self.SwitchButton_jsonmode)
        box1_jsonmode.setLayout(layout1_jsonmode)




        # -----创建第1.6个组(后来补的)，添加多个组件-----
        box1_conversion_toggle = QGroupBox()
        box1_conversion_toggle.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_conversion_toggle = QHBoxLayout()

        #设置“简繁转换开关”标签
        labe1_6 = QLabel(flags=Qt.WindowFlags())  
        labe1_6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_6.setText("简繁体自动转换")

       #设置“简繁体自动转换”选择开关
        self.SwitchButton_conversion_toggle = SwitchButton(parent=self)    



        layout1_conversion_toggle.addWidget(labe1_6)
        layout1_conversion_toggle.addStretch(1)  # 添加伸缩项
        layout1_conversion_toggle.addWidget(self.SwitchButton_conversion_toggle)
        box1_conversion_toggle.setLayout(layout1_conversion_toggle)



        # -----创建第1.6个组(后来补的)，添加多个组件-----
        box1_line_breaks = QGroupBox()
        box1_line_breaks.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_line_breaks = QHBoxLayout()

        #设置“换行符保留”标签
        labe1_6 = QLabel(flags=Qt.WindowFlags())  
        labe1_6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_6.setText("保留换行符")

       #设置“换行符保留”选择开关
        self.SwitchButton_line_breaks = SwitchButton(parent=self)    



        layout1_line_breaks.addWidget(labe1_6)
        layout1_line_breaks.addStretch(1)  # 添加伸缩项
        layout1_line_breaks.addWidget(self.SwitchButton_line_breaks)
        box1_line_breaks.setLayout(layout1_line_breaks)




        # -----创建第1.7个组(后来补的)，添加多个组件-----
        box1_thread_count = QGroupBox()
        box1_thread_count.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_thread_count = QHBoxLayout()

        #设置“最大线程数”标签
        label1_7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1_7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label1_7.setText("最大线程数")

        #设置“说明”显示
        label2_7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label2_7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        label2_7.setText("(0是自动根据电脑设置线程数)")  

       #设置“最大线程数”数值输入框
        self.spinBox_thread_count = SpinBox(self)
        #设置最大最小值
        self.spinBox_thread_count.setRange(0, 1000)    
        self.spinBox_thread_count.setValue(0)

        layout1_thread_count.addWidget(label1_7)
        layout1_thread_count.addWidget(label2_7)
        layout1_thread_count.addStretch(1)  # 添加伸缩项
        layout1_thread_count.addWidget(self.spinBox_thread_count)
        box1_thread_count.setLayout(layout1_thread_count)



        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_Lines)
        container.addWidget(box1_line_breaks)
        container.addWidget(box1_jsonmode)
        container.addWidget(box1_conversion_toggle)
        container.addWidget(box1_thread_count)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

    #设置“回复json格式”选择开关绑定函数
    def onjsonmode(self, isChecked: bool):
        if isChecked:
            user_interface_prompter.createWarningInfoBar("该设置现在仅支持openai接口的gpt-3.5-turbo-1106与gpt-4-1106-preview模型开启")




class Widget_start_translation(QFrame):  # 开始翻译主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_start_translation_A('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_start_translation_B('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '开始翻译')
        self.addSubInterface(self.B_settings, 'B_settings', '备份功能')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_start_translation_A(QFrame):#  开始翻译子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_project = QGroupBox()
        box_project.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")  # 分别设置了边框大小，边框颜色，边框圆角
        layout_project = QHBoxLayout()

        # 第一组水平布局
        layout_horizontal_1 = QHBoxLayout()

        self.label111 = QLabel(flags=Qt.WindowFlags())
        self.label111.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label111.setText("项目类型 :")

        self.translation_project = QLabel(flags=Qt.WindowFlags())
        self.translation_project.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.translation_project.setText("无")

        layout_horizontal_1.addWidget(self.label111)
        layout_horizontal_1.addStretch(1)  # 添加伸缩项
        layout_horizontal_1.addWidget(self.translation_project)
        layout_horizontal_1.addStretch(1)  # 添加伸缩项

        # 第二组水平布局
        layout_horizontal_2 = QHBoxLayout()

        self.label222 = QLabel(flags=Qt.WindowFlags())
        self.label222.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label222.setText("项目ID :")

        self.project_id = QLabel(flags=Qt.WindowFlags())
        self.project_id.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.project_id.setText("无")

        layout_horizontal_2.addWidget(self.label222)
        layout_horizontal_2.addStretch(1)  # 添加伸缩项
        layout_horizontal_2.addWidget(self.project_id)
        layout_horizontal_2.addStretch(1)  # 添加伸缩项

        # 将两个水平布局放入最外层水平布局
        layout_project.addLayout(layout_horizontal_1)
        layout_project.addLayout(layout_horizontal_2)

        box_project.setLayout(layout_project)


        # -----创建第2个组，添加多个组件-----
        box_text_line_count = QGroupBox()
        box_text_line_count.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")  # 分别设置了边框大小，边框颜色，边框圆角
        layout_text_line_count = QHBoxLayout()

        # 第三组水平布局
        layout_horizontal_3 = QHBoxLayout()

        self.label333 = QLabel(flags=Qt.WindowFlags())
        self.label333.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label333.setText("总文本行数 :")

        self.total_text_line_count = QLabel(flags=Qt.WindowFlags())
        self.total_text_line_count.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.total_text_line_count.setText("无")

        layout_horizontal_3.addWidget(self.label333)
        layout_horizontal_3.addStretch(1)  # 添加伸缩项
        layout_horizontal_3.addWidget(self.total_text_line_count)
        layout_horizontal_3.addStretch(1)  # 添加伸缩项

        # 第四组水平布局
        layout_horizontal_4 = QHBoxLayout()

        self.label444 = QLabel(flags=Qt.WindowFlags())
        self.label444.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label444.setText("已翻译行数 :")

        self.translated_line_count = QLabel(flags=Qt.WindowFlags())
        self.translated_line_count.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.translated_line_count.setText("无")

        layout_horizontal_4.addWidget(self.label444)
        layout_horizontal_4.addStretch(1)  # 添加伸缩项
        layout_horizontal_4.addWidget(self.translated_line_count)
        layout_horizontal_4.addStretch(1)  # 添加伸缩项

        # 将第三组和第四组水平布局放入最外层水平布局
        layout_text_line_count.addLayout(layout_horizontal_3)
        layout_text_line_count.addLayout(layout_horizontal_4)

        box_text_line_count.setLayout(layout_text_line_count)





        # -----创建第3个组，添加多个组件-----
        box_spent = QGroupBox()
        box_spent.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")  # 分别设置了边框大小，边框颜色，边框圆角
        layout_spent = QHBoxLayout()

        # 第五组水平布局
        layout_horizontal_5 = QHBoxLayout()

        self.labelx1 = QLabel(flags=Qt.WindowFlags())
        self.labelx1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.labelx1.setText("已花费tokens :")

        self.tokens_spent = QLabel(flags=Qt.WindowFlags())
        self.tokens_spent.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.tokens_spent.setText("无")

        layout_horizontal_5.addWidget(self.labelx1)
        layout_horizontal_5.addStretch(1)  # 添加伸缩项
        layout_horizontal_5.addWidget(self.tokens_spent)
        layout_horizontal_5.addStretch(1)  # 添加伸缩项

        # 第六组水平布局
        layout_horizontal_6 = QHBoxLayout()

        self.labelx2 = QLabel(flags=Qt.WindowFlags())
        self.labelx2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.labelx2.setText("已花费金额(＄) :")

        self.amount_spent = QLabel(flags=Qt.WindowFlags())
        self.amount_spent.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.amount_spent.setText("无")

        layout_horizontal_6.addWidget(self.labelx2)
        layout_horizontal_6.addStretch(1)  # 添加伸缩项
        layout_horizontal_6.addWidget(self.amount_spent)
        layout_horizontal_6.addStretch(1)  # 添加伸缩项

        # 将第五组和第六组水平布局放入最外层水平布局
        layout_spent.addLayout(layout_horizontal_5)
        layout_spent.addLayout(layout_horizontal_6)

        box_spent.setLayout(layout_spent)



        # -----创建第4个组，添加多个组件-----
        box_progressRing = QGroupBox()
        box_progressRing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_progressRing = QHBoxLayout()


        #设置“翻译进度”标签
        self.label_progressRing = QLabel( flags=Qt.WindowFlags())  
        self.label_progressRing.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.label_progressRing.setText("翻译进度")

        #设置翻译进度条
        self.progressRing = ProgressRing(self)
        self.progressRing.setValue(0)
        self.progressRing.setTextVisible(True)
        self.progressRing.setFixedSize(80, 80)


        layout_progressRing.addWidget(self.label_progressRing)
        layout_progressRing.addStretch(1)  # 添加伸缩项
        layout_progressRing.addWidget(self.progressRing)
        box_progressRing.setLayout(layout_progressRing)





        # -----创建第5个组，添加多个组件-----
        box_start_translation = QGroupBox()
        box_start_translation.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_start_translation = QHBoxLayout()


        #设置“开始翻译”的按钮
        self.primaryButton_start_translation = PrimaryPushButton('开始翻译', self, FIF.UPDATE)
        self.primaryButton_start_translation.clicked.connect(self.Start_translation_mtool) #按钮绑定槽函数


        layout_start_translation.addStretch(1)  # 添加伸缩项
        layout_start_translation.addWidget(self.primaryButton_start_translation)
        layout_start_translation.addStretch(1)  # 添加伸缩项
        box_start_translation.setLayout(layout_start_translation)


        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_project)
        container.addWidget(box_text_line_count)
        container.addWidget(box_spent)
        container.addWidget(box_progressRing)
        container.addWidget(box_start_translation)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下



    #开始翻译按钮绑定函数
    def Start_translation_mtool(self):
        global Running_status

        if Running_status == 0:
            #创建子线程
            thread = background_executor("执行翻译任务")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")


class Widget_start_translation_B(QFrame):#  开始翻译子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_switch = QGroupBox()
        box_switch.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_switch = QHBoxLayout()


        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("自动备份缓存文件到输出文件夹")


        self.checkBox_switch = CheckBox('启用功能')
        self.checkBox_switch.stateChanged.connect(self.checkBoxChanged1)

        layout_switch.addWidget(label1)
        layout_switch.addStretch(1)  # 添加伸缩项
        layout_switch.addWidget(self.checkBox_switch)
        box_switch.setLayout(layout_switch)



        # -----创建第1个组，添加多个组件-----
        box_export_cache_file_path = QGroupBox()
        box_export_cache_file_path.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_export_cache_file_path = QHBoxLayout()

        #设置“导出当前任务的缓存文件”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("导出当前任务的缓存文件")


        #设置导出当前任务的缓存文件按钮
        self.pushButton_export_cache_file_path = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_export_cache_file_path.clicked.connect(self.output_cachedata) #按钮绑定槽函数



        layout_export_cache_file_path.addWidget(label4)
        layout_export_cache_file_path.addStretch(1)  # 添加伸缩项
        layout_export_cache_file_path.addWidget(self.pushButton_export_cache_file_path)
        box_export_cache_file_path.setLayout(layout_export_cache_file_path)


        # -----创建第2个组，添加多个组件-----
        box_export_translated_file_path = QGroupBox()
        box_export_translated_file_path.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_export_translated_file_path = QHBoxLayout()

        #设置“导出当前任务的已翻译文本”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("导出当前任务的已翻译文本")


        #设置导出当前任务的已翻译文本按钮
        self.pushButton_export_translated_file_path = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_export_translated_file_path.clicked.connect(self.output_data) #按钮绑定槽函数


        

        layout_export_translated_file_path.addWidget(label6)
        layout_export_translated_file_path.addStretch(1)  # 添加伸缩项
        layout_export_translated_file_path.addWidget(self.pushButton_export_translated_file_path)
        box_export_translated_file_path.setLayout(layout_export_translated_file_path)







        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项        
        container.addWidget(box_switch)
        container.addWidget(box_export_cache_file_path)
        container.addWidget(box_export_translated_file_path)
        container.addStretch(1)  # 添加伸缩项

    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启自动备份功能")

    # 缓存文件输出
    def output_cachedata(self):
        global cache_list

        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)

            if len(cache_list)>= 3:
                File_Outputter.output_cache_file(self,cache_list,Output_Folder)
                user_interface_prompter.createSuccessInfoBar("已输出缓存文件")
                print('[INFO]  已输出缓存文件')
            else:
                print('[INFO]  未存在缓存文件')
                return  # 直接返回，不执行后续操作
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 缓存文件输出
    def output_data(self):
        global cache_list

        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)

            if len(cache_list)>= 3:
                File_Outputter.output_translated_content(self,cache_list,Output_Folder)
                user_interface_prompter.createSuccessInfoBar("已输出已翻译文件")
                print('[INFO]  已输出已翻译文件')
            else:
                print('[INFO]  未存在缓存文件')
                return  # 直接返回，不执行后续操作
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作



class Widget_check(QFrame):# 错行检查界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第0-1个组，添加多个组件-----
        box_weight = QGroupBox()
        box_weight.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_weight = QHBoxLayout()

        #设置“语义权重”标签
        label0_1 = QLabel( flags=Qt.WindowFlags())  
        label0_1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label0_1.setText("语义权重")

        #设置“语义权重”输入
        self.doubleSpinBox_semantic_weight = DoubleSpinBox(self)
        self.doubleSpinBox_semantic_weight.setMaximum(1.0)
        self.doubleSpinBox_semantic_weight.setMinimum(0.0)
        self.doubleSpinBox_semantic_weight.setValue(0.6)

        #设置“符号权重”标签
        label0_2 = QLabel( flags=Qt.WindowFlags())  
        label0_2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label0_2.setText("符号权重")

        #设置“符号权重”输入
        self.doubleSpinBox_symbol_weight = DoubleSpinBox(self)
        self.doubleSpinBox_symbol_weight.setMaximum(1.0)
        self.doubleSpinBox_symbol_weight.setMinimum(0.0)
        self.doubleSpinBox_symbol_weight.setValue(0.2)

        #设置“字数权重”标签
        label0_3 = QLabel( flags=Qt.WindowFlags())  
        label0_3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label0_3.setText("字数权重")

        #设置“字数权重”输入
        self.doubleSpinBox_word_count_weight = DoubleSpinBox(self)
        self.doubleSpinBox_word_count_weight.setMaximum(1.0)
        self.doubleSpinBox_word_count_weight.setMinimum(0.0)
        self.doubleSpinBox_word_count_weight.setValue(0.2)


        layout_weight.addWidget(label0_1)
        layout_weight.addWidget(self.doubleSpinBox_semantic_weight)
        layout_weight.addStretch(1)  # 添加伸缩项
        layout_weight.addWidget(label0_2)
        layout_weight.addWidget(self.doubleSpinBox_symbol_weight)
        layout_weight.addStretch(1)  # 添加伸缩项
        layout_weight.addWidget(label0_3)
        layout_weight.addWidget(self.doubleSpinBox_word_count_weight)

        box_weight.setLayout(layout_weight)


        # -----创建第0-2个组，添加多个组件-----
        box_similarity_threshold = QGroupBox()
        box_similarity_threshold.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_similarity_threshold = QHBoxLayout()

        #设置“相似度阈值”标签
        label0_4 = QLabel( flags=Qt.WindowFlags())  
        label0_4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label0_4.setText("相似度阈值")

        #设置“相似度阈值”输入
        self.spinBox_similarity_threshold = SpinBox(self)
        self.spinBox_similarity_threshold.setMaximum(100)
        self.spinBox_similarity_threshold.setMinimum(0)
        self.spinBox_similarity_threshold.setValue(50)

        layout_similarity_threshold.addWidget(label0_4)
        layout_similarity_threshold.addStretch(1)  # 添加伸缩项
        layout_similarity_threshold.addWidget(self.spinBox_similarity_threshold)
        box_similarity_threshold.setLayout(layout_similarity_threshold)


        # -----创建第1个组，添加多个组件-----
        box_translation_platform = QGroupBox()
        box_translation_platform.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_platform = QGridLayout()

        #设置“翻译平台”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("重翻平台")


        #设置“翻译平台”下拉选择框
        self.comboBox_translation_platform = ComboBox() #以demo为父类
        self.comboBox_translation_platform.addItems(['Openai官方',  'Openai代理'])
        self.comboBox_translation_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translation_platform.setFixedSize(150, 35)


        layout_translation_platform.addWidget(self.labelx, 0, 0)
        layout_translation_platform.addWidget(self.comboBox_translation_platform, 0, 1)
        box_translation_platform.setLayout(layout_translation_platform)


        # -----创建第1个组，添加多个组件-----
        box_translation_project = QGroupBox()
        box_translation_project.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_project = QGridLayout()

        #设置“翻译项目”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("检查项目")


        #设置“翻译项目”下拉选择框
        self.comboBox_translation_project = ComboBox() #以demo为父类
        self.comboBox_translation_project.addItems(['Mtool导出文件',  'T++导出文件'])
        self.comboBox_translation_project.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translation_project.setFixedSize(150, 35)


        layout_translation_project.addWidget(self.labelx, 0, 0)
        layout_translation_project.addWidget(self.comboBox_translation_project, 0, 1)
        box_translation_project.setLayout(layout_translation_project)


        # -----创建第2个组，添加多个组件-----
        box_input = QGroupBox()
        box_input.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("输入文件夹")

        #设置“输入文件夹”显示
        self.label_input_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_input_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_input_path.setText("(请选择已翻译文件所在的文件夹)")  

        #设置打开文件按钮
        self.pushButton_input = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_input.clicked.connect(File_Reader.Select_project_folder_check) #按钮绑定槽函数



        layout_input.addWidget(label4)
        layout_input.addWidget(self.label_input_path)
        layout_input.addStretch(1)  # 添加伸缩项
        layout_input.addWidget(self.pushButton_input)
        box_input.setLayout(layout_input)


        # -----创建第3个组，添加多个组件-----
        box_output = QGroupBox()
        box_output.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output = QHBoxLayout()

        #设置“输出文件夹”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("输出文件夹")

        #设置“输出文件夹”显示
        self.label_output_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_output_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_output_path.setText("(请选择检查重翻后文件存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_output = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_output.clicked.connect(File_Reader.Select_output_folder_check) #按钮绑定槽函数


        

        layout_output.addWidget(label6)
        layout_output.addWidget(self.label_output_path)
        layout_output.addStretch(1)  # 添加伸缩项
        layout_output.addWidget(self.pushButton_output)
        box_output.setLayout(layout_output)





        # -----创建第3个组，添加多个组件-----
        box_check = QGroupBox()
        box_check.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_check = QHBoxLayout()

        #设置“保存配置”的按钮
        self.primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        self.primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        #设置“开始检查”的按钮
        self.primaryButton1 = PrimaryPushButton('开始检查错行', self, FIF.UPDATE)
        self.primaryButton1.clicked.connect(self.Start_check) #按钮绑定槽函数
        

        layout_check.addStretch(1)  # 添加伸缩项
        layout_check.addWidget(self.primaryButton_save) 
        layout_check.addStretch(1)  # 添加伸缩项
        layout_check.addWidget(self.primaryButton1)
        layout_check.addStretch(1)  # 添加伸缩项
        box_check.setLayout(layout_check)


        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_weight)
        container.addWidget(box_similarity_threshold)
        container.addWidget(box_translation_platform)
        container.addWidget(box_translation_project)
        container.addWidget(box_input)
        container.addWidget(box_output)
        container.addWidget(box_check)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


    #开始翻译按钮绑定函数
    def Start_check(self):
        global Running_status

        if Running_status == 0:
            #创建子线程
            thread = background_executor("执行检查任务")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")




class Widget18(QFrame):#AI实时调教界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        # 最外层的垂直布局
        container = QVBoxLayout()


        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        #设置“启用实时参数”标签
        label0 = QLabel(flags=Qt.WindowFlags())  
        label0.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label0.setText("启用调教功能（OpenAI）")

        #设置官方文档说明链接按钮
        hyperlinkButton = HyperlinkButton(
            url='https://platform.openai.com/docs/api-reference/chat/create',
            text='(官方文档)'
        )

        #设置“启用实时参数”开关
        self.checkBox = CheckBox('实时设置AI参数', self)
        self.checkBox.stateChanged.connect(self.checkBoxChanged)


        layout1.addWidget(label0)
        layout1.addWidget(hyperlinkButton)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox)
        box1.setLayout(layout1)



        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        #设置“官方文档”标签
        label01 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label01.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label01.setText("官方文档说明")

        #设置官方文档说明链接按钮
        pushButton1 = PushButton('文档链接', self)
        pushButton1.clicked.connect(lambda: QDesktopServices.openUrl(QUrl('https://platform.openai.com/docs/api-reference/chat/create')))


        layout2.addWidget(label01)
        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(pushButton1)
        box2.setLayout(layout2)



        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“温度”标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label1.setText("温度")

        #设置“温度”副标签
        label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label11.setText("(官方默认值为1)")

        #设置“温度”滑动条
        self.slider1 = Slider(Qt.Horizontal, self)
        self.slider1.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label2 = QLabel(str(self.slider1.value()), self)
        self.label2.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider1.valueChanged.connect(lambda value: self.label2.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label2显示正确的值
        self.slider1.setMinimum(0)
        self.slider1.setMaximum(20)
        self.slider1.setValue(0)

        

        layout3.addWidget(label1)
        layout3.addWidget(label11)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.slider1)
        layout3.addWidget(self.label2)
        box3.setLayout(layout3)





        # -----创建第4个组，添加多个组件-----
        box4 = QGroupBox()
        box4.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout4 = QHBoxLayout()


        #设置“温度”说明文档
        label3 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px;  color: black")
        label3.setText("Temperature：控制结果的随机性。如果希望结果更有创意可以尝试 0.9，或者希望有固定结果可以尝试0.0\n官方建议不要与Top_p一同改变 ")


        layout4.addWidget(label3)
        box4.setLayout(layout4)




        # -----创建第5个组，添加多个组件-----
        box5 = QGroupBox()
        box5.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout5 = QHBoxLayout()

        #设置“top_p”标签
        label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label4.setText("概率阈值")

        #设置“top_p”副标签
        label41 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label41.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label41.setText("(官方默认值为1)")


        #设置“top_p”滑动条
        self.slider2 = Slider(Qt.Horizontal, self)
        self.slider2.setFixedWidth(200)


        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label5 = QLabel(str(self.slider2.value()), self)
        self.label5.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider2.valueChanged.connect(lambda value: self.label5.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放在后面是为了让上面的label5显示正确的值和格式
        self.slider2.setMinimum(0)
        self.slider2.setMaximum(10)
        self.slider2.setValue(10)



        layout5.addWidget(label4)
        layout5.addWidget(label41)
        layout5.addStretch(1)  # 添加伸缩项
        layout5.addWidget(self.slider2)
        layout5.addWidget(self.label5)
        box5.setLayout(layout5)




        # -----创建第6个组，添加多个组件-----
        box6 = QGroupBox()
        box6.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout6 = QHBoxLayout()


        #设置“top_p”说明文档
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px;  color: black")
        label6.setText("Top_p：用于控制生成文本的多样性，与Temperature的作用相同。如果希望结果更加多样可以尝试 0.9\n或者希望有固定结果可以尝试 0.0。官方建议不要与Temperature一同改变 ")



        layout6.addWidget(label6)
        box6.setLayout(layout6)





        # -----创建第7个组，添加多个组件-----
        box7 = QGroupBox()
        box7.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout7 = QHBoxLayout()

        #设置“presence_penalty”标签
        label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label7.setText("主题惩罚")

        #设置“presence_penalty”副标签
        label71 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label71.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label71.setText("(官方默认值为0)")


        #设置“presence_penalty”滑动条
        self.slider3 = Slider(Qt.Horizontal, self)
        self.slider3.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label8 = QLabel(str(self.slider3.value()), self)
        self.label8.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label8.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider3.valueChanged.connect(lambda value: self.label8.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label8显示正确的值和格式
        self.slider3.setMinimum(-20)
        self.slider3.setMaximum(20)
        self.slider3.setValue(5)



        layout7.addWidget(label7)
        layout7.addWidget(label71)
        layout7.addStretch(1)  # 添加伸缩项
        layout7.addWidget(self.slider3)
        layout7.addWidget(self.label8)
        box7.setLayout(layout7)



        # -----创建第8个组，添加多个组件-----
        box8 = QGroupBox()
        box8.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout8 = QHBoxLayout()


        #设置“presence_penalty”说明文档
        label82 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label82.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px;  color: black")
        label82.setText("Presence_penalty：用于控制主题的重复度。AI生成新内容时，会根据到目前为止已经出现在文本中的语句  \n负值是增加生成的新内容，正值是减少生成的新内容，从而改变AI模型谈论新主题内容的可能性")


        layout8.addWidget(label82)
        box8.setLayout(layout8)




        # -----创建第9个组，添加多个组件-----
        box9 = QGroupBox()
        box9.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout9 = QHBoxLayout()

        #设置“frequency_penalty”标签
        label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label9.setText("频率惩罚")

        #设置“presence_penalty”副标签
        label91 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label91.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label91.setText("(官方默认值为0)")

        #设置“frequency_penalty”滑动条
        self.slider4 = Slider(Qt.Horizontal, self)
        self.slider4.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label10 = QLabel(str(self.slider4.value()), self)
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label10.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.slider4.valueChanged.connect(lambda value: self.label10.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label10显示正确的值和格式
        self.slider4.setMinimum(-20)
        self.slider4.setMaximum(20)
        self.slider4.setValue(0)


        layout9.addWidget(label9)
        layout9.addWidget(label91)
        layout9.addStretch(1)  # 添加伸缩项
        layout9.addWidget(self.slider4)
        layout9.addWidget(self.label10)
        box9.setLayout(layout9)


        # -----创建第10个组，添加多个组件-----
        box10 = QGroupBox()
        box10.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout10 = QHBoxLayout()


        #设置“frequency_penalty”说明文档
        label11 = QLabel(parent=self, flags=Qt.WindowFlags())
        label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 14px;  color: black")
        label11.setText("Frequency_penalty：用于控制词语的频率。AI在生成新词时，会根据该词在文本中的现有频率 \n负值进行奖励，增加出现频率；正值进行惩罚，降低出现频率；以便增加或降低逐字重复同一行的可能性")


        layout10.addWidget(label11)
        box10.setLayout(layout10)






        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        #container.addWidget(box2)
        container.addWidget(box3)
        container.addWidget(box4)
        container.addWidget(box5)
        container.addWidget(box6)
        container.addWidget(box7)
        container.addWidget(box8)
        container.addWidget(box9)
        container.addWidget(box10)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(20) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

    # 勾选事件
    def checkBoxChanged(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已启用实时调教功能")


class Widget21(QFrame):#原文替换字典界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # self.scrollWidget = QWidget() #创建滚动窗口
        # #self.scrollWidget.resize(500, 400)    #设置滚动窗口大小
        # #self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff) #设置水平滚动条不可见
        # self.setViewportMargins(0, 0, 0, 0)   #设置滚动窗口的边距      
        # self.setWidget(self.scrollWidget)  #设置滚动窗口的内容  
        # self.setWidgetResizable(True)   #设置滚动窗口的内容可调整大小
        # self.verticalScrollBar().sliderPressed.connect(self.scrollContents) #滚动条滚动时，调用的函数
        #设置各个控件-----------------------------------------------------------------------------------------

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['原文', '译文']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序


        # songInfos = [
        #     ['かばん', 'aiko']
        # ]
        # for i, songInfo in enumerate(songInfos): #遍历数据
        #     for j in range(2): #遍历每一列
        #         self.tableView.setItem(i, j, QTableWidgetItem(songInfo[j])) #设置每个单元格的内容


        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        #设置“译前替换”标签
        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("原文本替换翻译")

        #设置“译前替换”显示
        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(在翻译开始前，根据表格中内容，将游戏文本中出现的原文全部替换成为译文，再发送过去翻译)")


        #设置“译前替换”开
        self.checkBox1 = CheckBox('启用功能')
        self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout2.addWidget(label1)
        layout2.addWidget(self.label2)
        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.checkBox1)
        box2.setLayout(layout2)


        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“译时提示”标签
        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("译时提示")

        #设置“译时提示”显示
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(在翻译进行时，如果该次游戏文本出现了部分字典原文，则会将这部分字典内容作为AI的翻译示例，一并发过去翻译)")


        #设置“译时提示”开
        self.checkBox2 = CheckBox('启用功能')
        #self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox2)
        box3.setLayout(layout3)


        # 把内容添加到容器中 
        container.addWidget(box2)   
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(20)     
        container.setContentsMargins(50, 70, 50, 30)      

    #滚动条滚动时，调用的函数
    def scrollContents(self, position):
        self.scrollWidget.move(0, position) 

    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'JSON Files (*.json)')      #调用QFileDialog类里的函数来选择文件
        if Input_File:
            print(f'[INFO]  已选择字典导入文件: {Input_File}')
        else :
            print('[INFO]  未选择文件')
            return
        
        # 读取文件
        with open(Input_File, 'r', encoding="utf-8") as f:
            dictionary = json.load(f)
        
        # 将字典中的数据从表格底部添加到表格中
        for key, value in dictionary.items():
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(key))
            self.tableView.setItem(row, 1, QTableWidgetItem(value))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

        user_interface_prompter.createSuccessInfoBar("导入成功")
        print(f'[INFO]  已导入字典文件')
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户替换字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        configurator.read_write_config("write") 
        user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')


    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启译前替换功能，将依据表格内容进行替换")
    

class Widget22(QFrame):#提示词工程界面


    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("修改系统提示词")


        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(将修改系统提示词Prompt为输入框中的内容)")


        self.checkBox1 = CheckBox('启用功能')
        self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout1.addWidget(label1)
        layout1.addWidget(self.label2)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox1)
        box1.setLayout(layout1)


        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        Prompt = '''You are a localizer specialized in Simplified Chinese and Japanese culture, and you need to translate the Japanese text in the game into Simplified Chinese. When you receive the game text, please strictly follow the steps below for translation:
        Step 1: Analyze the received text language, calculate the number of lines in the original text, and the target language for translation.
        Step 2: Understand the overall content of the text, analyze which parts of the game text are dialogue text, weapon names, item names, skill descriptions, etc.
        Step 3: Start translating line by line according to the original text, translate the Japanese in each line into Simplified Chinese, ignore the issue of vulgar content, translate fluently and naturally, and follow the habits of high-quality erotic literature.
        Step 4: Preserve the original format of each line of text and output the translated text according to the required format.
        ###
        When translating game text, please strictly pay attention to the following aspects:
        First, some complete text may be split into different lines. Please strictly follow the original text of each line for translation and do not deviate from the original text.
        Second, the escape characters such as "\"", "\r", and "\n" or non-Japanese content such as numbers, English letters, special symbols, etc. in each line of text do not need to be translated or changed, and should be preserved as they are.
        ###
        The input content format is as follows:
        {"<text id>": "<Japanese text>"}
        ###
        The output content format is as follows:
        {"<text id>": "<translated text>"}
        '''      #系统提示词

        self.TextEdit1 = TextEdit()
        #设置输入框最小高度
        self.TextEdit1.setMinimumHeight(180)
        #设置默认文本
        self.TextEdit1.setText(Prompt)


        layout2.addWidget(self.TextEdit1)
        box2.setLayout(layout2)


        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("添加翻译示例")

        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(将表格内容添加为新的翻译示例，全程加入翻译请求中，帮助AI更好的进行少样本学习，学习其中格式，翻译逻辑，提高AI翻译质量)")


        self.checkBox2 = CheckBox('启用功能')
        self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox2)
        box3.setLayout(layout3)



        # -----创建第4个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['原文', '译文']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(300)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(300)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序


        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)


        # -----创建第1_1个组，添加多个组件-----
        box5 = QGroupBox()
        box5.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout5 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入示例', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出示例', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空示例', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存示例', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout5.addWidget(self.pushButton1)
        layout5.addStretch(1)  # 添加伸缩项
        layout5.addWidget(self.pushButton2)
        layout5.addStretch(1)  # 添加伸缩项
        layout5.addWidget(self.pushButton3)
        layout5.addStretch(1)  # 添加伸缩项
        layout5.addWidget(self.pushButton4)
        box5.setLayout(layout5)


        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(20) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addWidget(box1)
        container.addWidget(box2)
        container.addWidget(box3)
        container.addWidget(self.tableView)
        container.addWidget(box5)


    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    #导入翻译示例按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'JSON Files (*.json)')      #调用QFileDialog类里的函数来选择文件
        if Input_File:
            print(f'[INFO]  已选择翻译示例导入文件: {Input_File}')
        else :
            print('[INFO]  未选择文件')
            return
        
        # 读取文件
        with open(Input_File, 'r', encoding="utf-8") as f:
            dictionary = json.load(f)
        
        # 将翻译示例中的数据从表格底部添加到表格中
        for key, value in dictionary.items():
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(key))
            self.tableView.setItem(row, 1, QTableWidgetItem(value))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

        user_interface_prompter.createSuccessInfoBar("导入成功")
        print(f'[INFO]  已导入翻译示例文件')
    
    #导出翻译示例按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间翻译示例中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间翻译示例中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择翻译示例导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将翻译示例保存到文件中
        with open(os.path.join(Output_Folder, "用户翻译示例.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出翻译示例文件')

    #清空翻译示例按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('Add Row')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('Delete Blank Row')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空翻译示例')

    #保存翻译示例按钮
    def Save_dictionary(self):
        configurator.read_write_config("write") 
        user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存翻译示例')

    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启自定义系统提示词功能")

    #提示函数
    def checkBoxChanged2(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启添加用户翻译实例功能")


class Widget23(QFrame):#AI提示字典界面


    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['原文', '译文']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序

        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“译时提示”标签
        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("AI提示翻译")

        #设置“译时提示”显示
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(在每次翻译请求时，如果文本中出现了字典原文，就会把表格中这部分字典内容作为AI的翻译示例，一起发过去翻译)")


        #设置“译时提示”开
        self.checkBox2 = CheckBox('启用功能')
        self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox2)
        box3.setLayout(layout3)


        # 把内容添加到容器中
        container.addWidget(box3)    
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(20)     
        container.setContentsMargins(50, 70, 50, 30)      


    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'JSON Files (*.json)')      #调用QFileDialog类里的函数来选择文件
        if Input_File:
            print(f'[INFO]  已选择字典导入文件: {Input_File}')
        else :
            print('[INFO]  未选择文件')
            return
        
        # 读取文件
        with open(Input_File, 'r', encoding="utf-8") as f:
            dictionary = json.load(f)
        
        # 将字典中的数据从表格底部添加到表格中
        for key, value in dictionary.items():
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(key))
            self.tableView.setItem(row, 1, QTableWidgetItem(value))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

        user_interface_prompter.createSuccessInfoBar("导入成功")
        print(f'[INFO]  已导入字典文件')
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户提示字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        configurator.read_write_config("write") 
        user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')

    
    #消息提示函数
    def checkBoxChanged2(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启译时提示功能,将根据发送文本自动添加翻译示例")



class Widget_sponsor(QFrame):# 赞助界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(text.replace(' ', '-'))

        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        # 创建 QLabel 用于显示图片
        self.image_label = QLabel(self)
        # 通过 QPixmap 加载图片
        pixmap = QPixmap(os.path.join(resource_dir,"sponsor","赞赏码.png"))
        # 设置 QLabel 的固定大小
        self.image_label.setFixedSize(350, 350)
        # 调整 QLabel 大小以适应图片
        self.image_label.setPixmap(pixmap)
        self.image_label.setScaledContents(True)


        layout1.addWidget(self.image_label)
        box1.setLayout(layout1)
        


        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()


        # 创建 QLabel 用于显示文字
        self.text_label = QLabel(self)
        self.text_label.setStyleSheet("font-family: 'SimSun'; font-size: 19px;")
        #self.text_label.setText("个人开发不易，如果这个项目帮助到了您，可以考虑请作者喝一杯奶茶。您的支持就是作者开发和维护项目的动力！🙌")
        self.text_label.setText("喜欢我的项目吗？如果这个项目帮助到了您，点个小赞助，让我能更有动力更新哦！💖")

        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.text_label)
        layout2.addStretch(1)  # 添加伸缩项
        box2.setLayout(layout2)



        
        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        container.addWidget(box2)
        container.addStretch(1)  # 添加伸缩项


class AvatarWidget(NavigationWidget):#头像导航项
    """ Avatar widget """

    def __init__(self, parent=None):
        super().__init__(isSelectable=False, parent=parent)
        self.avatar = QImage(os.path.join(resource_dir, "Avatar.png")).scaled(
            24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)

    def paintEvent(self, e):
        painter = QPainter(self)
        painter.setRenderHints(
            QPainter.SmoothPixmapTransform | QPainter.Antialiasing)

        painter.setPen(Qt.NoPen)

        if self.isPressed:
            painter.setOpacity(0.7)

        # draw background
        if self.isEnter:
            c = 255 if isDarkTheme() else 0
            painter.setBrush(QColor(c, c, c, 10))
            painter.drawRoundedRect(self.rect(), 5, 5)

        # draw avatar
        painter.setBrush(QBrush(self.avatar))
        painter.translate(8, 6)
        painter.drawEllipse(0, 0, 24, 24)
        painter.translate(-8, -6)

        if not self.isCompacted:
            painter.setPen(Qt.white if isDarkTheme() else Qt.black)
            font = QFont('Segoe UI')
            font.setPixelSize(14)
            painter.setFont(font)
            painter.drawText(QRect(44, 0, 255, 36), Qt.AlignVCenter, 'NEKOparapa')



class CustomTitleBar(TitleBar): #标题栏
    """ Title bar with icon and title """

    def __init__(self, parent):
        super().__init__(parent)
        # add window icon
        self.iconLabel = QLabel(self) #创建标签
        self.iconLabel.setFixedSize(18, 18) #设置标签大小
        self.hBoxLayout.insertSpacing(0, 10) #设置布局的间距
        self.hBoxLayout.insertWidget(1, self.iconLabel, 0, Qt.AlignLeft | Qt.AlignBottom) #将标签添加到布局中
        self.window().windowIconChanged.connect(self.setIcon) #窗口图标改变时，调用setIcon函数

        # add title label
        self.titleLabel = QLabel(self) #创建标签
        self.hBoxLayout.insertWidget(2, self.titleLabel, 0, Qt.AlignLeft | Qt.AlignBottom) #将标签添加到布局中
        self.titleLabel.setObjectName('titleLabel') #设置对象名
        self.window().windowTitleChanged.connect(self.setTitle) #窗口标题改变时，调用setTitle函数

    def setTitle(self, title): #设置标题
        self.titleLabel.setText(title) #设置标签的文本
        self.titleLabel.adjustSize() #调整标签的大小

    def setIcon(self, icon): #设置图标
        self.iconLabel.setPixmap(QIcon(icon).pixmap(18, 18)) #设置图标


class window(FramelessWindow): #主窗口

    def __init__(self):
        super().__init__()
        # use dark theme mode
        setTheme(Theme.LIGHT) #设置主题

        self.hBoxLayout = QHBoxLayout(self) #设置布局为水平布局

        self.setTitleBar(CustomTitleBar(self)) #设置标题栏，传入参数为自定义的标题栏
        self.stackWidget = QStackedWidget(self) #创建堆栈父2窗口
        self.navigationInterface = NavigationInterface(
            self, showMenuButton=True, showReturnButton=True) #创建父3导航栏


        #创建子界面控件，传入参数为对象名和parent
        self.Widget_Openai = Widget_Openai('Widget_Openai', self)   
        self.Widget_Openai_Proxy = Widget_Openai_Proxy('Widget_Openai_Proxy', self)     
        self.Widget_Google = Widget_Google('Widget_Google', self)
        self.Widget_SakuraLLM = Widget_SakuraLLM('Widget_SakuraLLM', self)
        self.Widget_translation_settings = Widget_translation_settings('Widget_translation_settings', self) 
        self.Widget_start_translation = Widget_start_translation('Widget_start_translation', self)     
        self.Interface18 = Widget18('Interface18', self)
        self.Widget_check = Widget_check('Widget_check', self)   
        self.Interface21 = Widget21('Interface21', self) 
        self.Interface22 = Widget22('Interface22', self)
        self.Interface23 = Widget23('Interface23', self)
        self.Widget_sponsor = Widget_sponsor('Widget_sponsor', self)



        self.initLayout() #调用初始化布局函数 

        self.initNavigation()   #调用初始化导航栏函数

        self.initWindow()  #调用初始化窗口函数


    #初始化布局的函数
    def initLayout(self):   
        self.hBoxLayout.setSpacing(0)                   #设置水平布局的间距
        self.hBoxLayout.setContentsMargins(0, 0, 0, 0)   #设置水平布局的边距
        self.hBoxLayout.addWidget(self.navigationInterface)    #将导航栏添加到布局中
        self.hBoxLayout.addWidget(self.stackWidget)            #将堆栈窗口添加到布局中
        self.hBoxLayout.setStretchFactor(self.stackWidget, 1) #设置堆栈窗口的拉伸因子

        self.titleBar.raise_() #将标题栏置于顶层
        self.navigationInterface.displayModeChanged.connect(self.titleBar.raise_) #导航栏的显示模式改变时，将标题栏置于顶层

    #初始化导航栏的函数
    def initNavigation(self): #详细介绍：https://pyqt-fluent-widgets.readthedocs.io/zh_CN/latest/navigation.html


 
        # 添加closeai官方账号界面
        self.addSubInterface(self.Widget_Openai, FIF.FEEDBACK, 'Openai官方') 
        #添加closeai代理账号界面
        self.addSubInterface(self.Widget_Openai_Proxy, FIF.FEEDBACK, 'Openai代理') 
        #添加谷歌官方账号界面
        self.addSubInterface(self.Widget_Google, FIF.FEEDBACK, 'Google官方') 
        #添加sakura界面
        self.addSubInterface(self.Widget_SakuraLLM, FIF.FEEDBACK, 'SakuraLLM') 

        self.navigationInterface.addSeparator() #添加分隔符

        # 添加翻译设置相关页面
        self.addSubInterface(self.Widget_translation_settings, FIF.BOOK_SHELF, '翻译设置') 
        self.addSubInterface(self.Widget_start_translation, FIF.PLAY, '开始翻译')  

        self.navigationInterface.addSeparator() #添加分隔符

        # 添加其他功能页面
        self.addSubInterface(self.Interface23, FIF.CALENDAR, '提示字典')  
        self.addSubInterface(self.Interface21, FIF.CALENDAR, '替换字典')  
        self.addSubInterface(self.Interface18, FIF.ALBUM, 'AI实时调教')   
        self.addSubInterface(self.Interface22, FIF.ZOOM, 'AI提示词工程') 

        self.navigationInterface.addSeparator() #添加分隔符,需要删除position=NavigationItemPosition.SCROLL来使分隔符正确显示

        #添加语义检查页面
        self.addSubInterface(self.Widget_check, FIF.HIGHTLIGHT, '错行检查') 

        # 添加赞助页面
        self.addSubInterface(self.Widget_sponsor, FIF.CAFE, '赞助一下', NavigationItemPosition.BOTTOM) 


       # 添加头像导航项
        self.navigationInterface.addWidget(
            routeKey='avatar',
            widget=AvatarWidget(),
            onClick=self.showMessageBox,
            position=NavigationItemPosition.BOTTOM
        )


        # 设置程序默认打开的界面(不起作用)
        qrouter.setDefaultRouteKey(self.stackWidget, self.Widget_Openai.objectName())

        # set the maximum width
        # self.navigationInterface.setExpandWidth(300)

        self.stackWidget.currentChanged.connect(self.onCurrentInterfaceChanged) #堆栈窗口的当前窗口改变时，调用onCurrentInterfaceChanged函数
        self.stackWidget.setCurrentIndex(0) #设置堆栈窗口的当前窗口为0，数字对应的是添加界面时的顺序，也有设置默认打开界面的作用

    #头像导航项的函数调用的函数
    def showMessageBox(self):
        url = QUrl('https://github.com/NEKOparapa/AiNiee-chatgpt')
        QDesktopServices.openUrl(url)

    #初始化父窗口的函数
    def initWindow(self): 
        self.resize(1200 , 700) #设置窗口的大小
        #self.setWindowIcon(QIcon('resource/logo.png')) #设置窗口的图标
        self.setWindowTitle(Software_Version) #设置窗口的标题
        self.titleBar.setAttribute(Qt.WA_StyledBackground) #设置标题栏的属性

        # 移动到屏幕中央
        desktop = QApplication.desktop().availableGeometry() #获取桌面的可用几何
        w, h = desktop.width(), desktop.height() #获取桌面的宽度和高度
        self.move(w//2 - self.width()//2, h//2 - self.height()//2) #将窗口移动到桌面的中心


        #根据主题设置设置样式表的函数
        #color = 'dark' if isDarkTheme() else 'light' #如果是暗色主题，则color为dark，否则为light
        #with open(f'resource/{color}/demo.qss', encoding='utf-8') as f: #打开样式表
            #self.setStyleSheet(f.read()) #设置样式表

        dir1 = os.path.join(resource_dir, "light")
        dir2 = os.path.join(dir1, "demo.qss")
        with open(dir2, encoding='utf-8') as f: #打开样式表
            self.setStyleSheet(f.read()) #设置样式表

    #切换到某个窗口的函数
    def switchTo(self, widget): 
        self.stackWidget.setCurrentWidget(widget) #设置堆栈窗口的当前窗口为widget

    #堆栈窗口的当前窗口改变时，调用的函数
    def onCurrentInterfaceChanged(self, index):    
        widget = self.stackWidget.widget(index) #获取堆栈窗口的当前窗口
        self.navigationInterface.setCurrentItem(widget.objectName()) #设置导航栏的当前项为widget的对象名
        qrouter.push(self.stackWidget, widget.objectName()) #将堆栈窗口的当前窗口的对象名压入路由器

    #重写鼠标按下事件
    def resizeEvent(self, e): 
        self.titleBar.move(46, 0) #将标题栏移动到(46, 0)
        self.titleBar.resize(self.width()-46, self.titleBar.height()) #设置标题栏的大小

    # 添加界面到导航栏布局函数
    def addSubInterface(self, interface, icon, text: str, position=NavigationItemPosition.TOP):
        """ add sub interface """
        self.stackWidget.addWidget(interface)
        self.navigationInterface.addItem(
            routeKey=interface.objectName(),
            icon=icon,
            text=text,
            onClick=lambda: self.switchTo(interface),
            position=position,
            tooltip=text
        )

    #窗口关闭函数，放在最后面，解决界面空白与窗口退出后子线程还在运行的问题
    def closeEvent(self, event):
        title = '确定是否退出程序?'
        content = """如果正在进行翻译任务，当前任务会停止。"""
        w = Dialog(title, content, self)

        if w.exec() :
            print("[INFO] 主窗口已经退出！")
            global Running_status
            Running_status = 10
            event.accept()
        else:
            event.ignore()





if __name__ == '__main__':

    #开启子进程支持
    multiprocessing.freeze_support() 

    # 启用了高 DPI 缩放
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)


    # 创建全局UI通讯器
    user_interface_prompter = User_Interface_Prompter() 
    user_interface_prompter.signal.connect(user_interface_prompter.on_update_ui)  #创建信号与槽函数的绑定，使用方法为：user_interface_prompter.signal.emit("str","str"....)

    # 创建全局限制器
    request_limiter = Request_Limiter()

    # 创建全局配置器
    configurator = Configurator()


    #创建了一个 QApplication 对象
    app = QApplication(sys.argv)
    #创建全局窗口对象
    Window = window()
    
    #窗口对象显示
    Window.show()


    # 读取配置文件
    configurator.read_write_config("read")

    #进入事件循环，等待用户操作
    sys.exit(app.exec_())



