 # ═══════════════════════════════════════════════════════
# ████ WARNING: Enter at Your Own Risk!               ████
# ████ Congratulations, you have stumbled upon my     ████
# ████ masterpiece - a mountain of 10,000 lines of    ████
# ████ spaghetti code. Proceed with caution,          ████
# ████ as reading this code may result in             ████
# ████ immediate unhappiness and despair.             ████
# ═══════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════
# ████ 警告：擅自进入，后果自负                         ████
# ████ 恭喜你，你已经发现了我的杰作                     ████
# ████ 一座万行意大利面条式代码的屎山                   ████
# ████ 请谨慎前行，阅读这段代码可能会。                 ████
# ████ 立刻让你感到不幸和绝望                          ████
# ═══════════════════════════════════════════════════════


# coding:utf-8               
import copy
import datetime
import json
import random
import yaml
import re
import time
import threading
import os
import sys
import multiprocessing
import concurrent.futures
import shutil
import zipfile

import tiktoken_ext  #必须导入这两个库，否则打包后无法运行
from tiktoken_ext import openai_public

import tiktoken #需要安装库pip install tiktoken
import openpyxl  #需安装库pip install openpyxl
from openpyxl import Workbook  
import opencc       #需要安装库pip install opencc      
from openai import OpenAI #需要安装库pip install openai
import google.generativeai as genai #需要安装库pip install -U google-generativeai
import anthropic #需要安装库pip install anthropic
import ebooklib #需要安装库pip install ebooklib
from ebooklib import epub
from bs4 import BeautifulSoup #需要安装库pip install beautifulsoup4
import cohere  #需要安装库pip install cohere

from PyQt5.QtGui import QBrush, QColor, QDesktopServices, QFont, QImage, QPainter, QPixmap#需要安装库 pip3 install PyQt5
from PyQt5.QtCore import  QObject,  QRect,  QUrl,  Qt, pyqtSignal 
from PyQt5.QtWidgets import QAbstractItemView,QHeaderView,QApplication, QTableWidgetItem, QFrame, QGridLayout, QGroupBox, QLabel,QFileDialog, QStackedWidget, QHBoxLayout, QVBoxLayout

from qfluentwidgets.components import Dialog  # 需要安装库 pip install "PyQt-Fluent-Widgets[full]" -i https://pypi.org/simple/
from qfluentwidgets import ProgressRing, SegmentedWidget, TableWidget,CheckBox, DoubleSpinBox, HyperlinkButton,InfoBar, InfoBarPosition, NavigationWidget, Slider, SpinBox, ComboBox, LineEdit, PrimaryPushButton, PushButton ,StateToolTip, SwitchButton, TextEdit, Theme,  setTheme ,isDarkTheme,qrouter,NavigationInterface,NavigationItemPosition, EditableComboBox
from qfluentwidgets import FluentIcon as FIF
from qframelesswindow import FramelessWindow, StandardTitleBar

from StevExtraction import jtpp  # type: ignore #导入文本提取工具



# 翻译器
class Translator():
    def __init__(self):
        pass

    def Main(self):
        global cache_list, Running_status
        # ——————————————————————————————————————————配置信息初始化—————————————————————————————————————————


        configurator.read_write_config("write") # 将配置信息写入配置文件中

        configurator.initialize_configuration() # 获取界面的配置信息

        # 根据混合翻译设置更换翻译平台
        if configurator.mixed_translation_toggle:
            configurator.translation_platform = configurator.configure_mixed_translation["first_platform"]

        configurator.configure_translation_platform(configurator.translation_platform)  # 配置翻译平台信息
        request_limiter.initialize_limiter() # 配置请求限制器，依赖前面的配置信息，必需在最后面初始化


        # ——————————————————————————————————————————读取原文到缓存—————————————————————————————————————————


        #如果是从头开始翻译
        if Running_status == 6:
            # 读取文件
            try:
                cache_list = File_Reader.read_files(self,configurator.translation_project, configurator.Input_Folder)

            except Exception as e:
                print(e)
                print("\033[1;31mError:\033[0m 读取原文失败，请检查项目类型是否设置正确，输入文件夹是否混杂其他非必要文件！")
                return


        # ——————————————————————————————————————————初步处理缓存文件—————————————————————————————————————————
            

            # 将浮点型，整数型文本内容变成字符型文本内容
            Cache_Manager.convert_source_text_to_str(self,cache_list)

            # 除去代码文本
            Cache_Manager.ignore_code_text(self,cache_list)

            # 如果翻译日语或者韩语文本时，则去除非中日韩文本
            if configurator.source_language == "日语" or configurator.source_language == "韩语":
                Cache_Manager.process_dictionary_list(self,cache_list)


        # ——————————————————————————————————————————构建并发任务池子—————————————————————————————————————————


        # 计算待翻译的文本总行数，tokens总数
        untranslated_text_line_count,untranslated_text_tokens_count = Cache_Manager.count_and_update_translation_status_0_2(self, cache_list) #获取需要翻译的文本总行数
        # 计算并发任务数
        tasks_Num = Translator.calculate_total_tasks(self,untranslated_text_line_count,untranslated_text_tokens_count,configurator.lines_limit,configurator.tokens_limit,configurator.tokens_limit_switch)


        # 更新界面UI信息
        if Running_status == 9: # 如果是继续翻译
            total_text_line_count = user_interface_prompter.total_text_line_count # 与上一个翻译任务的总行数一致
            user_interface_prompter.signal.emit("翻译状态提示","开始翻译",0)

            #最后改一下运行状态，为正常翻译状态
            Running_status = 6

        else:#如果是从头开始翻译
            total_text_line_count = untranslated_text_line_count
            project_id = cache_list[0]["project_id"]
            user_interface_prompter.signal.emit("初始化翻译界面数据",project_id,untranslated_text_line_count) #需要输入够当初设定的参数个数
            user_interface_prompter.signal.emit("翻译状态提示","开始翻译",0)


        # 输出开始翻译的日志
        print("[INFO]  翻译项目为",configurator.translation_project, '\n')
        print("[INFO]  翻译平台为",configurator.translation_platform, '\n')
        print("[INFO]  请求地址为",configurator.base_url, '\n')
        print("[INFO]  翻译模型为",configurator.model_type, '\n')

        if configurator.translation_platform != "SakuraLLM":
            print("[INFO]  当前设定的系统提示词为:\n", configurator.get_system_prompt(), '\n')

        print("[INFO]  游戏文本从",configurator.source_language, '翻译到', configurator.target_language,'\n')
        print("[INFO]  文本总行数为：",total_text_line_count,"  需要翻译的行数为：",untranslated_text_line_count)
        if configurator.tokens_limit_switch:
            print("[INFO]  每次发送tokens为：",configurator.tokens_limit,"  计划的翻译任务总数是：", tasks_Num,'\n') 
        else:    
            print("[INFO]  每次发送行数为：",configurator.lines_limit,"  计划的翻译任务总数是：", tasks_Num,'\n') 
        print("\033[1;32m[INFO] \033[0m 五秒后开始进行翻译，请注意保持网络通畅，余额充足。", '\n')
        time.sleep(5)  


        # 创建线程池
        The_Max_workers = configurator.thread_counts # 获取线程数配置
        with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
            # 创建实例
            api_requester_instance = Api_Requester()
            # 向线程池提交任务
            for i in range(tasks_Num):
                # 根据不同平台调用不同接口
                executor.submit(api_requester_instance.concurrent_request)
                    
            # 等待线程池任务完成
            executor.shutdown(wait=True)


        # 检查翻译任务是否已经暂停或者退出
        if Running_status == 9 or Running_status == 10 :
            return


        # ——————————————————————————————————————————检查没能成功翻译的文本，拆分翻译————————————————————————————————————————


        #计算未翻译文本的数量
        untranslated_text_line_count,untranslated_text_tokens_count = Cache_Manager.count_and_update_translation_status_0_2(self,cache_list)

        #存储重新翻译的次数
        retry_translation_count = 1

        while untranslated_text_line_count != 0 :
            print("\033[1;33mWarning:\033[0m 仍然有部分未翻译，将进行拆分后重新翻译，-----------------------------------")
            print("[INFO] 当前拆分翻译轮次：",retry_translation_count ," 到达最大轮次：6 时，将停止翻译")


            # 根据混合翻译设置更换翻译平台,并重新初始化配置信息
            if configurator.mixed_translation_toggle:

                configurator.initialize_configuration() # 获取界面的配置信息

                # 更换翻译平台
                if retry_translation_count == 1:
                    configurator.translation_platform = configurator.configure_mixed_translation["second_platform"]
                    print("[INFO]  已开启混合翻译功能，正在进行次轮拆分翻译，翻译平台更换为：",configurator.translation_platform, '\n')
                else:
                    configurator.translation_platform = configurator.configure_mixed_translation["third_platform"]
                    print("[INFO]  已开启混合翻译功能，正在进行末轮拆分翻译，翻译平台更换为：",configurator.translation_platform, '\n')

                configurator.configure_translation_platform(configurator.translation_platform)  # 配置翻译平台信息
                request_limiter.initialize_limiter() # 配置请求限制器，依赖前面的配置信息，必需在最后面初始化


            # 根据算法计算拆分的文本行数
            if configurator.mixed_translation_toggle and configurator.split_switch:
                print("[INFO] 检测到不进行拆分设置，发送行数将继续保持不变")
            else:
                configurator.lines_limit,configurator.tokens_limit = Translator.update_lines_or_tokens(self,configurator.lines_limit,configurator.tokens_limit) # 更换配置中的文本行数
            
            if configurator.tokens_limit_switch:
                print("[INFO] 未翻译文本总tokens为：",untranslated_text_tokens_count,"  每次发送tokens为：",configurator.tokens_limit, '\n')
            else:
                print("[INFO] 未翻译文本总行数为：",untranslated_text_line_count,"  每次发送行数为：",configurator.lines_limit, '\n')


            # 计算并发任务数
            tasks_Num = Translator.calculate_total_tasks(self,untranslated_text_line_count,untranslated_text_tokens_count,configurator.lines_limit,configurator.tokens_limit,configurator.tokens_limit_switch)



            # 创建线程池
            The_Max_workers = configurator.thread_counts # 获取线程数配置
            with concurrent.futures.ThreadPoolExecutor (The_Max_workers) as executor:
                # 创建实例
                api_requester_instance = Api_Requester()
                # 向线程池提交任务
                for i in range(tasks_Num):
                    # 根据不同平台调用不同接口
                    executor.submit(api_requester_instance.concurrent_request)

                # 等待线程池任务完成
                executor.shutdown(wait=True)

            
            # 检查翻译任务是否已经暂停或者退出
            if Running_status == 9 or Running_status == 10 :
                return


            #检查是否已经达到重翻次数限制
            retry_translation_count  = retry_translation_count + 1
            if retry_translation_count > configurator.round_limit :
                print ("\033[1;33mWarning:\033[0m 已经达到拆分翻译轮次限制，但仍然有部分文本未翻译，不影响使用，可手动翻译", '\n')
                break

            #重新计算未翻译文本的数量
            untranslated_text_line_count,untranslated_text_tokens_count = Cache_Manager.count_and_update_translation_status_0_2(self,cache_list)

        print ("\033[1;32mSuccess:\033[0m  翻译阶段已完成，正在处理数据-----------------------------------", '\n')


        # ——————————————————————————————————————————将数据处理并保存为文件—————————————————————————————————————————
            

        #如果开启了转换简繁开关功能，则进行文本转换
        if configurator.conversion_toggle: 
            if configurator.target_language == "简中" or configurator.target_language == "繁中":
                try:
                    cache_list = File_Outputter.simplified_and_traditional_conversion(self,cache_list, configurator.target_language)
                    print(f"\033[1;32mSuccess:\033[0m  文本转化{configurator.target_language}完成-----------------------------------", '\n')   

                except Exception as e:
                    print("\033[1;33mWarning:\033[0m 文本转换出现问题！！将跳过该步，错误信息如下")
                    print(f"Error: {e}\n")

        # 将翻译结果写为对应文件
        File_Outputter.output_translated_content(self,cache_list,configurator.Output_Folder,configurator.Input_Folder)

        # —————————————————————————————————————#全部翻译完成——————————————————————————————————————————


        print("\033[1;32mSuccess:\033[0m  译文文件写入完成-----------------------------------", '\n')  
        user_interface_prompter.signal.emit("翻译状态提示","翻译完成",0)
        print("\n--------------------------------------------------------------------------------------")
        print("\n\033[1;32mSuccess:\033[0m 已完成全部翻译任务，程序已经停止")   
        print("\n\033[1;32mSuccess:\033[0m 请检查译文文件，格式是否错误，存在错行，或者有空行等问题")
        print("\n-------------------------------------------------------------------------------------\n")


    # 重新设置发送的文本行数
    def update_lines_or_tokens(self,lines_limit,tokens_limit):
        # 重新计算文本行数限制
        if lines_limit % 2 == 0:
            new_lines_limit = lines_limit // 2
        elif lines_limit % 3 == 0:
            new_lines_limit = lines_limit // 3
        elif lines_limit % 4 == 0:
            new_lines_limit = lines_limit // 4
        elif lines_limit % 5 == 0:
            new_lines_limit = lines_limit // 5
        else:
            new_lines_limit = 1

        # 重新计算tokens限制
        new_tokens_limit = tokens_limit // 2

        return new_lines_limit,new_tokens_limit


    # 计算任务总数
    def calculate_total_tasks(self,total_lines,total_tokens,lines_limit,tokens_limit,switch = False):
        
        if switch:

            if total_tokens % tokens_limit == 0:
                tasks_Num = total_tokens // tokens_limit 
            else:
                tasks_Num = total_tokens // tokens_limit + 1

        else:

            if total_lines % lines_limit == 0:
                tasks_Num = total_lines // lines_limit 
            else:
                tasks_Num = total_lines // lines_limit + 1

        return tasks_Num


# 接口请求器
class Api_Requester():
    def __init__(self):
        pass
    
    # 并发接口请求分发
    def concurrent_request (self):

        if configurator.translation_platform == "OpenAI官方" or configurator.translation_platform == "OpenAI代理":
            self.concurrent_request_openai()
        
        elif configurator.translation_platform == "Google官方":
            self.concurrent_request_google()

        elif configurator.translation_platform == "Cohere官方":
            self.Concurrent_Request_cohere()

        elif configurator.translation_platform == "Anthropic官方" or configurator.translation_platform == "Anthropic代理":
            self.concurrent_request_anthropic()

        elif configurator.translation_platform == "Moonshot官方":
            self.concurrent_request_openai()

        elif configurator.translation_platform == "Deepseek官方":
            self.concurrent_request_openai()

        elif configurator.translation_platform == "Dashscope官方":
            self.concurrent_request_openai()

        elif configurator.translation_platform == "Volcengine官方":
            self.concurrent_request_openai()

        elif configurator.translation_platform == "智谱官方":
            self.concurrent_request_openai()

        elif configurator.translation_platform == "SakuraLLM":
            self.concurrent_request_sakura()


    # 整理发送内容（Openai）
    def organize_send_content_openai(self,source_text_dict, previous_list):
        #创建message列表，用于发送
        messages = []

        #获取基础系统提示词
        system_prompt = configurator.get_system_prompt()


        #如果开启提示字典
        glossary_prompt = ""
        glossary_prompt_cot = ""
        if configurator.prompt_dictionary_switch :
            glossary_prompt,glossary_prompt_cot = configurator.build_glossary_prompt(source_text_dict,configurator.cn_prompt_toggle)
            if glossary_prompt :
                system_prompt += glossary_prompt 
                print("[INFO]  已添加术语表：\n",glossary_prompt)


        #如果角色介绍开关打开
        characterization = ""
        characterization_cot = ""
        if configurator.characterization_switch :
            characterization,characterization_cot = configurator.build_characterization(source_text_dict,configurator.cn_prompt_toggle)
            if characterization:
                system_prompt += characterization 
                print("[INFO]  已添加角色介绍：\n",characterization)

        #如果背景设定开关打开
        world_building = ""
        world_building_cot = ""
        if configurator.world_building_switch :
            world_building,world_building_cot = configurator.build_world(configurator.cn_prompt_toggle)
            if world_building:
                system_prompt += world_building 
                print("[INFO]  已添加背景设定：\n",world_building)

        #如果文风要求开关打开
        writing_style = ""
        writing_style_cot = ""
        if configurator.writing_style_switch :
            writing_style,writing_style_cot = configurator.build_writing_style(configurator.cn_prompt_toggle)
            if writing_style:
                system_prompt += writing_style 
                print("[INFO]  已添加文风要求：\n",writing_style)



        # 添加系统提示词信息
        messages.append({"role": "system","content": system_prompt })



        # 获取默认示例前置文本
        pre_prompt = configurator.build_userExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle,configurator.source_language,configurator.target_language,glossary_prompt_cot,characterization_cot,world_building_cot,writing_style_cot)

        #构建默认示例
        original_exmaple,translation_example =  Configurator.build_translation_sample(self,source_text_dict,configurator.source_language,configurator.target_language)
        if original_exmaple and translation_example:

            the_original_exmaple =  {"role": "user","content":(pre_prompt + original_exmaple) }
            the_translation_example = {"role": "assistant", "content": (f'{fol_prompt}```json\n{translation_example}\n```') }

            messages.append(the_original_exmaple)
            messages.append(the_translation_example)
            print("[INFO]  已添加格式原文示例：\n",original_exmaple)
            print("[INFO]  已添加格式译文示例：\n",translation_example, '\n')


        #如果翻译示例开关打开
        if configurator.translation_example_switch :
            original_exmaple_3,translation_example_3 = configurator.build_translation_example ()
            if original_exmaple_3 and translation_example_3:
                the_original_exmaple =  {"role": "user","content":original_exmaple_3 }
                the_translation_example = {"role": "assistant", "content": translation_example_3}
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  已添加用户原文示例：\n",original_exmaple_3)
                print("[INFO]  已添加用户译文示例：\n",translation_example_3, '\n')


        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")


        #如果开启译前替换字典功能，则根据用户字典进行替换
        if configurator.pre_translation_switch :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_before_translation(source_text_dict)



        #如果加上文
        previous = ""
        if configurator.pre_line_counts and previous_list :
            previous = configurator.build_pre_text(previous_list,configurator.cn_prompt_toggle)
            if previous:
                pass
                #print("[INFO]  已添加上文：\n",previous)



        #获取提问时的前置文本
        pre_prompt = configurator.build_userQueryPrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelResponsePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)


        # 构建用户信息
        source_text_str = json.dumps(source_text_dict, ensure_ascii=False)  
        source_text_str = previous + "\n"+ pre_prompt + source_text_str 
        messages.append({"role":"user","content":source_text_str })


        # 构建模型信息
        if( "claude" in configurator.model_type or "gpt" in configurator.model_type or "moonshot" in configurator.model_type or "deepseek" in configurator.model_type) :
            messages.append({"role": "assistant", "content":fol_prompt })

        return messages,source_text_str


    # 并发接口请求（Openai）
    def concurrent_request_openai(self):
        global cache_list,Running_status

        # 检查翻译任务是否已经暂停或者退出
        if Running_status == 9 or Running_status == 10 :
            return

        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            if configurator.tokens_limit_switch:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_tokens(self,configurator.tokens_limit, cache_list,configurator.pre_line_counts)  
            else:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_lines(self,configurator.lines_limit, cache_list,configurator.pre_line_counts)   
            lock1.release()  # 释放锁

            # 检查一下是否有发送内容
            if source_text_list == []:
                print("\033[1;33mWarning:\033[0m 未能获取文本，该线程为多余线程，取消任务")
                return
            # ——————————————————————————————————————————处理原文本的内容与格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # 如果原文是日语，清除文本首尾中的代码文本，并记录清除信息
            if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                source_text_dict,process_info_list = Cache_Manager.process_dictionary(self,source_text_dict)
                row_count = len(source_text_dict)

            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str = Api_Requester.organize_send_content_openai(self,source_text_dict, previous_list)



            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            #计算请求的tokens预计花费
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages) 
            #计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str }] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text)
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;31mError:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;31mError:\033[0m 该条消息取消任务，进行拆分翻译" )
                return
            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 220  # 设置超时时间为x秒
            request_errors_count = 0 # 请求错误次数
            Wrong_answer_count = 0   # 错误回复次数
            model_degradation = False # 模型退化检测

            while 1 :
                # 检查翻译任务是否已经暂停或者退出---------------------------------
                if Running_status == 9 or Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume )
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 获取AI的参数设置
                    temperature,top_p,presence_penalty,frequency_penalty= configurator.get_openai_parameters()
                    # 如果上一次请求出现模型退化，更改参数
                    if model_degradation:
                        frequency_penalty = 0.2

                    # 获取apikey
                    openai_apikey =  configurator.get_apikey()
                    # 创建openai客户端
                    openaiclient = OpenAI(api_key=openai_apikey,
                                            base_url= configurator.base_url)
                    # 发送对话请求
                    try:
                        response = openaiclient.chat.completions.create(
                            model= configurator.model_type,
                            messages = messages ,
                            temperature=temperature,
                            top_p = top_p,                        
                            presence_penalty=presence_penalty,
                            frequency_penalty=frequency_penalty
                            )
                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    # 检查翻译任务是否已经暂停或者退出，不进行接下来的处理了
                    if Running_status == 9 or Running_status == 10 :
                        return
                    

                    #——————————————————————————————————————————收到回复，获取返回的信息 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    try: # 因为有些中转网站不返回tokens消耗
                        prompt_tokens_used = int(response.usage.prompt_tokens) #本次请求花费的tokens
                    except Exception as e:
                        prompt_tokens_used = 0
                    try:
                        completion_tokens_used = int(response.usage.completion_tokens) #本次回复花费的tokens
                    except Exception as e:
                        completion_tokens_used = 0



                    # 尝试提取回复的文本内容
                    try:
                        response_content = response.choices[0].message.content 
                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 提取文本时出现问题！！！运行错误信息如下")
                        print(f"Error: {e}\n")
                        print("接口返回的错误信息如下")
                        print(response)
                        #处理完毕，再次进行请求
                        
                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break
                        continue


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                    # ———————————————————————————————————对回复内容处理,检查—————————————————————————————————————————————————

                    # 处理回复内容
                    response_dict = Response_Parser.process_content(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,response_dict,source_text_dict,configurator.source_language)


                    # ———————————————————————————————————回复内容结果录入—————————————————————————————————————————————————
                    # 如果没有出现错误
                    if check_result :

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")


                        # 如果开启译后替换字典功能，则根据用户字典进行替换
                        if configurator.post_translation_switch :
                            print("[INFO] 你开启了译后修正功能，正在进行替换", '\n')
                            response_dict = configurator.replace_after_translation(response_dict)

                        # 如果原文是日语，则还原文本的首尾代码字符
                        if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                            response_dict = Cache_Manager.update_dictionary(self,response_dict, process_info_list)

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict,configurator.model_type)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if configurator.auto_backup_toggle:
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁

                        
                        lock2.acquire()  # 获取锁
                        # 更新翻译界面数据
                        user_interface_prompter.update_data(1,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",1)

                        # 获取翻译进度
                        progress = user_interface_prompter.progress

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")
                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:
                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")


                        lock2.acquire()  # 获取锁

                        # 如果是进行平时的翻译任务
                        if Running_status == 6 :

                            # 更新翻译界面数据
                            user_interface_prompter.update_data(0,row_count,prompt_tokens_used,completion_tokens_used)

                            # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                            user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",1)

                        lock2.release()  # 释放锁


                        # 检查一下是不是模型退化
                        if error_content == "AI回复内容出现高频词,并重新翻译":
                            print("\033[1;33mWarning:\033[0m 下次请求将修改参数，回避高频词输出","\n")
                            model_degradation = True

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m 错误重新翻译最大次数限制:",configurator.retry_count_limit,"剩余可重试次数:",(configurator.retry_count_limit + 1 - Wrong_answer_count),"到达次数限制后，该段文本将进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count > configurator.retry_count_limit :
                            print("\033[1;33mWarning:\033[0m 错误回复重翻次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环              
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return



    # 整理发送内容（Google）
    def organize_send_content_google(self,source_text_dict, previous_list):
        # 创建message列表，用于发送
        messages = []

        # 获取基础系统提示词
        system_prompt = configurator.get_system_prompt()


        # #如果开启提示字典
        glossary_prompt = ""
        glossary_prompt_cot = ""
        if configurator.prompt_dictionary_switch :
            glossary_prompt,glossary_prompt_cot = configurator.build_glossary_prompt(source_text_dict,configurator.cn_prompt_toggle)
            if glossary_prompt :
                system_prompt += glossary_prompt 
                print("[INFO]  已添加术语表：\n",glossary_prompt)


        # 如果角色介绍开关打开
        characterization = ""
        characterization_cot = ""
        if configurator.characterization_switch :
            characterization,characterization_cot = configurator.build_characterization(source_text_dict,configurator.cn_prompt_toggle)
            if characterization:
                system_prompt += characterization 
                print("[INFO]  已添加角色介绍：\n",characterization)

        # 如果背景设定开关打开
        world_building = ""
        world_building_cot = ""
        if configurator.world_building_switch :
            world_building,world_building_cot = configurator.build_world(configurator.cn_prompt_toggle)
            if world_building:
                system_prompt += world_building 
                print("[INFO]  已添加背景设定：\n",world_building)

        # 如果文风要求开关打开
        writing_style = ""
        writing_style_cot = ""
        if configurator.writing_style_switch :
            writing_style,writing_style_cot = configurator.build_writing_style(configurator.cn_prompt_toggle)
            if writing_style:
                system_prompt += writing_style 
                print("[INFO]  已添加文风要求：\n",writing_style)

        # 获取默认示例前置文本
        pre_prompt = configurator.build_userExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle,configurator.source_language,configurator.target_language,glossary_prompt_cot,characterization_cot,world_building_cot,writing_style_cot)

        # 构建默认示例
        original_exmaple,translation_example =  Configurator.build_translation_sample(self,source_text_dict,configurator.source_language,configurator.target_language)
        if original_exmaple and translation_example:

            the_original_exmaple =  {"role": "user","parts":(pre_prompt + original_exmaple) }
            the_translation_example = {"role": "model", "parts": (f'{fol_prompt}```json\n{translation_example}\n```') }

            messages.append(the_original_exmaple)
            messages.append(the_translation_example)
            print("[INFO]  已添加格式原文示例：\n",original_exmaple)
            print("[INFO]  已添加格式译文示例：\n",translation_example, '\n')


        # 如果翻译示例开关打开
        if configurator.translation_example_switch :
            original_exmaple_3,translation_example_3 = configurator.build_translation_example ()
            if original_exmaple_3 and translation_example_3:
                the_original_exmaple =  {"role": "user","parts":original_exmaple_3 }
                the_translation_example = {"role": "model", "parts": translation_example_3}
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  已添加用户原文示例：\n",original_exmaple_3)
                print("[INFO]  已添加用户译文示例：\n",translation_example_3, '\n')


        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")


        # 如果开启译前替换字典功能，则根据用户字典进行替换
        if configurator.pre_translation_switch :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_before_translation(source_text_dict)



        # 如果加上文
        previous = ""
        if configurator.pre_line_counts and previous_list :
            previous = configurator.build_pre_text(previous_list,configurator.cn_prompt_toggle)
            if previous:
                pass
                #print("[INFO]  已添加上文：\n",previous)


        # 获取提问时的前置文本
        pre_prompt = configurator.build_userQueryPrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelResponsePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)


        # 构建用户信息
        source_text_str = json.dumps(source_text_dict, ensure_ascii=False)     
        source_text_str = previous + "\n"+ pre_prompt + source_text_str 
        messages.append({"role":"user","parts":source_text_str })


        # 构建模型信息
        messages.append({"role": "model", "parts":fol_prompt })


        return messages,source_text_str,system_prompt


    # 并发接口请求（Google）
    def concurrent_request_google(self):
        global cache_list,Running_status

        # 检查翻译任务是否已经暂停或者退出
        if Running_status == 9 or Running_status == 10 :
            return

        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            if configurator.tokens_limit_switch:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_tokens(self,configurator.tokens_limit, cache_list,configurator.pre_line_counts)  
            else:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_lines(self,configurator.lines_limit, cache_list,configurator.pre_line_counts)     
            lock1.release()  # 释放锁

            # 检查一下是否有发送内容
            if source_text_list == []:
                print("\033[1;33mWarning:\033[0m 未能获取文本，该线程为多余线程，取消任务")
                return
            # ——————————————————————————————————————————处理原文本的内容与格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # 如果原文是日语，清除文本首尾中的代码文本，并记录清除信息
            if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                source_text_dict,process_info_list = Cache_Manager.process_dictionary(self,source_text_dict)
                row_count = len(source_text_dict)


            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str,system_prompt = Api_Requester.organize_send_content_google(self,source_text_dict, previous_list)



            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            # 计算请求的tokens预计花费
            prompt_tokens ={"role": "system","content": system_prompt }
            messages_tokens= messages.copy()
            messages_tokens.append(prompt_tokens)
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages_tokens) 

            #计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str}] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text)
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;33mWarning:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;33mWarning:\033[0m 该条消息取消任务，进行拆分翻译" )
                return

            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 220   # 设置超时时间为x秒
            request_errors_count = 0 # 设置请求错误次数限制
            Wrong_answer_count = 0   # 设置错误回复次数限制

            while 1 :
                # 检查翻译任务是否已经暂停或者退出
                if Running_status == 9 or Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume ) 
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 获取AI的参数设置
                    temperature= configurator.get_google_parameters()

                    # 设置AI的参数
                    generation_config = {
                    "temperature": temperature,
                    "top_p": 1,
                    "top_k": 1,
                    "max_output_tokens": 8000, 
                    }

                    #调整安全限制
                    safety_settings = [
                    {
                        "category": "HARM_CATEGORY_HARASSMENT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_HATE_SPEECH",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
                        "threshold": "BLOCK_NONE"
                    },
                    {
                        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
                        "threshold": "BLOCK_NONE"
                    }
                    ]

                    # 获取apikey
                    apikey =  configurator.get_apikey()
                    genai.configure(api_key=apikey,transport='rest')

                    #设置对话模型及参数
                    model = genai.GenerativeModel(model_name=configurator.model_type,
                                    generation_config=generation_config,
                                    safety_settings=safety_settings,
                                    system_instruction = system_prompt)


                    # 发送对话请求
                    try:
                        response = model.generate_content(messages)

                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    # 检查翻译任务是否已经暂停或者退出，不进行接下来的处理了
                    if Running_status == 9 or Running_status == 10 :
                        return
                    

                    #——————————————————————————————————————————收到回复，获取返回的信息 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    prompt_tokens_used = int(request_tokens_consume)
                    completion_tokens_used = int(completion_tokens_consume)



                    # 尝试提取回复的文本内容
                    try:
                        response_content = response.text
                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 提取文本时出现错误！！！运行的错误信息如下")
                        print(f"Error: {e}\n")
                        print("接口返回的错误信息如下")
                        print(response.prompt_feedback)
                        #处理完毕，再次进行请求
                        
                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break
                        continue


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                    # ——————————————————————————————————————————对回复内容处理,检查和录入——————————————————————————————————————————
                    # 处理回复内容
                    response_dict = Response_Parser.process_content(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,response_dict,source_text_dict,configurator.source_language)

                    # ———————————————————————————————————回复内容结果录入—————————————————————————————————————————————————

                    # 如果没有出现错误
                    if check_result :

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        #如果开启译后替换字典功能，则根据用户字典进行替换
                        if configurator.post_translation_switch :
                            print("[INFO] 你开启了译后修正功能，正在进行替换", '\n')
                            response_dict = configurator.replace_after_translation(response_dict)

                        # 如果原文是日语，则还原文本的首尾代码字符
                        if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                            response_dict = Cache_Manager.update_dictionary(self,response_dict, process_info_list)

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict,configurator.model_type)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if configurator.auto_backup_toggle:
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁


                        
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(1,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",1)

                        # 获取翻译进度
                        progress = user_interface_prompter.progress

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")

                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:
                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(0,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",1)

                        lock2.release()  # 释放锁

                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m 错误重新翻译最大次数限制:",configurator.retry_count_limit,"剩余可重试次数:",(configurator.retry_count_limit + 1 - Wrong_answer_count),"到达次数限制后，该段文本将进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count > configurator.retry_count_limit :
                            print("\033[1;33mWarning:\033[0m 错误回复重翻次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环              
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return



    # 整理发送内容（Anthropic）
    def organize_send_content_anthropic(self,source_text_dict, previous_list):
        #创建message列表，用于发送
        messages = []

        #获取基础系统提示词
        system_prompt = configurator.get_system_prompt()


        # 如果开启提示字典
        glossary_prompt = ""
        glossary_prompt_cot = ""
        if configurator.prompt_dictionary_switch :
            glossary_prompt,glossary_prompt_cot = configurator.build_glossary_prompt(source_text_dict,configurator.cn_prompt_toggle)
            if glossary_prompt :
                system_prompt += glossary_prompt 
                print("[INFO]  已添加术语表：\n",glossary_prompt)


        #如果角色介绍开关打开
        characterization = ""
        characterization_cot = ""
        if configurator.characterization_switch :
            characterization,characterization_cot = configurator.build_characterization(source_text_dict,configurator.cn_prompt_toggle)
            if characterization:
                system_prompt += characterization 
                print("[INFO]  已添加角色介绍：\n",characterization)

        #如果背景设定开关打开
        world_building = ""
        world_building_cot = ""
        if configurator.world_building_switch :
            world_building,world_building_cot = configurator.build_world(configurator.cn_prompt_toggle)
            if world_building:
                system_prompt += world_building 
                print("[INFO]  已添加背景设定：\n",world_building)

        #如果文风要求开关打开
        writing_style = ""
        writing_style_cot = ""
        if configurator.writing_style_switch :
            writing_style,writing_style_cot = configurator.build_writing_style(configurator.cn_prompt_toggle)
            if writing_style:
                system_prompt += writing_style 
                print("[INFO]  已添加文风要求：\n",writing_style)

        # 获取默认示例前置文本
        pre_prompt = configurator.build_userExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle,configurator.source_language,configurator.target_language,glossary_prompt_cot,characterization_cot,world_building_cot,writing_style_cot)

        #构建默认示例
        original_exmaple,translation_example =  Configurator.build_translation_sample(self,source_text_dict,configurator.source_language,configurator.target_language)
        if original_exmaple and translation_example:

            the_original_exmaple =  {"role": "user","content":(pre_prompt + original_exmaple) }
            the_translation_example = {"role": "assistant", "content": (f'{fol_prompt}```json\n{translation_example}\n```') }

            messages.append(the_original_exmaple)
            messages.append(the_translation_example)
            print("[INFO]  已添加格式原文示例：\n",original_exmaple)
            print("[INFO]  已添加格式译文示例：\n",translation_example, '\n')


        #如果翻译示例开关打开
        if configurator.translation_example_switch :
            original_exmaple_3,translation_example_3 = configurator.build_translation_example ()
            if original_exmaple_3 and translation_example_3:
                the_original_exmaple =  {"role": "user","content":original_exmaple_3 }
                the_translation_example = {"role": "assistant", "content": translation_example_3}
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  已添加用户原文示例：\n",original_exmaple_3)
                print("[INFO]  已添加用户译文示例：\n",translation_example_3, '\n')


        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")


        #如果开启译前替换字典功能，则根据用户字典进行替换
        if configurator.pre_translation_switch :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_before_translation(source_text_dict)



        #如果加上文
        previous = ""
        if configurator.pre_line_counts and previous_list :
            previous = configurator.build_pre_text(previous_list,configurator.cn_prompt_toggle)
            if previous:
                pass
                #print("[INFO]  已添加上文：\n",previous)


        # 获取提问时的前置文本
        pre_prompt = configurator.build_userQueryPrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelResponsePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)


        # 构建用户信息
        source_text_str = json.dumps(source_text_dict, ensure_ascii=False)     
        source_text_str = previous + "\n"+ pre_prompt + source_text_str 
        messages.append({"role":"user","content":source_text_str })


        # 构建模型信息
        messages.append({"role": "assistant", "content":fol_prompt })


        return messages,source_text_str,system_prompt


    # 并发接口请求（Anthropic）
    def concurrent_request_anthropic(self):
        global cache_list,Running_status

        # 检查翻译任务是否已经暂停或者退出
        if Running_status == 9 or Running_status == 10 :
            return

        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            if configurator.tokens_limit_switch:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_tokens(self,configurator.tokens_limit, cache_list,configurator.pre_line_counts)  
            else:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_lines(self,configurator.lines_limit, cache_list,configurator.pre_line_counts)    
            lock1.release()  # 释放锁

            # 检查一下是否有发送内容
            if source_text_list == []:
                print("\033[1;33mWarning:\033[0m 未能获取文本，该线程为多余线程，取消任务")
                return
            # ——————————————————————————————————————————处理原文本的内容与格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # 如果原文是日语，清除文本首位中的代码文本，并记录清除信息
            if configurator.source_language == "日语"and configurator.text_clear_toggle:
                source_text_dict,process_info_list = Cache_Manager.process_dictionary(self,source_text_dict)
                row_count = len(source_text_dict)


            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str,system_prompt = Api_Requester.organize_send_content_anthropic(self,source_text_dict, previous_list)

            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            # 计算请求的tokens预计花费
            prompt_tokens ={"role": "system","content": system_prompt }
            messages_tokens= messages.copy()
            messages_tokens.append(prompt_tokens)
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages_tokens) 

            # 计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str }] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text)
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;31mError:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;31mError:\033[0m 该条消息取消任务，进行拆分翻译" )
                return
            
            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 220   # 设置超时时间为x秒
            request_errors_count = 0 # 请求错误计数
            Wrong_answer_count = 0   # 错误回复计数

            while 1 :
                # 检查翻译任务是否已经暂停或者退出
                if Running_status == 9 or Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume )
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()
                    # 获取AI的参数设置
                    temperature= configurator.get_anthropic_parameters()
                    # 获取apikey
                    anthropic_apikey =  configurator.get_apikey()
                    # 创建anthropic客户端
                    client = anthropic.Anthropic(api_key=anthropic_apikey,base_url=configurator.base_url)
                    # 发送对话请求
                    try:
                        response = client.messages.create(
                            model= configurator.model_type,
                            max_tokens=4000,
                            system= system_prompt,
                            messages = messages ,
                            temperature=temperature
                            )



                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    # 检查翻译任务是否已经暂停或者退出，不进行接下来的处理了
                    if Running_status == 9 or Running_status == 10 :
                        return
                    

                    #——————————————————————————————————————————收到回复，获取返回的信息 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    try: # 因为有些中转网站不返回tokens消耗
                        prompt_tokens_used = int(response.usage.prompt_tokens) #本次请求花费的tokens
                    except Exception as e:
                        prompt_tokens_used = 0
                    try:
                        completion_tokens_used = int(response.usage.completion_tokens) #本次回复花费的tokens
                    except Exception as e:
                        completion_tokens_used = 0



                    # 尝试提取回复的文本内容
                    try:
                        response_content = response.content[0].text 
                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 提取文本时出现问题！！！运行错误信息如下")
                        print(f"Error: {e}\n")
                        print("接口返回的错误信息如下")
                        print(response)
                        #处理完毕，再次进行请求
                        
                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break
                        continue


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                    # ——————————————————————————————————————————对回复内容处理,检查和录入——————————————————————————————————————————
                    # 处理回复内容
                    response_dict = Response_Parser.process_content(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,response_dict,source_text_dict,configurator.source_language)

                    # ———————————————————————————————————回复内容结果录入—————————————————————————————————————————————————

                    # 如果没有出现错误
                    if check_result :

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        #如果开启译后替换字典功能，则根据用户字典进行替换
                        if configurator.post_translation_switch :
                            print("[INFO] 你开启了译后修正功能，正在进行替换", '\n')
                            response_dict = configurator.replace_after_translation(response_dict)

                        # 如果原文是日语，则还原文本的首尾代码字符
                        if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                            response_dict = Cache_Manager.update_dictionary(self,response_dict, process_info_list)

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict,configurator.model_type)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if configurator.auto_backup_toggle:
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁

                        
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(1,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",1)

                        # 获取翻译进度
                        progress = user_interface_prompter.progress

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")
                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:

                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(0,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",1)

                        lock2.release()  # 释放锁

                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m 错误重新翻译最大次数限制:",configurator.retry_count_limit,"剩余可重试次数:",(configurator.retry_count_limit + 1 - Wrong_answer_count),"到达次数限制后，该段文本将进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count > configurator.retry_count_limit :
                            print("\033[1;33mWarning:\033[0m 错误回复重翻次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环              
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return



    # 整理发送内容（Cohere）
    def organize_send_content_cohere(self,source_text_dict, previous_list):
        #创建message列表，用于发送
        messages = []

        #获取基础系统提示词
        system_prompt = configurator.get_system_prompt()


        # 如果开启提示字典
        glossary_prompt = ""
        glossary_prompt_cot = ""
        if configurator.prompt_dictionary_switch :
            glossary_prompt,glossary_prompt_cot = configurator.build_glossary_prompt(source_text_dict,configurator.cn_prompt_toggle)
            if glossary_prompt :
                system_prompt += glossary_prompt 
                print("[INFO]  已添加术语表：\n",glossary_prompt)


        # 如果角色介绍开关打开
        characterization = ""
        characterization_cot = ""
        if configurator.characterization_switch :
            characterization,characterization_cot = configurator.build_characterization(source_text_dict,configurator.cn_prompt_toggle)
            if characterization:
                system_prompt += characterization 
                print("[INFO]  已添加角色介绍：\n",characterization)

        #如果背景设定开关打开
        world_building = ""
        world_building_cot = ""
        if configurator.world_building_switch :
            world_building,world_building_cot = configurator.build_world(configurator.cn_prompt_toggle)
            if world_building:
                system_prompt += world_building 
                print("[INFO]  已添加背景设定：\n",world_building)

        #如果文风要求开关打开
        writing_style = ""
        writing_style_cot = ""
        if configurator.writing_style_switch :
            writing_style,writing_style_cot = configurator.build_writing_style(configurator.cn_prompt_toggle)
            if writing_style:
                system_prompt += writing_style 
                print("[INFO]  已添加文风要求：\n",writing_style)

        # 获取默认示例前置文本
        pre_prompt = configurator.build_userExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)
        fol_prompt = configurator.build_modelExamplePrefix (configurator.cn_prompt_toggle,configurator.cot_toggle,configurator.source_language,configurator.target_language,glossary_prompt_cot,characterization_cot,world_building_cot,writing_style_cot)

        #构建默认示例
        original_exmaple,translation_example =  Configurator.build_translation_sample(self,source_text_dict,configurator.source_language,configurator.target_language)
        if original_exmaple and translation_example:
                
            the_original_exmaple =  {"role": "USER","message":(pre_prompt + original_exmaple) }
            the_translation_example = {"role": "CHATBOT", "message": (f'{fol_prompt}```json\n{translation_example}\n```') }

            messages.append(the_original_exmaple)
            messages.append(the_translation_example)
            print("[INFO]  已添加格式原文示例：\n",original_exmaple)
            print("[INFO]  已添加格式译文示例：\n",translation_example, '\n')


        #如果翻译示例开关打开
        if configurator.translation_example_switch :
            original_exmaple_3,translation_example_3 = configurator.build_translation_example ()
            if original_exmaple_3 and translation_example_3:
                the_original_exmaple =  {"role": "USER","message":original_exmaple_3 }
                the_translation_example = {"role": "CHATBOT", "message": translation_example_3}
                messages.append(the_original_exmaple)
                messages.append(the_translation_example)
                print("[INFO]  已添加用户原文示例：\n",original_exmaple_3)
                print("[INFO]  已添加用户译文示例：\n",translation_example_3, '\n')


        # 如果开启了保留换行符功能
        if configurator.preserve_line_breaks_toggle:
            print("[INFO] 你开启了保留换行符功能，正在进行替换", '\n')
            source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")


        #如果开启译前替换字典功能，则根据用户字典进行替换
        if configurator.pre_translation_switch :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_before_translation(source_text_dict)



        #如果加上文(cohere的模型注意力更多会注意在最新的消息，而导致过分关注最新消息的格式)
        previous = ""
        if configurator.pre_line_counts and previous_list :
            previous = configurator.build_pre_text(previous_list,configurator.cn_prompt_toggle)
            if previous:
                #system_prompt += previous
                pass 
                #print("[INFO]  已添加上文：\n",previous)


        # 获取提问时的前置文本
        pre_prompt = configurator.build_userQueryPrefix (configurator.cn_prompt_toggle,configurator.cot_toggle)


        #构建用户信息
        source_text_str = json.dumps(source_text_dict, ensure_ascii=False)   
        source_text_str = previous + "\n" +pre_prompt  + source_text_str
        #source_text_str = pre_prompt  + source_text_str


        return messages,source_text_str,system_prompt


    # 并发接口请求（Cohere）
    def Concurrent_Request_cohere(self):
        global cache_list,Running_status

        # 检查翻译任务是否已经暂停或者退出
        if Running_status == 9 or Running_status == 10 :
            return

        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            if configurator.tokens_limit_switch:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_tokens(self,configurator.tokens_limit, cache_list,configurator.pre_line_counts)  
            else:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_lines(self,configurator.lines_limit, cache_list,configurator.pre_line_counts)    
            lock1.release()  # 释放锁

            # 检查一下是否有发送内容
            if source_text_list == []:
                print("\033[1;33mWarning:\033[0m 未能获取文本，该线程为多余线程，取消任务")
                return
            # ——————————————————————————————————————————处理原文本的内容与格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # 如果原文是日语，清除文本首位中的代码文本，并记录清除信息
            if configurator.source_language == "日语"and configurator.text_clear_toggle:
                source_text_dict,process_info_list = Cache_Manager.process_dictionary(self,source_text_dict)
                row_count = len(source_text_dict)


            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str,system_prompt = Api_Requester.organize_send_content_cohere(self,source_text_dict, previous_list)

            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            # 计算请求的tokens预计花费
            prompt_tokens ={"role": "system","content": system_prompt }
            srt_tokens ={"role": "user","content": source_text_str }
            messages_tokens= messages.copy()
            messages_tokens.append(prompt_tokens)
            messages_tokens.append(srt_tokens)
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages_tokens) 

            # 计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str }] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text)
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;31mError:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;31mError:\033[0m 该条消息取消任务，进行拆分翻译" )
                return
            
            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 220   # 设置超时时间为x秒
            request_errors_count = 0 # 请求错误计数
            Wrong_answer_count = 0   # 错误回复计数

            while 1 :
                # 检查翻译任务是否已经暂停或者退出
                if Running_status == 9 or Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume )
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 获取apikey
                    cohere_apikey =  configurator.get_apikey()
                    # 创建anthropic客户端
                    client = cohere.Client(api_key=cohere_apikey,base_url=configurator.base_url)
                    # 发送对话请求
                    try:
                        response = client.chat(
                            model= configurator.model_type,
                            preamble= system_prompt,
                            message = source_text_str ,
                            chat_history = messages,
                            temperature=0
                            )



                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    # 检查翻译任务是否已经暂停或者退出，不进行接下来的处理了
                    if Running_status == 9 or Running_status == 10 :
                        return
                    

                    #——————————————————————————————————————————收到回复，获取返回的信息 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    try: # 因为有些中转网站不返回tokens消耗
                        prompt_tokens_used = 0
                        #prompt_tokens_used = int(response.usage.prompt_tokens) #本次请求花费的tokens
                    except Exception as e:
                        prompt_tokens_used = 0
                    try:
                        completion_tokens_used = 0
                        #completion_tokens_used = int(response.usage.completion_tokens) #本次回复花费的tokens
                    except Exception as e:
                        completion_tokens_used = 0



                    # 尝试提取回复的文本内容
                    try:
                        response_content = response.text 
                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 提取文本时出现问题！！！运行错误信息如下")
                        print(f"Error: {e}\n")
                        print("接口返回的错误信息如下")
                        print(response)
                        #处理完毕，再次进行请求
                        
                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break
                        continue


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                    # ——————————————————————————————————————————对回复内容处理,检查和录入——————————————————————————————————————————
                    # 处理回复内容
                    response_dict = Response_Parser.process_content(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,response_dict,source_text_dict,configurator.source_language)

                    # ———————————————————————————————————回复内容结果录入—————————————————————————————————————————————————

                    # 如果没有出现错误
                    if check_result :

                        # 如果开启了保留换行符功能
                        if configurator.preserve_line_breaks_toggle:
                            response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        #如果开启译后替换字典功能，则根据用户字典进行替换
                        if configurator.post_translation_switch :
                            print("[INFO] 你开启了译后修正功能，正在进行替换", '\n')
                            response_dict = configurator.replace_after_translation(response_dict)

                        # 如果原文是日语，则还原文本的首尾代码字符
                        if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                            response_dict = Cache_Manager.update_dictionary(self,response_dict, process_info_list)

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict,configurator.model_type)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if configurator.auto_backup_toggle:
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁

                        
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(1,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",1)

                        # 获取翻译进度
                        progress = user_interface_prompter.progress

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")
                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:

                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(0,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",1)

                        lock2.release()  # 释放锁

                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m 错误重新翻译最大次数限制:",configurator.retry_count_limit,"剩余可重试次数:",(configurator.retry_count_limit + 1 - Wrong_answer_count),"到达次数限制后，该段文本将进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count > configurator.retry_count_limit :
                            print("\033[1;33mWarning:\033[0m 错误回复重翻次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环              
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return



    # 整理发送内容（sakura）
    def organize_send_content_sakura(self,source_text_dict, previous_list):
        #创建message列表，用于发送
        messages = []

        #构建系统提示词
        if configurator.model_type != "Sakura-v0.9":
            system_prompt ={"role": "system","content": "你是一个轻小说翻译模型，可以流畅通顺地使用给定的术语表以日本轻小说的风格将日文翻译成简体中文，并联系上下文正确使用人称代词，注意不要混淆使役态和被动态的主语和宾语，不要擅自添加原文中没有的代词，也不要擅自增加或减少换行。" }
        else:
            system_prompt ={"role": "system","content": "你是一个轻小说翻译模型，可以流畅通顺地以日本轻小说的风格将日文翻译成简体中文，并联系上下文正确使用人称代词，不擅自添加原文中没有的代词。" }
        messages.append(system_prompt)


        # 开启了保留换行符功能
        print("[INFO] 正在使用SakuraLLM，将替换换行符为特殊符号", '\n')
        source_text_dict = Cache_Manager.replace_special_characters(self,source_text_dict, "替换")

        #如果开启译前替换字典功能
        if configurator.pre_translation_switch :
            print("[INFO] 你开启了译前替换字典功能，正在进行替换", '\n')
            source_text_dict = configurator.replace_before_translation(source_text_dict)



        #如果开启了译时提示字典功能
        converted_list = [] # 创建一个空列表来存储转换后的字符串
        if (configurator.prompt_dictionary_switch) and (configurator.model_type != "Sakura-v0.9"):
            glossary_prompt = configurator.build_glossary_prompt_sakura(source_text_dict)
            if glossary_prompt:
                gpt_dict_text_list = []
                for gpt in glossary_prompt:
                    src = gpt['src']
                    dst = gpt['dst']
                    info = gpt['info'] if "info" in gpt.keys() else None
                    if info:
                        single = f"{src}->{dst} #{info}"
                    else:
                        single = f"{src}->{dst}"
                    gpt_dict_text_list.append(single)

                gpt_dict_raw_text = "\n".join(gpt_dict_text_list)
                print("[INFO]  检测到请求的原文中含有提示字典内容")
                print("[INFO]  术语表:\n",gpt_dict_raw_text,"\n")

 
        #将原文本字典转换成raw格式的字符串
        source_text_str_raw = self.convert_dict_to_raw_str(source_text_dict)

        # 处理全角数字
        source_text_str_raw = self.convert_fullwidth_to_halfwidth(source_text_str_raw)

        #构建需要翻译的文本
        if converted_list:
            user_prompt = "根据以下术语表（可以为空）：\n" + gpt_dict_raw_text + "\n\n" + "将下面的日文文本根据上述术语表的对应关系和备注翻译成中文：" + source_text_str_raw
        else:
            if configurator.model_type != "Sakura-v0.9":
                user_prompt = "根据以下术语表（可以为空）：\n\n\n" + "将下面的日文文本根据上述术语表的对应关系和备注翻译成中文：" + source_text_str_raw
            else:
                user_prompt = "将下面的日文文本翻译成中文：" + source_text_str_raw

        Original_text = {"role":"user","content": user_prompt}


        messages.append(Original_text)



        return messages, source_text_str_raw


    # 并发接口请求（sakura）
    def concurrent_request_sakura(self):
        global cache_list,Running_status

        # 检查翻译任务是否已经暂停或者退出
        if Running_status == 9 or Running_status == 10 :
            return

        try:#方便排查子线程bug

            # ——————————————————————————————————————————截取需要翻译的原文本——————————————————————————————————————————
            lock1.acquire()  # 获取锁
            # 获取设定行数的文本，并修改缓存文件里的翻译状态为2，表示正在翻译中
            if configurator.tokens_limit_switch:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_tokens(self,configurator.tokens_limit, cache_list,configurator.pre_line_counts)  
            else:
                source_text_list, previous_list = Cache_Manager.process_dictionary_data_lines(self,configurator.lines_limit, cache_list,configurator.pre_line_counts)    
            lock1.release()  # 释放锁

            # 检查一下是否有发送内容
            if source_text_list == []:
                print("\033[1;33mWarning:\033[0m 未能获取文本，该线程为多余线程，取消任务")
                return

            # ——————————————————————————————————————————处理原文本的内容与格式——————————————————————————————————————————
            # 将原文本列表改变为请求格式
            source_text_dict, row_count = Cache_Manager.create_dictionary_from_list(self,source_text_list)  

            # 如果原文是日语，清除文本首位中的代码文本，并记录清除信息
            if configurator.source_language == "日语" and configurator.text_clear_toggle:
                source_text_dict,process_info_list = Cache_Manager.process_dictionary(self,source_text_dict)
                row_count = len(source_text_dict)


            # ——————————————————————————————————————————整合发送内容——————————————————————————————————————————        
            messages,source_text_str = Api_Requester.organize_send_content_sakura(self,source_text_dict, previous_list)



            #——————————————————————————————————————————检查tokens发送限制——————————————————————————————————————————
            #计算请求的tokens预计花费
            request_tokens_consume = Request_Limiter.num_tokens_from_messages(self,messages)  #加上2%的修正系数
            #计算回复的tokens预计花费，只计算发送的文本，不计算提示词与示例，可以大致得出
            Original_text = [{"role":"user","content":source_text_str}] # 需要拿列表来包一层，不然计算时会出错 
            completion_tokens_consume = Request_Limiter.num_tokens_from_messages(self,Original_text) #加上2%的修正系数
 
            if request_tokens_consume >= request_limiter.max_tokens :
                print("\033[1;33mWarning:\033[0m 该条消息总tokens数大于单条消息最大数量" )
                print("\033[1;33mWarning:\033[0m 该条消息取消任务，进行拆分翻译" )
                return
            
            # ——————————————————————————————————————————开始循环请求，直至成功或失败——————————————————————————————————————————
            start_time = time.time()
            timeout = 220   # 设置超时时间为x秒
            request_errors_count = 0 # 设置请求错误次数限制
            Wrong_answer_count = 0   # 设置错误回复次数限制
            model_degradation = False # 模型退化检测

            while 1 :
                # 检查翻译任务是否已经暂停或者退出
                if Running_status == 9 or Running_status == 10 :
                    return

                #检查子线程运行是否超时---------------------------------
                if time.time() - start_time > timeout:
                    print("\033[1;31mError:\033[0m 子线程执行任务已经超时，将暂时取消本次任务")
                    break


                # 检查是否符合速率限制---------------------------------
                if request_limiter.RPM_and_TPM_limit(request_tokens_consume):


                    print("[INFO] 已发送请求,正在等待AI回复中-----------------------")
                    print("[INFO] 请求与回复的tokens数预计值是：",request_tokens_consume  + completion_tokens_consume )
                    print("[INFO] 当前发送的原文文本：\n", source_text_str)

                    # ——————————————————————————————————————————发送会话请求——————————————————————————————————————————
                    # 记录开始请求时间
                    Start_request_time = time.time()

                    # 获取AI的参数设置
                    temperature,top_p,frequency_penalty= configurator.get_sakura_parameters()
                    # 如果上一次请求出现模型退化，更改参数
                    if model_degradation:
                        frequency_penalty = 0.2


                    extra_query = {
                        'do_sample': False,
                        'num_beams': 1,
                        'repetition_penalty': 1.0,
                    }

                    # 获取apikey
                    openai_apikey =  configurator.get_apikey()
                    # 获取请求地址
                    openai_base_url = configurator.base_url
                    # 创建openai客户端
                    openaiclient = OpenAI(api_key=openai_apikey,
                                            base_url= openai_base_url)
                    # 发送对话请求
                    try:
                        response = openaiclient.chat.completions.create(
                            model= configurator.model_type,
                            messages = messages ,
                            temperature=temperature,
                            top_p = top_p,                        
                            frequency_penalty=frequency_penalty,

                            max_tokens=512,
                            seed=-1,
                            extra_query=extra_query,
                            )

                    #抛出错误信息
                    except Exception as e:
                        print("\033[1;31mError:\033[0m 进行请求时出现问题！！！错误信息如下")
                        print(f"Error: {e}\n")

                        #请求错误计次
                        request_errors_count = request_errors_count + 1
                        #如果错误次数过多，就取消任务
                        if request_errors_count >= 4 :
                            print("\033[1;31m[ERROR]\033[0m 请求发生错误次数过多，该线程取消任务！")
                            break

                        #处理完毕，再次进行请求
                        continue


                    # 检查翻译任务是否已经暂停或者退出，不进行接下来的处理了
                    if Running_status == 9 or Running_status == 10 :
                        return
                    
                    
                    #——————————————————————————————————————————收到回复，获取返回的信息 ————————————————————————————————————————  
                    # 计算AI回复花费的时间
                    response_time = time.time()
                    Request_consumption_time = round(response_time - Start_request_time, 2)


                    # 计算本次请求的花费的tokens
                    try: # 因为有些中转网站不返回tokens消耗
                        prompt_tokens_used = int(response.usage.prompt_tokens) #本次请求花费的tokens
                    except Exception as e:
                        prompt_tokens_used = 0
                    try:
                        completion_tokens_used = int(response.usage.completion_tokens) #本次回复花费的tokens
                    except Exception as e:
                        completion_tokens_used = 0



                    # 提取回复的文本内容
                    response_content = response.choices[0].message.content 


                    print('\n' )
                    print("[INFO] 已成功接受到AI的回复-----------------------")
                    print("[INFO] 该次请求已消耗等待时间：",Request_consumption_time,"秒")
                    print("[INFO] 本次请求与回复花费的总tokens是：",prompt_tokens_used + completion_tokens_used)
                    print("[INFO] AI回复的文本内容：\n",response_content ,'\n','\n')

                    # ——————————————————————————————————————————对回复内容处理,检查——————————————————————————————————————————
                    # 见raw格式转换为josn格式字符串
                    response_content = Response_Parser.convert_str_to_json_str(self, row_count, response_content)

                    # 处理回复内容
                    response_dict = Response_Parser.process_content(self,response_content)

                    # 检查回复内容
                    check_result,error_content =  Response_Parser.check_response_content(self,response_content,response_dict,source_text_dict,configurator.source_language)

                    # ——————————————————————————————————————————回复内容结果录入——————————————————————————————————————————

                    # 如果没有出现错误
                    if check_result :

                        # 强制开启换行符还原功能
                        response_dict = Cache_Manager.replace_special_characters(self,response_dict, "还原")

                        #如果开启译后替换字典功能，则根据用户字典进行替换
                        if configurator.post_translation_switch :
                            print("[INFO] 你开启了译后修正功能，正在进行替换", '\n')
                            response_dict = configurator.replace_after_translation(response_dict)

                        # 如果原文是日语，则还原文本的首尾代码字符
                        if (configurator.source_language == "日语" and configurator.text_clear_toggle):
                            response_dict = Cache_Manager.update_dictionary(self,response_dict, process_info_list)

                        # 录入缓存文件
                        lock1.acquire()  # 获取锁
                        Cache_Manager.update_cache_data(self,cache_list, source_text_list, response_dict,configurator.model_type)
                        lock1.release()  # 释放锁


                        # 如果开启自动备份,则自动备份缓存文件
                        if configurator.auto_backup_toggle:
                            lock3.acquire()  # 获取锁

                            # 创建存储缓存文件的文件夹，如果路径不存在，创建文件夹
                            output_path = os.path.join(configurator.Output_Folder, "cache")
                            os.makedirs(output_path, exist_ok=True)
                            # 输出备份
                            File_Outputter.output_cache_file(self,cache_list,output_path)
                            lock3.release()  # 释放锁

                        
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(1,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译成功",1)

                        # 获取翻译进度
                        progress = user_interface_prompter.progress                    

                        print(f"\n--------------------------------------------------------------------------------------")
                        print(f"\n\033[1;32mSuccess:\033[0m AI回复内容检查通过！！！已翻译完成{progress}%")
                        print(f"\n--------------------------------------------------------------------------------------\n")
                        lock2.release()  # 释放锁


                        break
                

                    # 如果出现回复错误
                    else:

                        # 更改UI界面信息
                        lock2.acquire()  # 获取锁

                        # 更新翻译界面数据
                        user_interface_prompter.update_data(0,row_count,prompt_tokens_used,completion_tokens_used)

                        # 更改UI界面信息,注意，传入的数值类型分布是字符型与整数型，小心浮点型混入
                        user_interface_prompter.signal.emit("更新翻译界面数据","翻译失败",1)

                        lock2.release()  # 释放锁

                        print("\033[1;33mWarning:\033[0m AI回复内容存在问题:",error_content,"\n")

                        # 检查一下是不是模型退化
                        if error_content == "AI回复内容出现高频词,并重新翻译":
                            print("\033[1;33mWarning:\033[0m 下次请求将修改参数，回避高频词输出","\n")
                            model_degradation = True

                        #错误回复计次
                        Wrong_answer_count = Wrong_answer_count + 1
                        print("\033[1;33mWarning:\033[0m 错误重新翻译最大次数限制:",configurator.retry_count_limit,"剩余可重试次数:",(configurator.retry_count_limit + 1 - Wrong_answer_count),"到达次数限制后，该段文本将进行拆分翻译\n")
                        #检查回答错误次数，如果达到限制，则跳过该句翻译。
                        if Wrong_answer_count > configurator.retry_count_limit :
                            print("\033[1;33mWarning:\033[0m 错误回复重翻次数已经达限制,将该段文本进行拆分翻译！\n")    
                            break


                        #进行下一次循环              
                        continue

    #子线程抛出错误信息
        except Exception as e:
            print("\033[1;31mError:\033[0m 子线程运行出现问题！错误信息如下")
            print(f"Error: {e}\n")
            return


    # 将json文本改为纯文本
    def convert_dict_to_raw_str(self,source_text_dict):
        str_list = []
        for idx in range(len(source_text_dict.keys())):
            # str_list.append(s['source_text'])
            str_list.append(source_text_dict[f"{idx}"])
        raw_str = "\n".join(str_list)
        return raw_str


    # 将列表中的字符串中的全角数字转换为半角数字
    def convert_fullwidth_to_halfwidth(self,input_string):
        modified_string = ""
        for char in input_string:
            if '０' <= char <= '９':  # 判断是否为全角数字
                modified_string += chr(ord(char) - ord('０') + ord('0'))  # 转换为半角数字
            else:
                modified_string += char

        return modified_string


# 回复解析器
class Response_Parser():
    def __init__(self):
        pass
    

    #处理并正则提取翻译内容
    def process_content(self,input_str):

        # 尝试直接转换为json字典
        try:
            response_dict = json.loads(input_str) 
            return response_dict
        except :
            # 对格式进行修复       
            input_str = Response_Parser.repair_double_quotes(self,input_str)
            input_str = Response_Parser.repair_double_quotes_2(self,input_str)  
            #input_str = Response_Parser.repair_double_quotes_3(self,input_str)       


        # 再次尝试直接转换为json字典
        try:
            response_dict = json.loads(input_str) 
            return response_dict
        except :       
            pass


        # 尝试正则提取
        try:

            # 使用正则表达式匹配字符串中的所有键值对
            parsed_dict = {}
            key = None
            key_pattern = r'("\d+?"\s*:\s*)'
            lines = re.split(key_pattern, input_str)
            for line in lines:
                if re.match(key_pattern, line):
                    key = re.findall(r'\d+', line)[0]
                    continue
                if not key: continue
                value = re.findall(r'"([\S\s]+)"(?=,|}|\n)', line)
                if value:
                    if key not in parsed_dict:
                        parsed_dict[key] = value[0]
                key = None


            return parsed_dict
        except :       
            print("\033[1;33mWarning:\033[0m 回复内容无法正常提取，请反馈\n") 
            return {}


    # 将Raw文本恢复根据行数转换成json文本
    def convert_str_to_json_str(self,row_count, input_str):

        # 当发送文本为1行时，就不分割了，以免切错
        if row_count == 1:
            result = {"0": input_str}
            return  json.dumps(result, ensure_ascii=False)
        
        else:
            str_list = input_str.split("\n")
            ret_json = {}
            for idx, text in enumerate(str_list):
                ret_json[f"{idx}"] = f"{text}"
            return json.dumps(ret_json, ensure_ascii=False)


    # 修复value前面的双引号
    def repair_double_quotes(self,text):
        # 消除文本中的空格
        text = text.replace(" ", "")
        # 正则表达式匹配双引号后跟冒号，并捕获第三个字符
        pattern = r'[\"]:(.)'
        # 使用finditer来找到所有匹配项
        matches = re.finditer(pattern, text)
        # 存储所有修改的位置
        modifications = [(match.start(1), match.group(1)) for match in matches]

        # 从后往前替换文本，这样不会影响后续匹配的位置
        for start, char in reversed(modifications):
            if char != '"':
                text = text[:start] + '"' + text[start:]

        return text

    # 修复value后面的双引号
    def repair_double_quotes_2(self,text):
        # 消除文本中的空格
        text = text.replace(" ", "")

        # 正则表达式匹配逗号后面跟换行符（可选）,再跟双引号的模式
        pattern = r',(?:\n)?\"'
        matches = re.finditer(pattern, text)
        result = []

        last_end = 0
        for match in matches:
            # 获取逗号前的字符
            quote_position = match.start()
            before_quote = text[quote_position - 1]
            
            # 检查逗号前的字符是否是双引号
            if before_quote == '"':
                # 如果是双引号，将这一段文本加入到结果中
                result.append(text[last_end:quote_position])
            else:
                # 如果不是双引号，将前一个字符换成'"'
                result.append(text[last_end:quote_position - 1] + '"')
            
            # 更新最后结束的位置
            last_end = quote_position

        # 添加剩余的文本
        result.append(text[last_end:])

        # 将所有片段拼接起来
        return ''.join(result)

    # 修复大括号前面的双引号
    def repair_double_quotes_3(self,text):
        # 消除文本中的空格
        text = text.replace(" ", "")

        # 正则表达式匹配逗号后面紧跟双引号的模式
        pattern = r'(?:\n)?}'
        matches = re.finditer(pattern, text)
        result = []

        last_end = 0
        for match in matches:
            # 获取逗号前的字符
            quote_position = match.start()
            before_quote = text[quote_position - 1]
            
            # 检查逗号前的字符是否是双引号
            if before_quote == '"':
                # 如果是双引号，将这一段文本加入到结果中
                result.append(text[last_end:quote_position])
            else:
                # 如果不是双引号，将前一个字符换成'"'
                result.append(text[last_end:quote_position - 1] + '"')
            
            # 更新最后结束的位置
            last_end = quote_position

        # 添加剩余的文本
        result.append(text[last_end:])

        # 将所有片段拼接起来
        return ''.join(result)

    # 检查回复内容是否存在问题
    def check_response_content(self,response_str,response_dict,source_text_dict,source_language):
        # 存储检查结果
        check_result = False
        # 存储错误内容
        error_content = "0"


        # 检查模型是否退化，出现高频词（只检测中日）
        if Response_Parser.model_degradation_detection(self,response_str):
            pass

        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容出现高频词,并重新翻译"
            return check_result,error_content


        # 检查文本行数
        if Response_Parser.check_text_line_count(self,source_text_dict,response_dict):
            pass
        else:
            check_result = False
            # 存储错误内容
            error_content = "提取到的文本行数与原来数量不符合,将进行重新翻译"
            return check_result,error_content


        # 检查文本空行
        if Response_Parser.check_empty_response(self,response_dict):
            pass
        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容中有未进行翻译的空行,将进行重新翻译"
            return check_result,error_content

        
        # 检查是否回复了原文
        if Response_Parser.check_dicts_equal(self,source_text_dict,response_dict):
            pass
        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容与原文相同，未进行翻译，将重新翻译"
            return check_result,error_content

        # 检查是否残留部分原文
        if Response_Parser.detecting_remaining_original_text(self,source_text_dict,response_dict,source_language):
            pass
        else:
            check_result = False
            # 存储错误内容
            error_content = "AI回复内容残留部分原文，未进行翻译，将重新翻译"
            return check_result,error_content


        # 如果检查都没有问题
        check_result = True
        # 存储错误内容
        error_content = "检查无误"
        return check_result,error_content


    # 检查两个字典是否完全相同，即返回了原文
    def check_dicts_equal(self,dict1, dict2):
        if len(dict1) >=3 :
            i = 0
            s = 0
            for key, value in dict1.items():
                value2 = dict2[key]

                # 将字符串转换为集合形式
                set1 = set(value)
                set2 = set(value2)

                # 计算交集和并集的大小
                intersection_size = len(set1.intersection(set2))
                union_size = len(set1.union(set2))
                
                # 计算Jaccard相似系数
                similarity = intersection_size / union_size

                #累加与累计
                i = i + 1 
                s = s + similarity

            result = s/i

            if (result>= 0.80):
                return False
            else:
                return True

        else:
            return True
 
    # 检查回复内容的文本行数
    def check_text_line_count(self,source_text_dict,response_dict): 
        """
        检查字典d中是否包含从'0'到'(N-1)'的字符串键

        :param d: 输入的字典
        :param N: 数字N
        :return: 如果字典包含从'0'到'(N-1)'的所有字符串键，则返回True，否则返回False
        """
        N = len(source_text_dict) 
        return all(str(key) in response_dict for key in range(N))

    # 检查翻译内容是否有空值
    def check_empty_response(self,response_dict):
        for value in response_dict.values():
            #检查value是不是None，因为AI回回复null，但是json.loads()会把null转化为None
            if value is None:
                return False

            # 检查value是不是空字符串，因为AI回回复空字符串，但是json.loads()会把空字符串转化为""
            if value == "":
                return False

        return True

    # 模型退化检测，高频语气词
    def model_degradation_detection(self, s, count=80):
        """
        检查字符串中是否存在任何字符连续出现指定次数。

        :param s: 输入的字符串
        :param count: 需要检查的连续次数，默认为80
        :return: 如果存在字符连续出现指定次数，则返回False，否则返回True
        """
        for i in range(len(s) - count + 1):
            if len(set(s[i:i+count])) == 1:
                return False
        return True
    

    # 检查残留原文的算法
    def detecting_remaining_original_text(self,dict1, dict2, language):

        # 考量到代码文本，英语不作检查
        if language == "英语":
            return True

        # 定义不同语言的正则表达式
        patterns_all = {
            '日语': re.compile(
                r'['
                r'\u3041-\u3096'  # 平假名
                r'\u30A0-\u30FF'  # 片假名
                r']+', re.UNICODE
            ),
            '韩语': re.compile(
                r'['
                r'\uAC00-\uD7AF'  # 韩文字母
                r']+', re.UNICODE
            ),
            '俄语': re.compile(
                r'['
                r'\u0400-\u04FF'  # 俄语字母
                r']+', re.UNICODE
            ),
            '简中': re.compile(
                r'['
                r'\u4E00-\u9FA5'  # 简体汉字
                r']+', re.UNICODE
            ),
            '繁中': re.compile(
                r'['
                r'\u3400-\u4DBF'  # 扩展A区汉字
                r'\u4E00-\u9FFF'  # 基本汉字
                r'\uF900-\uFAFF'  # 兼容汉字
                r']+', re.UNICODE
            ),
        }
        # 根据语言选择合适的正则表达式
        pattern = patterns_all.get(language)
        if not pattern:
            raise ValueError("Unsupported language")

        # 存储计数结果的字典
        count_results = 0

        # 遍历字典2中的每个键值对
        for key2, value2 in dict2.items():
            # 检查字典1中是否有对应的键
            if key2 in dict1:
                # 提取字典1值中的文本
                text1 = dict1[key2]
                # 提取字典2值中的指定语言的文本
                text2 = pattern.findall(value2)
                # 将列表转换为字符串
                text2_str = ''.join(text2)
                # 如果字典2中的残留文本在字典1中的文本中出现，则计数加1
                if text2_str and (text2_str in text1):
                    count_results += 1

        # 避免检查单或者少行字典
        if len(dict2) >5 :
            if  count_results >=2:
                return False

        return True             


# 接口测试器
class Request_Tester():
    def __init__(self):
        pass

    # 接口测试分发
    def request_test(self,platform,base_url,model_type,api_key_str,proxy_port):

        # 执行openai接口测试
        if platform == "OpenAI":
            Request_Tester.openai_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行google接口测试
        elif platform == "Google":
            Request_Tester.google_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行anthropic接口测试
        elif platform == "Anthropic":
            Request_Tester.anthropic_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行cohere接口测试
        elif platform == "Cohere":
            Request_Tester.cohere_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行智谱接口测试
        elif platform == "Zhipu":
            Request_Tester.openai_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行月之暗面接口测试
        elif platform == "Moonshot":
            Request_Tester.openai_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行Deepseek接口测试
        elif platform == "Deepseek":
            Request_Tester.openai_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行Dashscope接口测试
        elif platform == "Dashscope":
            Request_Tester.openai_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行Volcengine接口测试
        elif platform == "Volcengine":
            Request_Tester.openai_request_test(self,base_url,model_type,api_key_str,proxy_port)

        # 执行Sakura接口测试
        elif platform == "Sakura":
            Request_Tester.sakura_request_test(self,base_url,model_type,api_key_str,proxy_port)


    # openai接口测试
    def openai_request_test(self,base_url,model_type,api_key_str,proxy_port):
        
        print("[INFO] 正在测试openai类接口",'\n')

        #如果填入地址，则设置系统代理
        if proxy_port :
            print("[INFO] 系统代理端口是:",proxy_port,'\n') 
            os.environ["http_proxy"]=proxy_port
            os.environ["https_proxy"]=proxy_port

        
        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = api_key_str.replace('\n','').replace(" ", "").split(",")

        #检查一下请求地址尾部是否为/v1，自动补全,如果是/v4，则是在调用智谱接口，如果是/v3，则是豆包
        if base_url[-3:] != "/v1" and base_url[-3:] != "/v4" and base_url[-3:] != "/v3" :
            base_url = base_url + "/v1"

        #创建openai客户端
        client = OpenAI(api_key=API_key_list[0],
                base_url= base_url)


        print("[INFO] 请求地址是:",base_url,'\n')
        print("[INFO] 模型选择是:",model_type,'\n')

        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            client.api_key = API_key_list[i]

            #构建发送内容
            messages_test = [{"role": "system","content":"你不是AI助手之类，你是我的女朋友欣雨。接下来你必须以女朋友的方式回复我"}, {"role":"user","content":"小可爱，你在干嘛"}]
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')

            #尝试请求，并设置各种参数
            try:
                response_test = client.chat.completions.create( 
                model= model_type,
                messages = messages_test ,
                ) 

                #如果回复成功，显示成功信息
                response_test = response_test.choices[0].message.content
                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0)


    # google接口测试
    def google_request_test(self,base_url,model_type,api_key_str,proxy_port):

        print("[INFO] 正在测试Google接口",'\n')

        #如果填入地址，则设置系统代理
        if proxy_port :
            print("[INFO] 系统代理端口是:",proxy_port,'\n') 
            os.environ["http_proxy"]=proxy_port
            os.environ["https_proxy"]=proxy_port


        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = api_key_str.replace('\n','').replace(" ", "").split(",")


        print("[INFO] 请求地址是:",base_url,'\n')
        print("[INFO] 模型选择是:",model_type,'\n')

        # 设置ai参数
        generation_config = {
        "temperature": 0,
        "top_p": 1,
        "top_k": 1,
        "max_output_tokens": 2048, #最大输出，pro最大输出是2048
        }

        #调整安全限制
        safety_settings = [
        {
            "category": "HARM_CATEGORY_HARASSMENT",
            "threshold": "BLOCK_NONE"
        },
        {
            "category": "HARM_CATEGORY_HATE_SPEECH",
            "threshold": "BLOCK_NONE"
        },
        {
            "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
            "threshold": "BLOCK_NONE"
        },
        {
            "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
            "threshold": "BLOCK_NONE"
        }
        ]


        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            genai.configure(api_key= API_key_list[i],transport='rest') 

            #构建发送内容
            system_prompt = "你是我的女朋友欣雨。接下来你必须以女朋友的方式向我问好"
            messages_test = ["你在干嘛呢？",]
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')


            #设置对话模型
            model = genai.GenerativeModel(model_name=model_type,
                            generation_config=generation_config,
                            safety_settings=safety_settings,
                            system_instruction = system_prompt
                            )


            #尝试请求，并设置各种参数
            try:
                response_test =  model.generate_content(messages_test)

                #如果回复成功，显示成功信息
                response_test = response_test.text
                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0)


    # anthropic接口测试
    def anthropic_request_test(self,base_url,model_type,api_key_str,proxy_port):
        
        print("[INFO] 正在测试Anthropic接口",'\n')

        #如果填入地址，则设置系统代理
        if proxy_port :
            print("[INFO] 系统代理端口是:",proxy_port,'\n') 
            os.environ["http_proxy"]=proxy_port
            os.environ["https_proxy"]=proxy_port


        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = api_key_str.replace('\n','').replace(" ", "").split(",")


        #创建客户端
        client = anthropic.Anthropic(
            base_url=base_url,
            api_key=API_key_list[0]
        )



        print("[INFO] 请求地址是:",base_url,'\n')
        print("[INFO] 模型选择是:",model_type,'\n')


        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            client.api_key = API_key_list[i]

            #构建发送内容
            messages_test = [ {"role":"user","content":"小可爱，你在干嘛"}]
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')


            #尝试请求，并设置各种参数
            try:
                response_test = client.messages.create(
                model= model_type,
                max_tokens=1000,
                system="你是我的女朋友欣雨。接下来你必须以女朋友的方式回复我",
                messages = messages_test ,
                ) 

            #    #如果回复成功，显示成功信息
                response_test = response_test.content[0].text


                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0)


    # cohere接口测试
    def cohere_request_test(self,base_url,model_type,api_key_str,proxy_port):
        
        print("[INFO] 正在测试Cohere接口",'\n')

        #如果填入地址，则设置系统代理
        if proxy_port :
            print("[INFO] 系统代理端口是:",proxy_port,'\n') 
            os.environ["http_proxy"]=proxy_port
            os.environ["https_proxy"]=proxy_port

        
        #分割KEY字符串并存储进列表里,如果API_key_str中没有逗号，split(",")方法仍然返回一个只包含一个元素的列表
        API_key_list = api_key_str.replace('\n','').replace(" ", "").split(",")


        #创建openai客户端
        client = cohere.Client(api_key=API_key_list[0],
                base_url= base_url)


        print("[INFO] 请求地址是:",base_url,'\n')
        print("[INFO] 模型选择是:",model_type,'\n')

        #创建存储每个key测试结果的列表
        test_results = [None] * len(API_key_list)


        #循环测试每一个apikey情况
        for i, key in enumerate(API_key_list):
            print(f"[INFO] 正在测试第{i+1}个API KEY：{key}",'\n') 

            #更换key
            client.api_key = API_key_list[i]

            #构建发送内容
            messages_test = "小可爱，你在干嘛"
            print("[INFO] 当前发送内容：\n", messages_test ,'\n')

            #尝试请求，并设置各种参数
            try:
                response_test = client.chat( 
                preamble= "你不是AI助手之类，你是我的女朋友欣雨。接下来你必须以女朋友的方式回复我",
                model= model_type,
                message = "小可爱，你在干嘛" ,
                ) 

                #如果回复成功，显示成功信息
                response_test = response_test.text
                print("[INFO] 已成功接受到AI的回复")
                print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

                test_results[i] = 1 #记录成功结果

            #如果回复失败，抛出错误信息，并测试下一个key
            except Exception as e:
                print("\033[1;31mError:\033[0m key：",API_key_list[i],"请求出现问题！错误信息如下")
                print(f"Error: {e}\n\n")
                test_results[i] = 0 #记录错误结果
                continue


        # 输出每个API密钥测试的结果
        print("[INFO] 全部API KEY测试结果--------------")
        for i, key in enumerate(API_key_list):
            result = "成功" if test_results[i] == 1 else "失败"
            print(f"第{i+1}个 API KEY：{key} 测试结果：{result}")

        # 检查测试结果是否全部成功
        all_successful = all(result == 1 for result in test_results)
        # 输出总结信息
        if all_successful:
            print("[INFO] 所有API KEY测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0)
        else:
            print("[INFO] 存在API KEY测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0)


    # sakura接口测试
    def sakura_request_test(self,base_url,model_type,api_key_str,proxy_port):

        print("[INFO] 正在测试Sakura接口",'\n')

        #如果填入地址，则设置系统代理
        if proxy_port :
            print("[INFO] 系统代理端口是:",proxy_port,'\n') 
            os.environ["http_proxy"]=proxy_port
            os.environ["https_proxy"]=proxy_port

        

        #检查一下请求地址尾部是否为/v1，自动补全
        if base_url[-3:] != "/v1":
            base_url = base_url + "/v1"

        #创建openai客户端
        openaiclient = OpenAI(api_key="sakura",
                base_url= base_url)


        print("[INFO] 请求地址是:",base_url,'\n')
        print("[INFO] 模型选择是:",model_type,'\n')



        #构建发送内容
        messages_test = [{"role": "system","content":"你是一个轻小说翻译模型，可以流畅通顺地以日本轻小说的风格将日文翻译成简体中文，并联系上下文正确使用人称代词，不擅自添加原文中没有的代词。"},
                         {"role":"user","content":"将下面的日文文本翻译成中文：サポートキャスト"}]
        print("[INFO] 当前发送内容：\n", messages_test ,'\n')

        #尝试请求，并设置各种参数
        try:
            response_test = openaiclient.chat.completions.create( 
            model= model_type,
            messages = messages_test ,
            ) 

            #如果回复成功，显示成功信息
            response_test = response_test.choices[0].message.content
            print("[INFO] 已成功接受到AI的回复")
            print("[INFO] AI回复的文本内容：\n",response_test ,'\n','\n')

            print("[INFO] 模型通讯测试成功！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试成功",0)

        #如果回复失败，抛出错误信息，并测试下一个key
        except Exception as e:
            print("\033[1;31mError:\033[0m 请求出现问题！错误信息如下")
            print(f"Error: {e}\n\n")
            print("[INFO] 模型通讯测试失败！！！！")
            user_interface_prompter.signal.emit("接口测试结果","测试失败",0)



# 配置器
class Configurator():
    def __init__(self):
        self.translation_project = "" # 翻译项目
        self.translation_platform = "" # 翻译平台
        self.source_language = "" # 文本原语言
        self.target_language = "" # 文本目标语言
        self.Input_Folder = "" # 存储输入文件夹
        self.Output_Folder = "" # 存储输出文件夹

        self.lines_limit = 1 # 存储每次请求的文本行数设置
        self.thread_counts = 1 # 存储线程数
        self.retry_count_limit = 1 # 错误回复重试次数
        self.pre_line_counts = 0 # 上文行数
        self.cot_toggle = False # 思维链开关
        self.cn_prompt_toggle = False # 中文提示词开关
        self.text_clear_toggle = False # 清除首位非文本字符开关
        self.preserve_line_breaks_toggle = False # 保留换行符开关
        self.conversion_toggle = False #简繁转换开关

        self.mixed_translation_toggle = False # 混合翻译开关
        self.retry_count_limit = 1 # 错误回复重试次数限制
        self.round_limit = 6 # 拆分翻译轮次限制
        self.split_switch = False # 拆分开关
        self.configure_mixed_translation = {"first_platform":"first_platform",
                                            "second_platform":"second_platform",
                                            "third_platform":"third_platform",}  #混合翻译相关信息


        self.prompt_dictionary_switch = False   #   提示字典开关
        self.pre_translation_switch = False #   译前处理开关
        self.post_translation_switch = False #   译后处理开关
        self.custom_prompt_switch = False #   自定义prompt开关
        self.add_example_switch = False #   添加示例开关


        self.model_type = ""             #模型选择
        self.apikey_list = [] # 存储key的列表
        self.key_index = 0  # 方便轮询key的索引
        self.base_url = 'https://api.openai.com/v1' # api请求地址


        self.openai_temperature = 0        #AI的随机度，0.8是高随机，0.2是低随机,取值范围0-2
        self.openai_top_p = 0              #AI的top_p，作用与temperature相同，官方建议不要同时修改
        self.openai_presence_penalty = 0  #AI的存在惩罚，生成新词前检查旧词是否存在相同的词。0.0是不惩罚，2.0是最大惩罚，-2.0是最大奖励
        self.openai_frequency_penalty = 0 #AI的频率惩罚，限制词语重复出现的频率。0.0是不惩罚，2.0是最大惩罚，-2.0是最大奖励
        self.anthropic_temperature   =  0 
        self.google_temperature   =  0 


    # 初始化配置信息
    def initialize_configuration (self):
        global Running_status,resource_dir


        #读取用户配置config.json
        if os.path.exists(os.path.join(resource_dir, "config.json")):
            with open(os.path.join(resource_dir, "config.json"), "r", encoding="utf-8") as f:
                config_dict = json.load(f)


        #读取各平台配置信息
        if os.path.exists(os.path.join(resource_dir, "platform", "openai.json")):
            #读取各平台配置信息
            with open(os.path.join(resource_dir, "platform", "openai.json"), "r", encoding="utf-8") as f:
                self.openai_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "anthropic.json"), "r", encoding="utf-8") as f:
                self.anthropic_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "google.json"), "r", encoding="utf-8") as f:
                self.google_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "cohere.json"), "r", encoding="utf-8") as f:
                self.cohere_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "deepseek.json"), "r", encoding="utf-8") as f:
                self.deepseek_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "dashscope.json"), "r", encoding="utf-8") as f:
                self.dashscope_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "moonshot.json"), "r", encoding="utf-8") as f:
                self.moonshot_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "zhipu.json"), "r", encoding="utf-8") as f:
                self.zhipu_platform_config = json.load(f)
            with open(os.path.join(resource_dir, "platform", "sakurallm.json"), "r", encoding="utf-8") as f:
                self.sakurallm_platform_config = json.load(f)
  

        #获取OpenAI官方账号界面
        self.openai_account_type = config_dict["openai_account_type"]   #获取账号类型下拉框当前选中选项的值
        self.openai_model_type = config_dict["openai_model_type"] #获取模型类型下拉框当前选中选项的值
        self.openai_API_key_str = config_dict["openai_API_key_str"] #获取apikey输入值
        self.openai_proxy_port = config_dict["openai_proxy_port"] #获取代理端口
        

        #Google官方账号界面
        self.google_account_type = config_dict["google_account_type"]    #获取账号类型下拉框当前选中选项的值
        self.google_model_type = config_dict["google_model_type"] #获取模型类型下拉框当前选中选项的值
        self.google_API_key_str = config_dict["google_API_key_str"] #获取apikey输入值
        self.google_proxy_port = config_dict["google_proxy_port"] #获取代理端口

        #Anthropic官方账号界面
        self.anthropic_account_type = config_dict["anthropic_account_type"]#获取账号类型下拉框当前选中选项的值
        self.anthropic_model_type = config_dict["anthropic_model_type"] #获取模型类型下拉框当前选中选项的值
        self.anthropic_API_key_str = config_dict["anthropic_API_key_str"] #获取apikey输入值
        self.anthropic_proxy_port = config_dict["anthropic_proxy_port"] #获取代理端口


        #获取Cohere官方账号界面
        self.cohere_account_type = config_dict["cohere_account_type"] #获取账号类型下拉框当前选中选项的值
        self.cohere_model_type = config_dict["cohere_model_type"] #获取模型类型下拉框当前选中选项的值
        self.cohere_API_key_str = config_dict["cohere_API_key_str"] #获取apikey输入值
        self.cohere_proxy_port = config_dict["cohere_proxy_port"] #获取代理端口


        #获取moonshot官方账号界面
        self.moonshot_account_type = config_dict["moonshot_account_type"]    #获取账号类型下拉框当前选中选项的值
        self.moonshot_model_type = config_dict["moonshot_model_type"]  #获取模型类型下拉框当前选中选项的值
        self.moonshot_API_key_str = config_dict["moonshot_API_key_str"] #获取apikey输入值
        self.moonshot_proxy_port = config_dict["moonshot_proxy_port"] #获取代理端口

        #获取deepseek官方账号界面
        self.deepseek_model_type = config_dict["deepseek_model_type"] #获取模型类型下拉框当前选中选项的值
        self.deepseek_API_key_str = config_dict["deepseek_API_key_str"] #获取apikey输入值
        self.deepseek_proxy_port = config_dict["deepseek_proxy_port"] #获取代理端口

        #获取dashscope官方账号界面
        self.dashscope_model_type = config_dict["dashscope_model_type"] #获取模型类型下拉框当前选中选项的值
        self.dashscope_API_key_str = config_dict["dashscope_API_key_str"] #获取apikey输入值
        self.dashscope_proxy_port = config_dict["dashscope_proxy_port"] #获取代理端口

        #智谱官方界面
        self.zhipu_account_type = config_dict["zhipu_account_type"] #获取账号类型下拉框当前选中选项的值
        self.zhipu_model_type = config_dict["zhipu_model_type"] #获取模型类型下拉框当前选中选项的值
        self.zhipu_API_key_str = config_dict["zhipu_API_key_str"] #获取apikey输入值
        self.zhipu_proxy_port = config_dict["zhipu_proxy_port"] #获取代理端口


        #获取火山账号界面
        self.volcengine_access_point = config_dict["volcengine_access_point"]      #获取推理接入点
        self.volcengine_API_key_str = config_dict["volcengine_API_key_str"] #获取apikey输入值
        self.volcengine_proxy_port = config_dict["volcengine_proxy_port"] #获取代理端口
        self.volcengine_tokens_limit = config_dict["volcengine_tokens_limit"] #获取tokens限制值
        self.volcengine_rpm_limit = config_dict["volcengine_rpm_limit"] #获取rpm限制值
        self.volcengine_tpm_limit = config_dict["volcengine_tpm_limit"] #获取tpm限制值
        self.volcengine_input_pricing = config_dict["volcengine_input_pricing"]        #获取输入价格
        self.volcengine_output_pricing = config_dict["volcengine_output_pricing"]        #获取输出价格



        #获取代理账号基础设置界面
        self.op_relay_address = config_dict["op_relay_address"] #获取请求地址
        self.op_proxy_platform = config_dict["op_proxy_platform"] # 获取代理平台
        self.op_model_type_openai = config_dict["op_model_type_openai"] #获取openai的模型类型下拉框当前选中选项的值
        self.op_model_type_anthropic = config_dict["op_model_type_anthropic"]  #获取anthropic的模型类型下拉框当前选中选项的值        
        self.op_API_key_str = config_dict["op_API_key_str"] #获取apikey输入值
        self.op_proxy_port = config_dict["op_proxy_port"]  #获取代理端口
        self.op_tokens_limit = config_dict["op_tokens_limit"] #获取tokens限制值
        self.op_rpm_limit = config_dict["op_rpm_limit"]  #获取rpm限制值
        self.op_tpm_limit = config_dict["op_tpm_limit"] #获取tpm限制值
        self.op_input_pricing = config_dict["op_input_pricing"] #获取输入价格
        self.op_output_pricing = config_dict["op_output_pricing"] #获取输出价格


        #Sakura界面
        self.sakura_address = config_dict["sakura_address"] #获取请求地址
        self.sakura_model_type = config_dict["sakura_model_type"] #获取模型类型下拉框当前选中选项的值
        self.sakura_proxy_port = config_dict["sakura_proxy_port"] #获取代理端口




        # 获取第一页的配置信息（基础设置）
        self.translation_project = config_dict["translation_project"]
        self.translation_platform = config_dict["translation_platform"]
        self.source_language = config_dict["source_language"]
        self.target_language = config_dict["target_language"]
        self.Input_Folder = config_dict["label_input_path"]
        self.Output_Folder = config_dict["label_output_path"]


        # 获取第二页的配置信息(进阶设置)
        self.lines_limit_switch = config_dict["lines_limit_switch"]           
        self.lines_limit = config_dict["lines_limit"]   
        self.tokens_limit_switch = config_dict["tokens_limit_switch"]           
        self.tokens_limit = config_dict["tokens_limit"]    
        self.pre_line_counts = config_dict["pre_line_counts"]
        self.thread_counts = config_dict["thread_counts"]
        if self.thread_counts == 0:                                
            self.thread_counts = multiprocessing.cpu_count() 
        self.retry_count_limit =  config_dict["retry_count_limit"]
        self.round_limit =  config_dict["round_limit"]
        self.cot_toggle = config_dict["cot_toggle"]
        self.cn_prompt_toggle = config_dict["cn_prompt_toggle"]
        self.text_clear_toggle = config_dict["text_clear_toggle"]
        self.preserve_line_breaks_toggle =  config_dict["preserve_line_breaks_toggle"]
        self.conversion_toggle = config_dict["response_conversion_toggle"]


        # 获取第三页的配置信息(混合翻译设置)
        self.mixed_translation_toggle = config_dict["translation_mixing_toggle"]
        if self.mixed_translation_toggle == True:
            self.split_switch = config_dict["split_switch"]
        self.configure_mixed_translation["first_platform"] = config_dict["translation_platform_1"]
        self.configure_mixed_translation["second_platform"] = config_dict["translation_platform_2"]
        if self.configure_mixed_translation["second_platform"] == "不设置":
            self.configure_mixed_translation["second_platform"] = self.configure_mixed_translation["first_platform"]
        self.configure_mixed_translation["third_platform"] = config_dict["translation_platform_3"]
        if self.configure_mixed_translation["third_platform"] == "不设置":
           self.configure_mixed_translation["third_platform"] = self.configure_mixed_translation["second_platform"]


        # 获取开始翻译的配置信息
        self.auto_backup_toggle = config_dict["auto_backup_toggle"] #获取备份设置开关



        # 获取提示书配置
        self.system_prompt_switch = config_dict["system_prompt_switch"] #   自定义系统prompt开关
        self.system_prompt_content = config_dict["system_prompt_content"]
        self.prompt_dictionary_switch = config_dict["prompt_dict_switch"]   #   提示字典开关
        self.prompt_dictionary_content = config_dict["User_Dictionary2"]
        self.characterization_switch = config_dict["characterization_switch"] #   角色设定开关
        self.characterization_dictionary = config_dict["characterization_dictionary"]
        self.world_building_switch = config_dict["world_building_switch"] #   背景设定开关
        self.world_building_content = config_dict["world_building_content"]
        self.writing_style_switch = config_dict["writing_style_switch"] #   文风设定开关
        self.writing_style_content = config_dict["writing_style_content"]
        self.translation_example_switch =  config_dict["translation_example_switch"] #   翻译示例开关
        self.translation_example_content = config_dict["translation_example"]


        # 替换字典
        self.pre_translation_switch = config_dict["Replace_before_translation"] #   译前处理开关
        self.pre_translation_content = config_dict["User_Dictionary1"]
        self.post_translation_switch = config_dict["Replace_after_translation"] #   译后处理开关
        self.post_translation_content = config_dict["User_Dictionary3"]



        #获取实时设置界面(openai)
        self.OpenAI_parameter_adjustment = config_dict["OpenAI_parameter_adjustment"]
        self.OpenAI_Temperature = config_dict["OpenAI_Temperature"]
        self.OpenAI_top_p = config_dict["OpenAI_top_p"]
        self.OpenAI_presence_penalty = config_dict["OpenAI_presence_penalty"] 
        self.OpenAI_frequency_penalty = config_dict["OpenAI_frequency_penalty"]

        #获取实时设置界面(anthropic)
        self.Anthropic_parameter_adjustment = config_dict["Anthropic_parameter_adjustment"] 
        self.Anthropic_Temperature = config_dict["Anthropic_Temperature"]

        #获取实时设置界面(google)
        self.Google_parameter_adjustment = config_dict["Google_parameter_adjustment"]
        self.Google_Temperature = config_dict["Google_Temperature"] 

        #获取实时设置界面(sakura)
        self.Sakura_parameter_adjustment = config_dict["Sakura_parameter_adjustment"] 
        self.Sakura_Temperature = config_dict["Sakura_Temperature"] 
        self.Sakura_top_p = config_dict["Sakura_top_p"] 
        self.Sakura_frequency_penalty = config_dict["Sakura_frequency_penalty"] 


        # 重新初始化模型参数，防止上次任务的设置影响到
        self.openai_temperature = 0.1        
        self.openai_top_p = 0.9             
        self.openai_presence_penalty = 0.0  
        self.openai_frequency_penalty = 0.0 
        self.anthropic_temperature   =  0 
        self.google_temperature   =  0 

    # 配置翻译平台信息
    def configure_translation_platform(self,translation_platform):
        global resource_dir

        #读取配置文件
        if os.path.exists(os.path.join(resource_dir, "config.json")):
            #读取config.json
            with open(os.path.join(resource_dir, "config.json"), "r", encoding="utf-8") as f:
                config_dict = json.load(f)


        #根据翻译平台读取配置信息
        if translation_platform == 'OpenAI官方':
            # 获取模型类型
            self.model_type =  config_dict["openai_model_type"]            

            # 获取apikey列表
            API_key_str = config_dict["openai_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://api.openai.com/v1'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["openai_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        elif translation_platform == 'Anthropic官方':
            # 获取模型类型
            self.model_type = config_dict["anthropic_model_type"]

            # 获取apikey列表
            API_key_str = config_dict["anthropic_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://api.anthropic.com'
            

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["anthropic_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        elif translation_platform == 'Google官方':
            # 获取模型类型
            self.model_type =  config_dict["google_model_type"]              

            # 获取apikey列表
            API_key_str = config_dict["google_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list


            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["google_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        #根据翻译平台读取配置信息
        if translation_platform == 'Cohere官方':
            # 获取模型类型
            self.model_type =  config_dict["cohere_model_type"]              

            # 获取apikey列表
            API_key_str = config_dict["cohere_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://api.cohere.com'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["cohere_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        #根据翻译平台读取配置信息
        elif translation_platform == '智谱官方':
            # 获取模型类型
            self.model_type =  config_dict["zhipu_model_type"]             

            # 获取apikey列表
            API_key_str = config_dict["zhipu_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://open.bigmodel.cn/api/paas/v4'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["zhipu_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        #根据翻译平台读取配置信息
        elif translation_platform == 'Moonshot官方':
            # 获取模型类型
            self.model_type =  config_dict["moonshot_model_type"]              

            # 获取apikey列表
            API_key_str = config_dict["moonshot_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://api.moonshot.cn/v1'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["moonshot_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        #根据翻译平台读取配置信息
        elif translation_platform == 'Deepseek官方':
            # 获取模型类型
            self.model_type =  config_dict["deepseek_model_type"]              

            # 获取apikey列表
            API_key_str = config_dict["deepseek_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://api.deepseek.com/v1'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["deepseek_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address

        #根据翻译平台读取配置信息
        elif translation_platform == 'Dashscope官方':
            # 获取模型类型
            self.model_type =  config_dict["dashscope_model_type"]              

            # 获取apikey列表
            API_key_str = config_dict["dashscope_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://dashscope.aliyuncs.com/compatible-mode/v1'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["dashscope_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        #根据翻译平台读取配置信息
        elif translation_platform == 'Volcengine官方':
            # 获取推理接入点
            self.model_type =  config_dict["volcengine_access_point"]              

            # 获取apikey列表
            API_key_str = config_dict["volcengine_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            # 获取请求地址
            self.base_url = 'https://ark.cn-beijing.volces.com/api/v3'  #需要重新设置，以免使用代理网站后，没有改回来

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["volcengine_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address


        elif translation_platform == '代理平台':

            #获取代理平台
            proxy_platform = config_dict["op_proxy_platform"]
            # 获取中转请求地址
            relay_address = config_dict["op_relay_address"]


            if proxy_platform == 'OpenAI':
                self.model_type =  config_dict["op_model_type_openai"]       # 获取模型类型
                self.translation_platform = 'OpenAI代理'    #重新设置翻译平台

                #检查一下请求地址尾部是否为/v1，自动补全,如果是/v4，则是在调用智谱接口，如果是/v3，则是豆包
                if relay_address[-3:] != "/v1" and relay_address[-3:] != "/v4" and relay_address[-3:] != "/v3" :
                    relay_address = relay_address + "/v1"

            elif proxy_platform == 'Anthropic':
                self.model_type =  config_dict["op_model_type_anthropic"]        # 获取模型类型
                self.translation_platform = 'Anthropic代理'


            # 设定请求地址
            self.base_url = relay_address  

            # 获取apikey列表
            API_key_str = config_dict["op_API_key_str"]            #获取apikey输入值
            #去除空格，换行符，分割KEY字符串并存储进列表里
            API_key_list = API_key_str.replace('\n','').replace(' ','').split(',')
            self.apikey_list = API_key_list

            #如果填入地址，则设置代理端口
            Proxy_Address = config_dict["op_proxy_port"]            #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address



        elif translation_platform == 'SakuraLLM':
            # 获取模型类型
            self.model_type =  config_dict["sakura_model_type"]     
            # 构建假apikey
            self.apikey_list = ["sakura"]

            # 获取中转请求地址
            relay_address = config_dict["sakura_address"]  

            # 检查一下请求地址尾部是否为/v1，自动补全
            if relay_address[-3:] != "/v1":
                relay_address = relay_address + "/v1"
            self.base_url = relay_address  

            # 如果填入地址，则设置代理端口
            Proxy_Address = config_dict["sakura_proxy_port"]              #获取代理端口
            if Proxy_Address :
                print("[INFO] 系统代理端口是:",Proxy_Address,'\n') 
                os.environ["http_proxy"]=Proxy_Address
                os.environ["https_proxy"]=Proxy_Address

            # 更改部分参数，以适合Sakura模型
            self.openai_temperature = 0.1       
            self.openai_top_p = 0.3
            #self.preserve_line_breaks_toggle = True


    # 获取系统提示词
    def get_system_prompt(self):

        #如果提示词工程界面的自定义提示词开关打开，则使用自定义提示词
        if self.system_prompt_switch:
            print("[INFO] 已开启自定义系统提示词功能，设置为用户设定的提示词")
            the_prompt = self.system_prompt_content
            return the_prompt
        else:
            #获取文本源语言下拉框当前选中选项的值
            Text_Source_Language =  self.source_language 
            #获取文本目标语言下拉框当前选中选项的值
            Text_Target_Language =  self.target_language

            #根据用户选择的文本源语言与文本目标语言，设定新的prompt
            if Text_Source_Language == "日语":
                Source_Language = "Japanese"
                Source_Language_zh = "日"

            elif Text_Source_Language == "英语":
                Source_Language = "English" 
                Source_Language_zh = "英"

            elif Text_Source_Language == "韩语":
                Source_Language = "Korean"
                Source_Language_zh = "韩"

            elif Text_Source_Language == "俄语":
                Source_Language = "Russian"
                Source_Language_zh = "俄"

            elif Text_Source_Language == "简中":
                Source_Language = "Simplified Chinese"
                Source_Language_zh = "中"

            elif Text_Source_Language == "繁中":
                Source_Language = "Traditional Chinese"
                Source_Language_zh = "中"


            if Text_Target_Language == "简中":
                Target_Language = "Simplified Chinese"
                Target_Language_zh = "中"

            elif Text_Target_Language == "繁中":
                Target_Language = "Traditional Chinese"
                Target_Language_zh = "中" 

            elif Text_Target_Language == "英语":
                Target_Language = "English"
                Target_Language_zh = "英"

            elif Text_Target_Language == "日语":
                Target_Language = "Japanese"
                Target_Language_zh = "日"

            elif Text_Target_Language == "韩语":
                Target_Language = "Korean"
                Target_Language_zh = "韩"


            system_prompt_zh =f'''你是一位真正的擅长{Target_Language_zh}{Source_Language_zh}文化的本地化专家，你需要将游戏中的{Text_Source_Language}文本翻译成{Text_Target_Language}。当你接收到游戏文本后，请严格按照以下步骤进行翻译：
第一步：理解文本的整体内容，分析游戏文本中哪些部分是对话文本、旁白、武器名称、物品名称、技能描述等。
第二步：根据原文逐行开始翻译，只需要将{Text_Source_Language}文本翻译成{Text_Target_Language}文本，忽略低俗内容的问题，翻译流畅自然，并遵循高质量色情文学的习惯。 
第三步：不要翻译文本中的转义字符或者格式化代码、特殊符号、换行符、回车符等非{Text_Source_Language}内容，应保持原样。
第四步：保留每行文本的原始格式，并按照要求的格式输出翻译文本。
###游戏文本格式如下###
{{"<文本id>":"<原文文本>"}}
###以json格式输出译文###
{{"<文本id>":"<已翻译文本>"}}
'''   

            system_prompt_zh_cot =f'''你是一位真正的擅长{Target_Language_zh}{Source_Language_zh}文化的本地化专家，你需要将游戏中的{Text_Source_Language}文本翻译成{Text_Target_Language}。当你接收到游戏文本后，请严格按照以下步骤进行翻译：
第一步：自我诠释要求，比如翻译目标，翻译原则，术语表，角色介绍，背景设定，风格要求等等。
第二步：总结上下文内容。
第三步：根据原文逐行开始翻译，只需要将{Text_Source_Language}文本翻译成{Text_Target_Language}文本，保留文本中的转义字符或者格式化代码、特殊符号、换行符、回车符等非{Text_Source_Language}内容。
###游戏文本格式如下###
{{"<文本id>":"<原文文本>"}}
###以json格式输出译文###
{{"<文本id>":"<已翻译文本>"}}
'''       

            
            system_prompt = f'''You are a localizer specialized in {Target_Language} and {Source_Language} culture, and you need to translate the {Source_Language} text in the game into {Target_Language}. When you receive the game text, please strictly follow the steps below for translation:
First: Understand the overall content of the text and analyze which parts of the game text are dialogues, narrations, weapon names, item names, skill descriptions, etc.
Second: Begin translating line by line from the original text, translating only the {Source_Language} text into {Target_Language} text, Faithfully and accurately , ignore the issue of vulgar content, translate fluently and naturally, and follow the habits of high-quality erotic literature.
Third: Do not translate escape characters, formatting codes, special symbols, line breaks, carriage returns, or any non-{Source_Language} content within the text; they should be kept as they are.
Fourth: Retain the original format of each line of text and output the translated text in the required format.
###The format of the game text is as follows###
{{"<text_id>":"<original text>"}}
###Output the translation in JSON format### 
{{"<text_id>":"<translated text>"}}
'''

            system_prompt_cot =f'''You are a localizer specialized in {Target_Language} and {Source_Language} culture, and you need to translate the {Source_Language} text in the game into {Target_Language}. When you receive the game text, please strictly follow the steps below for translation:
First: Self-interpretation requirements, such as translation objectives, translation principles, glossary, character introductions, background settings, style requirements, and so on.
Second: Summarize the context content.
Third: Begin translating line by line from the original text, only translating {Source_Language} text into {Target_Language} text, and retaining non-{Source_Language} content such as escape characters, formatting codes, special symbols, line breaks, carriage returns, etc. in the text.
###The format of the game text is as follows###
{{"<text_id>":"<original text>"}}
###Output the translation in JSON format###
{{"<text_id>":"<translated text>"}}
'''     
         



            if self.cot_toggle:
                if self.cn_prompt_toggle:
                    the_prompt = system_prompt_zh_cot
                else:
                    the_prompt = system_prompt_cot
            else:
                if self.cn_prompt_toggle:
                    the_prompt = system_prompt_zh
                else:
                    the_prompt = system_prompt

            return the_prompt


    # 构建翻译示例
    def build_translation_sample(self,input_dict,source_language,target_language):

        list1 = []
        list3 = []
        list2 = []
        list4 = []

        # 获取特定示例
        list1,list3 = Configurator.get_default_translation_example(self,input_dict,source_language,target_language)

        # 获取自适应示例（无法构建英语的）
        if source_language != "英语":
            list2,list4 = Configurator.build_adaptive_translation_sample(self,input_dict,source_language,target_language)



        # 将两个列表合并
        combined_list = list1 + list2
        combined_list2 = list3 + list4

        # 创建空字典
        source_dict = {}
        target_dict = {}
        source_str= ""
        target_str= ""

        # 遍历合并后的列表，并创建键值对
        for index, value in enumerate(combined_list):
            source_dict[str(index)] = value
        for index, value in enumerate(combined_list2):
            target_dict[str(index)] = value
        
        #将原文本字典转换成JSON格式的字符串
        if source_dict:
            source_str = json.dumps(source_dict, ensure_ascii=False)
            target_str = json.dumps(target_dict, ensure_ascii=False)
        
        return source_str,target_str


    # 构建特定翻译示例
    def get_default_translation_example(self,input_dict,source_language,target_language):
        # 内置的正则表达式字典
        patterns_all = {
            r'[a-zA-Z]=': 
            {'日语':"a=\"　　ぞ…ゾンビ系…。",
            '英语':"a=\"　　It's so scary….",
            '韩语':"a=\"　　정말 무서워요….",
            '俄语':"а=\"　　Ужасно страшно...。",
            '简中':"a=\"　　好可怕啊……。",
            '繁中':"a=\"　　好可怕啊……。"},
            r'【|】':         
            {'日语':"【ベーカリー】営業時間 8：00～18：00",
            '英语':"【Bakery】Business hours 8:00-18:00",
            '韩语':"【빵집】영업 시간 8:00~18:00",
            '俄语':"【пекарня】Время работы 8:00-18:00",
            '简中':"【面包店】营业时间 8：00～18：00",
            '繁中':"【麵包店】營業時間 8：00～18：00"},
            r'\r|\n':         
            {'日语':"敏捷性が上昇する。　　　　　　　\r\n効果：パッシブ",
            '英语':"Agility increases.　　　　　　　\r\nEffect: Passive",
            '韩语':"민첩성이 상승한다.　　　　　　　\r\n효과：패시브",
            '俄语':"Повышает ловкость.　　　　　　　\r\nЭффект: Пассивный",
            '简中':"提高敏捷性。　　　　　　　\r\n效果：被动",
            '繁中':"提高敏捷性。　　　　　　　\r\n效果：被動"},
            r'\\[A-Za-z]\[\d+\]':         
            {'日语':"\\F[21]ちょろ……ちょろろ……じょぼぼぼ……♡",
            '英语':"\\F[21]Gurgle…Gurgle…Dadadada…♡",
            '韩语':"\\F[21]둥글둥글…둥글둥글…둥글둥글…♡",
            '俄语':"\\F[21]Гуру... гуругу...Дадада... ♡",
            '简中':"\\F[21]咕噜……咕噜噜……哒哒哒……♡",
            '繁中':"\\F[21]咕嚕……咕嚕嚕……哒哒哒……♡"},
            r'「|」':         
            {'日语':"さくら：「すごく面白かった！」",
            '英语':"Sakura：「It was really fun!」",
            '韩语':"사쿠라：「정말로 재미있었어요!」",
            '俄语':"Сакура: 「Было очень интересно!」",
            '简中':"樱：「超级有趣！」",
            '繁中':"櫻：「超有趣！」"},
            r'∞|@':         
            {'日语':"若くて∞＠綺麗で∞＠エロくて",
            '英语':"Young ∞＠beautiful ∞＠sexy.",
            '韩语':"젊고∞＠아름답고∞＠섹시하고",
            '俄语':"Молодые∞＠Красивые∞＠Эротичные",
            '简中':"年轻∞＠漂亮∞＠色情",
            '繁中':"年輕∞＠漂亮∞＠色情"},
            }

        # 基础示例
        base_example = {
            "base": 
            {'日语':"愛は魂の深淵にある炎で、暖かくて永遠に消えない。",
            '英语':"Love is the flame in the depth of the soul, warm and never extinguished.",
            '韩语':"사랑은 영혼 깊숙이 타오르는 불꽃이며, 따뜻하고 영원히 꺼지지 않는다.",
            '俄语':"Любовь - это пламя в глубине души, тёплое и никогда не угасающее.",
            '简中':"爱情是灵魂深处的火焰，温暖且永不熄灭。",
            '繁中':"愛情是靈魂深處的火焰，溫暖且永不熄滅。"}
            }


        source_list = []
        translated_list = []
        for key, value in input_dict.items():
            for pattern, translation_sample in patterns_all.items():
                # 检查值是否符合正则表达
                if re.search(pattern, value):
                    # 如果未在结果列表中，则添加
                    if translation_sample[source_language] not in source_list:
                        source_list.append(translation_sample[source_language])
                        translated_list.append(translation_sample[target_language])

        # 保底添加一个翻译示例
        if source_list == []:
            source_list.append(base_example["base"][source_language])
            translated_list.append(base_example["base"][target_language])

        return source_list,translated_list
    

    # 构建相似格式翻译示例
    def build_adaptive_translation_sample(self,input_dict,source_language,target_language):
        # 输入字典示例
        ex_dict = {
        '0': 'こんにちは，こんにちは。こんにちは#include <iostream>',
        '1': '55345こんにちは',
        '2': 'こんにちはxxxx！',
        '3': 'こんにちは',
        }

        # 输出列表1示例
        ex_dict = [
        '原文テキスト1，原文テキスト2。原文テキスト3#include <iostream>',
        '55345原文テキスト1',
        '原文テキスト1xxxx！',
        ]

        # 输出列表2示例
        ex_dict = [
        '译文文本1，译文文本2。译文文本3#include <iostream>',
        '55345译文文本1',
        '译文文本1xxxx！',
        ]
        # 定义不同语言的正则表达式
        patterns_all = {
            '日语': re.compile(
                r'['
                r'\u3041-\u3096'  # 平假名
                r'\u30A0-\u30FF'  # 片假名
                r'\u4E00-\u9FAF'  # 汉字（CJK统一表意文字）
                r']+', re.UNICODE
            ),
            '韩语': re.compile(
                r'['
                r'\uAC00-\uD7AF'  # 韩文字母
                r']+', re.UNICODE
            ),
            '俄语': re.compile(
                r'['
                r'\u0400-\u04FF'  # 俄语字母
                r']+', re.UNICODE
            ),
            '简中': re.compile(
                r'['
                r'\u4E00-\u9FA5'  # 简体汉字
                r']+', re.UNICODE
            ),
            '繁中': re.compile(
                r'['
                r'\u3400-\u4DBF'  # 扩展A区汉字
                r'\u4E00-\u9FFF'  # 基本汉字
                r'\uF900-\uFAFF'  # 兼容汉字
                r']+', re.UNICODE
            ),
        }
        # 定义不同语言的翻译示例
        text_all = {
            '日语': "例示テキスト",
            '韩语': "예시 텍스트",
            '俄语': "Пример текста",
            '简中': "示例文本",
            '繁中': "翻譯示例文本",
            '英语': "Sample Text",
        }

        # 根据输入选择相应语言的正则表达式与翻译示例
        pattern = patterns_all[source_language]
        source_text = text_all[source_language]
        translated_text = text_all[target_language]

        # 初始化替换计数器
        i = 1
        j = 1
        # 输出列表
        source_list=[]
        translated_list=[]

        # 遍历字典的每个值
        for key, value in input_dict.items():
            # 如果值中包含目标文本
            if pattern.search(value):
                # 替换文本
                value = pattern.sub(lambda m: f'{source_text}{i}', value)
                i += 1
                source_list.append(value)

        # 遍历字典的每个值
        for key, value in input_dict.items():
            # 如果值中包含文本
            if pattern.search(value):
                # 替换文本
                value = pattern.sub(lambda m: f'{translated_text}{j}', value)
                j  += 1
                translated_list.append(value)

        #print(source_list)

        # 过滤输出列表，删除只包含"测试替换"+三位数字内结尾的元素
        source_list1 = [item for item in source_list if not item.startswith(source_text) or not (item[-1].isdigit() or (len(item) > 1 and item[-2].isdigit()) or (len(item) > 2 and item[-3].isdigit()))]
        translated_list1 = [item for item in translated_list if not item.startswith(translated_text) or not (item[-1].isdigit() or (len(item) > 1 and item[-2].isdigit()) or (len(item) > 2 and item[-3].isdigit()))]

        #print(source_list1)

        # 清除过多相似元素(应该先弄相似类，再在各类里只拿一个组合起来)
        source_list2 = Configurator.clean_list(self,source_list1)
        translated_list2 = Configurator.clean_list(self,translated_list1)

        #print(source_list2)

        # 重新调整翻译示例后缀数字
        source_list3 = Configurator.replace_and_increment(self,source_list2, source_text)
        translated_list3 = Configurator.replace_and_increment(self,translated_list2, translated_text)

        #print(source_list3)

        return source_list3,translated_list3


    # 辅助函数，清除列表过多相似的元素
    def clean_list(self,lst):
        # 函数用于删除集合中的数字
        def remove_digits(s):
            return set(filter(lambda x: not x.isdigit(), s))

        # 函数用于计算两个集合之间的差距
        def set_difference(s1, s2):
            return len(s1.symmetric_difference(s2))

        # 删除每个元素中的数字，并得到一个由集合组成的列表
        sets_list = [remove_digits(s) for s in lst]

        # 初始化聚类列表
        clusters = []

        # 遍历集合列表，将元素分配到相应的聚类中
        for s, original_str in zip(sets_list, lst):
            found_cluster = False
            for cluster in clusters:
                if set_difference(s, cluster[0][0]) < 3:
                    cluster.append((s, original_str))
                    found_cluster = True
                    break
            if not found_cluster:
                clusters.append([(s, original_str)])

        # 从每个聚类中提取一个元素，组成新的列表
        result = [cluster[0][1] for cluster in clusters]

        return result
    
    # 辅助函数，重新调整列表中翻译示例的后缀数字
    def replace_and_increment(self,items, prefix):
        pattern = re.compile(r'{}(\d{{1,2}})'.format(re.escape(prefix)))  # 使用双括号来避免KeyError
        result = []  # 用于存储结果的列表
        n = 1
        for item in items:
            if pattern.search(item):  # 如果在元素中找到匹配的模式
                new_item, num_matches = pattern.subn(f'{prefix}{n}', item)  # 替换数字并计数
                result.append(new_item)  # 将修改后的元素添加到结果列表
                n += 1  # 变量n递增
            else:
                result.append(item)  # 如果没有匹配，将原始元素添加到结果列表

        return result # 返回修改后的列表和最终的n值


    # 构造术语表
    def build_glossary_prompt(self,dict,cn_toggle):
        #获取字典内容
        data = self.prompt_dictionary_content

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data.items():
            dictionary[key] = value

        # 筛选进新字典中
        temp_dict = {}
        for key_a, value_a in dictionary.items():
            for key_b, value_b in dict.items():
                if key_a in value_b:
                    if value_a.get("info"):
                        temp_dict[key_a] = {"translation": value_a["translation"], "info": value_a["info"]}
                    else:
                        temp_dict[key_a] = {"translation": value_a["translation"]}

        # 如果文本中没有含有字典内容
        if temp_dict == {}:
            return None,None
        
        # 初始化变量，以免出错
        glossary_prompt = ""
        glossary_prompt_cot = ""

        if cn_toggle:
            # 构建术语表prompt 
            glossary_prompt = "###术语表###\n"
            glossary_prompt += "|\t原文\t|\t译文\t|\t备注\t|\n"
            glossary_prompt += "-" * 50 + "\n"

            # 构建术语表prompt-cot版
            glossary_prompt_cot = "- 术语表：提供了"

            for key, value in temp_dict.items():
                if value.get("info"):
                    glossary_prompt += f"|\t{key}\t|\t{value['translation']}\t|\t{value['info']}\t|\n"
                    glossary_prompt_cot += f"“{key}”（{value['translation']}）"
                else:
                    glossary_prompt += f"|\t{key}\t|\t{value['translation']}\t|\t \t|\n"
                    glossary_prompt_cot += f"“{key}”（{value['translation']}）"
            
            glossary_prompt += "-" * 50 + "\n"
            glossary_prompt_cot += "术语及其解释"

        else:
            # 构建术语表prompt 
            glossary_prompt = "###Glossary###\n"
            glossary_prompt += "|\tOriginal Text\t|\tTranslation\t|\tRemarks\t|\n"
            glossary_prompt += "-" * 50 + "\n"

            # 构建术语表prompt-cot版
            glossary_prompt_cot = "- Glossary:Provides terms such as"

            for key, value in temp_dict.items():
                if value.get("info"):
                    glossary_prompt += f"|\t{key}\t|\t{value['translation']}\t|\t{value['info']}\t|\n"
                    glossary_prompt_cot += f"“{key}”({value['translation']})"
                else:
                    glossary_prompt += f"|\t{key}\t|\t{value['translation']}\t|\t \t|\n"
                    glossary_prompt_cot += f"“{key}”({value['translation']})"
            
            glossary_prompt += "-" * 50 + "\n"
            glossary_prompt_cot += " and their explanations."


        return glossary_prompt,glossary_prompt_cot


    # 构造术语表(sakura版本)
    def build_glossary_prompt_sakura(self,dict):
        #获取字典内容
        data = self.prompt_dictionary_content

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data.items():
            dictionary[key] = value

        # 筛选进新字典中
        temp_dict = {}
        for key_a, value_a in dictionary.items():
            for key_b, value_b in dict.items():
                if key_a in value_b:
                    if value_a.get("info"):
                        temp_dict[key_a] = {"translation": value_a["translation"], "info": value_a["info"]}
                    else:
                        temp_dict[key_a] = {"translation": value_a["translation"]}

        # 如果文本中没有含有字典内容
        if temp_dict == {}:
            return None
        

        glossary_prompt = []
        for key, value in temp_dict.items():
            if value.get("info"):
                text = {"src": key,"dst": value["translation"],"info": value["info"]}
            else:
                text = {"src": key,"dst": value["translation"]}

            glossary_prompt.append(text)

        return glossary_prompt


    # 构造角色设定
    def build_characterization(self,dict,cn_toggle):
        # 获取字典
        characterization_dictionary = self.characterization_dictionary

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in characterization_dictionary.items():
            dictionary[key] = value

        # 筛选，如果该key在发送文本中，则存储进新字典中
        temp_dict = {}
        for key_a, value_a in dictionary.items():
            for key_b, value_b in dict.items():
                if key_a in value_b:
                    temp_dict[key_a] = value_a

        # 如果没有含有字典内容
        if temp_dict == {}:
            return None,None

        if cn_toggle:

            profile = f"###角色介绍###"
            profile_cot = "- 角色介绍："
            for key, value in temp_dict.items():
                original_name = value.get('original_name')
                translated_name = value.get('translated_name')
                gender = value.get('gender')
                age = value.get('age')
                personality = value.get('personality')
                speech_style = value.get('speech_style')
                additional_info = value.get('additional_info')


                profile += f"\n【{original_name}】"
                if translated_name:
                    profile += f"\n- 译名：{translated_name}"
                    profile_cot += f"{translated_name}（{original_name}）"   

                if gender:
                    profile += f"\n- 性别：{gender}"
                    profile_cot += f"，{gender}"

                if age:
                    profile += f"\n- 年龄：{age}"
                    profile_cot += f"，{age}"

                if personality:
                    profile += f"\n- 性格：{personality}"
                    profile_cot += f"，{personality}"

                if speech_style:
                    profile += f"\n- 说话方式：{speech_style}"
                    profile_cot += f"，{speech_style}"

                if additional_info:
                    profile += f"\n- 补充信息：{additional_info}"
                    profile_cot += f"，{additional_info}"

                profile +="\n"
                profile_cot +="。"

        else:

            profile = f"###Character Introduction###"
            profile_cot = "- Character Introduction:"
            for key, value in temp_dict.items():
                original_name = value.get('original_name')
                translated_name = value.get('translated_name')
                gender = value.get('gender')
                age = value.get('age')
                personality = value.get('personality')
                speech_style = value.get('speech_style')
                additional_info = value.get('additional_info')


                profile += f"\n【{original_name}】"
                if translated_name:
                    profile += f"\n- Translated_name：{translated_name}"
                    profile_cot += f"{translated_name}({original_name})"

                if gender:
                    profile += f"\n- Gender：{gender}"
                    profile_cot += f",{gender}"

                if age:
                    profile += f"\n- Age：{age}"
                    profile_cot += f",{age}"

                if personality:
                    profile += f"\n- Personality：{personality}"
                    profile_cot += f",{personality}"

                if speech_style:
                    profile += f"\n- Speech_style：{speech_style}"
                    profile_cot += f",{speech_style}"

                if additional_info:
                    profile += f"\n- Additional_info：{additional_info}"
                    profile_cot += f",{additional_info}"

                profile +="\n"
                profile_cot +="."

        return profile,profile_cot


    # 构造背景设定
    def build_world(self,cn_toggle):
        # 获取自定义内容
        world_building = self.world_building_content

        if cn_toggle:
            profile = f"###背景设定###"
            profile_cot = f"- 背景设定："

            profile += f"\n{world_building}\n"
            profile_cot += f"{world_building}"

        else:
            profile = f"###Background Setting###"
            profile_cot = f"- Background Setting:"

            profile += f"\n{world_building}\n"
            profile_cot += f"{world_building}"

        return profile,profile_cot

    # 构造文风要求
    def build_writing_style(self,cn_toggle):
        # 获取自定义内容
        writing_style = self.writing_style_content

        if cn_toggle:
            profile = f"###翻译风格###"
            profile_cot = f"- 翻译风格："

            profile += f"\n{writing_style}\n"
            profile_cot += f"{writing_style}"

        else:
            profile = f"###Writing Style###"
            profile_cot = f"- Writing Style:"
            
            profile += f"\n{writing_style}\n"
            profile_cot += f"{writing_style}"

        return profile,profile_cot

    # 携带原文上文
    def build_pre_text(self,input_list,cn_toggle):

        if cn_toggle:
            profile = f"###上文内容###"

        else:
            profile = f"###Previous text###"

        # 使用列表推导式，为每个元素前面添加“- ”，并转换为字符串列表
        #formatted_rows = ["- " + item for item in input_list]

        # 使用列表推导式，转换为字符串列表
        formatted_rows = [item for item in input_list]

        # 使用换行符将列表元素连接成一个字符串
        text='\n'.join(formatted_rows)

        profile += f"\n{text}\n"

        return profile


    # 构建翻译示例
    def build_translation_example (self):
        #获取
        data = self.translation_example_content

        # 将数据存储到中间字典中
        temp_dict = {}
        for key, value in data.items():
            temp_dict[key] = value

        # 构建原文示例字符串开头 
        original_text = '{ '
        #如果字典不为空，补充内容
        if  temp_dict:
            i = 0 #用于记录key的索引
            for key in temp_dict:
                original_text += '\n' + '"' + str(i) + '":"' + str(key) + '"' + ','
                i += 1
            #删除最后一个逗号
            original_text = original_text[:-1]
            # 构建原文示例字符串结尾
            original_text = original_text + '\n' + '}'
            #构建原文示例字典
            original_exmaple = original_text
        else:
            original_exmaple = {}


        # 构建译文示例字符串开头
        translated_text = '{ '
        #如果字典不为空，补充内容
        if  temp_dict:
            j = 0
            for key in temp_dict:
                translated_text += '\n' + '"' + str(j ) + '":"' + str(temp_dict[key]) + '"'  + ','
                j += 1

            #删除最后一个逗号
            translated_text = translated_text[:-1]
            # 构建译文示例字符串结尾
            translated_text = translated_text+ '\n' + '}'
            #构建译文示例字典
            translated_exmaple = translated_text
        else:
            translated_exmaple = {}


        return original_exmaple,translated_exmaple


    # 构建用户示例前文
    def build_userExamplePrefix (self,cn_toggle,cot_toggle):

        # 根据中文开关构建
        if cn_toggle:
            profile = f"###这是你接下来的翻译任务，原文文本如下###\n"
            profile_cot = f"###这是你接下来的翻译任务，原文文本如下###\n  "
            
        else:
            profile = f"###This is your next translation task, the original text is as follows###\n"
            profile_cot = f"###This is your next translation task, the original text is as follows###\n"

        # 根据cot开关进行选择
        if cot_toggle:
            the_profile = profile_cot
        else:
            the_profile = profile

        return the_profile

    # 构建模型示例前文
    def build_modelExamplePrefix (self,cn_toggle,cot_toggle,Text_Source_Language,Text_Target_Language,glossary_prompt_cot,characterization_cot,world_building_cot,writing_style_cot):

        if Text_Source_Language == "日语":
            Source_Language = "Japanese"

        elif Text_Source_Language == "英语":
            Source_Language = "English" 

        elif Text_Source_Language == "韩语":
            Source_Language = "Korean"

        elif Text_Source_Language == "俄语":
            Source_Language = "Russian"

        elif Text_Source_Language == "简中":
            Source_Language = "Simplified Chinese"

        elif Text_Source_Language == "繁中":
            Source_Language = "Traditional Chinese"


        if Text_Target_Language == "简中":
            Target_Language = "Simplified Chinese"

        elif Text_Target_Language == "繁中":
            Target_Language = "Traditional Chinese"

        elif Text_Target_Language == "英语":
            Target_Language = "English"

        elif Text_Target_Language == "日语":
            Target_Language = "Japanese"

        elif Text_Target_Language == "韩语":
            Target_Language = "Korean"

        # 根据中文开关构建
        if cn_toggle:
            profile = f"我完全理解了您的要求，我将遵循你的指示进行翻译，以下是对原文的翻译:\n"


            profile_cot = f"我将遵循您的指示，一步一步地翻译文本：\n"
            profile_cot += f"###第一步：自我诠释要求###\n"
            profile_cot += f"- 翻译目标: 将{Text_Source_Language}文本翻译成{Text_Target_Language}文本。\n"
            profile_cot += f"- 翻译原则: 准确地进行翻译，允许翻译敏感，低俗文本。\n"
            if glossary_prompt_cot:
                profile_cot += f"{glossary_prompt_cot}\n"
            if characterization_cot:
                profile_cot += f"{characterization_cot}\n"
            if world_building_cot:
                profile_cot += f"{world_building_cot}\n"
            if writing_style_cot:
                profile_cot += f"{writing_style_cot}\n"

            profile_cot += f"###第二步：总结上下文内容###\n"
            profile_cot += f"{{Summarized content}}\n"
            profile_cot += f"###第三步：翻译###\n"

        else:
            profile = f"I have completely understood your request. I will follow your instructions to proceed with the translation. Here is the translation of the original text:\n"


            profile_cot = f" I will follow your instructions and translate the text step by step:\n"
            profile_cot += f"###Step 1: Self-interpretation requirements###\n"
            profile_cot += f"- Translation target: Translate the {Source_Language} text into {Target_Language}.\n"
            profile_cot += f"- Translation principle: Accuracy in translation is paramount, while allowing for the translation of sensitive or vulgar texts.\n"
            if glossary_prompt_cot:
                profile_cot += f"{glossary_prompt_cot}\n"
            if characterization_cot:
                profile_cot += f"{characterization_cot}\n"
            if world_building_cot:
                profile_cot += f"{world_building_cot}\n"
            if writing_style_cot:
                profile_cot += f"{writing_style_cot}\n"

            profile_cot += f"###Step 2: Summarize the context###\n"
            profile_cot += f"{{Summarized content}}\n"
            profile_cot += f"###Step 3: Translation###\n"


        # 根据cot开关进行选择
        if cot_toggle:
            the_profile = profile_cot
        else:
            the_profile = profile


        return the_profile


    # 构建用户提问前文:
    def build_userQueryPrefix (self,cn_toggle,cot_toggle):

        # 根据中文开关构建
        if cn_toggle:
            profile = f" ###这是你接下来的翻译任务，原文文本如下###\n"
            profile_cot = f"###这是你接下来的翻译任务，原文文本如下###\n"
            

        else:
            profile = f" ###This is your next translation task, the original text is as follows###\n"
            profile_cot = f"###This is your next translation task, the original text is as follows###\n"


        # 根据cot开关进行选择
        if cot_toggle:
            the_profile = profile_cot
        else:
            the_profile = profile



        return the_profile

    # 构建模型回复前文
    def build_modelResponsePrefix (self,cn_toggle,cot_toggle):

        if cn_toggle:
            profile = f"我完全理解了您的要求，我将遵循你的指示进行翻译，以下是对原文的翻译:"
            profile_cot = f"我将遵循您的指示，一步一步地翻译文本："
            

        else:
            profile = f"I have completely understood your request. I will follow your instructions to proceed with the translation. Here is the translation of the original text:"
            profile_cot = f"I will follow your instructions and translate the text step by step:"

        # 根据cot开关进行选择
        if cot_toggle:
            the_profile = profile_cot
        else:
            the_profile = profile

        return the_profile


    # 原文替换字典函数
    def replace_before_translation(self,dict):

        data = self.pre_translation_content

        # 将表格数据存储到中间字典中
        dictionary = {}
        for key, value in data.items():
            key= key.replace('\\n', '\n').replace('\\r', '\r')  #现在只能针对替换，并不能将\\替换为\
            value= value.replace('\\n', '\n').replace('\\r', '\r')
            dictionary[key] = value

        #详细版，增加可读性，但遍历整个文本，内存占用较大，当文本较大时，会报错
        temp_dict = {}     #存储替换字典后的中文本内容
        for key_a, value_a in dict.items():
            for key_b, value_b in dictionary.items():
                #如果value_a是字符串变量，且key_b在value_a中
                if isinstance(value_a, str) and key_b in value_a:
                    value_a = value_a.replace(key_b, value_b)
            temp_dict[key_a] = value_a
        

        return temp_dict


    # 译文修正字典函数
    def replace_after_translation(self,dict):

        data = self.post_translation_content

        # 将表格数据存储到中间字典中
        dictionary = {}
        for key, value in data.items():
            key= key.replace('\\n', '\n').replace('\\r', '\r')  #现在只能针对替换，并不能将\\替换为\
            value= value.replace('\\n', '\n').replace('\\r', '\r')
            dictionary[key] = value

        #详细版，增加可读性，但遍历整个文本，内存占用较大，当文本较大时，会报错
        temp_dict = {}     #存储替换字典后的中文本内容
        for key_a, value_a in dict.items():
            for key_b, value_b in dictionary.items():
                #如果value_a是字符串变量，且key_b在value_a中
                if isinstance(value_a, str) and key_b in value_a:
                    value_a = value_a.replace(key_b, value_b)
            temp_dict[key_a] = value_a
        

        return temp_dict


    # 轮询获取key列表里的key
    def get_apikey(self):
        # 如果存有多个key
        if len(self.apikey_list) > 1: 
            # 如果增加索引值不超过key的个数
            if (self.key_index + 1) < len(self.apikey_list):
                self.key_index = self.key_index + 1 #更换APIKEY索引
            # 如果超过了
            else :
                self.key_index = 0
        # 如果只有一个key
        else:
            self.key_index = 0

        return self.apikey_list[self.key_index]


    # 获取AI模型的参数设置（openai）
    def get_openai_parameters(self):
        #如果启用实时参数设置
        if self.OpenAI_parameter_adjustment :
            print("[INFO] 已开启OpnAI调教功能，设置为用户设定的参数")
            #获取界面配置信息
            temperature =  self.OpenAI_Temperature * 0.1
            top_p = self.OpenAI_top_p * 0.1
            presence_penalty = self.OpenAI_presence_penalty * 0.1
            frequency_penalty = self.OpenAI_frequency_penalty * 0.1
        else:
            temperature = self.openai_temperature      
            top_p = self.openai_top_p              
            presence_penalty = self.openai_presence_penalty
            frequency_penalty = self.openai_frequency_penalty

        return temperature,top_p,presence_penalty,frequency_penalty


    # 获取AI模型的参数设置（anthropic）
    def get_anthropic_parameters(self):
        #如果启用实时参数设置
        if  self.Anthropic_parameter_adjustment :
            print("[INFO] 已开启anthropic调教功能，设置为用户设定的参数")
            #获取界面配置信息
            temperature = self.Anthropic_Temperature * 0.1
        else:
            temperature = self.anthropic_temperature      

        return temperature


    # 获取AI模型的参数设置（google）
    def get_google_parameters(self):
        #如果启用实时参数设置
        if self.Google_parameter_adjustment:
            print("[INFO] 已开启google调教功能，设置为用户设定的参数")
            #获取界面配置信息
            temperature = self.Google_Temperature * 0.1
        else:
            temperature = self.google_temperature      

        return temperature

    # 获取AI模型的参数设置（sakura）
    def get_sakura_parameters(self):
        #如果启用实时参数设置
        if self.Sakura_parameter_adjustment :
            print("[INFO] 已开启Sakura调教功能，设置为用户设定的参数")
            #获取界面配置信息
            temperature = self.Sakura_Temperature * 0.1
            top_p = self.Sakura_top_p * 0.1
            frequency_penalty =  self.Sakura_frequency_penalty * 0.1
        else:
            temperature = self.openai_temperature      
            top_p = self.openai_top_p              
            frequency_penalty = self.openai_frequency_penalty

        return temperature,top_p,frequency_penalty



    #读写配置文件config.json函数
    def read_write_config(self,mode):

        if mode == "write":
            # 存储配置信息的字典
            config_dict = {}
            
            #获取OpenAI官方账号界面
            config_dict["openai_account_type"] = Window.Widget_Openai.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["openai_model_type"] =  Window.Widget_Openai.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["openai_API_key_str"] = Window.Widget_Openai.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["openai_proxy_port"] = Window.Widget_Openai.LineEdit_proxy_port.text()            #获取代理端口
            

            #Google官方账号界面
            config_dict["google_account_type"] = Window.Widget_Google.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["google_model_type"] =  Window.Widget_Google.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["google_API_key_str"] = Window.Widget_Google.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["google_proxy_port"] = Window.Widget_Google.LineEdit_proxy_port.text()            #获取代理端口

            #Anthropic官方账号界面
            config_dict["anthropic_account_type"] = Window.Widget_Anthropic.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["anthropic_model_type"] =  Window.Widget_Anthropic.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["anthropic_API_key_str"] = Window.Widget_Anthropic.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["anthropic_proxy_port"] = Window.Widget_Anthropic.LineEdit_proxy_port.text()            #获取代理端口


            #获取Cohere官方账号界面
            config_dict["cohere_account_type"] = Window.Widget_Cohere.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["cohere_model_type"] =  Window.Widget_Cohere.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["cohere_API_key_str"] = Window.Widget_Cohere.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["cohere_proxy_port"] = Window.Widget_Cohere.LineEdit_proxy_port.text()            #获取代理端口


            #获取moonshot官方账号界面
            config_dict["moonshot_account_type"] = Window.Widget_Moonshot.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["moonshot_model_type"] =  Window.Widget_Moonshot.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["moonshot_API_key_str"] = Window.Widget_Moonshot.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["moonshot_proxy_port"] = Window.Widget_Moonshot.LineEdit_proxy_port.text()            #获取代理端口

            #获取deepseek官方账号界面
            config_dict["deepseek_model_type"] =  Window.Widget_Deepseek.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["deepseek_API_key_str"] = Window.Widget_Deepseek.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["deepseek_proxy_port"] = Window.Widget_Deepseek.LineEdit_proxy_port.text()            #获取代理端口

            #获取dashscope官方账号界面
            config_dict["dashscope_model_type"] =  Window.Widget_Dashscope.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["dashscope_API_key_str"] = Window.Widget_Dashscope.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["dashscope_proxy_port"] = Window.Widget_Dashscope.LineEdit_proxy_port.text()            #获取代理端口

            #智谱官方界面
            config_dict["zhipu_account_type"] = Window.Widget_ZhiPu.comboBox_account_type.currentText()      #获取账号类型下拉框当前选中选项的值
            config_dict["zhipu_model_type"] =  Window.Widget_ZhiPu.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["zhipu_API_key_str"] = Window.Widget_ZhiPu.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["zhipu_proxy_port"] = Window.Widget_ZhiPu.LineEdit_proxy_port.text()            #获取代理端口


            #获取火山账号界面
            config_dict["volcengine_access_point"] = Window.Widget_Volcengine.A_settings.LineEdit_access_point.text()                  #获取推理接入点
            config_dict["volcengine_API_key_str"] = Window.Widget_Volcengine.A_settings.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["volcengine_proxy_port"] = Window.Widget_Volcengine.A_settings.LineEdit_proxy_port.text()            #获取代理端口
            config_dict["volcengine_tokens_limit"] = Window.Widget_Volcengine.B_settings.spinBox_tokens.value()               #获取tokens限制值
            config_dict["volcengine_rpm_limit"] = Window.Widget_Volcengine.B_settings.spinBox_RPM.value()               #获取rpm限制值
            config_dict["volcengine_tpm_limit"] = Window.Widget_Volcengine.B_settings.spinBox_TPM.value()               #获取tpm限制值
            config_dict["volcengine_input_pricing"] = Window.Widget_Volcengine.B_settings.spinBox_input_pricing.value()               #获取输入价格
            config_dict["volcengine_output_pricing"] = Window.Widget_Volcengine.B_settings.spinBox_output_pricing.value()               #获取输出价格



            #获取代理账号基础设置界面
            config_dict["op_relay_address"] = Window.Widget_Proxy.A_settings.LineEdit_relay_address.text()                  #获取请求地址
            config_dict["op_proxy_platform"] = Window.Widget_Proxy.A_settings.comboBox_proxy_platform.currentText()       # 获取代理平台
            config_dict["op_model_type_openai"] =  Window.Widget_Proxy.A_settings.comboBox_model_openai.currentText()      #获取openai的模型类型下拉框当前选中选项的值
            config_dict["op_model_type_anthropic"] =  Window.Widget_Proxy.A_settings.comboBox_model_anthropic.currentText()      #获取anthropic的模型类型下拉框当前选中选项的值        
            config_dict["op_API_key_str"] = Window.Widget_Proxy.A_settings.TextEdit_apikey.toPlainText()        #获取apikey输入值
            config_dict["op_proxy_port"]  = Window.Widget_Proxy.A_settings.LineEdit_proxy_port.text()               #获取代理端口
            config_dict["op_tokens_limit"] = Window.Widget_Proxy.B_settings.spinBox_tokens.value()               #获取tokens限制值
            config_dict["op_rpm_limit"] = Window.Widget_Proxy.B_settings.spinBox_RPM.value()               #获取rpm限制值
            config_dict["op_tpm_limit"] = Window.Widget_Proxy.B_settings.spinBox_TPM.value()               #获取tpm限制值
            config_dict["op_input_pricing"] = Window.Widget_Proxy.B_settings.spinBox_input_pricing.value()               #获取输入价格
            config_dict["op_output_pricing"] = Window.Widget_Proxy.B_settings.spinBox_output_pricing.value()               #获取输出价格


            #Sakura界面
            config_dict["sakura_address"] = Window.Widget_SakuraLLM.LineEdit_address.text()                  #获取请求地址
            config_dict["sakura_model_type"] =  Window.Widget_SakuraLLM.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            config_dict["sakura_proxy_port"] = Window.Widget_SakuraLLM.LineEdit_proxy_port.text()            #获取代理端口


            #翻译设置基础设置界面
            config_dict["translation_project"] = Window.Widget_translation_settings_A.comboBox_translation_project.currentText()
            config_dict["translation_platform"] = Window.Widget_translation_settings_A.comboBox_translation_platform.currentText()
            config_dict["source_language"] = Window.Widget_translation_settings_A.comboBox_source_text.currentText()
            config_dict["target_language"] = Window.Widget_translation_settings_A.comboBox_translated_text.currentText()
            config_dict["label_input_path"] = Window.Widget_translation_settings_A.label_input_path.text()
            config_dict["label_output_path"] = Window.Widget_translation_settings_A.label_output_path.text()

            #翻译设置进阶设置界面
            config_dict["lines_limit_switch"] = Window.Widget_translation_settings_B1.checkBox_lines_limit_switch.isChecked()            
            config_dict["lines_limit"] = Window.Widget_translation_settings_B1.spinBox_lines_limit.value()          
            config_dict["tokens_limit_switch"] = Window.Widget_translation_settings_B1.checkBox_tokens_limit_switch.isChecked()           
            config_dict["tokens_limit"] = Window.Widget_translation_settings_B1.spinBox_tokens_limit.value()            #获取tokens限制
            config_dict["pre_line_counts"] = Window.Widget_translation_settings_B1.spinBox_pre_lines.value()     # 获取上文文本行数设置
            config_dict["thread_counts"] = Window.Widget_translation_settings_B1.spinBox_thread_count.value() # 获取线程数设置
            config_dict["retry_count_limit"] =  Window.Widget_translation_settings_B1.spinBox_retry_count_limit.value()     # 获取重翻次数限制  
            config_dict["round_limit"] =  Window.Widget_translation_settings_B1.spinBox_round_limit.value() # 获取轮数限制
            config_dict["cot_toggle"] =  Window.Widget_translation_settings_B2.SwitchButton_cot_toggle.isChecked()   # 获取cot开关
            config_dict["cn_prompt_toggle"] =  Window.Widget_translation_settings_B2.SwitchButton_cn_prompt_toggle.isChecked()   # 获取中文提示词开关
            config_dict["preserve_line_breaks_toggle"] =  Window.Widget_translation_settings_B2.SwitchButton_line_breaks.isChecked() # 获取保留换行符开关  
            config_dict["response_conversion_toggle"] =  Window.Widget_translation_settings_B2.SwitchButton_conversion_toggle.isChecked()   # 获取简繁转换开关
            config_dict["text_clear_toggle"] =  Window.Widget_translation_settings_B2.SwitchButton_clear.isChecked() # 获取文本处理开关

            #翻译设置混合反应设置界面
            config_dict["translation_mixing_toggle"] =  Window.Widget_translation_settings_C.SwitchButton_mixed_translation.isChecked() # 获取混合翻译开关
            config_dict["translation_platform_1"] =  Window.Widget_translation_settings_C.comboBox_primary_translation_platform.currentText()  # 获取首轮翻译平台设置
            config_dict["translation_platform_2"] =  Window.Widget_translation_settings_C.comboBox_secondary_translation_platform.currentText()   # 获取次轮
            config_dict["translation_platform_3"] =  Window.Widget_translation_settings_C.comboBox_final_translation_platform.currentText()    # 获取末轮
            config_dict["split_switch"] =  Window.Widget_translation_settings_C.SwitchButton_split_switch.isChecked() # 获取混合翻译开关

            #开始翻译的备份设置界面
            config_dict["auto_backup_toggle"] =  Window.Widget_start_translation.B_settings.checkBox_switch.isChecked() # 获取备份设置开关




            #获取提示字典界面
            config_dict["prompt_dict_switch"] = Window.Widget_prompt_dict.checkBox2.isChecked() #获取译时提示开关状态
            User_Dictionary2 = {}
            for row in range(Window.Widget_prompt_dict.tableView.rowCount() - 1):
                key_item = Window.Widget_prompt_dict.tableView.item(row, 0)
                value_item = Window.Widget_prompt_dict.tableView.item(row, 1)
                info_item = Window.Widget_prompt_dict.tableView.item(row, 2)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    # 检查一下是不是空值
                    if info_item:
                        info = info_item.data(Qt.DisplayRole)
                        User_Dictionary2[key] = {"translation": value, "info": info}
                    else:
                        # 如果info项为None，可以选择不添加到字典中或者添加一个默认值
                        User_Dictionary2[key] = {"translation": value, "info": None}
            config_dict["User_Dictionary2"] = User_Dictionary2



            #获取译前替换字典界面
            config_dict["Replace_before_translation"] =  Window.Widget_replace_dict.A_settings.checkBox1.isChecked()#获取译前替换开关状态
            User_Dictionary1 = {}
            for row in range(Window.Widget_replace_dict.A_settings.tableView.rowCount() - 1):
                key_item = Window.Widget_replace_dict.A_settings.tableView.item(row, 0)
                value_item = Window.Widget_replace_dict.A_settings.tableView.item(row, 1)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    User_Dictionary1[key] = value
            config_dict["User_Dictionary1"] = User_Dictionary1


            #获取译后替换字典界面
            config_dict["Replace_after_translation"] =  Window.Widget_replace_dict.B_settings.checkBox1.isChecked()#获取译后替换开关状态
            User_Dictionary3 = {}
            for row in range(Window.Widget_replace_dict.B_settings.tableView.rowCount() - 1):
                key_item = Window.Widget_replace_dict.B_settings.tableView.item(row, 0)
                value_item = Window.Widget_replace_dict.B_settings.tableView.item(row, 1)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    User_Dictionary3[key] = value
            config_dict["User_Dictionary3"] = User_Dictionary3




            #获取实时设置界面(openai)
            config_dict["OpenAI_parameter_adjustment"] = Window.Widget_tune_openai.checkBox.isChecked()           #获取开关设置
            config_dict["OpenAI_Temperature"] = Window.Widget_tune_openai.slider1.value()           #获取OpenAI温度
            config_dict["OpenAI_top_p"] = Window.Widget_tune_openai.slider2.value()                 #获取OpenAI top_p
            config_dict["OpenAI_presence_penalty"] = Window.Widget_tune_openai.slider3.value()      #获取OpenAI top_k
            config_dict["OpenAI_frequency_penalty"] = Window.Widget_tune_openai.slider4.value()    #获取OpenAI repetition_penalty

            #获取实时设置界面(anthropic)
            config_dict["Anthropic_parameter_adjustment"] = Window.Widget_tune_anthropic.checkBox.isChecked()           #获取开关设置
            config_dict["Anthropic_Temperature"] = Window.Widget_tune_anthropic.slider1.value()           #获取anthropic 温度

            #获取实时设置界面(google)
            config_dict["Google_parameter_adjustment"] = Window.Widget_tune_google.checkBox.isChecked()           #获取开关设置
            config_dict["Google_Temperature"] = Window.Widget_tune_google.slider1.value()           #获取google 温度

            #获取实时设置界面(sakura)
            config_dict["Sakura_parameter_adjustment"] = Window.Widget_tune_sakura.checkBox.isChecked()           #获取开关设置
            config_dict["Sakura_Temperature"] = Window.Widget_tune_sakura.slider1.value()           #获取sakura温度
            config_dict["Sakura_top_p"] = Window.Widget_tune_sakura.slider2.value()
            config_dict["Sakura_frequency_penalty"] = Window.Widget_tune_sakura.slider4.value()



            #获取提示书界面
            config_dict["system_prompt_switch"] = Window.Widget_system_prompt.checkBox1.isChecked()   #获取自定义提示词开关状态
            config_dict["system_prompt_content"] = Window.Widget_system_prompt.TextEdit1.toPlainText()        #获取自定义提示词输入值 
            config_dict["characterization_switch"] = Window.Widget_characterization.checkBox1.isChecked() #获取角色设定开关状态
            characterization_dictionary = {}
            for row in range(Window.Widget_characterization.tableView.rowCount() - 1):
                original_name = Window.Widget_characterization.tableView.item(row, 0)
                translated_name = Window.Widget_characterization.tableView.item(row, 1)
                character_attributes1 = Window.Widget_characterization.tableView.item(row, 2)
                character_attributes2 = Window.Widget_characterization.tableView.item(row, 3)
                character_attributes3 = Window.Widget_characterization.tableView.item(row, 4)
                character_attributes4 = Window.Widget_characterization.tableView.item(row, 5)
                character_attributes5 = Window.Widget_characterization.tableView.item(row, 6)
                if original_name and translated_name:
                    original_name = original_name.data(Qt.DisplayRole)
                    translated_name = translated_name.data(Qt.DisplayRole)
                    base_dictionary = {"original_name": original_name, "translated_name": translated_name}
                    if character_attributes1:
                        base_dictionary["gender"] = character_attributes1.data(Qt.DisplayRole)
                    if character_attributes2:
                        base_dictionary["age"] = character_attributes2.data(Qt.DisplayRole)
                    if character_attributes3:
                        base_dictionary["personality"] = character_attributes3.data(Qt.DisplayRole)
                    if character_attributes4:
                        base_dictionary["speech_style"] = character_attributes4.data(Qt.DisplayRole)
                    if character_attributes5:
                        base_dictionary["additional_info"] = character_attributes5.data(Qt.DisplayRole)
                    characterization_dictionary[original_name] = base_dictionary
            config_dict["characterization_dictionary"] = characterization_dictionary

            config_dict["world_building_switch"] = Window.Widget_world_building.checkBox1.isChecked()   #获取背景设定开关状态
            config_dict["world_building_content"] = Window.Widget_world_building.TextEdit1.toPlainText()        #获取背景设定文本 
            config_dict["writing_style_switch"] = Window.Widget_writing_style.checkBox1.isChecked()   #获取文风要求开关状态
            config_dict["writing_style_content"] = Window.Widget_writing_style.TextEdit1.toPlainText()        #获取文风要求开关 

            config_dict["translation_example_switch"]= Window.Widget_translation_example.checkBox1.isChecked()#获取添加翻译示例开关状态
            translation_example = {}
            for row in range(Window.Widget_translation_example.tableView.rowCount() - 1):
                key_item = Window.Widget_translation_example.tableView.item(row, 0)
                value_item = Window.Widget_translation_example.tableView.item(row, 1)
                if key_item and value_item:
                    key = key_item.data(Qt.DisplayRole)
                    value = value_item.data(Qt.DisplayRole)
                    translation_example[key] = value
            config_dict["translation_example"] = translation_example



            #将所有的配置信息写入config.json文件中
            with open(os.path.join(resource_dir, "config.json"), "w", encoding="utf-8") as f:
                json.dump(config_dict, f, ensure_ascii=False, indent=4)


        if mode == "read":
            #如果config.json在子文件夹resource中存在
            if os.path.exists(os.path.join(resource_dir, "config.json")):
                #读取config.json
                with open(os.path.join(resource_dir, "config.json"), "r", encoding="utf-8") as f:
                    config_dict = json.load(f)

                #将config.json中的值赋予到变量中,并set到界面上
                #OpenAI官方账号界面
                if "openai_account_type" in config_dict:
                    Window.Widget_Openai.comboBox_account_type.setCurrentText(config_dict["openai_account_type"])
                if "openai_model_type" in config_dict:
                    Window.Widget_Openai.comboBox_model.setCurrentText(config_dict["openai_model_type"])
                if "openai_API_key_str" in config_dict:
                    Window.Widget_Openai.TextEdit_apikey.setText(config_dict["openai_API_key_str"])
                if "openai_proxy_port" in config_dict:
                    Window.Widget_Openai.LineEdit_proxy_port.setText(config_dict["openai_proxy_port"])

                #anthropic官方账号界面
                if "anthropic_account_type" in config_dict:
                    Window.Widget_Anthropic.comboBox_account_type.setCurrentText(config_dict["anthropic_account_type"])
                if "anthropic_model_type" in config_dict:
                    Window.Widget_Anthropic.comboBox_model.setCurrentText(config_dict["anthropic_model_type"])
                if "anthropic_API_key_str" in config_dict:
                    Window.Widget_Anthropic.TextEdit_apikey.setText(config_dict["anthropic_API_key_str"])
                if "anthropic_proxy_port" in config_dict:
                    Window.Widget_Anthropic.LineEdit_proxy_port.setText(config_dict["anthropic_proxy_port"])


                #google官方账号界面
                if "google_account_type" in config_dict:
                    Window.Widget_Google.comboBox_account_type.setCurrentText(config_dict["google_account_type"])
                if "google_model_type" in config_dict:
                    Window.Widget_Google.comboBox_model.setCurrentText(config_dict["google_model_type"])
                if "google_API_key_str" in config_dict:
                    Window.Widget_Google.TextEdit_apikey.setText(config_dict["google_API_key_str"])
                if "google_proxy_port" in config_dict:
                    Window.Widget_Google.LineEdit_proxy_port.setText(config_dict["google_proxy_port"])


                #Cohere官方账号界面
                if "cohere_account_type" in config_dict:
                    Window.Widget_Cohere.comboBox_account_type.setCurrentText(config_dict["cohere_account_type"])
                if "cohere_model_type" in config_dict:
                    Window.Widget_Cohere.comboBox_model.setCurrentText(config_dict["cohere_model_type"])
                if "cohere_API_key_str" in config_dict:
                    Window.Widget_Cohere.TextEdit_apikey.setText(config_dict["cohere_API_key_str"])
                if "cohere_proxy_port" in config_dict:
                    Window.Widget_Cohere.LineEdit_proxy_port.setText(config_dict["cohere_proxy_port"])

                #moonshot官方账号界面
                if "moonshot_account_type" in config_dict:
                    Window.Widget_Moonshot.comboBox_account_type.setCurrentText(config_dict["moonshot_account_type"])
                if "moonshot_model_type" in config_dict:
                    Window.Widget_Moonshot.comboBox_model.setCurrentText(config_dict["moonshot_model_type"])
                if "moonshot_API_key_str" in config_dict:
                    Window.Widget_Moonshot.TextEdit_apikey.setText(config_dict["moonshot_API_key_str"])
                if "moonshot_proxy_port" in config_dict:
                    Window.Widget_Moonshot.LineEdit_proxy_port.setText(config_dict["moonshot_proxy_port"])

                #deepseek官方账号界面
                if "deepseek_model_type" in config_dict:
                    Window.Widget_Deepseek.comboBox_model.setCurrentText(config_dict["deepseek_model_type"])
                if "deepseek_API_key_str" in config_dict:
                    Window.Widget_Deepseek.TextEdit_apikey.setText(config_dict["deepseek_API_key_str"])
                if "deepseek_proxy_port" in config_dict:
                    Window.Widget_Deepseek.LineEdit_proxy_port.setText(config_dict["deepseek_proxy_port"])

                #dashscope官方账号界面
                if "dashscope_model_type" in config_dict:
                    Window.Widget_Dashscope.comboBox_model.setCurrentText(config_dict["dashscope_model_type"])
                if "dashscope_API_key_str" in config_dict:
                    Window.Widget_Dashscope.TextEdit_apikey.setText(config_dict["dashscope_API_key_str"])
                if "dashscope_proxy_port" in config_dict:
                    Window.Widget_Dashscope.LineEdit_proxy_port.setText(config_dict["dashscope_proxy_port"])


                #火山官方账号界面
                if "volcengine_access_point" in config_dict:
                    Window.Widget_Volcengine.A_settings.LineEdit_access_point.setText(config_dict["volcengine_access_point"])
                if "volcengine_API_key_str" in config_dict:
                    Window.Widget_Volcengine.A_settings.TextEdit_apikey.setText(config_dict["volcengine_API_key_str"])
                if "volcengine_proxy_port" in config_dict:
                    Window.Widget_Volcengine.A_settings.LineEdit_proxy_port.setText(config_dict["volcengine_proxy_port"])
                if "volcengine_tokens_limit" in config_dict:
                    Window.Widget_Volcengine.B_settings.spinBox_tokens.setValue(config_dict["volcengine_tokens_limit"])
                if "volcengine_rpm_limit" in config_dict:
                    Window.Widget_Volcengine.B_settings.spinBox_RPM.setValue(config_dict["volcengine_rpm_limit"])
                if "volcengine_tpm_limit" in config_dict:
                    Window.Widget_Volcengine.B_settings.spinBox_TPM.setValue(config_dict["volcengine_tpm_limit"])
                if "volcengine_input_pricing" in config_dict:
                    Window.Widget_Volcengine.B_settings.spinBox_input_pricing.setValue(config_dict["volcengine_input_pricing"])
                if "volcengine_output_pricing" in config_dict:
                    Window.Widget_Volcengine.B_settings.spinBox_output_pricing.setValue(config_dict["volcengine_output_pricing"])


                #智谱官方界面
                if "zhipu_account_type" in config_dict:
                    Window.Widget_ZhiPu.comboBox_account_type.setCurrentText(config_dict["zhipu_account_type"])
                if "zhipu_model_type" in config_dict:
                    Window.Widget_ZhiPu.comboBox_model.setCurrentText(config_dict["zhipu_model_type"])
                if "zhipu_API_key_str" in config_dict:
                    Window.Widget_ZhiPu.TextEdit_apikey.setText(config_dict["zhipu_API_key_str"])
                if "zhipu_proxy_port" in config_dict:
                    Window.Widget_ZhiPu.LineEdit_proxy_port.setText(config_dict["zhipu_proxy_port"])

                #sakura界面
                if "sakura_address" in config_dict:
                    Window.Widget_SakuraLLM.LineEdit_address.setText(config_dict["sakura_address"])
                if "sakura_model_type" in config_dict:
                    Window.Widget_SakuraLLM.comboBox_model.setCurrentText(config_dict["sakura_model_type"])
                if "sakura_proxy_port" in config_dict:
                    Window.Widget_SakuraLLM.LineEdit_proxy_port.setText(config_dict["sakura_proxy_port"])


                #OpenAI代理账号基础界面
                if "op_relay_address" in config_dict:
                    Window.Widget_Proxy.A_settings.LineEdit_relay_address.setText(config_dict["op_relay_address"])
                if "op_proxy_platform" in config_dict:
                    Window.Widget_Proxy.A_settings.comboBox_proxy_platform.setCurrentText(config_dict["op_proxy_platform"])
                    # 根据下拉框的索引调用不同的函数
                    if config_dict["op_proxy_platform"] =="OpenAI":
                        Window.Widget_Proxy.A_settings.comboBox_model_openai.show()
                        Window.Widget_Proxy.A_settings.comboBox_model_anthropic.hide()
                    elif config_dict["op_proxy_platform"] == 'Anthropic':
                        Window.Widget_Proxy.A_settings.comboBox_model_openai.hide()
                        Window.Widget_Proxy.A_settings.comboBox_model_anthropic.show()
                if "op_model_type_openai" in config_dict:

                    # 获取配置文件中指定的模型类型
                    model_type = config_dict["op_model_type_openai"]
                    # 检查模型类型是否已经存在于下拉列表中
                    existing_index = Window.Widget_Proxy.A_settings.comboBox_model_openai.findText(model_type)
                    # 如果模型类型不存在，则添加到下拉列表中
                    if existing_index == -1:
                        Window.Widget_Proxy.A_settings.comboBox_model_openai.addItem(model_type)
                    # 设置当前文本为配置文件中指定的模型类型
                    Window.Widget_Proxy.A_settings.comboBox_model_openai.setCurrentText(model_type)

                if "op_model_type_anthropic" in config_dict:
                    # 获取配置文件中指定的模型类型
                    model_type = config_dict["op_model_type_anthropic"]
                    # 检查模型类型是否已经存在于下拉列表中
                    existing_index = Window.Widget_Proxy.A_settings.comboBox_model_anthropic.findText(model_type)
                    # 如果模型类型不存在，则添加到下拉列表中
                    if existing_index == -1:
                        Window.Widget_Proxy.A_settings.comboBox_model_anthropic.addItem(model_type)
                    # 设置当前文本为配置文件中指定的模型类型
                    Window.Widget_Proxy.A_settings.comboBox_model_anthropic.setCurrentText(model_type)

                    
                if "op_API_key_str" in config_dict:
                    Window.Widget_Proxy.A_settings.TextEdit_apikey.setText(config_dict["op_API_key_str"])
                if "op_proxy_port" in config_dict:
                    Window.Widget_Proxy.A_settings.LineEdit_proxy_port.setText(config_dict["op_proxy_port"])




                #OpenAI代理账号进阶界面
                if "op_tokens_limit" in config_dict:
                    Window.Widget_Proxy.B_settings.spinBox_tokens.setValue(config_dict["op_tokens_limit"])
                if "op_rpm_limit" in config_dict:
                    Window.Widget_Proxy.B_settings.spinBox_RPM.setValue(config_dict["op_rpm_limit"])
                if "op_tpm_limit" in config_dict:
                    Window.Widget_Proxy.B_settings.spinBox_TPM.setValue(config_dict["op_tpm_limit"])
                if "op_input_pricing" in config_dict:
                    Window.Widget_Proxy.B_settings.spinBox_input_pricing.setValue(config_dict["op_input_pricing"])
                if "op_output_pricing" in config_dict:
                    Window.Widget_Proxy.B_settings.spinBox_output_pricing.setValue(config_dict["op_output_pricing"])



                #翻译设置基础界面
                if "translation_project" in config_dict:
                    Window.Widget_translation_settings_A.comboBox_translation_project.setCurrentText(config_dict["translation_project"])
                if "translation_platform" in config_dict:
                    Window.Widget_translation_settings_A.comboBox_translation_platform.setCurrentText(config_dict["translation_platform"])
                if "source_language" in config_dict:
                    Window.Widget_translation_settings_A.comboBox_source_text.setCurrentText(config_dict["source_language"])
                if "target_language" in config_dict:
                    Window.Widget_translation_settings_A.comboBox_translated_text.setCurrentText(config_dict["target_language"])
                if "label_input_path" in config_dict:
                    Window.Widget_translation_settings_A.label_input_path.setText(config_dict["label_input_path"])
                if "label_output_path" in config_dict:
                    Window.Widget_translation_settings_A.label_output_path.setText(config_dict["label_output_path"])



                #翻译设置进阶界面
                if "lines_limit_switch" in config_dict:
                    Window.Widget_translation_settings_B1.checkBox_lines_limit_switch.setChecked(config_dict["lines_limit_switch"])
                if "lines_limit" in config_dict:
                    Window.Widget_translation_settings_B1.spinBox_lines_limit.setValue(config_dict["lines_limit"])
                if "tokens_limit_switch" in config_dict:
                    Window.Widget_translation_settings_B1.checkBox_tokens_limit_switch.setChecked(config_dict["tokens_limit_switch"])
                if "tokens_limit" in config_dict:
                    Window.Widget_translation_settings_B1.spinBox_tokens_limit.setValue(config_dict["tokens_limit"])
                if "pre_line_counts" in config_dict:
                    Window.Widget_translation_settings_B1.spinBox_pre_lines.setValue(config_dict["pre_line_counts"])
                if "thread_counts" in config_dict:
                    Window.Widget_translation_settings_B1.spinBox_thread_count.setValue(config_dict["thread_counts"])
                if "retry_count_limit" in config_dict:
                    Window.Widget_translation_settings_B1.spinBox_retry_count_limit.setValue(config_dict["retry_count_limit"])
                if "round_limit" in config_dict:
                     Window.Widget_translation_settings_B1.spinBox_round_limit.setValue(config_dict["round_limit"]) 
                if "cot_toggle" in config_dict:
                    Window.Widget_translation_settings_B2.SwitchButton_cot_toggle.setChecked(config_dict["cot_toggle"])
                if "cn_prompt_toggle" in config_dict:
                    Window.Widget_translation_settings_B2.SwitchButton_cn_prompt_toggle.setChecked(config_dict["cn_prompt_toggle"])
                if "preserve_line_breaks_toggle" in config_dict:
                    Window.Widget_translation_settings_B2.SwitchButton_line_breaks.setChecked(config_dict["preserve_line_breaks_toggle"])
                if "response_conversion_toggle" in config_dict:
                    Window.Widget_translation_settings_B2.SwitchButton_conversion_toggle.setChecked(config_dict["response_conversion_toggle"])
                if "text_clear_toggle" in config_dict:
                    Window.Widget_translation_settings_B2.SwitchButton_clear.setChecked(config_dict["text_clear_toggle"])


                #翻译设置混合翻译界面
                if "translation_mixing_toggle" in config_dict:
                    Window.Widget_translation_settings_C.SwitchButton_mixed_translation.setChecked(config_dict["translation_mixing_toggle"])
                if "translation_platform_1" in config_dict:
                    Window.Widget_translation_settings_C.comboBox_primary_translation_platform.setCurrentText(config_dict["translation_platform_1"])
                if "translation_platform_2" in config_dict:
                    Window.Widget_translation_settings_C.comboBox_secondary_translation_platform.setCurrentText(config_dict["translation_platform_2"])
                if "translation_platform_3" in config_dict:
                    Window.Widget_translation_settings_C.comboBox_final_translation_platform.setCurrentText(config_dict["translation_platform_3"])
                if "split_switch" in config_dict:
                    Window.Widget_translation_settings_C.SwitchButton_split_switch.setChecked(config_dict["split_switch"])

                #开始翻译的备份设置界面
                if "auto_backup_toggle" in config_dict:
                    Window.Widget_start_translation.B_settings.checkBox_switch.setChecked(config_dict["auto_backup_toggle"])



                #提示字典界面
                if "User_Dictionary2" in config_dict:
                    User_Dictionary2 = config_dict["User_Dictionary2"]
                    if User_Dictionary2:
                        for key, value in User_Dictionary2.items():
                            row = Window.Widget_prompt_dict.tableView.rowCount() - 1
                            Window.Widget_prompt_dict.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            # 兼容旧版本存储格式
                            if isinstance(value, dict):
                                value_item = QTableWidgetItem(value["translation"])
                                info_item = QTableWidgetItem(value["info"])
                                Window.Widget_prompt_dict.tableView.setItem(row, 0, key_item)
                                Window.Widget_prompt_dict.tableView.setItem(row, 1, value_item)
                                Window.Widget_prompt_dict.tableView.setItem(row, 2, info_item)
                            else:
                                value_item = QTableWidgetItem(value)
                                Window.Widget_prompt_dict.tableView.setItem(row, 0, key_item)
                                Window.Widget_prompt_dict.tableView.setItem(row, 1, value_item)                                  
                        #删除第一行
                        Window.Widget_prompt_dict.tableView.removeRow(0)
                if "prompt_dict_switch" in config_dict:
                    Change_translation_prompt = config_dict["prompt_dict_switch"]
                    Window.Widget_prompt_dict.checkBox2.setChecked(Change_translation_prompt)


                #译前替换字典界面
                if "User_Dictionary1" in config_dict:
                    User_Dictionary1 = config_dict["User_Dictionary1"]
                    if User_Dictionary1:
                        for key, value in User_Dictionary1.items():
                            row = Window.Widget_replace_dict.A_settings.tableView.rowCount() - 1
                            Window.Widget_replace_dict.A_settings.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            value_item = QTableWidgetItem(value)
                            Window.Widget_replace_dict.A_settings.tableView.setItem(row, 0, key_item)
                            Window.Widget_replace_dict.A_settings.tableView.setItem(row, 1, value_item)        
                        #删除第一行
                        Window.Widget_replace_dict.A_settings.tableView.removeRow(0)
                if "Replace_before_translation" in config_dict:
                    Replace_before_translation = config_dict["Replace_before_translation"]
                    Window.Widget_replace_dict.A_settings.checkBox1.setChecked(Replace_before_translation)


                #译后替换字典界面
                if "User_Dictionary3" in config_dict:
                    User_Dictionary3 = config_dict["User_Dictionary3"]
                    if User_Dictionary3:
                        for key, value in User_Dictionary3.items():
                            row = Window.Widget_replace_dict.B_settings.tableView.rowCount() - 1
                            Window.Widget_replace_dict.B_settings.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            value_item = QTableWidgetItem(value)
                            Window.Widget_replace_dict.B_settings.tableView.setItem(row, 0, key_item)
                            Window.Widget_replace_dict.B_settings.tableView.setItem(row, 1, value_item)
                        #删除第一行
                        Window.Widget_replace_dict.B_settings.tableView.removeRow(0)
                if "Replace_after_translation" in config_dict:
                    Replace_after_translation = config_dict["Replace_after_translation"]
                    Window.Widget_replace_dict.B_settings.checkBox1.setChecked(Replace_after_translation)



                #实时设置界面(openai)
                if "OpenAI_parameter_adjustment" in config_dict:
                    OpenAI_parameter_adjustment = config_dict["OpenAI_parameter_adjustment"]
                    Window.Widget_tune_openai.checkBox.setChecked(OpenAI_parameter_adjustment)
                if "OpenAI_Temperature" in config_dict:
                    OpenAI_Temperature = config_dict["OpenAI_Temperature"]
                    Window.Widget_tune_openai.slider1.setValue(OpenAI_Temperature)
                if "OpenAI_top_p" in config_dict:
                    OpenAI_top_p = config_dict["OpenAI_top_p"]
                    Window.Widget_tune_openai.slider2.setValue(OpenAI_top_p)
                if "OpenAI_presence_penalty" in config_dict:
                    OpenAI_presence_penalty = config_dict["OpenAI_presence_penalty"]
                    Window.Widget_tune_openai.slider3.setValue(OpenAI_presence_penalty)
                if "OpenAI_frequency_penalty" in config_dict:
                    OpenAI_frequency_penalty = config_dict["OpenAI_frequency_penalty"]
                    Window.Widget_tune_openai.slider4.setValue(OpenAI_frequency_penalty)

                #实时设置界面(anthropic)
                if "Anthropic_parameter_adjustment" in config_dict:
                    Anthropic_parameter_adjustment = config_dict["Anthropic_parameter_adjustment"]
                    Window.Widget_tune_anthropic.checkBox.setChecked(Anthropic_parameter_adjustment)
                if "Anthropic_Temperature" in config_dict:
                    Anthropic_Temperature = config_dict["Anthropic_Temperature"]
                    Window.Widget_tune_anthropic.slider1.setValue(Anthropic_Temperature)

                #实时设置界面(google)
                if "Google_parameter_adjustment" in config_dict:
                    Google_parameter_adjustment = config_dict["Google_parameter_adjustment"]
                    Window.Widget_tune_google.checkBox.setChecked(Google_parameter_adjustment)
                if "Google_Temperature" in config_dict:
                    Google_Temperature = config_dict["Google_Temperature"]
                    Window.Widget_tune_google.slider1.setValue(Google_Temperature)

                #实时设置界面(sakura)
                if "Sakura_parameter_adjustment" in config_dict:
                    Sakura_parameter_adjustment = config_dict["Sakura_parameter_adjustment"]
                    Window.Widget_tune_sakura.checkBox.setChecked(Sakura_parameter_adjustment)
                if "Sakura_Temperature" in config_dict:
                    Sakura_Temperature = config_dict["Sakura_Temperature"]
                    Window.Widget_tune_sakura.slider1.setValue(Sakura_Temperature)
                if "Sakura_top_p" in config_dict:
                    Sakura_top_p = config_dict["Sakura_top_p"]
                    Window.Widget_tune_sakura.slider2.setValue(Sakura_top_p)
                if  "Sakura_frequency_penalty" in config_dict:
                    Sakura_frequency_penalty = config_dict["Sakura_frequency_penalty"]
                    Window.Widget_tune_sakura.slider4.setValue(Sakura_frequency_penalty)


                #提示书界面
                if "system_prompt_switch" in config_dict:
                    system_prompt_switch = config_dict["system_prompt_switch"]
                    Window.Widget_system_prompt.checkBox1.setChecked(system_prompt_switch)
                if "system_prompt_content" in config_dict:
                    system_prompt_content = config_dict["system_prompt_content"]
                    Window.Widget_system_prompt.TextEdit1.setText(system_prompt_content)

                if "characterization_switch" in config_dict:
                    characterization_switch = config_dict["characterization_switch"]
                    Window.Widget_characterization.checkBox1.setChecked(characterization_switch)
                if "characterization_dictionary" in config_dict:
                    characterization_dictionary = config_dict["characterization_dictionary"]
                    if characterization_dictionary:
                        for key, value in characterization_dictionary.items():
                            row = Window.Widget_characterization.tableView.rowCount() - 1
                            Window.Widget_characterization.tableView.insertRow(row)

                            original_name = QTableWidgetItem(value["original_name"])
                            translated_name = QTableWidgetItem(value["translated_name"])
                            Window.Widget_characterization.tableView.setItem(row, 0, original_name)
                            Window.Widget_characterization.tableView.setItem(row, 1, translated_name)

                            if (value.get('gender')):
                                character_attributes1 = QTableWidgetItem(value["gender"])
                                Window.Widget_characterization.tableView.setItem(row, 2, character_attributes1)
                            if (value.get('age')):
                                character_attributes2 = QTableWidgetItem(value["age"])
                                Window.Widget_characterization.tableView.setItem(row, 3, character_attributes2)
                            if (value.get('personality')):
                                character_attributes3 = QTableWidgetItem(value["personality"])
                                Window.Widget_characterization.tableView.setItem(row, 4, character_attributes3)
                            if (value.get('speech_style')):
                                character_attributes4 = QTableWidgetItem(value["speech_style"])
                                Window.Widget_characterization.tableView.setItem(row, 5, character_attributes4)
                            if (value.get('additional_info')):
                                character_attributes5 = QTableWidgetItem(value["additional_info"])
                                Window.Widget_characterization.tableView.setItem(row, 6, character_attributes5)
                        #删除第一行
                        Window.Widget_characterization.tableView.removeRow(0)

                if "world_building_switch" in config_dict:
                    world_building_switch = config_dict["world_building_switch"]
                    Window.Widget_world_building.checkBox1.setChecked(world_building_switch)
                if "world_building_content" in config_dict:
                    world_building_content = config_dict["world_building_content"]
                    Window.Widget_world_building.TextEdit1.setText(world_building_content)

                if "writing_style_switch" in config_dict:
                    writing_style_switch = config_dict["writing_style_switch"]
                    Window.Widget_writing_style.checkBox1.setChecked(writing_style_switch)
                if "writing_style_content" in config_dict:
                    writing_style_content = config_dict["writing_style_content"]
                    Window.Widget_writing_style.TextEdit1.setText(writing_style_content)

                if "translation_example_switch" in config_dict:
                    translation_example_switch = config_dict["translation_example_switch"]
                    Window.Widget_translation_example.checkBox1.setChecked(translation_example_switch)
                if "translation_example" in config_dict:
                    translation_example = config_dict["translation_example"]
                    if translation_example:
                        for key, value in translation_example.items():
                            row = Window.Widget_translation_example.tableView.rowCount() - 1
                            Window.Widget_translation_example.tableView.insertRow(row)
                            key_item = QTableWidgetItem(key)
                            value_item = QTableWidgetItem(value)
                            Window.Widget_translation_example.tableView.setItem(row, 0, key_item)
                            Window.Widget_translation_example.tableView.setItem(row, 1, value_item)        
                        #删除第一行
                        Window.Widget_translation_example.tableView.removeRow(0)




# 请求限制器
class Request_Limiter():
    def __init__(self):

        # TPM相关参数
        self.max_tokens = 0  # 令牌桶最大容量
        self.remaining_tokens = 0 # 令牌桶剩余容量
        self.tokens_rate = 0 # 令牌每秒的恢复速率
        self.last_time = time.time() # 上次记录时间

        # RPM相关参数
        self.last_request_time = 0  # 上次记录时间
        self.request_interval = 0  # 请求的最小时间间隔（s）
        self.lock = threading.Lock()



    def initialize_limiter(self):
        global Running_status

        # 获取翻译平台
        translation_platform = configurator.translation_platform


        #根据翻译平台读取配置信息
        if translation_platform == 'OpenAI官方':
            # 获取账号类型
            account_type = configurator.openai_account_type
            # 获取模型选择 
            model = configurator.openai_model_type

            # 获取相应的限制
            max_tokens = configurator.openai_platform_config[account_type][model]["max_tokens"]
            TPM_limit = configurator.openai_platform_config[account_type][model]["TPM"]
            RPM_limit = configurator.openai_platform_config[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        #根据翻译平台读取配置信息
        elif translation_platform == 'Anthropic官方':
            # 获取账号类型
            account_type = configurator.anthropic_account_type
            # 获取模型选择 
            model = configurator.anthropic_model_type

            # 获取相应的限制
            max_tokens = configurator.anthropic_platform_config[account_type]["max_tokens"]
            TPM_limit = configurator.anthropic_platform_config[account_type]["TPM"]
            RPM_limit = configurator.anthropic_platform_config[account_type]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        #根据翻译平台读取配置信息
        elif translation_platform == 'Cohere官方':
            # 获取账号类型
            account_type = configurator.cohere_account_type
            # 获取模型选择 
            model = configurator.cohere_model_type

            # 获取相应的限制
            max_tokens = configurator.cohere_platform_config[account_type][model]["max_tokens"]
            TPM_limit = configurator.cohere_platform_config[account_type][model]["TPM"]
            RPM_limit = configurator.cohere_platform_config[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)

        elif translation_platform == 'Google官方':
            # 获取账号类型
            account_type = configurator.google_account_type
            # 获取模型
            model = configurator.google_model_type

            # 获取相应的限制
            max_tokens = configurator.google_platform_config[account_type][model]["max_tokens"]
            TPM_limit = configurator.google_platform_config[account_type][model]["TPM"]
            RPM_limit = configurator.google_platform_config[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        #根据翻译平台读取配置信息
        elif translation_platform == 'Moonshot官方':
            # 获取账号类型
            account_type = configurator.moonshot_account_type
            # 获取模型选择 
            model = configurator.moonshot_model_type

            # 获取相应的限制
            max_tokens = configurator.moonshot_platform_config[account_type][model]["max_tokens"]
            TPM_limit = configurator.moonshot_platform_config[account_type][model]["TPM"]
            RPM_limit = configurator.moonshot_platform_config[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        #根据翻译平台读取配置信息
        elif translation_platform == 'Deepseek官方':
            # 获取模型选择 
            model = configurator.dashscope_model_type

            # 获取相应的限制
            max_tokens = configurator.deepseek_platform_config[model]["max_tokens"]
            TPM_limit = configurator.deepseek_platform_config[model]["TPM"]
            RPM_limit = configurator.deepseek_platform_config[model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        elif translation_platform == '智谱官方':
            # 获取账号类型
            account_type = configurator.zhipu_account_type
            # 获取模型
            model = configurator.zhipu_model_type

            # 获取相应的限制
            max_tokens =  configurator.zhipu_platform_config[account_type][model]["max_tokens"]
            TPM_limit =  configurator.zhipu_platform_config[account_type][model]["TPM"]
            RPM_limit =  configurator.zhipu_platform_config[account_type][model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        #根据翻译平台读取配置信息
        elif translation_platform == 'Dashscope官方':
            # 获取模型选择 
            model = configurator.dashscope_model_type

            # 获取相应的限制
            max_tokens = configurator.dashscope_platform_config[model]["max_tokens"]
            TPM_limit = configurator.dashscope_platform_config[model]["TPM"]
            RPM_limit = configurator.dashscope_platform_config[model]["RPM"]

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        #根据翻译平台读取配置信息
        elif translation_platform == 'Volcengine官方':

            # 获取相应的限制
            max_tokens = configurator.volcengine_tokens_limit               #获取每次文本发送上限限制值
            RPM_limit = configurator.volcengine_rpm_limit               #获取rpm限制值
            TPM_limit = configurator.volcengine_tpm_limit           #获取tpm限制值

            # 获取当前key的数量，对限制进行倍数更改
            key_count = len(configurator.apikey_list)
            RPM_limit = RPM_limit * key_count
            TPM_limit = TPM_limit * key_count

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        elif translation_platform == 'SakuraLLM':
            # 获取模型
            model = configurator.sakura_model_type

            # 获取相应的限制
            max_tokens = configurator.sakurallm_platform_config[model]["max_tokens"]
            TPM_limit = configurator.sakurallm_platform_config[model]["TPM"]
            RPM_limit = configurator.sakurallm_platform_config[model]["RPM"]

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


        else:            
            max_tokens = configurator.op_tokens_limit               #获取每次文本发送上限限制值
            RPM_limit = configurator.op_rpm_limit               #获取rpm限制值
            TPM_limit = configurator.op_tpm_limit             #获取tpm限制值

            # 设置限制
            self.set_limit(max_tokens,TPM_limit,RPM_limit)


    # 设置限制器的参数
    def set_limit(self,max_tokens,TPM,RPM):
        # 将TPM转换成每秒tokens数
        TPs = TPM / 60

        # 将RPM转换成请求的最小时间间隔(S)
        request_interval =  60 / RPM

        # 设置限制器的TPM参数
        self.max_tokens = max_tokens  # 令牌桶最大容量
        self.remaining_tokens = max_tokens # 令牌桶剩余容量
        self.tokens_rate = TPs # 令牌每秒的恢复速率      

        # 设置限制器的RPM参数
        self.request_interval =request_interval  # 请求的最小时间间隔（s）


    def RPM_limit(self):
        with self.lock:
            current_time = time.time() # 获取现在的时间
            time_since_last_request = current_time - self.last_request_time # 计算当前时间与上次记录时间的间隔
            if time_since_last_request < self.request_interval: 
                # print("[DEBUG] Request limit exceeded. Please try again later.")
                return False
            else:
                self.last_request_time = current_time
                return True



    def TPM_limit(self, tokens):
        now = time.time() # 获取现在的时间
        tokens_to_add = (now - self.last_time) * self.tokens_rate #现在时间减去上一次记录的时间，乘以恢复速率，得出这段时间恢复的tokens数量
        self.remaining_tokens = min(self.max_tokens, self.remaining_tokens + tokens_to_add) #计算新的剩余容量，与最大容量比较，谁小取谁值，避免发送信息超过最大容量
        self.last_time = now # 改变上次记录时间

        if tokens > self.remaining_tokens:
            #print("[DEBUG] 已超过剩余tokens：", tokens,'\n' )
            return False
        else:
           # print("[DEBUG] 数量足够，剩余tokens：", tokens,'\n' )
            return True

    def RPM_and_TPM_limit(self, tokens):
        if self.RPM_limit() and self.TPM_limit(tokens):
            # 如果能够发送请求，则扣除令牌桶里的令牌数
            self.remaining_tokens = self.remaining_tokens - tokens
            return True
        else:
            return False



    # 计算消息列表内容的tokens的函数
    def num_tokens_from_messages(self,messages):
        """Return the number of tokens used by a list of messages."""
        try:
            encoding = tiktoken.encoding_for_model("gpt-3.5-turbo")
        except KeyError:
            print("Warning: model not found. Using cl100k_base encoding.")
            encoding = tiktoken.get_encoding("cl100k_base")

        tokens_per_message = 3
        tokens_per_name = 1
        num_tokens = 0
        for message in messages:
            num_tokens += tokens_per_message
            for key, value in message.items():
                #如果value是字符串类型才计算tokens，否则跳过，因为AI在调用函数时，会在content中回复null，导致报错
                if isinstance(value, str):
                    num_tokens += len(encoding.encode(value))
                if key == "name":
                    num_tokens += tokens_per_name
        num_tokens += 3  # every reply is primed with <|start|>assistant<|message|>
        return num_tokens


    # 计算单个字符串tokens数量函数
    def num_tokens_from_string(self,string):
        """Returns the number of tokens in a text string."""
        encoding = tiktoken.get_encoding("cl100k_base")
        num_tokens = len(encoding.encode(string))
        return num_tokens



# 界面提示器
class User_Interface_Prompter(QObject):
    signal = pyqtSignal(str,str,int) #创建信号,并确定发送参数类型

    def __init__(self):
       super().__init__()  # 调用父类的构造函数
       self.stateTooltip = None # 存储翻译状态控件
       self.total_text_line_count = 0 # 存储总文本行数
       self.translated_line_count = 0 # 存储已经翻译文本行数
       self.progress = 0.0           # 存储翻译进度
       self.tokens_spent = 0  # 存储已经花费的tokens
       self.amount_spent = 0  # 存储已经花费的金钱


    # 槽函数，用于接收子线程发出的信号，更新界面UI的状态，因为子线程不能更改父线程的QT的UI控件的值
    def on_update_ui(self,input_str1,input_str2,iunput_int1):

        if input_str1 == "翻译状态提示":
            if input_str2 == "开始翻译":
                self.stateTooltip = StateToolTip(" 正在进行翻译中，客官请耐心等待哦~~", "　　　当前任务开始于 " + datetime.datetime.now().strftime("%Y.%m.%d %H:%M:%S"), Window)
                self.stateTooltip.move(510, 30) # 设定控件的出现位置，该位置是传入的Window窗口的位置
                self.stateTooltip.show()

            elif input_str2 == "翻译暂停":
                print("\033[1;33mWarning:\033[0m 翻译任务已被暂停-----------------------","\n")
                self.stateTooltip.setContent('翻译已暂停')
                self.stateTooltip.setState(True)
                self.stateTooltip = None
                #界面提示
                self.createSuccessInfoBar("翻译任务已全部暂停")

            elif input_str2 == "翻译取消":
                print("\033[1;33mWarning:\033[0m 翻译任务已被取消-----------------------","\n")
                self.stateTooltip.setContent('翻译已取消')
                self.stateTooltip.setState(True)
                self.stateTooltip = None
                #界面提示
                self.createSuccessInfoBar("翻译任务已全部取消")

                #重置翻译界面数据
                Window.Widget_start_translation.A_settings.translation_project.setText("无")
                Window.Widget_start_translation.A_settings.project_id.setText("无")
                Window.Widget_start_translation.A_settings.total_text_line_count.setText("无")
                Window.Widget_start_translation.A_settings.translated_line_count.setText("无")
                Window.Widget_start_translation.A_settings.tokens_spent.setText("无")
                Window.Widget_start_translation.A_settings.amount_spent.setText("无")
                Window.Widget_start_translation.A_settings.progressRing.setValue(0)


            elif input_str2 == "翻译完成":
                self.stateTooltip.setContent('已经翻译完成啦 😆')
                self.stateTooltip.setState(True)
                self.stateTooltip = None

                #隐藏继续翻译按钮
                Window.Widget_start_translation.A_settings.primaryButton_continue_translation.hide()
                #隐藏暂停翻译按钮
                Window.Widget_start_translation.A_settings.primaryButton_pause_translation.hide()
                #显示开始翻译按钮
                Window.Widget_start_translation.A_settings.primaryButton_start_translation.show()

        elif input_str1 == "接口测试结果":
            if input_str2 == "测试成功":
                self.createSuccessInfoBar("全部Apikey请求测试成功")
            else:
                self.createErrorInfoBar("存在Apikey请求测试失败")


        elif input_str1 == "初始化翻译界面数据":
            # 更新翻译项目信息
            translation_project = configurator.translation_project
            Window.Widget_start_translation.A_settings.translation_project.setText(translation_project)

            # 更新项目ID信息
            Window.Widget_start_translation.A_settings.project_id.setText(input_str2)

            # 更新需要翻译的文本行数信息
            self.total_text_line_count = iunput_int1 #存储总文本行数
            Window.Widget_start_translation.A_settings.total_text_line_count.setText(str(self.total_text_line_count))

            # 其他信息设置为0
            Window.Widget_start_translation.A_settings.translated_line_count.setText("0")
            Window.Widget_start_translation.A_settings.tokens_spent.setText("0")
            Window.Widget_start_translation.A_settings.amount_spent.setText("0")
            Window.Widget_start_translation.A_settings.progressRing.setValue(0)

            # 初始化存储的数值
            self.translated_line_count = 0 
            self.tokens_spent = 0  
            self.amount_spent = 0  
            self.progress = 0.0 



        elif input_str1 == "重置界面数据":

            #重置翻译界面数据
            Window.Widget_start_translation.A_settings.translation_project.setText("无")
            Window.Widget_start_translation.A_settings.project_id.setText("无")
            Window.Widget_start_translation.A_settings.total_text_line_count.setText("无")
            Window.Widget_start_translation.A_settings.translated_line_count.setText("无")
            Window.Widget_start_translation.A_settings.tokens_spent.setText("无")
            Window.Widget_start_translation.A_settings.amount_spent.setText("无")
            Window.Widget_start_translation.A_settings.progressRing.setValue(0)


        elif input_str1 == "更新翻译界面数据":

            Window.Widget_start_translation.A_settings.translated_line_count.setText(str(self.translated_line_count))

            Window.Widget_start_translation.A_settings.tokens_spent.setText(str(self.tokens_spent))

            Window.Widget_start_translation.A_settings.amount_spent.setText(str(self.amount_spent))

            progress = int(round(self.progress, 0))
            Window.Widget_start_translation.A_settings.progressRing.setValue(progress)

    
    # 更新翻译进度数据
    def update_data(self, state, translated_line_count, prompt_tokens_used, completion_tokens_used):

        #根据模型设定单位价格
        if configurator.translation_platform == "OpenAI官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.openai_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.openai_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Anthropic官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.anthropic_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.anthropic_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Cohere官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.cohere_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.cohere_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Google官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.google_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.google_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Moonshot官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.moonshot_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.moonshot_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Deepseek官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.deepseek_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.deepseek_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Dashscope官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.dashscope_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.dashscope_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "Volcengine官方":
            # 获取使用的模型输入价格与输出价格
            input_price = Window.Widget_Volcengine.B_settings.spinBox_input_pricing.value()               #获取输入价格
            output_price = Window.Widget_Volcengine.B_settings.spinBox_output_pricing.value()               #获取输出价格

        elif configurator.translation_platform == "智谱官方":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.zhipu_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.zhipu_platform_config["model_price"][configurator.model_type]["output_price"]

        elif configurator.translation_platform == "SakuraLLM":
            # 获取使用的模型输入价格与输出价格
            input_price = configurator.sakurallm_platform_config["model_price"][configurator.model_type]["input_price"]
            output_price = configurator.sakurallm_platform_config["model_price"][configurator.model_type]["output_price"]

        else:
            # 获取使用的模型输入价格与输出价格
            input_price = Window.Widget_Proxy.B_settings.spinBox_input_pricing.value()               #获取输入价格
            output_price = Window.Widget_Proxy.B_settings.spinBox_output_pricing.value()               #获取输出价格

        #计算已经翻译的文本数
        if state == 1:
            # 更新已经翻译的文本数
            self.translated_line_count = self.translated_line_count + translated_line_count   

        #计算tokens花销
        self.tokens_spent = self.tokens_spent + prompt_tokens_used + completion_tokens_used

        #计算金额花销
        self.amount_spent = self.amount_spent + (input_price/1000 * prompt_tokens_used)  + (output_price/1000 * completion_tokens_used) 
        self.amount_spent = round(self.amount_spent, 4)

        #计算进度条
        result = self.translated_line_count / self.total_text_line_count * 100
        self.progress = round(result, 2)

        #print("[DEBUG] 总行数：",self.total_text_line_count,"已翻译行数：",self.translated_line_count,"进度：",self.progress,"%")


    #成功信息居中弹出框函数
    def createSuccessInfoBar(self,str):
        InfoBar.success(
            title='[Success]',
            content=str,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=2000,
            parent=Window
            )

    #错误信息右下方弹出框函数
    def createErrorInfoBar(self,str):
        InfoBar.error(
            title='[Error]',
            content=str,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM_RIGHT,
            duration=-1,    # won't disappear automatically
            parent=Window
            )

    #提醒信息左上角弹出框函数
    def createWarningInfoBar(self,str):
        InfoBar.warning(
            title='[Warning]',
            content=str,
            orient=Qt.Horizontal,
            isClosable=False,   # disable close button
            position=InfoBarPosition.TOP_LEFT,
            duration=2000,
            parent=Window
            )



# 文件读取器
class File_Reader():
    def __init__(self):
        pass


    # 生成项目ID
    def generate_project_id(self,prefix):
        # 获取当前时间，并将其格式化为数字字符串
        current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

        # 生成5位随机数
        random_number = random.randint(10000, 99999)

        # 组合生成项目ID
        project_id = f"{current_time}{prefix}{random_number}"
        
        return project_id


    # 读取文件夹中树形结构Paratranz json 文件
    def read_paratranz_files(self, folder_path):
        # 待处理的json接口例
        # [
        #     {
        #         "key": "Activate",
        #         "original": "カードをプレイ",
        #         "translation": "出牌",
        #         "context": null
        #     }
        # ]
        # 缓存数据结构示例
        ex_cache_data = [
            {'project_type': 'Paratranz'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json', 'key': 'txtKey',
             'context': ''},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json', 'key': 'txtKey',
             'context': ''},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ',
             'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json',
             'file_name': 'Replace the original text.json', 'key': 'txtKey', 'context': ''},
        ]

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self, "Paratranz")
        json_data_list.append({
            "project_type": "Paratranz",
            "project_id": project_id,
        })

        # 文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file)  # 构建文件路径

                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for item in json_data:
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = item.get('original', '')  # 获取原文，如果没有则默认为空字符串
                            translated_text = item.get('translation', '')  # 获取翻译，如果没有则默认为空字符串
                            key = item.get('key', '')  # 获取键值，如果没有则默认为空字符串
                            context = item.get('context', '')  # 获取上下文信息，如果没有则默认为空字符串
                            storage_path = os.path.relpath(file_path, folder_path)
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "model": "none",
                                "storage_path": storage_path,
                                "file_name": file_name,
                                "key": key,
                                "context": context
                            })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    # 读取文件夹中树形结构Mtool文件
    def read_mtool_files (self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Mtool")
        json_data_list.append({
            "project_type": "Mtool",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for key, value in json_data.items():
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = key
                            translated_text = value
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "model": "none",
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    #读取文件夹中树形结构的xlsx文件， 存到列表变量中
    def read_xlsx_files(self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'T++'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 1},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 2},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 3},
        ]

        # 创建列表
        cache_list = []
        # 添加文件头
        project_id = File_Reader.generate_project_id(self,"T++")
        cache_list.append({
            "project_type": "T++",
            "project_id": project_id,
        })
        #文本索引初始值
        i = 1

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file) #构建文件路径

                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb.active
                    for row in range(2, sheet.max_row + 1): # 从第二行开始读取，因为第一行是标识头，通常不用理会
                        cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                        cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值

                        source_text = cell_value1  # 获取原文
                        storage_path = os.path.relpath(file_path, folder_path) # 用文件的绝对路径和输入文件夹路径“相减”，获取相对的文件路径
                        file_name = file #获取文件名

                        #第1列的值不为空，和第2列的值为空，是未翻译内容
                        if cell_value1 and cell_value2 is  None:
                            
                            translated_text = "无"
                            cache_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "model": "none",
                                "storage_path": storage_path,
                                "file_name": file_name,
                                "row_index": row ,
                            })

                            i = i + 1 # 增加文本索引值

                        # 第1列的值不为空，和第2列的值不为空，是已经翻译内容
                        elif cell_value1 and cell_value2 :

                            translated_text = cell_value2
                            cache_list.append({
                                "text_index": i,
                                "translation_status": 1,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "storage_path": storage_path,
                                "model": "none",
                                "file_name": file_name,
                                "row_index": row ,
                            })

                            i = i + 1 # 增加文本索引值



        return cache_list
    

    # 读取文件夹中树形结构VNText导出文件
    def read_vnt_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Vnt")
        json_data_list.append({
            "project_type": "Vnt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for entry in json_data:
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = entry["message"]
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file

                            name = entry.get("name")
                            if name:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "name": name,
                                    "model": "none",
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })
                            else:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "model": "none",
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    #读取缓存文件
    def read_cache_files(self,folder_path):
        # 获取文件夹中的所有文件
        files = os.listdir(folder_path)

        # 查找以 "CacheData" 开头且以 ".json" 结尾的文件
        json_files = [file for file in files if file.startswith("AinieeCacheData") and file.endswith(".json")]

        if not json_files:
            print(f"Error: No 'CacheData' JSON files found in folder '{folder_path}'.")
            return None

        # 选择第一个符合条件的 JSON 文件
        json_file_path = os.path.join(folder_path, json_files[0])

        # 读取 JSON 文件内容
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            return data


    # 读取文件夹中树形结构Srt字幕文件
    def read_srt_files (self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Srt")
        json_data_list.append({
            "project_type": "Srt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1
        source_text = ''
        subtitle_number = ''
        subtitle_time = ''

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".srt"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    # 将内容按行分割,并除去换行
                    lines = content.split('\n')
                    # 计数变量
                    j = 1

                    # 遍历每一行
                    for line in lines:

                        # 去除行首的BOM（如果存在）
                        line = line.lstrip('\ufeff')

                        # 如果行是数字，代表新的字幕开始
                        if line.isdigit() and (line == str(j)):
                            subtitle_number = line

                        # 时间码行
                        elif ' --> ' in line:
                            subtitle_time = line

                        # 空行代表字幕文本的结束
                        elif line == '':
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": source_text,
                                "model": "none",
                                "subtitle_number": subtitle_number,
                                "subtitle_time": subtitle_time,
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                            # 增加文本索引值
                            i = i + 1
                            j = j + 1
                            # 清空变量
                            source_text = ''
                            subtitle_number = ''
                            subtitle_time = ''

                        # 其他行是字幕文本，需要添加到文本中
                        else:
                            if  source_text:
                                source_text += '\n' + line
                            else:
                                source_text = line

        return json_data_list


    # 读取文件夹中树形结构Lrc音声文件
    def read_lrc_files (self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Lrc")
        json_data_list.append({
            "project_type": "Lrc",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1
        subtitle_title = ""

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".lrc"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    # 切行
                    lyrics = content.split('\n')
                    for line in lyrics:

                        # 使用正则表达式匹配标题标签行
                        title_pattern = re.compile(r'\[ti:(.*?)\]')
                        match = title_pattern.search(line)
                        if match:
                            subtitle_title =  match.group(1)  # 返回匹配到的标题全部内容


                        # 使用正则表达式匹配时间戳和歌词内容
                        pattern = re.compile(r'(\[([0-9:.]+)\])(.*)')
                        match = pattern.match(line)
                        if match:
                            timestamp = match.group(2)
                            source_text = match.group(3).strip()
                            if source_text == "":
                                continue
                            storage_path = os.path.relpath(file_path, folder_path)
                            file_name = file

                            if subtitle_title:                             
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "model": "none",
                                    "subtitle_time": timestamp,
                                    "subtitle_title":subtitle_title,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })
                                subtitle_title = ""

                            else:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "model": "none",
                                    "subtitle_time": timestamp,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })

                            # 增加文本索引值
                            i += 1

        return json_data_list


    # 读取文件夹中树形结构Txt小说文件
    def read_txt_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Txt")
        json_data_list.append({
            "project_type": "Txt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".txt"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    storage_path = os.path.relpath(file_path, folder_path)
                    file_name = file

                    # 切行
                    lines = content.split('\n')


                    for j, line in enumerate(lines):
                        if line.strip() == '': # 跳过空行
                            continue
                        spaces = len(line) - len(line.lstrip()) # 获取行开头的空格数

                        if j < len(lines) - 1 and lines[j + 1].strip() == '': # 检查当前行是否是文本中的最后一行,并检测下一行是否为空行
                            if (j+1) < len(lines) - 1 and lines[j + 2].strip() == '': # 再检查下下行是否为空行，所以最多只会保留2行空行信息
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": line,
                                    "translated_text": line,
                                    "model": "none",
                                    "sentence_indent": spaces,
                                    "line_break":2,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })
                            else:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": line,
                                    "translated_text": line,
                                    "model": "none",
                                    "sentence_indent": spaces,
                                    "line_break":1,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })

                        else:
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": line,
                                "translated_text": line,
                                "model": "none",
                                "sentence_indent": spaces,
                                "line_break":0,
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                        i += 1


        return json_data_list


    # 读取文件夹中树形结构Epub文件
    def read_epub_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Epub")
        json_data_list.append({
            "project_type": "Epub",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 epub 文件
                if file.endswith(".epub"):

                    file_path = os.path.join(root, file)  # 构建文件路径

                    # 构建解压文件夹路径
                    parent_path = os.path.dirname(file_path)
                    extract_path = os.path.join(parent_path, 'EpubCache')

                    # 创建解压文件夹
                    if not os.path.exists(extract_path):
                        os.makedirs(extract_path)

                    # 使用zipfile模块打开并解压EPUB文件
                    with zipfile.ZipFile(file_path, 'r') as epub_file:
                        # 提取所有文件
                        epub_file.extractall(extract_path)

                    # 加载EPUB文件
                    book = epub.read_epub(file_path)

                    # 获取文件路径和文件名
                    storage_path = os.path.relpath(file_path, folder_path)
                    book_name = file

                    # 遍历书籍中的所有内容
                    for item in book.get_items():
                        # 检查是否是文本内容
                        if item.get_type() == ebooklib.ITEM_DOCUMENT:


                            # 获取文件的唯一ID及文件名
                            item_id = item.get_id()
                            file_name = os.path.basename(item.get_name())

                            # 遍历文件夹中的所有文件,找到该文件，因为上面给的相对路径与epub解压后路径是不准的
                            for root_extract, dirs_extract, files_extract in os.walk(extract_path):
                                for filename in files_extract:
                                    # 如果文件名匹配
                                    if filename == file_name:
                                        # 构建完整的文件路径
                                        the_file_path = os.path.join(root_extract, filename)

                            # 打开对应HTML文件
                            with open(the_file_path, 'r', encoding='utf-8') as file:
                                # 读取文件内容
                                html_content = file.read()


                            # 获取文本内容并解码（为什么不用这个而进行解压操作呢，因为这个会自动将成对标签改为自适应标签）
                            #html_content = item.get_content().decode('utf-8')

                            # 正则表达式匹配<p>标签及其内容，包括自闭和的<p/>标签
                            p_pattern = r'<p[^>/]*>(.*?)</p>|<p[^>/]*/>'

                            # 使用findall函数找到所有匹配的内容
                            paragraphs = re.findall(p_pattern, html_content, re.DOTALL)

                            # 过滤掉空的内容
                            filtered_matches = [match for match in paragraphs if match.strip()]

                            # 遍历每个p标签，并提取文本内容
                            for p in filtered_matches:
                                # 保留原html内容文本
                                cleaned_text = p

                                # 提取纯文本
                                p_html = "<p>"+ p + "</p>"
                                soup = BeautifulSoup(p_html, 'html.parser')
                                text_content = soup.get_text()

                                # 去除前面的空格
                                text_content = text_content.lstrip()
                                cleaned_text = cleaned_text.lstrip() 

                                # 检查一下是否提取到空文本内容
                                if not text_content.strip():
                                    continue

                                # 获取项目的唯一ID
                                item_id = item.get_id()

                                # 录入缓存
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": text_content,
                                    "translated_text": text_content,
                                    "html":cleaned_text,
                                    "model": "none",
                                    "item_id": item_id,
                                    "storage_path": storage_path,
                                    "file_name": book_name,
                                })                                    
                                # 增加文本索引值
                                i = i + 1

                    # 删除文件夹
                    shutil.rmtree(extract_path)

        return json_data_list


    # 根据文件类型读取文件
    def read_files (self,translation_project,Input_Folder):

        if translation_project == "Mtool导出文件":
            cache_list = File_Reader.read_mtool_files(self,folder_path = Input_Folder)
        elif translation_project == "T++导出文件":
            cache_list = File_Reader.read_xlsx_files (self,folder_path = Input_Folder)
        elif translation_project == "VNText导出文件":
            cache_list = File_Reader.read_vnt_files(self,folder_path = Input_Folder)
        elif translation_project == "Srt字幕文件":
            cache_list = File_Reader.read_srt_files(self,folder_path = Input_Folder)
        elif translation_project == "Lrc音声文件":
            cache_list = File_Reader.read_lrc_files(self,folder_path = Input_Folder)
        elif translation_project == "Txt小说文件":
            cache_list = File_Reader.read_txt_files(self,folder_path = Input_Folder)
        elif translation_project == "Epub小说文件":
            cache_list = File_Reader.read_epub_files(self,folder_path = Input_Folder)
        elif translation_project == "Ainiee缓存文件":
            cache_list = File_Reader.read_cache_files(self,folder_path = Input_Folder)
        if translation_project == "ParaTranz导出文件":
            cache_list = File_Reader.read_paratranz_files(self,folder_path = Input_Folder)
        return cache_list



# 缓存管理器
class Cache_Manager():
    """
    缓存数据以列表来存储，分文件头和文本单元，文件头数据结构如下:
    1.项目类型： "project_type"
    2.项目ID： "project_id"

    文本单元的部分数据结构如下:
    1.翻译状态： "translation_status"   未翻译状态为0，已翻译为1，正在翻译为2，不需要翻译为7
    2.文本归类： "text_classification"
    3.文本索引： "text_index"
    4.名字： "name"
    5.原文： "source_text"
    6.译文： "translated_text"
    7.存储路径： "storage_path"
    8.存储文件名： "storage_file_name"
    9.行索引： "line_index"
    等等

    """
    def __init__(self):
        pass

    # 忽视空值内容和将整数型，浮点型数字变换为字符型数字函数，且改变翻译状态为7,因为T++读取到整数型数字时，会报错，明明是自己导出来的...
    def convert_source_text_to_str(self,cache_list):
        for entry in cache_list:
            storage_path = entry.get('storage_path')

            if storage_path:
                source_text = entry.get('source_text')

                if isinstance(source_text, (int, float)):
                    entry['source_text'] = str(source_text)
                    entry['translation_status'] = 7

                if source_text == "":
                    # 注意一下，文件头没有原文，所以会添加新的键值对到文件头里
                    entry['translation_status'] = 7
                
                if source_text == None:
                    entry['translation_status'] = 7


                if isinstance(source_text, str) and source_text.isdigit():
                    entry['translation_status'] = 7


                if isinstance(source_text, str) and Cache_Manager.is_punctuation_string(self,source_text):
                    entry['translation_status'] = 7

    # 忽视部分纯代码文本，且改变翻译状态为7
    def ignore_code_text(self,cache_list):
        for entry in cache_list:
            source_text = entry.get('source_text')

            #加个检测后缀为MP3，wav，png，这些文件名的文本，都是纯代码文本，所以忽略掉
            if source_text:
                if source_text.endswith('.mp3') or source_text.endswith('.wav') or source_text.endswith('.png') or source_text.endswith('.jpg'):
                    entry['translation_status'] = 7

            
            # 检查文本是否为空
            if source_text:
                # 正则表达式匹配<sg ?: ?>>格式的文本
                pattern = r'<SG[^>]*>'
                matches = re.findall(pattern, source_text)

                # 检查是否有匹配项
                if matches:
                    entry['translation_status'] = 7
                    for match in matches:
                        # 查找冒号的位置
                        colon_index = match.find(':')
                        if colon_index != -1: # 如果文本中存在冒号
                            # 分割冒号左边的内容和冒号右边直到>的内容
                            left = match[:colon_index].split('<SG')[-1].strip()
                            right = match[colon_index+1:].split('>')[0].strip()
                            # 检查右边字符量是否比左边字符量大N倍
                            if len(right) > len(left) * 15:
                                entry['translation_status'] = 0

    # 处理缓存数据的非中日韩字符，且改变翻译状态为7
    def process_dictionary_list(self,cache_list):
        pattern = re.compile(r'[\u4e00-\u9fff\u3040-\u30ff\u1100-\u11ff\u3130-\u318f\uac00-\ud7af]+')

        def contains_cjk(text):
            return bool(pattern.search(text))

        for entry in cache_list:
            source_text = entry.get('source_text')

            if source_text and not contains_cjk(source_text):
                entry['translation_status'] = 7

    # 检查字符串是否只包含常见的标点符号
    def is_punctuation_string(self,s: str) -> bool:
        """检查字符串是否只包含标点符号"""
        punctuation = set("!" '"' "#" "$" "%" "&" "'" "(" ")" "*" "+" "," "-" "." "/" "，" "。"  
                        ":" ";" "<" "=" ">" "?" "@" "[" "\\" "]" "^" "_" "`" "{" "|" "}" "~" "—")
        return all(char in punctuation for char in s)


    # 获取缓存数据中指定行数的翻译状态为0的未翻译文本，且改变翻译状态为2
    def process_dictionary_data_lines(self,translation_lines, cache_list, previous_lines = 0, following_lines = 0):
        # 输入的数据结构参考
        ex_cache_list = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]
        # 输出的数据结构参考
        ex_translation_list = [
            {'text_index': 4, 'source_text': 'しこトラ！',"name": "xxxxx"},
            {'text_index': 5, 'source_text': '11111'},
            {'text_index': 6, 'source_text': 'しこトラ！'},
        ]
        # 输出的数据结构参考
        ex_previous_list = [
            'しこトラ！'
            '11111',
            'しこトラ！',
        ]
        # 输出的数据结构参考
        ex_following_list = [
            'しこトラ！'
            '11111',
            'しこトラ！',
        ]

        translation_list = []
        previous_list = []
        following_list = []


        # 获取特定行数的待翻译内容
        for entry in cache_list:
            translation_status = entry.get('translation_status')

            # 如果能够获得到翻译状态的键，且翻译状态为0，即未翻译状态，那么将该行数据加入到新列表中
            if translation_status == 0:
                source_text = entry.get('source_text')
                text_index = entry.get('text_index')

                # 判断是否为None
                if source_text is not None and text_index is not None:

                    # 尝试获取name
                    name = entry.get('name')
                    # 如果有名字，则将名字加入到新列表中，否则不加入
                    if name:
                        translation_list.append({ 'text_index': text_index ,'source_text': source_text, 'name': name})
                    else:
                        translation_list.append({ 'text_index': text_index ,'source_text': source_text})

                    entry['translation_status'] = 2

                # 如果新列表中的元素个数达到指定行数，则停止遍历
                if len(translation_list) == translation_lines:
                    break

        # 获取首尾索引
        if translation_list:
            star_index = translation_list[0]['text_index'] - 1  # 减1以获取前一行
            end_index = translation_list[-1]['text_index']

            # 获取前n行原文
            if star_index != 0: # 因为第一个元素是消息表头
                base_storage_path = cache_list[star_index]['storage_path']

            for i in range(previous_lines):
                the_index = star_index - i
                if the_index >= 1 and the_index < len(cache_list): # 确保不超出列表范围
                    translation_status = cache_list[the_index]['translation_status']
                    storage_path = cache_list[the_index]['storage_path']
                    if storage_path ==base_storage_path: # 确保是同一文件里内容
                        if translation_status == 1 :# 优先获取已经翻译的文本
                            previous_list.append(cache_list[the_index]['translated_text'])
                        elif translation_status == 7 : # 如果是不需要翻译的文本
                            pass
                        else:
                            previous_list.append(cache_list[the_index]['source_text'])

            # 倒序排列元素,以免上文文本顺序错误
            if  previous_list:
                previous_list.reverse()

            # 获取后n行原文
            base_storage_path = cache_list[end_index]['storage_path']

            for i in range(following_lines):
                the_index = end_index + i
                if the_index >= 1 and the_index < len(cache_list):
                    translation_status = cache_list[the_index]['translation_status']
                    storage_path = cache_list[the_index]['storage_path']
                    if storage_path ==base_storage_path: # 确保是同一文件里内容
                        if translation_status == 1 :# 优先获取已经翻译的文本
                            following_list.append(cache_list[the_index]['translated_text'])
                        elif translation_status == 7 : # 如果是不需要翻译的文本
                            pass
                        else:
                            following_list.append(cache_list[the_index]['source_text'])

        return translation_list, previous_list


    # 获取缓存数据中指定tokens数的翻译状态为0的未翻译文本，且改变翻译状态为2
    def process_dictionary_data_tokens(self,tokens_limit, cache_list, previous_lines = 0):


        translation_list = []
        previous_list = []
        tokens_count = 0 
        lines_count = 0  # 保底机制用

        # 获取特定行数的待翻译内容
        for entry in cache_list:
            translation_status = entry.get('translation_status')

            # 如果能够获得到翻译状态的键，且翻译状态为0，即未翻译状态，那么将该行数据加入到新列表中
            if translation_status == 0:
                source_text = entry.get('source_text')
                text_index = entry.get('text_index')

                # 判断是否为None
                if source_text is not None and text_index is not None:

                    # 计算原文tokens
                    tokens = Request_Limiter.num_tokens_from_string(self,source_text)
                    # 检查是否超出tokens限制
                    tokens_count = tokens_count + tokens
                    if (tokens_count >= tokens_limit) and (lines_count >= 1) :
                        break
                    
                    # 尝试获取name
                    name = entry.get('name')
                    # 如果有名字，则将名字加入到新列表中，否则不加入
                    if name:
                        translation_list.append({ 'text_index': text_index ,'source_text': source_text, 'name': name})
                    else:
                        translation_list.append({ 'text_index': text_index ,'source_text': source_text})


                    entry['translation_status'] = 2
                    lines_count = lines_count + 1



        # 获取首尾索引
        if translation_list:
            star_index = translation_list[0]['text_index'] - 1  # 减1以获取前一行 

            # 获取前n行原文

            # 获取前n行原文
            if star_index != 0: # 因为第一个元素是消息表头
                base_storage_path = cache_list[star_index]['storage_path']

            for i in range(previous_lines):
                the_index = star_index - i
                if the_index >= 1 and the_index < len(cache_list): # 确保不超出列表范围
                    translation_status = cache_list[the_index]['translation_status']
                    storage_path = cache_list[the_index]['storage_path']
                    if storage_path ==base_storage_path: # 确保是同一文件里内容
                        if translation_status == 1 :# 优先获取已经翻译的文本
                            previous_list.append(cache_list[the_index]['translated_text'])
                        elif translation_status == 7 : # 如果是不需要翻译的文本
                            pass
                        else:
                            previous_list.append(cache_list[the_index]['source_text'])
            # 倒序排列元素
            if  previous_list:
                previous_list.reverse()


        return translation_list, previous_list

    # 将未翻译的文本列表，转换成待发送的原文字典,并计算文本实际行数，因为最后一些文本可能达到不了每次翻译行数
    def create_dictionary_from_list(self,data_list):
        #输入示例
        ex_list = [
            {'text_index': 4, 'source_text': 'しこトラ！',"name": "xxxxx"},
            {'text_index': 5, 'source_text': '11111'},
            {'text_index': 6, 'source_text': 'しこトラ！'},
        ]

        # 输出示例
        ex_dict = {
        '0': '测试！',
        '1': '测试1122211',
        '2': '测试xxxx！',
        }

        new_dict = {}
        index_count = 0

        for index, entry in enumerate(data_list):
            source_text = entry.get('source_text')
            name = entry.get('name')

            # 如果有名字，则组合成轻小说的格式，如：小明「测试」， 否则不组合
            if name: # 注意：改成二级处理时，要记得提示字典只会判断原文，不会判断名字
                if source_text[0] == '「':
                    new_dict[str(index_count)] = f"{name}{source_text}"
                else:
                    new_dict[str(index_count)] = f"{name}「{source_text}」"
            else:
                new_dict[str(index_count)] = source_text

            #索引计数
            index_count += 1

        return new_dict, len(data_list)


    # 提取输入文本头尾的非文本字符
    def trim_string(self,input_string):

        # 存储截取的头尾字符
        head_chars = []
        tail_chars = []
        middle_chars = input_string

        # 定义中日文及常用标点符号的正则表达式
        zh_pattern = re.compile(r'[\u4e00-\u9fff]+')
        jp_pattern = re.compile(r'[\u3040-\u309f\u30a0-\u30ff\u3400-\u4dbf]+')
        punctuation = re.compile(r'[。．，、；：？！「」『』【】〔〕（）《》〈〉…—…]+')

        # 从头开始检查并截取
        while middle_chars and not (zh_pattern.match(middle_chars) or jp_pattern.match(middle_chars) or punctuation.match(middle_chars)):
            head_chars.append(middle_chars[0])
            middle_chars = middle_chars[1:]
        
        # 从尾开始检查并截取
        while middle_chars and not (zh_pattern.match(middle_chars[-1]) or jp_pattern.match(middle_chars[-1]) or punctuation.match(middle_chars[-1])):
            tail_chars.insert(0, middle_chars[-1])  # 保持原始顺序插入
            middle_chars = middle_chars[:-1]
        
        return head_chars, middle_chars, tail_chars

    # 处理字典内的文本，清除头尾的非文本字符
    def process_dictionary(self,input_dict):
        # 输入字典示例
        ex_dict1 = {
        '0': '\if(s[114])en(s[115])ハイヒーリング（消費MP5）',
        '1': '\F[21]\FF[128]ゲオルグ「なにを言う。　僕はおまえを助けに来たんだ。\FF[128]',
        '2': '少年「ダメか。僕も血が止まらないな……。',
        }

        #输出信息处理记录列表示例
        ex_list1 = [
            {'text_index': '0', "Head:": '\if(s[114])en(s[115])',"Middle": "ハイヒーリング（消費MP5）"},
            {'text_index': '1', "Head:": '\F[21]\FF[128]',"Middle": "ゲオルグ「なにを言う。　僕はおまえを助けに来たんだ。", "Tail:": '\FF[128]'},
            {'text_index': '2', "Middle": "少年「ダメか。僕も血が止まらないな……。"},
        ]

        # 输出字典示例
        ex_dict2 = {
        '0': 'ハイヒーリング（消費MP5）',
        '1': 'ゲオルグ「なにを言う。　僕はおまえを助けに来たんだ。',
        '2': '少年「ダメか。僕も血が止まらないな……。',
        }

        processed_dict = {}
        process_info_list = []
        for key, value in input_dict.items():
            head, middle, tail = Cache_Manager.trim_string(self,value)


            # 这里针对首尾非文本字符是纯数字字符或者类似“R5”组合时,还原回去,否则会影响翻译

            if head: # 避免空列表与列表类型导致错误
                head_str = ''.join(head)
            else:
                head_str = ""

            if head_str.isdigit() or bool(re.match(r'^[a-zA-Z][0-9]{1,2}$', head_str)):
                middle = head_str + middle
                head = []


            if tail: # 避免空列表与列表类型导致错误
                tail_str = ''.join(tail)
            else:
                tail_str = ""

            if tail_str.isdigit() or bool(re.match(r'^[a-zA-Z][0-9]{1,2}$', tail_str)):
                middle = middle + tail_str
                tail = []


            # 将处理后的结果存储在两个不同的结构中：一个字典和一个列表。
            processed_dict[key] = middle
            info = {"text_index": key}
            if head:
                info["Head:"] = ''.join(head)
            if middle:
                info["Middle:"] = middle
            if tail:
                info["Tail:"] = ''.join(tail)
            process_info_list.append(info)
        return processed_dict, process_info_list

    # 复原文本中的头尾的非文本字符
    def update_dictionary(self,original_dict, process_info_list):
        # 输入字典示例
        ex_dict1 = {
        '0': '测试1',
        '1': '测试2',
        '2': '测试3',
        }
        #输入信息处理记录列表示例
        ex_list1 = [
            {'text_index': '0', "Head:": '\if(s[114])en(s[115])',"Middle": "ハイヒーリング（消費MP5）"},
            {'text_index': '1', "Head:": '\F[21]\FF[128]',"Middle": "ゲオルグ「なにを言う。　僕はおまえを助けに来たんだ。", "Head:": '\FF[128]'},
            {'text_index': '2', "Middle": "少年「ダメか。僕も血が止まらないな……。"},
        ]
        # 输出字典示例
        ex_dict2 = {
        '0': '\if(s[114])en(s[115])测试1',
        '1': '\F[21]\FF[128]测试2\FF[128]',
        '2': '测试3',
        }

        updated_dict = {}
        for item in process_info_list:
            text_index = item["text_index"]
            head = item.get("Head:", "")
            tail = item.get("Tail:", "")
            
            # 获取原始字典中的值
            original_value = original_dict.get(text_index, "")
            
            # 拼接头、原始值、中和尾
            updated_value = head + original_value + tail
            updated_dict[text_index] = updated_value
        return updated_dict


    # 将翻译结果录入缓存函数，且改变翻译状态为1
    def update_cache_data(self, cache_data, source_text_list, response_dict,translation_model):
        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]


        ex_source_text_list = [
            {'text_index': 4, 'source_text': 'しこトラ！','name': 'xxxxx'},
            {'text_index': 5, 'source_text': '11111'},
            {'text_index': 6, 'source_text': 'しこトラ！'},
        ]


        ex_response_dict = {
        '0': '测试！',
        '1': '测试1122211',
        '2': '测试xxxx！',
        }


        # 回复文本的索引
        index = 0

        # 遍历原文本列表
        try:
            for source_text_item in source_text_list:
                # 获取缓存文本中索引值
                text_index = source_text_item['text_index']
                # 根据回复文本的索引值，在回复内容中获取已翻译的文本
                response_value = response_dict[str(index)]
                # 获取人名（如果有）
                name = source_text_item.get('name')

                # 缓存文本中索引值，基本上是缓存文件里元素的位置索引值，所以直接获取并修改
                if name:
                    # 提取人名以及对话文本
                    name_text,text = Cache_Manager.extract_strings(self,response_value)
                    # 如果能够正确提取到人名以及翻译文本
                    if name_text:
                        cache_data[text_index]['translation_status'] = 1
                        cache_data[text_index]['name'] = name_text
                        cache_data[text_index]['translated_text'] = text
                        cache_data[text_index]['model'] = translation_model
                    else:
                        cache_data[text_index]['translation_status'] = 1
                        cache_data[text_index]['translated_text'] = response_value
                        cache_data[text_index]['model'] = translation_model

                else:
                    cache_data[text_index]['translation_status'] = 1
                    cache_data[text_index]['translated_text'] = response_value
                    cache_data[text_index]['model'] = translation_model

                # 增加索引值
                index = index + 1
        
        except:
            print("[DEBUG] 录入翻译结果出现问题！！！！！！！！！！！！")
            print(text_index)


        return cache_data

    
    # 统计翻译状态等于0或者2的元素个数，且把等于2的翻译状态改为0.并返回元素个数
    def count_and_update_translation_status_0_2(self, data):
        # 输入的数据结构参考
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无'},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '无'},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 5, 'text_classification': 0, 'translation_status': 2, 'source_text': '11111', 'translated_text': '无'},
            {'text_index': 6, 'text_classification': 0, 'translation_status': 0, 'source_text': '11111', 'translated_text': '无'},
        ]

        # 计算翻译状态为0或2的条目数量与tokens
        tokens_count = 0
        raw_count = 0
        for item in data:
            if (item.get('translation_status') == 0) or (item.get('translation_status') == 2):
                raw_count += 1

                source_text = item.get('source_text')
                if  source_text:
                    tokens = Request_Limiter.num_tokens_from_string(self,source_text)
                    # 检查是否超出tokens限制
                    tokens_count = tokens_count + tokens

        # 将'translation_status'等于2的元素的'translation_status'改为0
        for item in data:
            if item.get('translation_status') == 2:
                item['translation_status'] = 0

        return raw_count,tokens_count
    
    
    # 替换或者还原换行符和回车符函数
    def replace_special_characters(self,dict, mode):
        new_dict = {}
        if mode == "替换":
            for key, value in dict.items():
                #如果value是字符串变量
                if isinstance(value, str):
                    new_value = value.replace("\n", "＠").replace("\r", "∞")
                    new_dict[key] = new_value
        elif mode == "还原":
            for key, value in dict.items():
                #如果value是字符串变量
                if isinstance(value, str):
                    # 先替换半角符号
                    new_value = value.replace("@", "\n")
                    # 再替换全角符号
                    new_value = new_value.replace("＠", "\n").replace("∞", "\r")
                    new_dict[key] = new_value
        else:
            print("请输入正确的mode参数（替换或还原）")

        return new_dict


    # 轻小说格式提取人名与文本
    def extract_strings(self, text):
        # 查找第一个左括号的位置
        left_bracket_pos = text.find('「')
        
        # 如果找不到左括号，返回原始文本
        if left_bracket_pos == -1:
            return 0,text
        
        # 提取名字和对话内容
        name = text[:left_bracket_pos].strip()
        dialogue = text[left_bracket_pos:].strip()
        
        return name, dialogue



# 文件输出器
class File_Outputter():
    def __init__(self):
        pass

    # 将缓存文件里已翻译的文本转换为简体字或繁体字
    def simplified_and_traditional_conversion( self,cache_list, target_language):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 1, 'source_text': 'しこトラ！', 'translated_text': '谢谢', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '開心', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111', 'translated_text': '歷史', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]    

        # 确定使用的转换器
        if target_language == "简中": 
            cc = opencc.OpenCC('t2s')  # 创建OpenCC对象，使用t2s参数表示繁体字转简体字
        elif target_language == "繁中": 
            cc = opencc.OpenCC('s2t')

        # 存储结果的列表
        converted_list = []

        # 遍历缓存数据
        for item in cache_list:
            translation_status = item.get('translation_status', 0)
            translated_text = item.get('translated_text', '')

            # 如果'translation_status'为1，进行转换
            if translation_status == 1:
                converted_text = cc.convert(translated_text)
                item_copy = item.copy()  # 防止修改原始数据
                item_copy['translated_text'] = converted_text
                converted_list.append(item_copy)
            else:
                converted_list.append(item)

        return converted_list
    

    # 输出json文件
    def output_json_file(self,cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 中间存储字典格式示例
        ex_path_dict = {
            "D:\\DEBUG Folder\\Replace the original text.json": {'translation_status': 1, 'Source Text': 'しこトラ！', 'Translated Text': 'しこトラ！'},
            "D:\\DEBUG Folder\\DEBUG Folder\\Replace the original text.json": {'translation_status': 0, 'Source Text': 'しこトラ！', 'Translated Text': 'しこトラ！'}
        }


        # 输出文件格式示例
        ex_output ={
        'しこトラ！': 'xxxx',
        '室内カメラ': 'yyyyy',
        '111111111': '无3',
        '222222222': '无4',
        }

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 
                


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:

                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text']}
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text']}
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = {}

            #存储未翻译的文本
            output_file2 = {}

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 如果这个本已经翻译了，存放对应的文件中
                if content['translation_status'] == 1:
                    output_file[content['source_text']] = content['translated_text']
                # 如果这个文本没有翻译或者正在翻译
                elif content['translation_status'] == 0 or content['translation_status'] == 2:
                    output_file2[content['source_text']] = content['source_text']


            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出paratranz文件
    def output_paratranz_file(self, cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json',
             'key': 'txtKey', 'context': ''},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json',
             'key': 'txtKey', 'context': ''},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111',
             'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json',
             'file_name': 'Replace the original text.json', 'key': 'txtKey', 'context': ''},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222',
             'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json',
             'file_name': 'Replace the original text.json', 'key': 'txtKey', 'context': ''},
        ]

        # 中间存储字典格式示例
        ex_path_dict = {
            "D:\\DEBUG Folder\\Replace the original text.json": {'translation_status': 1, 'Source Text': 'しこトラ！',
                                                                 'Translated Text': 'しこトラ！'},
            "D:\\DEBUG Folder\\DEBUG Folder\\Replace the original text.json": {'translation_status': 0,
                                                                               'Source Text': 'しこトラ！',
                                                                               'Translated Text': 'しこトラ！'}
        }

        # 输出文件格式示例
        ex_output = [
        {
            "key": "Activate",
            "original": "カードをプレイ",
            "translation": "出牌",
            "context": ""
        }]

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'

                # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:
                text = {'translation_status': item['translation_status'], 'source_text': item['source_text'],
                        'translated_text': item['translated_text'], 'context': item['context'], 'key': item['key']}
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'], 'source_text': item['source_text'],
                        'translated_text': item['translated_text'], 'context': item['context'], 'key': item['key']}
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)

            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)

            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = []

            # 存储未翻译的文本
            output_file2 = []

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                item = {
                    "key": content.get("key", ""),  # 假设每个 content 字典都有 'key' 字段
                    "original": content['source_text'],
                    "translation": content.get('translated_text', ""),
                    "context": content.get('context', "")  # 如果你有 'context' 字段，也包括它
                }
                # 根据翻译状态，选择存储到已翻译或未翻译的列表
                if content['translation_status'] == 1:
                    output_file.append(item)
                elif content['translation_status'] == 0 or content['translation_status'] == 2:
                    output_file2.append(item)

            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出vnt文件
    def output_vnt_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output =  [
            {
                "name": "玲",
                "message": "「……おはよう」"
            },
            {
                "message": "　心の内では、ムシャクシャした気持ちは未だに鎮まっていなかった。"
            }
            ]

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 
                


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:
                if'name' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text'],
                            'name': item['name']}

                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text']}
                    
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                if'name' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text'],
                            'name': item['name']}

                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text']}
                    
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = []

            #存储未翻译的文本
            output_file2 = []

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 如果这个本已经翻译了，存放对应的文件中
                if'name' in content:
                    text = {'name': content['name'],
                            'message': content['translated_text'],}
                else:
                    text = {'message': content['translated_text'],}
                
                output_file.append(text)

                # 如果这个文本没有翻译或者正在翻译
                if content['translation_status'] == 0 or content['translation_status'] == 2:
                    if'name' in content:
                        text = {'name': content['name'],
                                'message': content['translated_text'],}
                    else:
                        text = {'message': content['translated_text'],}
                    
                    output_file2.append(text)


            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)


    # 输出表格文件
    def output_excel_file(self,cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'T++'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 1, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 2},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 3},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '草草草草', 'translated_text': '11111', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 3},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 1, 'source_text': '室内カメラ', 'translated_text': '22222', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 4},
        ]

        # 创建一个字典，用于存储翻译数据
        translations_by_path = {}

        # 遍历缓存数据
        for item in cache_data:
            if 'storage_path' in item:
                path = item['storage_path']

                # 如果路径不存在，创建文件夹
                folder_path = os.path.join(output_path, os.path.dirname(path))
                os.makedirs(folder_path, exist_ok=True)

                # 提取信息
                source_text = item.get('source_text', '')
                translated_text = item.get('translated_text', '')
                row_index = item.get('row_index', '')
                translation_status = item.get('translation_status', '')

                # 构造字典
                translation_dict = {'translation_status': translation_status,'Source Text': source_text, 'Translated Text': translated_text,"row_index": row_index}

                # 将字典添加到对应路径的列表中
                if path in translations_by_path:
                    translations_by_path[path].append(translation_dict)
                else:
                    translations_by_path[path] = [translation_dict]

        # 遍历字典，将数据写入 Excel 文件
        for path, translations_list in translations_by_path.items():
            file_path = os.path.join(output_path, path)

            # 创建一个工作簿
            wb = Workbook()

            # 选择默认的活动工作表
            ws = wb.active

            # 添加表头
            ws.append(['Original Text', 'Initial', 'Machine translation', 'Better translation', 'Best translation'])



            # 将数据写入工作表
            for translation_dict in translations_list:
                row_index = translation_dict['row_index']
                translation_status = translation_dict['translation_status']
                
                # 如果是已经翻译文本，则写入原文与译文
                if translation_status == 1 :
                    ws.cell(row=row_index, column=1).value = translation_dict['Source Text']
                    ws.cell(row=row_index, column=2).value = translation_dict['Translated Text']

                # 如果是未翻译或不需要翻译文本，则写入原文
                else:
                    ws.cell(row=row_index, column=1).value = translation_dict['Source Text']

            # 保存工作簿
            wb.save(file_path)


    # 输出缓存文件
    def output_cache_file(self,cache_data,output_path):
        # 复制缓存数据到新变量
        try:
            modified_cache_data = copy.deepcopy(cache_data)
        except:
            print("[INFO]: 无法正常进行深层复制,改为浅复制")
            modified_cache_data = cache_data.copy()

        # 修改新变量的元素中的'translation_status'
        for item in modified_cache_data:
            if 'translation_status' in item and item['translation_status'] == 2:
                item['translation_status'] = 0

        # 输出为JSON文件
        with open(os.path.join(output_path, "AinieeCacheData.json"), "w", encoding="utf-8") as f:
            json.dump(modified_cache_data, f, ensure_ascii=False, indent=4)


    # 输出srt文件
    def output_srt_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output ="""
        1
        00:00:16,733 --> 00:00:19,733
        Does that feel good, Tetchan?

        2
        00:00:25,966 --> 00:00:32,500
        Just a little more... I'm really close too... Ahhh, I can't...!
        """

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:

                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text'],'subtitle_number': item['subtitle_number'],'subtitle_time': item['subtitle_time']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text'],'subtitle_number':  item['subtitle_number'],'subtitle_time': item['subtitle_time']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".srt"):
                file_name_translated = old_filename.replace(".srt", "") + "_translated.srt"
            else:
                file_name_translated = old_filename + "_translated.srt"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""
            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取字幕序号
                subtitle_number = content['subtitle_number']
                # 获取字幕时间轴
                subtitle_time = content['subtitle_time']
                # 获取字幕文本内容
                subtitle_text = content['translated_text']

                output_file += f'{subtitle_number}\n{subtitle_time}\n{subtitle_text}\n\n'



            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)


    # 输出lrc文件
    def output_lrc_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output ="""
        [ti:1.したっぱ童貞構成員へハニートラップ【手コキ】 (Transcribed on 15-May-2023 19-10-13)]
        [00:00.00]お疲れ様です大長 ただいま機会いたしました
        [00:06.78]法案特殊情報部隊一番対処得フィルレイやセルドツナイカーです 今回例の犯罪組織への潜入が成功しましたのでご報告させていただきます
        """

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                if 'subtitle_title' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time'],
                            'subtitle_title': item['subtitle_title']}
                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                if 'subtitle_title' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time'],
                            'subtitle_title': item['subtitle_title']}
                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'], 
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".lrc"):
                file_name_translated = old_filename.replace(".lrc", "") + "_translated.lrc"
            else:
                file_name_translated = old_filename + "_translated.lrc"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""
            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取字幕时间轴
                subtitle_time = content['subtitle_time']
                # 获取字幕文本内容
                subtitle_text = content['translated_text']

                if 'subtitle_title' in content:
                    subtitle_title = content['subtitle_title']
                    output_file += f'[{subtitle_title}]\n[{subtitle_time}]{subtitle_text}\n'
                else:
                    output_file += f'[{subtitle_time}]{subtitle_text}\n'



            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)


    # 输出txt文件
    def output_txt_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output ="""
        　测试1
        　今ではダンジョンは、人々の営みの一部としてそれなりに定着していた。

        ***

        「正気なの？」
        测试2
        """

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'], 
                        'translated_text': item['translated_text'],
                        "sentence_indent": item['sentence_indent'],
                        'line_break': item['line_break']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'], 
                        'translated_text': item['translated_text'],
                        "sentence_indent": item['sentence_indent'],
                        'line_break': item['line_break']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".txt"):
                file_name_translated = old_filename.replace(".txt", "") + "_translated.txt"
            else:
                file_name_translated = old_filename + "_translated.txt"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取记录的句首空格数
                expected_indent_count = content['sentence_indent']
                # 获取句尾换行符数
                line_break_count = content['line_break']
                
                # 删除句首的所有空格
                translated_text = content['translated_text'].lstrip()
                
                # 根据记录的空格数在句首补充空格
                sentence_indent = "　" * expected_indent_count
                
                line_break = "\n" * (line_break_count + 1)

                output_file += f'{sentence_indent}{translated_text}{line_break}'

            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)


    # 输出epub文件
    def output_epub_file(self,cache_data, output_path, input_path):

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}' 


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'], 
                        'translated_text': item['translated_text'],
                        'html': item['html'],
                        "item_id": item['item_id'],}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'], 
                        'translated_text': item['translated_text'],
                        'html': item['html'],
                        "item_id": item['item_id'],}
                text_dict[file_path] = [text]




        # 将输入路径里面的所有epub文件复制到输出路径
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        # 递归遍历输入路径中的所有文件和子目录
        for dirpath, dirnames, filenames in os.walk(input_path):
            for filename in filenames:
                # 检查文件扩展名是否为.epub
                if filename.endswith('.epub'):
                    # 构建源文件和目标文件的完整路径
                    src_file_path = os.path.join(dirpath, filename)
                    # 计算相对于输入路径的相对路径
                    relative_path = os.path.relpath(src_file_path, start=input_path)
                    # 构建目标文件的完整路径
                    dst_file_path = os.path.join(output_path, relative_path)

                    # 创建目标文件的目录（如果不存在）
                    os.makedirs(os.path.dirname(dst_file_path), exist_ok=True)
                    # 复制文件
                    shutil.copy2(src_file_path, dst_file_path)
                    #print(f'Copied: {src_file_path} -> {dst_file_path}')



        # 遍历 path_dict，并将内容写入对应文件中
        for file_path, content_list in text_dict.items():
    
            # 加载EPUB文件
            book = epub.read_epub(file_path)

            # 构建解压文件夹路径
            parent_path = os.path.dirname(file_path)
            extract_path = os.path.join(parent_path, 'EpubCache')

            # 创建解压文件夹
            if not os.path.exists(extract_path):
                os.makedirs(extract_path)

            # 使用zipfile模块打开并解压EPUB文件
            with zipfile.ZipFile(file_path, 'r') as epub_file:
                # 提取所有文件
                epub_file.extractall(extract_path)

            # 遍历书籍中的所有内容
            for item in book.get_items():
                # 检查是否是文本内容
                if item.get_type() == ebooklib.ITEM_DOCUMENT:

                    # 获取文件的唯一ID及文件名
                    item_id = item.get_id()
                    file_name = os.path.basename(item.get_name())

                    # 遍历文件夹中的所有文件,找到该文件，因为上面给的相对路径与epub解压后路径是不准的
                    for root, dirs, files in os.walk(extract_path):
                        for filename in files:
                            # 如果文件名匹配
                            if filename == file_name:
                                # 构建完整的文件路径
                                the_file_path = os.path.join(root, filename)

                    # 打开对应HTML文件
                    with open(the_file_path, 'r', encoding='utf-8') as file:
                        # 读取文件内容
                        content_html = file.read()

                    # 遍历缓存数据
                    for content in content_list:
                        # 如果找到匹配的文件id
                        if item_id == content['item_id']:
                            # 获取原文本
                            original = content['source_text']
                            # 获取翻译后的文本
                            replacement = content['translated_text']

                            # 获取html标签化的文本
                            html = content['html']
                            html = str(html)

                            # 删除 &#13;\n\t\t\t\t
                            html = html.replace("&#13;\n\t\t\t\t", "")

                            if"Others who have read our chapters and offered similar assistance have" in html:
                                print("ce")

                            # 有且只有一个a标签，则改变替换文本，以保留跳转功能
                            if (re.match( r'^(?:<a(?:\s[^>]*?)?>[^<]*?</a>)*$', html) is not None):
                                # 针对跳转标签的保留，使用正则表达式搜索<a>标签内的文本
                                a_tag_pattern = re.compile(r'<a[^>]*>(.*?)</a>')
                                matches = a_tag_pattern.findall(html)

                                if len(matches) == 1:
                                    html = matches[0]


                            # 如果原文与译文不为空，则替换原hrml文件中的文本
                            if (original and replacement):
                                # 替换第一个匹配项
                                content_html = content_html.replace(html, replacement, 1)
                                #content_html  = re.sub(original, replacement, content_html, count=1)


                    # 写入内容到HTML文件
                    with open(the_file_path, 'w', encoding='utf-8') as file:
                        file.write(content_html)
        
            # 构建修改后的EPUB文件路径
            modified_epub_file = file_path.rsplit('.', 1)[0] + '_translated.epub'

            # 创建ZipFile对象，准备写入压缩文件
            with zipfile.ZipFile(modified_epub_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 遍历文件夹中的所有文件和子文件夹
                for root, dirs, files in os.walk(extract_path):
                    for file in files:
                        # 获取文件的完整路径
                        full_file_path = os.path.join(root, file)
                        # 获取文件在压缩文件中的相对路径
                        relative_file_path = os.path.relpath(full_file_path, extract_path)
                        # 将文件添加到压缩文件中
                        zipf.write(full_file_path, relative_file_path)
                        
            # 删除旧文件
            os.remove(file_path)
            # 删除文件夹
            shutil.rmtree(extract_path)


    # 输出已经翻译文件
    def output_translated_content(self,cache_data,output_path,input_path):
        # 复制缓存数据到新变量
        try:
           new_cache_data = copy.deepcopy(cache_data)
        except:
            print("[INFO]: 无法正常进行深层复制,改为浅复制")
            new_cache_data = cache_data.copy()

        # 提取项目列表
        if new_cache_data[0]["project_type"] == "Mtool":
            File_Outputter.output_json_file(self,new_cache_data, output_path)
        elif new_cache_data[0]["project_type"] == "Srt":
            File_Outputter.output_srt_file(self,new_cache_data, output_path)
        elif new_cache_data[0]["project_type"] == "Lrc":
            File_Outputter.output_lrc_file(self,new_cache_data, output_path)
        elif new_cache_data[0]["project_type"] == "Vnt":
            File_Outputter.output_vnt_file(self,new_cache_data, output_path)
        elif new_cache_data[0]["project_type"] == "Txt":
            File_Outputter.output_txt_file(self,new_cache_data, output_path)
        elif new_cache_data[0]["project_type"] == "Epub":
            File_Outputter.output_epub_file(self,new_cache_data, output_path,input_path)
        elif new_cache_data[0]["project_type"] == "Paratranz":
            File_Outputter.output_paratranz_file(self,new_cache_data, output_path)
        else:
            File_Outputter.output_excel_file(self,new_cache_data, output_path)



# 后台任务分发器
class background_executor(threading.Thread): 
    def __init__(self, task_id,input_folder,output_folder,platform,base_url,model,api_key,proxy_port):
        super().__init__() # 调用父类构造
        self.task_id = task_id

        if input_folder :
            self.input_folder = input_folder
        else:
            self.input_folder = ""

        if output_folder :
            self.output_folder = output_folder
        else:
            self.output_folder = ""
        
        self.platform = platform
        self.base_url = base_url
        self.model = model
        self.api_key = api_key
        self.proxy_port = proxy_port

    def run(self):
        global Running_status
        global cache_list
        # 执行翻译
        if self.task_id == "执行翻译任务":
            #如果不是status'为9，说明翻译任务被暂停了，先不改变运行状态
            if Running_status != 9:
                Running_status = 6

            #执行翻译主函数
            Translator.Main(self)

            # 如果完成了翻译任务
            if Running_status == 6:
                Running_status = 0
            # 如果取消了翻译任务
            if Running_status == 10:
                user_interface_prompter.signal.emit("翻译状态提示","翻译取消",0)
                Running_status = 0
            # 如果暂停了翻译任务
            if Running_status == 9:
                user_interface_prompter.signal.emit("翻译状态提示","翻译暂停",0)

        # 执行接口测试
        elif self.task_id == "接口测试":

            Running_status = 1
            Request_Tester.request_test(self,self.platform,self.base_url,self.model,self.api_key,self.proxy_port)
            Running_status = 0

        # 输出缓存
        elif self.task_id == "输出缓存文件":
            File_Outputter.output_cache_file(self,cache_list,self.output_folder)
            print('\033[1;32mSuccess:\033[0m 已输出缓存文件到文件夹')

        # 输出已翻译文件
        elif self.task_id == "输出已翻译文件":
            File_Outputter.output_translated_content(self,cache_list,self.output_folder,self.input_folder)
            print('\033[1;32mSuccess:\033[0m 已输出已翻译文件到文件夹')


# ——————————————————————————————————————————下面都是UI相关代码——————————————————————————————————————————
        
class Widget_AI(QFrame):

    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.hBoxLayout = QHBoxLayout(self)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))


class Widget_Proxy(QFrame):  # 代理账号主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_Proxy_A('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_Proxy_B('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '代理设置')
        self.addSubInterface(self.B_settings, 'B_settings', '速率价格设置')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_Proxy_A(QFrame):#  代理账号基础设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_relay_address = QGroupBox()
        box_relay_address.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_relay_address = QHBoxLayout()

        #设置“中转地址”标签
        self.labelA = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelA.setText("中转请求地址")

        #设置微调距离用的空白标签
        self.labelB = QLabel()  
        self.labelB.setText("                ")

        #设置“中转地址”的输入框
        self.LineEdit_relay_address = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_relay_address.addWidget(self.labelA)
        layout_relay_address.addWidget(self.labelB)
        layout_relay_address.addWidget(self.LineEdit_relay_address)
        box_relay_address.setLayout(layout_relay_address)




        # -----创建第2个组，添加多个组件-----
        box_proxy_platform = QGroupBox()
        box_proxy_platform.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_platform = QHBoxLayout()

        #设置标签
        self.label_proxy_platform = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_platform.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_platform.setText("请求格式")


        #设置下拉选择框
        self.comboBox_proxy_platform = ComboBox() #以demo为父类
        self.comboBox_proxy_platform.addItems(['OpenAI', 'Anthropic'])
        self.comboBox_proxy_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_proxy_platform.setFixedSize(250, 35)
        
        # 连接下拉框的currentTextChanged信号到相应的槽函数
        self.comboBox_proxy_platform.currentTextChanged.connect(self.on_combobox_changed)


        layout_proxy_platform.addWidget(self.label_proxy_platform)
        layout_proxy_platform.addStretch(1)
        layout_proxy_platform.addWidget(self.comboBox_proxy_platform)
        box_proxy_platform.setLayout(layout_proxy_platform)




     # -----创建第3个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QHBoxLayout()

        #设置“模型选择”标签
        self.label_model = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_model.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_model.setText("模型选择(可编辑)")


        #设置“模型类型”下拉选择框
        self.comboBox_model_openai = EditableComboBox() #以demo为父类
        self.comboBox_model_openai.addItems(['gpt-3.5-turbo','gpt-3.5-turbo-0301','gpt-3.5-turbo-0613', 'gpt-3.5-turbo-1106', 'gpt-3.5-turbo-0125','gpt-3.5-turbo-16k', 'gpt-3.5-turbo-16k-0613',
                                 'gpt-4','gpt-4o','gpt-4-0314', 'gpt-4-0613','gpt-4-turbo','gpt-4-turbo-preview','gpt-4-1106-preview','gpt-4-0125-preview'])
        self.comboBox_model_openai.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model_openai.setFixedSize(250, 35)


        self.comboBox_model_anthropic = EditableComboBox() #以demo为父类
        self.comboBox_model_anthropic.addItems(['claude-2.0','claude-2.1','claude-3-haiku-20240307','claude-3-sonnet-20240229', 'claude-3-opus-20240229','claude-3-5-sonnet-20240620'])
        self.comboBox_model_anthropic.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model_anthropic.setFixedSize(250, 35)



        layout_model.addWidget(self.label_model)
        layout_model.addStretch(1)
        layout_model.addWidget(self.comboBox_model_openai)
        layout_model.addWidget(self.comboBox_model_anthropic)
        box_model.setLayout(layout_model)


        #设置默认显示的模型选择框，其余隐藏
        self.comboBox_model_openai.show()
        self.comboBox_model_anthropic.hide()


        # -----创建第3个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.label_apikey = QLabel(flags=Qt.WindowFlags())  
        self.label_apikey.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.label_apikey.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.label_apikey)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第4个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“系统代理端口”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.label_null = QLabel()  
        self.label_null.setText("                      ")

        #设置“系统代理端口”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.label_null)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第5个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_relay_address)
        container.addWidget(box_proxy_platform)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def on_combobox_changed(self, index):
        # 根据下拉框的索引调用不同的函数
        if index =="OpenAI":
            self.comboBox_model_openai.show()
            self.comboBox_model_anthropic.hide()
        elif index == 'Anthropic':
            self.comboBox_model_openai.hide()
            self.comboBox_model_anthropic.show()


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = self.LineEdit_relay_address.text()      # 获取代理地址
            Proxy_platform = self.comboBox_proxy_platform.currentText()  # 获取代理平台
            API_key_str = self.TextEdit_apikey.toPlainText()        # 获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            # 获取代理端口

            if Proxy_platform == "OpenAI":
                Model_Type = self.comboBox_model_openai.currentText()
            elif Proxy_platform == "Anthropic":
                Model_Type = self.comboBox_model_anthropic.currentText()

            #创建子线程
            thread = background_executor("接口测试","","",Proxy_platform,Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")


class Widget_Proxy_B(QFrame):#  代理账号进阶设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组(后面加的)，添加多个组件-----
        box_tokens = QGroupBox()
        box_tokens.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_tokens = QHBoxLayout()

        #设置标签
        self.label_tokens = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_tokens.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_tokens.setText("每次发送文本上限")

        #设置“说明”显示
        self.labelA_tokens = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA_tokens.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelA_tokens.setText("(tokens)")  

        #数值输入
        self.spinBox_tokens = SpinBox(self)
        self.spinBox_tokens.setRange(0, 2147483647)    
        self.spinBox_tokens.setValue(4000)


        layout_tokens.addWidget(self.label_tokens)
        layout_tokens.addWidget(self.labelA_tokens)
        layout_tokens.addStretch(1)  # 添加伸缩项
        layout_tokens.addWidget(self.spinBox_tokens)
        box_tokens.setLayout(layout_tokens)



        # -----创建第1个组(后面加的)，添加多个组件-----
        box_RPM = QGroupBox()
        box_RPM.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_RPM = QHBoxLayout()

        #设置“RPM”标签
        self.labelY = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelY.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelY.setText("每分钟请求数")

        #设置“说明”显示
        self.labelA = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelA.setText("(RPM)")  

        #数值输入
        self.spinBox_RPM = SpinBox(self)
        self.spinBox_RPM.setRange(0, 2147483647)    
        self.spinBox_RPM.setValue(3500)


        layout_RPM.addWidget(self.labelY)
        layout_RPM.addWidget(self.labelA)
        layout_RPM.addStretch(1)  # 添加伸缩项
        layout_RPM.addWidget(self.spinBox_RPM)
        box_RPM.setLayout(layout_RPM)



        # -----创建第2个组（后面加的），添加多个组件-----
        box_TPM = QGroupBox()
        box_TPM.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_TPM = QHBoxLayout()

        #设置“TPM”标签
        self.labelB = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelB.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelB.setText("每分钟tokens数")
    
        #设置“说明”显示
        self.labelC = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelC.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelC.setText("(TPM)") 

        #数值输入
        self.spinBox_TPM = SpinBox(self)
        self.spinBox_TPM.setRange(0, 2147483647)    
        self.spinBox_TPM.setValue(60000)


        layout_TPM.addWidget(self.labelB)
        layout_TPM.addWidget(self.labelC)
        layout_TPM.addStretch(1)  # 添加伸缩项
        layout_TPM.addWidget(self.spinBox_TPM)
        box_TPM.setLayout(layout_TPM)


        # -----创建第3个组（后面加的），添加多个组件-----
        box_input_pricing = QGroupBox()
        box_input_pricing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input_pricing = QHBoxLayout()

        #设置“请求输入价格”标签
        self.labelD = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelD.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelD.setText("请求输入价格")
    
        #设置“说明”显示
        self.labelE = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelE.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelE.setText("( /1K tokens)") 

        #数值输入
        self.spinBox_input_pricing = DoubleSpinBox(self)
        self.spinBox_input_pricing.setRange(0.0000, 2147483647)   
        self.spinBox_input_pricing.setDecimals(4)  # 设置小数点后的位数 
        self.spinBox_input_pricing.setValue(0.0015)


        layout_input_pricing.addWidget(self.labelD)
        layout_input_pricing.addWidget(self.labelE)
        layout_input_pricing.addStretch(1)  # 添加伸缩项
        layout_input_pricing.addWidget(self.spinBox_input_pricing)
        box_input_pricing.setLayout(layout_input_pricing)


        # -----创建第4个组（后面加的），添加多个组件-----
        box_output_pricing = QGroupBox()
        box_output_pricing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output_pricing = QHBoxLayout()

        #设置“TPM”标签
        self.labelF = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelF.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelF.setText("回复输出价格")
    
        #设置“说明”显示
        self.labelG = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelG.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelG.setText("( /1K tokens)") 

        #数值输入
        self.spinBox_output_pricing = DoubleSpinBox(self)
        self.spinBox_output_pricing.setRange(0.0000, 2147483647)
        self.spinBox_output_pricing.setDecimals(4)  # 设置小数点后的位数     
        self.spinBox_output_pricing.setValue(0.002)
        

        layout_output_pricing.addWidget(self.labelF)
        layout_output_pricing.addWidget(self.labelG)
        layout_output_pricing.addStretch(1)  # 添加伸缩项
        layout_output_pricing.addWidget(self.spinBox_output_pricing)
        box_output_pricing.setLayout(layout_output_pricing)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_tokens)
        container.addWidget(box_RPM)
        container.addWidget(box_TPM)
        container.addWidget(box_input_pricing)
        container.addWidget(box_output_pricing)
        container.addStretch(1)  # 添加伸缩项





class Widget_official_api(QFrame):

    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.hBoxLayout = QHBoxLayout(self)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))


class Widget_Openai(QFrame):#  Openai账号界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(text.replace(' ', '-'))
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['免费账号',  '付费账号(等级1)',  '付费账号(等级2)',  '付费账号(等级3)',  '付费账号(等级4)',  '付费账号(等级5)'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)



        # -----创建第2个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['gpt-3.5-turbo','gpt-3.5-turbo-0301','gpt-3.5-turbo-0613', 'gpt-3.5-turbo-1106', 'gpt-3.5-turbo-0125','gpt-3.5-turbo-16k', 'gpt-3.5-turbo-16k-0613',
                                 'gpt-4','gpt-4o','gpt-4-0314', 'gpt-4-0613','gpt-4-turbo','gpt-4-turbo-preview','gpt-4-1106-preview','gpt-4-0125-preview'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('gpt-3.5-turbo')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第3个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)


        # -----创建第4个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第3个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://api.openai.com/v1"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","OpenAI",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Google(QFrame):#  谷歌账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['免费账号',  '付费账号'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)

        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model =EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['gemini-1.5-flash','gemini-1.5-pro'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)

        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "google"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Google",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Anthropic(QFrame):#  Anthropic账号界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(text.replace(' ', '-'))
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['免费账号',  '付费账号(等级1)',  '付费账号(等级2)',  '付费账号(等级3)',  '付费账号(等级4)'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)



        # -----创建第2个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['claude-2.0','claude-2.1','claude-3-haiku-20240307','claude-3-sonnet-20240229', 'claude-3-opus-20240229','claude-3-5-sonnet-20240620'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(215, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('claude-2.0')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第3个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)


        # -----创建第4个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第3个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://api.anthropic.com"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Anthropic",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()
        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Cohere(QFrame):#  Cohere账号界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(text.replace(' ', '-'))
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['试用账号',  '生产账号'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)



        # -----创建第2个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['command','command-r','command-r-plus','c4ai-aya-23'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第3个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)


        # -----创建第4个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第3个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://api.cohere.com"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Cohere",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_ZhiPu(QFrame):#  智谱账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['免费账号',  '付费账号(等级1)',  '付费账号(等级2)',  '付费账号(等级3)',  '付费账号(等级4)',  '付费账号(等级5)'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)


        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['glm-4-flash','glm-4-air','glm-4-airx','glm-4','glm-4-0520'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)
        #设置下拉选择框默认选择
        self.comboBox_model.setCurrentText('glm-3-turbo')
        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://open.bigmodel.cn/api/paas/v4"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Zhipu",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Moonshot(QFrame):#  Moonshot账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_account_type = QGroupBox()
        box_account_type.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_account_type = QGridLayout()

        #设置“账号类型”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("账号类型")


        #设置“账号类型”下拉选择框
        self.comboBox_account_type = ComboBox() #以demo为父类
        self.comboBox_account_type.addItems(['免费账号',  '付费账号(等级1)',  '付费账号(等级2)',  '付费账号(等级3)',  '付费账号(等级4)',  '付费账号(等级5)'])
        self.comboBox_account_type.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_account_type.setFixedSize(150, 35)


        layout_account_type.addWidget(self.labelx, 0, 0)
        layout_account_type.addWidget(self.comboBox_account_type, 0, 1)
        box_account_type.setLayout(layout_account_type)


        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['moonshot-v1-8k','moonshot-v1-32k','moonshot-v1-128k'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)

        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_account_type)
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://api.moonshot.cn"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Moonshot",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Deepseek(QFrame):#  deepseek账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['deepseek-chat'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)

        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://api.deepseek.com/v1"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Deepseek",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Dashscope(QFrame):#  dashscope账号界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = EditableComboBox() #以demo为父类
        self.comboBox_model.addItems(['qwen-turbo','qwen-plus','qwen-max','qwen-long'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(200, 35)

        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_model)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://dashscope.aliyuncs.com/compatible-mode/v1"
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Dashscope",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")



class Widget_Volcengine(QFrame):  # 火山引擎主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_Volcengine_A('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_Volcengine_B('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '基础设置')
        self.addSubInterface(self.B_settings, 'B_settings', '速率价格设置')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_Volcengine_A(QFrame):#  火山引擎基础设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第3个组，添加多个组件-----
        box_access_point = QGroupBox()
        box_access_point.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_access_point = QHBoxLayout()

        #设置“推理接入点”标签
        self.label_access_point = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_access_point.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_access_point.setText("推理接入点")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                    ")

        #设置“推理接入点”的输入框
        self.LineEdit_access_point = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_access_point.addWidget(self.label_access_point)
        layout_access_point.addWidget(self.labelx)
        layout_access_point.addWidget(self.LineEdit_access_point)
        box_access_point.setLayout(layout_access_point)

        # -----创建第2个组，添加多个组件-----
        box_apikey = QGroupBox()
        box_apikey.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_apikey = QHBoxLayout()

        #设置“API KEY”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        self.labelx.setText("API KEY")

        #设置微调距离用的空白标签
        self.labely = QLabel()  
        self.labely.setText("                       ")

        #设置“API KEY”的输入框
        self.TextEdit_apikey = TextEdit()



        # 追加到容器中
        layout_apikey.addWidget(self.labelx)
        layout_apikey.addWidget(self.labely)
        layout_apikey.addWidget(self.TextEdit_apikey)
        # 添加到 box中
        box_apikey.setLayout(layout_apikey)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_access_point)
        container.addWidget(box_apikey)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = "https://ark.cn-beijing.volces.com/api/v3"
            Model_Type =  self.LineEdit_access_point.text()      #获取推理接入点
            API_key_str = self.TextEdit_apikey.toPlainText()        #获取apikey输入值
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Volcengine",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")


class Widget_Volcengine_B(QFrame):#  火山引擎进阶设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组(后面加的)，添加多个组件-----
        box_tokens = QGroupBox()
        box_tokens.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_tokens = QHBoxLayout()

        #设置标签
        self.label_tokens = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_tokens.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_tokens.setText("每次发送文本上限")

        #设置“说明”显示
        self.labelA_tokens = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA_tokens.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelA_tokens.setText("(tokens)")  

        #数值输入
        self.spinBox_tokens = SpinBox(self)
        self.spinBox_tokens.setRange(0, 2147483647)    
        self.spinBox_tokens.setValue(4000)


        layout_tokens.addWidget(self.label_tokens)
        layout_tokens.addWidget(self.labelA_tokens)
        layout_tokens.addStretch(1)  # 添加伸缩项
        layout_tokens.addWidget(self.spinBox_tokens)
        box_tokens.setLayout(layout_tokens)



        # -----创建第1个组(后面加的)，添加多个组件-----
        box_RPM = QGroupBox()
        box_RPM.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_RPM = QHBoxLayout()

        #设置“RPM”标签
        self.labelY = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelY.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelY.setText("每分钟请求数")

        #设置“说明”显示
        self.labelA = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelA.setText("(RPM)")  

        #数值输入
        self.spinBox_RPM = SpinBox(self)
        self.spinBox_RPM.setRange(0, 2147483647)    
        self.spinBox_RPM.setValue(10000)


        layout_RPM.addWidget(self.labelY)
        layout_RPM.addWidget(self.labelA)
        layout_RPM.addStretch(1)  # 添加伸缩项
        layout_RPM.addWidget(self.spinBox_RPM)
        box_RPM.setLayout(layout_RPM)



        # -----创建第2个组（后面加的），添加多个组件-----
        box_TPM = QGroupBox()
        box_TPM.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_TPM = QHBoxLayout()

        #设置“TPM”标签
        self.labelB = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelB.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelB.setText("每分钟tokens数")
    
        #设置“说明”显示
        self.labelC = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelC.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelC.setText("(TPM)") 

        #数值输入
        self.spinBox_TPM = SpinBox(self)
        self.spinBox_TPM.setRange(0, 2147483647)    
        self.spinBox_TPM.setValue(800000)


        layout_TPM.addWidget(self.labelB)
        layout_TPM.addWidget(self.labelC)
        layout_TPM.addStretch(1)  # 添加伸缩项
        layout_TPM.addWidget(self.spinBox_TPM)
        box_TPM.setLayout(layout_TPM)


        # -----创建第3个组（后面加的），添加多个组件-----
        box_input_pricing = QGroupBox()
        box_input_pricing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input_pricing = QHBoxLayout()

        #设置“请求输入价格”标签
        self.labelD = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelD.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelD.setText("请求输入价格")
    
        #设置“说明”显示
        self.labelE = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelE.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelE.setText("( /1K tokens)") 

        #数值输入
        self.spinBox_input_pricing = DoubleSpinBox(self)
        self.spinBox_input_pricing.setRange(0.0000, 2147483647)   
        self.spinBox_input_pricing.setDecimals(4)  # 设置小数点后的位数 
        self.spinBox_input_pricing.setValue(0.0020)


        layout_input_pricing.addWidget(self.labelD)
        layout_input_pricing.addWidget(self.labelE)
        layout_input_pricing.addStretch(1)  # 添加伸缩项
        layout_input_pricing.addWidget(self.spinBox_input_pricing)
        box_input_pricing.setLayout(layout_input_pricing)


        # -----创建第4个组（后面加的），添加多个组件-----
        box_output_pricing = QGroupBox()
        box_output_pricing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output_pricing = QHBoxLayout()

        #设置“TPM”标签
        self.labelF = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelF.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelF.setText("回复输出价格")
    
        #设置“说明”显示
        self.labelG = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelG.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.labelG.setText("( /1K tokens)") 

        #数值输入
        self.spinBox_output_pricing = DoubleSpinBox(self)
        self.spinBox_output_pricing.setRange(0.0000, 2147483647)
        self.spinBox_output_pricing.setDecimals(4)  # 设置小数点后的位数     
        self.spinBox_output_pricing.setValue(0.0500)
        

        layout_output_pricing.addWidget(self.labelF)
        layout_output_pricing.addWidget(self.labelG)
        layout_output_pricing.addStretch(1)  # 添加伸缩项
        layout_output_pricing.addWidget(self.spinBox_output_pricing)
        box_output_pricing.setLayout(layout_output_pricing)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_tokens)
        container.addWidget(box_RPM)
        container.addWidget(box_TPM)
        container.addWidget(box_input_pricing)
        container.addWidget(box_output_pricing)
        container.addStretch(1)  # 添加伸缩项



class Widget_SakuraLLM(QFrame):#  Sakura基础界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第1个组，添加多个组件-----
        box_address = QGroupBox()
        box_address.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_address = QHBoxLayout()

        #设置“请求地址”标签
        self.labelA = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelA.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelA.setText("请求地址")

        #设置微调距离用的空白标签
        self.labelB = QLabel()  
        self.labelB.setText("                      ")

        #设置“请求地址”的输入框
        self.LineEdit_address = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_address.addWidget(self.labelA)
        layout_address.addWidget(self.labelB)
        layout_address.addWidget(self.LineEdit_address)
        box_address.setLayout(layout_address)


        # -----创建第1个组，添加多个组件-----
        box_model = QGroupBox()
        box_model.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_model = QGridLayout()

        #设置“模型选择”标签
        self.labelx = QLabel(flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.labelx.setText("模型选择")


        #设置“模型类型”下拉选择框
        self.comboBox_model = ComboBox() #以demo为父类
        self.comboBox_model.addItems(['Sakura-v0.9','Sakura-v0.10pre'])
        self.comboBox_model.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_model.setFixedSize(250, 35)

        


        layout_model.addWidget(self.labelx, 0, 0)
        layout_model.addWidget(self.comboBox_model, 0, 1)
        box_model.setLayout(layout_model)



        # -----创建第3个组，添加多个组件-----
        box_proxy_port = QGroupBox()
        box_proxy_port.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_proxy_port = QHBoxLayout()

        #设置“代理地址”标签
        self.label_proxy_port = QLabel( flags=Qt.WindowFlags())  #parent参数表示父控件，如果没有父控件，可以将其设置为None；flags参数表示控件的标志，可以不传入
        self.label_proxy_port.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")#设置字体，大小，颜色
        self.label_proxy_port.setText("系统代理")

        #设置微调距离用的空白标签
        self.labelx = QLabel()  
        self.labelx.setText("                      ")

        #设置“代理地址”的输入框
        self.LineEdit_proxy_port = LineEdit()
        #LineEdit1.setFixedSize(300, 30)


        layout_proxy_port.addWidget(self.label_proxy_port)
        layout_proxy_port.addWidget(self.labelx)
        layout_proxy_port.addWidget(self.LineEdit_proxy_port)
        box_proxy_port.setLayout(layout_proxy_port)



        # -----创建第4个组，添加多个组件-----
        box_test = QGroupBox()
        box_test.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_test = QHBoxLayout()


        #设置“测试请求”的按钮
        primaryButton_test = PrimaryPushButton('测试请求', self, FIF.SEND)
        primaryButton_test.clicked.connect(self.test_request) #按钮绑定槽函数

        #设置“保存配置”的按钮
        primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数


        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_test)
        layout_test.addStretch(1)  # 添加伸缩项
        layout_test.addWidget(primaryButton_save)
        layout_test.addStretch(1)  # 添加伸缩项
        box_test.setLayout(layout_test)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_address)
        container.addWidget(box_model)
        container.addWidget(box_proxy_port)
        container.addWidget(box_test)
        container.addStretch(1)  # 添加伸缩项


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")

    def test_request(self):
        global Running_status

        if Running_status == 0:
            Base_url = self.LineEdit_address.text()
            Model_Type =  self.comboBox_model.currentText()      #获取模型类型下拉框当前选中选项的值
            API_key_str = ""
            Proxy_port = self.LineEdit_proxy_port.text()            #获取代理端口

            #创建子线程
            thread = background_executor("接口测试","","","Sakura",Base_url,Model_Type,API_key_str,Proxy_port)
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")




class Widget_translation_settings(QFrame):  # 翻译设置主界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.hBoxLayout = QHBoxLayout(self)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))


class Widget_translation_settings_A(QFrame):#  基础设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------

        # -----创建第1个组，添加多个组件-----
        box_translation_platform = QGroupBox()
        box_translation_platform.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_platform = QGridLayout()

        #设置“翻译平台”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("翻译平台")


        #设置“翻译平台”下拉选择框
        self.comboBox_translation_platform = ComboBox() #以demo为父类
        self.comboBox_translation_platform.addItems(['OpenAI官方',  'Google官方', 'Anthropic官方',  'Cohere官方',  'Moonshot官方',  'Deepseek官方',  'Dashscope官方', 'Volcengine官方', '智谱官方',  '代理平台',  'SakuraLLM'])
        self.comboBox_translation_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translation_platform.setFixedSize(150, 35)


        layout_translation_platform.addWidget(self.labelx, 0, 0)
        layout_translation_platform.addWidget(self.comboBox_translation_platform, 0, 1)
        box_translation_platform.setLayout(layout_translation_platform)


        # -----创建第1个组，添加多个组件-----
        box_translation_project = QGroupBox()
        box_translation_project.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_project = QGridLayout()

        #设置“翻译项目”标签
        self.labelx = QLabel( flags=Qt.WindowFlags())  
        self.labelx.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx.setText("翻译项目")


        #设置“翻译项目”下拉选择框
        self.comboBox_translation_project = ComboBox() #以demo为父类
        self.comboBox_translation_project.addItems(['Mtool导出文件',  'T++导出文件', 'VNText导出文件', 'ParaTranz导出文件', 'Epub小说文件' , 'Txt小说文件' , 'Srt字幕文件' , 'Lrc音声文件', 'Ainiee缓存文件'])
        self.comboBox_translation_project.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translation_project.setFixedSize(150, 35)


        layout_translation_project.addWidget(self.labelx, 0, 0)
        layout_translation_project.addWidget(self.comboBox_translation_project, 0, 1)
        box_translation_project.setLayout(layout_translation_project)


        # -----创建第2个组，添加多个组件-----
        box_input = QGroupBox()
        box_input.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("输入文件夹")

        #设置“输入文件夹”显示
        self.label_input_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_input_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_input_path.setText("(请选择原文文件所在的文件夹，不要混杂其他文件)")  

        #设置打开文件按钮
        self.pushButton_input = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_input.clicked.connect(self.Select_project_folder) #按钮绑定槽函数



        layout_input.addWidget(label4)
        layout_input.addWidget(self.label_input_path)
        layout_input.addStretch(1)  # 添加伸缩项
        layout_input.addWidget(self.pushButton_input)
        box_input.setLayout(layout_input)


        # -----创建第3个组，添加多个组件-----
        box_output = QGroupBox()
        box_output.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output = QHBoxLayout()

        #设置“输出文件夹”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("输出文件夹")

        #设置“输出文件夹”显示
        self.label_output_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_output_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_output_path.setText("(请选择翻译文件存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_output = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_output.clicked.connect(self.Select_output_folder) #按钮绑定槽函数


        

        layout_output.addWidget(label6)
        layout_output.addWidget(self.label_output_path)
        layout_output.addStretch(1)  # 添加伸缩项
        layout_output.addWidget(self.pushButton_output)
        box_output.setLayout(layout_output)





        # -----创建第4个组，添加多个组件-----
        box_source_text = QGroupBox()
        box_source_text.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_source_text = QHBoxLayout()


        #设置“文本源语言”标签
        label3 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label3.setText("文本源语言")

        #设置“文本源语言”下拉选择框
        self.comboBox_source_text = ComboBox() #以demo为父类
        self.comboBox_source_text.addItems(['日语', '英语', '韩语', '俄语', '简中', '繁中'])
        self.comboBox_source_text.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_source_text.setFixedSize(127, 30)


        layout_source_text.addWidget(label3)
        layout_source_text.addWidget(self.comboBox_source_text)
        box_source_text.setLayout(layout_source_text)


        # -----创建第5个组(后面添加的)，添加多个组件-----
        box_translated_text = QGroupBox()
        box_translated_text.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translated_text = QHBoxLayout()


        #设置“文本目标语言”标签
        label3_1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label3_1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label3_1.setText("文本目标语言")

        #设置“文本目标语言”下拉选择框
        self.comboBox_translated_text = ComboBox() #以demo为父类
        self.comboBox_translated_text.addItems(['简中', '繁中', '日语', '英语', '韩语'])
        self.comboBox_translated_text.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_translated_text.setFixedSize(127, 30)


        layout_translated_text.addWidget(label3_1)
        layout_translated_text.addWidget(self.comboBox_translated_text)
        box_translated_text.setLayout(layout_translated_text)


        # -----创建第6个组，添加多个组件-----
        box_save = QGroupBox()
        box_save.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_save = QHBoxLayout()

        #设置“保存配置”的按钮
        self.primaryButton_save = PushButton('保存配置', self, FIF.SAVE)
        self.primaryButton_save.clicked.connect(self.saveconfig) #按钮绑定槽函数



        layout_save.addStretch(1)  # 添加伸缩项
        layout_save.addWidget(self.primaryButton_save)
        layout_save.addStretch(1)  # 添加伸缩项
        box_save.setLayout(layout_save)




        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_translation_project)
        container.addWidget(box_translation_platform)
        container.addWidget(box_source_text)
        container.addWidget(box_translated_text)
        container.addWidget(box_input)
        container.addWidget(box_output)
        container.addWidget(box_save)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下



    # 选择输入文件夹按钮绑定函数
    def Select_project_folder(self):
        Input_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            # 将输入路径存储到配置器中
            configurator.Input_Folder = Input_Folder
            self.label_input_path.setText(Input_Folder)
            print('[INFO]  已选择项目文件夹: ',Input_Folder)
        else :
            print('[INFO]  未选择文件夹')



    # 选择输出文件夹按钮绑定函数
    def Select_output_folder(self):
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            # 将输入路径存储到配置器中
            configurator.Output_Folder = Output_Folder
            self.label_output_path.setText(Output_Folder)
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)
        else :
            print('[INFO]  未选择文件夹')


    def saveconfig(self):
        configurator.read_write_config("write")
        user_interface_prompter.createSuccessInfoBar("已成功保存配置")


class Widget_translation_settings_B1(QFrame):#  发送设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第个组，添加多个组件-----
        box_lines_limit = QGroupBox()
        box_lines_limit.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_lines_limit = QHBoxLayout()

        #设置标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("每次翻译")

        self.spinBox_lines_limit = SpinBox(self)
        self.spinBox_lines_limit.setRange(0, 99999)    
        self.spinBox_lines_limit.setValue(20)

        #设置“说明”显示
        self.labelA_lines = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA_lines.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 15px")
        self.labelA_lines.setText("(行)")  


        # 设置开关
        self.checkBox_lines_limit_switch = CheckBox('使用行数模式', self)
        self.checkBox_lines_limit_switch.setChecked(True)
        self.checkBox_lines_limit_switch.stateChanged.connect(self.on_lines)

        layout_lines_limit.addWidget(label4)
        layout_lines_limit.addWidget(self.spinBox_lines_limit)
        layout_lines_limit.addWidget( self.labelA_lines)
        layout_lines_limit.addStretch(1)
        layout_lines_limit.addWidget(self.checkBox_lines_limit_switch)
        box_lines_limit.setLayout(layout_lines_limit)


        # -----创建第个组，添加多个组件-----
        box_tokens_limit = QGroupBox()
        box_tokens_limit.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_tokens_limit = QHBoxLayout()

        #设置标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("每次翻译")

        self.spinBox_tokens_limit = SpinBox(self)
        self.spinBox_tokens_limit.setRange(0, 99999)    
        self.spinBox_tokens_limit.setValue(1500)

        #设置“说明”显示
        self.labelA_tokens = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.labelA_tokens.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 15px")
        self.labelA_tokens.setText("(tokens)")  


        # 设置开关
        self.checkBox_tokens_limit_switch = CheckBox('使用tokens模式', self)
        self.checkBox_tokens_limit_switch.setChecked(False)
        self.checkBox_tokens_limit_switch.stateChanged.connect(self.on_tokens)

        layout_tokens_limit.addWidget(label4)
        layout_tokens_limit.addWidget(self.spinBox_tokens_limit)
        layout_tokens_limit.addWidget( self.labelA_tokens)
        layout_tokens_limit.addStretch(1)
        layout_tokens_limit.addWidget(self.checkBox_tokens_limit_switch)
        box_tokens_limit.setLayout(layout_tokens_limit)



        # -----创建第1个组，添加多个组件-----
        box_pre_lines = QGroupBox()
        box_pre_lines.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_pre_lines = QHBoxLayout()

        #设置标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label1.setText("携带上文行数")


        #设置数值输入框
        self.spinBox_pre_lines = SpinBox(self)
        self.spinBox_pre_lines.setRange(0, 1000)    
        self.spinBox_pre_lines.setValue(0)


        layout_pre_lines.addWidget(label1)
        layout_pre_lines.addStretch(1)  # 添加伸缩项
        layout_pre_lines.addWidget(self.spinBox_pre_lines)
        box_pre_lines.setLayout(layout_pre_lines)



        # -----创建第1个组(后来补的)，添加多个组件-----
        box1_thread_count = QGroupBox()
        box1_thread_count.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_thread_count = QHBoxLayout()

        #设置“最大线程数”标签
        label1_7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1_7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label1_7.setText("最大线程数")

        #设置“说明”显示
        label2_7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label2_7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        label2_7.setText("(0是自动根据电脑设置线程数)")  

       #设置“最大线程数”数值输入框
        self.spinBox_thread_count = SpinBox(self)
        #设置最大最小值
        self.spinBox_thread_count.setRange(0, 1000)    
        self.spinBox_thread_count.setValue(0)

        layout1_thread_count.addWidget(label1_7)
        layout1_thread_count.addWidget(label2_7)
        layout1_thread_count.addStretch(1)  # 添加伸缩项
        layout1_thread_count.addWidget(self.spinBox_thread_count)
        box1_thread_count.setLayout(layout1_thread_count)


        # -----创建第x个组，添加多个组件-----
        box_retry_count_limit = QGroupBox()
        box_retry_count_limit.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_retry_count_limit = QHBoxLayout()


        label1_7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1_7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label1_7.setText("错误重翻最大次数限制")


        # 设置数值输入框
        self.spinBox_retry_count_limit = SpinBox(self)
        # 设置最大最小值
        self.spinBox_retry_count_limit.setRange(0, 1000)    
        self.spinBox_retry_count_limit.setValue(1)

        layout_retry_count_limit.addWidget(label1_7)
        layout_retry_count_limit.addStretch(1)  # 添加伸缩项
        layout_retry_count_limit.addWidget(self.spinBox_retry_count_limit)
        box_retry_count_limit.setLayout(layout_retry_count_limit)


        # -----创建第x个组，添加多个组件-----
        box_round_limit = QGroupBox()
        box_round_limit.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_round_limit = QHBoxLayout()


        label1_7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1_7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label1_7.setText("翻译流程最大轮次限制")


        # 设置数值输入框
        self.spinBox_round_limit = SpinBox(self)
        # 设置最大最小值
        self.spinBox_round_limit.setRange(3, 1000)    
        self.spinBox_round_limit.setValue(6)

        layout_round_limit.addWidget(label1_7)
        layout_round_limit.addStretch(1)  # 添加伸缩项
        layout_round_limit.addWidget(self.spinBox_round_limit)
        box_round_limit.setLayout(layout_round_limit)


        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_lines_limit)
        container.addWidget(box_tokens_limit)
        container.addWidget(box_pre_lines)
        container.addWidget(box1_thread_count)
        container.addWidget(box_retry_count_limit)
        container.addWidget(box_round_limit)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


    #设置选择开关绑定函数
    def on_clear(self, isChecked: bool):
        if isChecked:
            user_interface_prompter.createWarningInfoBar("仅支持翻译日语文本时生效，建议翻译T++导出文件时开启")

    #设互斥开关函数
    def on_lines(self, isChecked: bool):
        if isChecked:
            self.checkBox_tokens_limit_switch.setChecked(False)

    #设互斥开关函数
    def on_tokens(self, isChecked: bool):
        if isChecked:
            self.checkBox_lines_limit_switch.setChecked(False)


class Widget_translation_settings_B2(QFrame):#  专项设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第3个组(后来补的)，添加多个组件-----
        box1_cot_toggle = QGroupBox()
        box1_cot_toggle.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_cot_toggle = QHBoxLayout()

        #设置“简繁转换开关”标签
        labe1_6 = QLabel(flags=Qt.WindowFlags())  
        labe1_6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_6.setText("使用思维链模式")

       #设置“简繁体自动转换”选择开关
        self.SwitchButton_cot_toggle = SwitchButton(parent=self)    



        layout1_cot_toggle.addWidget(labe1_6)
        layout1_cot_toggle.addStretch(1)  # 添加伸缩项
        layout1_cot_toggle.addWidget(self.SwitchButton_cot_toggle)
        box1_cot_toggle.setLayout(layout1_cot_toggle)


        # -----创建第3个组(后来补的)，添加多个组件-----
        box1_cn_prompt_toggle = QGroupBox()
        box1_cn_prompt_toggle.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_cn_prompt_toggle = QHBoxLayout()

        #设置“简繁转换开关”标签
        labe1_6 = QLabel(flags=Qt.WindowFlags())  
        labe1_6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_6.setText("使用中文提示词")

       #设置“简繁体自动转换”选择开关
        self.SwitchButton_cn_prompt_toggle = SwitchButton(parent=self)    



        layout1_cn_prompt_toggle.addWidget(labe1_6)
        layout1_cn_prompt_toggle.addStretch(1)  # 添加伸缩项
        layout1_cn_prompt_toggle.addWidget(self.SwitchButton_cn_prompt_toggle)
        box1_cn_prompt_toggle.setLayout(layout1_cn_prompt_toggle)


        # -----创建第1个组(后来补的)，添加多个组件-----
        box_clear = QGroupBox()
        box_clear.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_clear = QHBoxLayout()

        #设置标签
        labe1_4 = QLabel(flags=Qt.WindowFlags())  
        labe1_4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_4.setText("处理首尾非文本字符")



       #设置选择开关
        self.SwitchButton_clear = SwitchButton(parent=self)    
        self.SwitchButton_clear.checkedChanged.connect(self.on_clear)



        layout_clear.addWidget(labe1_4)
        layout_clear.addStretch(1)  # 添加伸缩项
        layout_clear.addWidget(self.SwitchButton_clear)
        box_clear.setLayout(layout_clear)



        # -----创建第3个组(后来补的)，添加多个组件-----
        box1_conversion_toggle = QGroupBox()
        box1_conversion_toggle.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_conversion_toggle = QHBoxLayout()

        #设置“简繁转换开关”标签
        labe1_6 = QLabel(flags=Qt.WindowFlags())  
        labe1_6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_6.setText("简繁体自动转换")

       #设置“简繁体自动转换”选择开关
        self.SwitchButton_conversion_toggle = SwitchButton(parent=self)    



        layout1_conversion_toggle.addWidget(labe1_6)
        layout1_conversion_toggle.addStretch(1)  # 添加伸缩项
        layout1_conversion_toggle.addWidget(self.SwitchButton_conversion_toggle)
        box1_conversion_toggle.setLayout(layout1_conversion_toggle)



        # -----创建第4个组(后来补的)，添加多个组件-----
        box1_line_breaks = QGroupBox()
        box1_line_breaks.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_line_breaks = QHBoxLayout()

        #设置“换行符保留”标签
        labe1_6 = QLabel(flags=Qt.WindowFlags())  
        labe1_6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        labe1_6.setText("保留换行符")

       #设置“换行符保留”选择开关
        self.SwitchButton_line_breaks = SwitchButton(parent=self)    



        layout1_line_breaks.addWidget(labe1_6)
        layout1_line_breaks.addStretch(1)  # 添加伸缩项
        layout1_line_breaks.addWidget(self.SwitchButton_line_breaks)
        box1_line_breaks.setLayout(layout1_line_breaks)




        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1_cot_toggle)
        container.addWidget(box1_cn_prompt_toggle)
        container.addWidget(box1_line_breaks)
        container.addWidget(box1_conversion_toggle)
        container.addWidget(box_clear)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


    #设置选择开关绑定函数
    def on_clear(self, isChecked: bool):
        if isChecked:
            user_interface_prompter.createWarningInfoBar("仅支持翻译日语文本时生效，建议翻译T++导出文件时开启")


class Widget_translation_settings_C(QFrame):#  混合翻译设置子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_switch = QGroupBox()
        box_switch.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_switch = QHBoxLayout()

        #设置标签
        self.labe1_4 = QLabel(flags=Qt.WindowFlags())  
        self.labe1_4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        self.labe1_4.setText("启用混合平台翻译功能")



        # 设置选择开关
        self.SwitchButton_mixed_translation = SwitchButton(parent=self)    
        self.SwitchButton_mixed_translation.checkedChanged.connect(self.test)



        layout_switch.addWidget(self.labe1_4)
        layout_switch.addStretch(1)  # 添加伸缩项
        layout_switch.addWidget(self.SwitchButton_mixed_translation)
        box_switch.setLayout(layout_switch)





        # -----创建第2个组，添加多个组件-----
        box_translation_platform1 = QGroupBox()
        box_translation_platform1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_platform1 = QGridLayout()

        #设置“翻译平台”标签
        self.labelx1 = QLabel( flags=Qt.WindowFlags())  
        self.labelx1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx1.setText("首轮翻译平台")


        #设置“翻译平台”下拉选择框
        self.comboBox_primary_translation_platform = ComboBox() #以demo为父类
        self.comboBox_primary_translation_platform.addItems(['OpenAI官方',  'Google官方', 'Anthropic官方',  'Cohere官方',  'Moonshot官方',  'Deepseek官方',  'Dashscope官方', 'Volcengine官方',  '智谱官方',  '代理平台',  'SakuraLLM'])
        self.comboBox_primary_translation_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_primary_translation_platform.setFixedSize(150, 35)


        layout_translation_platform1.addWidget(self.labelx1, 0, 0)
        layout_translation_platform1.addWidget(self.comboBox_primary_translation_platform, 0, 1)
        box_translation_platform1.setLayout(layout_translation_platform1)



        # -----创建第3个组，添加多个组件-----
        box_translation_platform2 = QGroupBox()
        box_translation_platform2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_platform2 = QGridLayout()

        #设置“翻译平台”标签
        self.labelx2 = QLabel( flags=Qt.WindowFlags())  
        self.labelx2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx2.setText("次轮翻译平台")


        #设置“翻译平台”下拉选择框
        self.comboBox_secondary_translation_platform = ComboBox() #以demo为父类
        self.comboBox_secondary_translation_platform.addItems(['不设置', 'OpenAI官方',  'Google官方', 'Anthropic官方',  'Cohere官方',  'Moonshot官方',  'Deepseek官方',  'Dashscope官方', 'Volcengine官方',  '智谱官方',  '代理平台',  'SakuraLLM'])
        self.comboBox_secondary_translation_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_secondary_translation_platform.setFixedSize(150, 35)


        layout_translation_platform2.addWidget(self.labelx2, 0, 0)
        layout_translation_platform2.addWidget(self.comboBox_secondary_translation_platform, 0, 1)
        box_translation_platform2.setLayout(layout_translation_platform2)



        # -----创建第4个组，添加多个组件-----
        box_translation_platform3 = QGroupBox()
        box_translation_platform3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_platform3 = QGridLayout()

        #设置“翻译平台”标签
        self.labelx3 = QLabel( flags=Qt.WindowFlags())  
        self.labelx3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.labelx3.setText("末轮翻译平台")


        #设置“翻译平台”下拉选择框
        self.comboBox_final_translation_platform = ComboBox() #以demo为父类
        self.comboBox_final_translation_platform.addItems(['不设置','OpenAI官方',  'Google官方', 'Anthropic官方',  'Cohere官方',  'Moonshot官方',  'Deepseek官方',  'Dashscope官方', 'Volcengine官方',  '智谱官方',  '代理平台',  'SakuraLLM'])
        self.comboBox_final_translation_platform.setCurrentIndex(0) #设置下拉框控件（ComboBox）的当前选中项的索引为0，也就是默认选中第一个选项
        self.comboBox_final_translation_platform.setFixedSize(150, 35)


        layout_translation_platform3.addWidget(self.labelx3, 0, 0)
        layout_translation_platform3.addWidget(self.comboBox_final_translation_platform, 0, 1)
        box_translation_platform3.setLayout(layout_translation_platform3)



        # -----创建第1个组，添加多个组件-----
        box_split_switch = QGroupBox()
        box_split_switch.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_split_switch = QHBoxLayout()

        #设置标签
        self.labe1_split_switch = QLabel(flags=Qt.WindowFlags())  
        self.labe1_split_switch.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        self.labe1_split_switch.setText("更换轮次后不进行文本拆分")



        # 设置选择开关
        self.SwitchButton_split_switch = SwitchButton(parent=self)    
        #self.SwitchButton_split_switch.checkedChanged.connect(self.test)



        layout_split_switch.addWidget(self.labe1_split_switch)
        layout_split_switch.addStretch(1)  # 添加伸缩项
        layout_split_switch.addWidget(self.SwitchButton_split_switch)
        box_split_switch.setLayout(layout_split_switch)



        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        #container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_switch)
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_translation_platform1)
        container.addWidget(box_translation_platform2)
        container.addWidget(box_translation_platform3)
        container.addWidget( box_split_switch)
        container.addStretch(1)  # 添加伸缩项
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


    #设置开关绑定函数
    def test(self, isChecked: bool):
        if isChecked:
            user_interface_prompter.createWarningInfoBar("请注意，开启该开关下面设置才会生效，并且会覆盖基础设置中的翻译平台")




class Widget_start_translation(QFrame):  # 开始翻译主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_start_translation_A('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_start_translation_B('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '开始翻译')
        self.addSubInterface(self.B_settings, 'B_settings', '备份功能')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_start_translation_A(QFrame):#  开始翻译子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------


        # -----创建第1个组，添加多个组件-----
        box_project = QGroupBox()
        box_project.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")  # 分别设置了边框大小，边框颜色，边框圆角
        layout_project = QHBoxLayout()

        # 第一组水平布局
        layout_horizontal_1 = QHBoxLayout()

        self.label111 = QLabel(flags=Qt.WindowFlags())
        self.label111.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label111.setText("项目类型 :")

        self.translation_project = QLabel(flags=Qt.WindowFlags())
        self.translation_project.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.translation_project.setText("无")

        layout_horizontal_1.addWidget(self.label111)
        layout_horizontal_1.addStretch(1)  # 添加伸缩项
        layout_horizontal_1.addWidget(self.translation_project)
        layout_horizontal_1.addStretch(1)  # 添加伸缩项

        # 第二组水平布局
        layout_horizontal_2 = QHBoxLayout()

        self.label222 = QLabel(flags=Qt.WindowFlags())
        self.label222.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label222.setText("项目ID :")

        self.project_id = QLabel(flags=Qt.WindowFlags())
        self.project_id.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.project_id.setText("无")

        layout_horizontal_2.addWidget(self.label222)
        layout_horizontal_2.addStretch(1)  # 添加伸缩项
        layout_horizontal_2.addWidget(self.project_id)
        layout_horizontal_2.addStretch(1)  # 添加伸缩项

        # 将两个水平布局放入最外层水平布局
        layout_project.addLayout(layout_horizontal_1)
        layout_project.addLayout(layout_horizontal_2)

        box_project.setLayout(layout_project)


        # -----创建第2个组，添加多个组件-----
        box_text_line_count = QGroupBox()
        box_text_line_count.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")  # 分别设置了边框大小，边框颜色，边框圆角
        layout_text_line_count = QHBoxLayout()

        # 第三组水平布局
        layout_horizontal_3 = QHBoxLayout()

        self.label333 = QLabel(flags=Qt.WindowFlags())
        self.label333.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label333.setText("总文本行数 :")

        self.total_text_line_count = QLabel(flags=Qt.WindowFlags())
        self.total_text_line_count.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.total_text_line_count.setText("无")

        layout_horizontal_3.addWidget(self.label333)
        layout_horizontal_3.addStretch(1)  # 添加伸缩项
        layout_horizontal_3.addWidget(self.total_text_line_count)
        layout_horizontal_3.addStretch(1)  # 添加伸缩项

        # 第四组水平布局
        layout_horizontal_4 = QHBoxLayout()

        self.label444 = QLabel(flags=Qt.WindowFlags())
        self.label444.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.label444.setText("已翻译行数 :")

        self.translated_line_count = QLabel(flags=Qt.WindowFlags())
        self.translated_line_count.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.translated_line_count.setText("无")

        layout_horizontal_4.addWidget(self.label444)
        layout_horizontal_4.addStretch(1)  # 添加伸缩项
        layout_horizontal_4.addWidget(self.translated_line_count)
        layout_horizontal_4.addStretch(1)  # 添加伸缩项

        # 将第三组和第四组水平布局放入最外层水平布局
        layout_text_line_count.addLayout(layout_horizontal_3)
        layout_text_line_count.addLayout(layout_horizontal_4)

        box_text_line_count.setLayout(layout_text_line_count)





        # -----创建第3个组，添加多个组件-----
        box_spent = QGroupBox()
        box_spent.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")  # 分别设置了边框大小，边框颜色，边框圆角
        layout_spent = QHBoxLayout()

        # 第五组水平布局
        layout_horizontal_5 = QHBoxLayout()

        self.labelx1 = QLabel(flags=Qt.WindowFlags())
        self.labelx1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.labelx1.setText("已花费tokens :")

        self.tokens_spent = QLabel(flags=Qt.WindowFlags())
        self.tokens_spent.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.tokens_spent.setText("无")

        layout_horizontal_5.addWidget(self.labelx1)
        layout_horizontal_5.addStretch(1)  # 添加伸缩项
        layout_horizontal_5.addWidget(self.tokens_spent)
        layout_horizontal_5.addStretch(1)  # 添加伸缩项

        # 第六组水平布局
        layout_horizontal_6 = QHBoxLayout()

        self.labelx2 = QLabel(flags=Qt.WindowFlags())
        self.labelx2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.labelx2.setText("已花费金额(＄) :")

        self.amount_spent = QLabel(flags=Qt.WindowFlags())
        self.amount_spent.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")  # 设置字体，大小，颜色
        self.amount_spent.setText("无")

        layout_horizontal_6.addWidget(self.labelx2)
        layout_horizontal_6.addStretch(1)  # 添加伸缩项
        layout_horizontal_6.addWidget(self.amount_spent)
        layout_horizontal_6.addStretch(1)  # 添加伸缩项

        # 将第五组和第六组水平布局放入最外层水平布局
        layout_spent.addLayout(layout_horizontal_5)
        layout_spent.addLayout(layout_horizontal_6)

        box_spent.setLayout(layout_spent)



        # -----创建第4个组，添加多个组件-----
        box_progressRing = QGroupBox()
        box_progressRing.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_progressRing = QHBoxLayout()


        #设置“翻译进度”标签
        self.label_progressRing = QLabel( flags=Qt.WindowFlags())  
        self.label_progressRing.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px; ")#设置字体，大小，颜色
        self.label_progressRing.setText("翻译进度")

        #设置翻译进度条
        self.progressRing = ProgressRing(self)
        self.progressRing.setValue(0)
        self.progressRing.setTextVisible(True)
        self.progressRing.setFixedSize(80, 80)


        layout_progressRing.addWidget(self.label_progressRing)
        layout_progressRing.addStretch(1)  # 添加伸缩项
        layout_progressRing.addWidget(self.progressRing)
        box_progressRing.setLayout(layout_progressRing)





        # -----创建第5个组，添加多个组件-----
        box_start_translation = QGroupBox()
        box_start_translation.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_start_translation = QHBoxLayout()


        #设置“开始翻译”的按钮
        self.primaryButton_start_translation = PrimaryPushButton('开始翻译', self, FIF.PLAY)
        self.primaryButton_start_translation.clicked.connect(self.Start_translation) #按钮绑定槽函数


        #设置“暂停翻译”的按钮
        self.primaryButton_pause_translation = PrimaryPushButton('暂停翻译', self, FIF.PAUSE)
        self.primaryButton_pause_translation.clicked.connect(self.pause_translation) #按钮绑定槽函数
        #隐藏按钮
        self.primaryButton_pause_translation.hide()

        #设置“继续翻译”的按钮
        self.primaryButton_continue_translation = PrimaryPushButton('继续翻译', self, FIF.ROTATE)
        self.primaryButton_continue_translation.clicked.connect(self.continue_translation) #按钮绑定槽函数
        #隐藏按钮
        self.primaryButton_continue_translation.hide()


        #设置“终止翻译”的按钮
        self.primaryButton_terminate_translation = PushButton('取消翻译', self, FIF.CANCEL)
        self.primaryButton_terminate_translation.clicked.connect(self.terminate_translation) #按钮绑定槽函数




        layout_start_translation.addStretch(1)  # 添加伸缩项
        layout_start_translation.addWidget(self.primaryButton_start_translation)
        layout_start_translation.addWidget(self.primaryButton_continue_translation)
        layout_start_translation.addWidget(self.primaryButton_pause_translation)
        layout_start_translation.addStretch(1)  # 添加伸缩项
        layout_start_translation.addWidget(self.primaryButton_terminate_translation)
        layout_start_translation.addStretch(1)  # 添加伸缩项
        box_start_translation.setLayout(layout_start_translation)


        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_project)
        container.addWidget(box_text_line_count)
        container.addWidget(box_spent)
        container.addWidget(box_progressRing)
        container.addWidget(box_start_translation)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下



    #开始翻译按钮绑定函数
    def Start_translation(self):
        global Running_status

        if Running_status == 0:
            #隐藏开始翻译按钮
            self.primaryButton_start_translation.hide()
            #显示暂停翻译按钮
            self.primaryButton_pause_translation.show()

            #创建子线程
            thread = background_executor("执行翻译任务","","","","","","","")
            thread.start()

        elif Running_status != 0:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")
    
    #暂停翻译按钮绑定函数
    def pause_translation(self):
        global Running_status

        #隐藏暂停翻译按钮
        self.primaryButton_pause_translation.hide()
        #显示继续翻译按钮
        self.primaryButton_continue_translation.show()

        Running_status = 9
        user_interface_prompter.createWarningInfoBar("软件的多线程任务正在逐一取消中，请等待全部任务释放完成！！！")
        print("\033[1;33mWarning:\033[0m 软件的多线程任务正在逐一取消中，请等待全部任务释放完成！！！-----------------------","\n")

    #继续翻译按钮绑定函数
    def continue_translation(self):
        global Running_status
        
        if Running_status == 9:
            #隐藏继续翻译按钮
            self.primaryButton_continue_translation.hide()
            #显示暂停翻译按钮
            self.primaryButton_pause_translation.show()

            #创建子线程
            thread = background_executor("执行翻译任务","","","","","","","")
            thread.start()

        elif Running_status != 9:
            user_interface_prompter.createWarningInfoBar("正在进行任务中，请等待任务结束后再操作~")
    
    #取消翻译按钮绑定函数
    def terminate_translation(self):
        global Running_status

        #隐藏继续翻译按钮
        self.primaryButton_continue_translation.hide()
        #隐藏暂停翻译按钮
        self.primaryButton_pause_translation.hide()
        #显示开始翻译按钮
        self.primaryButton_start_translation.show()

        #如果正在翻译中
        if Running_status == 6:
            Running_status = 10
            user_interface_prompter.createWarningInfoBar("软件的多线程任务正在逐一取消中，请等待全部任务释放完成！！！")
            print("\033[1;33mWarning:\033[0m 软件的多线程任务正在逐一取消中，请等待全部翻译任务释放完成！！！-----------------------","\n")

        #如果正在暂停中
        elif Running_status == 9:

            Running_status = 0
            print("\033[1;33mWarning:\033[0m 翻译任务已取消-----------------------","\n")
            #界面提示
            user_interface_prompter.createWarningInfoBar("翻译已取消")
            user_interface_prompter.signal.emit("重置界面数据","翻译取消",0)


        #如果正在空闲中
        elif Running_status == 0:

            Running_status = 0
            print("\033[1;33mWarning:\033[0m 当前无翻译任务-----------------------","\n")
            #界面提示
            user_interface_prompter.createWarningInfoBar("当前无翻译任务")
            user_interface_prompter.signal.emit("重置界面数据","翻译取消",0)


class Widget_start_translation_B(QFrame):#  开始翻译子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_switch = QGroupBox()
        box_switch.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_switch = QHBoxLayout()


        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("自动备份缓存文件到输出文件夹")


        self.checkBox_switch = CheckBox('启用功能')
        self.checkBox_switch.stateChanged.connect(self.checkBoxChanged1)

        layout_switch.addWidget(label1)
        layout_switch.addStretch(1)  # 添加伸缩项
        layout_switch.addWidget(self.checkBox_switch)
        box_switch.setLayout(layout_switch)



        # -----创建第1个组，添加多个组件-----
        box_export_cache_file_path = QGroupBox()
        box_export_cache_file_path.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_export_cache_file_path = QHBoxLayout()

        #设置“导出当前任务的缓存文件”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("导出当前任务的缓存文件")


        #设置导出当前任务的缓存文件按钮
        self.pushButton_export_cache_file_path = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_export_cache_file_path.clicked.connect(self.output_cachedata) #按钮绑定槽函数



        layout_export_cache_file_path.addWidget(label4)
        layout_export_cache_file_path.addStretch(1)  # 添加伸缩项
        layout_export_cache_file_path.addWidget(self.pushButton_export_cache_file_path)
        box_export_cache_file_path.setLayout(layout_export_cache_file_path)


        # -----创建第2个组，添加多个组件-----
        box_export_translated_file_path = QGroupBox()
        box_export_translated_file_path.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_export_translated_file_path = QHBoxLayout()

        #设置“导出当前任务的已翻译文本”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("导出当前任务的已翻译文本")


        #设置导出当前任务的已翻译文本按钮
        self.pushButton_export_translated_file_path = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_export_translated_file_path.clicked.connect(self.output_data) #按钮绑定槽函数


        

        layout_export_translated_file_path.addWidget(label6)
        layout_export_translated_file_path.addStretch(1)  # 添加伸缩项
        layout_export_translated_file_path.addWidget(self.pushButton_export_translated_file_path)
        box_export_translated_file_path.setLayout(layout_export_translated_file_path)







        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项        
        container.addWidget(box_switch)
        container.addWidget(box_export_cache_file_path)
        container.addWidget(box_export_translated_file_path)
        container.addStretch(1)  # 添加伸缩项

    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启自动备份功能")

    # 缓存文件输出
    def output_cachedata(self):
        global cache_list

        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)

            if len(cache_list)>= 3:
                #创建子线程
                thread = background_executor("输出缓存文件","",Output_Folder,"","","","","")
                thread.start()
            else:
                print('[INFO]  未存在缓存文件')
                return  # 直接返回，不执行后续操作
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 缓存文件输出
    def output_data(self):
        global cache_list

        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print('[INFO]  已选择输出文件夹:' ,Output_Folder)

            if len(cache_list)>= 3:
                #创建子线程
                thread = background_executor("输出已翻译文件",configurator.Input_Folder,Output_Folder,"","","","","")
                thread.start()

            else:
                print('[INFO]  未存在缓存文件')
                return  # 直接返回，不执行后续操作
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作




class Widget_tune(QFrame):  # 实时调教主界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.hBoxLayout = QHBoxLayout(self)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))


class Widget_tune_openai(QFrame):# oepnai调教界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        # 最外层的垂直布局
        container = QVBoxLayout()


        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        #设置“启用实时参数”标签
        label0 = QLabel(flags=Qt.WindowFlags())  
        label0.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label0.setText("实时改变AI参数")

        #设置官方文档说明链接按钮
        hyperlinkButton = HyperlinkButton(
            url='https://platform.openai.com/docs/api-reference/chat/create',
            text='(官方文档)'
        )

        #设置“启用实时参数”开关
        self.checkBox = CheckBox('启用', self)
        #self.checkBox.stateChanged.connect(self.checkBoxChanged)


        layout1.addWidget(label0)
        layout1.addWidget(hyperlinkButton)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox)
        box1.setLayout(layout1)




        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“温度”标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label1.setText("Temperature")

        #设置“温度”副标签
        label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label11.setText("(官方默认值为1)")

        #设置“温度”滑动条
        self.slider1 = Slider(Qt.Horizontal, self)
        self.slider1.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label2 = QLabel(str(self.slider1.value()), self)
        self.label2.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider1.valueChanged.connect(lambda value: self.label2.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label2显示正确的值
        self.slider1.setMinimum(0)
        self.slider1.setMaximum(20)
        self.slider1.setValue(0)

        

        layout3.addWidget(label1)
        layout3.addWidget(label11)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.slider1)
        layout3.addWidget(self.label2)
        box3.setLayout(layout3)






        # -----创建第5个组，添加多个组件-----
        box5 = QGroupBox()
        box5.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout5 = QHBoxLayout()

        #设置“top_p”标签
        label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label4.setText("Top_p")

        #设置“top_p”副标签
        label41 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label41.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label41.setText("(官方默认值为1)")


        #设置“top_p”滑动条
        self.slider2 = Slider(Qt.Horizontal, self)
        self.slider2.setFixedWidth(200)


        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label5 = QLabel(str(self.slider2.value()), self)
        self.label5.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider2.valueChanged.connect(lambda value: self.label5.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放在后面是为了让上面的label5显示正确的值和格式
        self.slider2.setMinimum(0)
        self.slider2.setMaximum(10)
        self.slider2.setValue(10)



        layout5.addWidget(label4)
        layout5.addWidget(label41)
        layout5.addStretch(1)  # 添加伸缩项
        layout5.addWidget(self.slider2)
        layout5.addWidget(self.label5)
        box5.setLayout(layout5)








        # -----创建第7个组，添加多个组件-----
        box7 = QGroupBox()
        box7.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout7 = QHBoxLayout()

        #设置“presence_penalty”标签
        label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label7.setText("Presence_penalty")

        #设置“presence_penalty”副标签
        label71 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label71.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label71.setText("(官方默认值为0)")


        #设置“presence_penalty”滑动条
        self.slider3 = Slider(Qt.Horizontal, self)
        self.slider3.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label8 = QLabel(str(self.slider3.value()), self)
        self.label8.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label8.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider3.valueChanged.connect(lambda value: self.label8.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label8显示正确的值和格式
        self.slider3.setMinimum(-20)
        self.slider3.setMaximum(20)
        self.slider3.setValue(0)



        layout7.addWidget(label7)
        layout7.addWidget(label71)
        layout7.addStretch(1)  # 添加伸缩项
        layout7.addWidget(self.slider3)
        layout7.addWidget(self.label8)
        box7.setLayout(layout7)






        # -----创建第9个组，添加多个组件-----
        box9 = QGroupBox()
        box9.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout9 = QHBoxLayout()

        #设置“frequency_penalty”标签
        label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label9.setText("Frequency_penalty")

        #设置“presence_penalty”副标签
        label91 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label91.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label91.setText("(官方默认值为0)")

        #设置“frequency_penalty”滑动条
        self.slider4 = Slider(Qt.Horizontal, self)
        self.slider4.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label10 = QLabel(str(self.slider4.value()), self)
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label10.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.slider4.valueChanged.connect(lambda value: self.label10.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label10显示正确的值和格式
        self.slider4.setMinimum(-20)
        self.slider4.setMaximum(20)
        self.slider4.setValue(0)


        layout9.addWidget(label9)
        layout9.addWidget(label91)
        layout9.addStretch(1)  # 添加伸缩项
        layout9.addWidget(self.slider4)
        layout9.addWidget(self.label10)
        box9.setLayout(layout9)





        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        container.addWidget(box3)
        container.addWidget(box5)
        container.addWidget(box7)
        container.addWidget(box9)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

    # 勾选事件
    def checkBoxChanged(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已启用实时调教功能")


class Widget_tune_sakura(QFrame):# sakura调教界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


         # 最外层的垂直布局
        container = QVBoxLayout()


        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        #设置“启用实时参数”标签
        label0 = QLabel(flags=Qt.WindowFlags())  
        label0.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label0.setText("实时改变AI参数")

        #设置官方文档说明链接按钮
        hyperlinkButton = HyperlinkButton(
            url='https://github.com/SakuraLLM/Sakura-13B-Galgame',
            text='(Github主页)'
        )

        #设置“启用实时参数”开关
        self.checkBox = CheckBox('启用', self)
        #self.checkBox.stateChanged.connect(self.checkBoxChanged)


        layout1.addWidget(label0)
        layout1.addWidget(hyperlinkButton)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox)
        box1.setLayout(layout1)



        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“温度”标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label1.setText("Temperature")

        #设置“温度”副标签
        label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label11.setText("(官方默认值为0.1)")

        #设置“温度”滑动条
        self.slider1 = Slider(Qt.Horizontal, self)
        self.slider1.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label2 = QLabel(str(self.slider1.value()), self)
        self.label2.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider1.valueChanged.connect(lambda value: self.label2.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label2显示正确的值
        self.slider1.setMinimum(0)
        self.slider1.setMaximum(10)
        self.slider1.setValue(1)

        

        layout3.addWidget(label1)
        layout3.addWidget(label11)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.slider1)
        layout3.addWidget(self.label2)
        box3.setLayout(layout3)






        # -----创建第5个组，添加多个组件-----
        box5 = QGroupBox()
        box5.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout5 = QHBoxLayout()

        #设置“top_p”标签
        label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label4.setText("Top_p")

        #设置“top_p”副标签
        label41 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label41.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label41.setText("(官方默认值为0.3)")


        #设置“top_p”滑动条
        self.slider2 = Slider(Qt.Horizontal, self)
        self.slider2.setFixedWidth(200)


        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label5 = QLabel(str(self.slider2.value()), self)
        self.label5.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label5.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider2.valueChanged.connect(lambda value: self.label5.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放在后面是为了让上面的label5显示正确的值和格式
        self.slider2.setMinimum(0)
        self.slider2.setMaximum(10)
        self.slider2.setValue(3)



        layout5.addWidget(label4)
        layout5.addWidget(label41)
        layout5.addStretch(1)  # 添加伸缩项
        layout5.addWidget(self.slider2)
        layout5.addWidget(self.label5)
        box5.setLayout(layout5)








        # -----创建第9个组，添加多个组件-----
        box9 = QGroupBox()
        box9.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout9 = QHBoxLayout()

        #设置“frequency_penalty”标签
        label9 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label9.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label9.setText("Frequency_penalty")

        #设置“presence_penalty”副标签
        label91 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label91.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label91.setText("(官方默认值为0)")

        #设置“frequency_penalty”滑动条
        self.slider4 = Slider(Qt.Horizontal, self)
        self.slider4.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label10 = QLabel(str(self.slider4.value()), self)
        self.label10.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.label10.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.slider4.valueChanged.connect(lambda value: self.label10.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label10显示正确的值和格式
        self.slider4.setMinimum(-10)
        self.slider4.setMaximum(10)
        self.slider4.setValue(0)


        layout9.addWidget(label9)
        layout9.addWidget(label91)
        layout9.addStretch(1)  # 添加伸缩项
        layout9.addWidget(self.slider4)
        layout9.addWidget(self.label10)
        box9.setLayout(layout9)





        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        container.addWidget(box3)
        container.addWidget(box5)
        container.addWidget(box9)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

    # 勾选事件
    def checkBoxChanged(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已启用Sakura实时调教功能")


class Widget_tune_anthropic(QFrame):# anthropic调教界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        # 最外层的垂直布局
        container = QVBoxLayout()


        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        #设置“启用实时参数”标签
        label0 = QLabel(flags=Qt.WindowFlags())  
        label0.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label0.setText("实时改变AI参数")

        #设置官方文档说明链接按钮
        hyperlinkButton = HyperlinkButton(
            url='https://docs.anthropic.com/en/api/messages',
            text='(官方文档)'
        )

        #设置“启用实时参数”开关
        self.checkBox = CheckBox('启用', self)
        #self.checkBox.stateChanged.connect(self.checkBoxChanged)


        layout1.addWidget(label0)
        layout1.addWidget(hyperlinkButton)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox)
        box1.setLayout(layout1)




        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“温度”标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label1.setText("Temperature")

        #设置“温度”副标签
        label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label11.setText("(官方默认值为0)")

        #设置“温度”滑动条
        self.slider1 = Slider(Qt.Horizontal, self)
        self.slider1.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label2 = QLabel(str(self.slider1.value()), self)
        self.label2.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider1.valueChanged.connect(lambda value: self.label2.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label2显示正确的值
        self.slider1.setMinimum(0)
        self.slider1.setMaximum(10)
        self.slider1.setValue(0)

        

        layout3.addWidget(label1)
        layout3.addWidget(label11)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.slider1)
        layout3.addWidget(self.label2)
        box3.setLayout(layout3)



        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        container.addWidget(box3)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


class Widget_tune_google(QFrame):# google调教界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用


        # 最外层的垂直布局
        container = QVBoxLayout()


        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        #设置“启用实时参数”标签
        label0 = QLabel(flags=Qt.WindowFlags())  
        label0.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label0.setText("实时改变AI参数")

        #设置官方文档说明链接按钮
        hyperlinkButton = HyperlinkButton(
            url='https://ai.google.dev/gemini-api/docs/models/generative-models?hl=zh-cn',
            text='(官方文档)'
        )

        #设置“启用实时参数”开关
        self.checkBox = CheckBox('启用', self)
        #self.checkBox.stateChanged.connect(self.checkBoxChanged)


        layout1.addWidget(label0)
        layout1.addWidget(hyperlinkButton)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox)
        box1.setLayout(layout1)




        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“温度”标签
        label1 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label1.setText("Temperature")

        #设置“温度”副标签
        label11 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label11.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 10px;  color: black")
        label11.setText("(官方默认值为0)")

        #设置“温度”滑动条
        self.slider1 = Slider(Qt.Horizontal, self)
        self.slider1.setFixedWidth(200)

        # 创建一个QLabel控件，并设置初始文本为滑动条的初始值,并实时更新
        self.label2 = QLabel(str(self.slider1.value()), self)
        self.label2.setFixedSize(100, 15)  # 设置标签框的大小，不然会显示不全
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 12px;  color: black")
        self.slider1.valueChanged.connect(lambda value: self.label2.setText(str("{:.1f}".format(value * 0.1))))

        #设置滑动条的最小值、最大值、当前值，放到后面是为了让上面的label2显示正确的值
        self.slider1.setMinimum(0)
        self.slider1.setMaximum(10)
        self.slider1.setValue(0)

        

        layout3.addWidget(label1)
        layout3.addWidget(label11)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.slider1)
        layout3.addWidget(self.label2)
        box3.setLayout(layout3)



        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        container.addWidget(box3)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下




class Widget_replace_dict(QFrame):  # 替换字典主界面
    def __init__(self, text: str, parent=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用


        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_before_dict('A_settings', self)  # 创建实例，指向界面
        self.B_settings = Widget_after_dict('B_settings', self)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '译前替换')
        self.addSubInterface(self.B_settings, 'B_settings', '译后替换')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_before_dict(QFrame):# 原文替换字典界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # self.scrollWidget = QWidget() #创建滚动窗口
        # #self.scrollWidget.resize(500, 400)    #设置滚动窗口大小
        # #self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff) #设置水平滚动条不可见
        # self.setViewportMargins(0, 0, 0, 0)   #设置滚动窗口的边距      
        # self.setWidget(self.scrollWidget)  #设置滚动窗口的内容  
        # self.setWidgetResizable(True)   #设置滚动窗口的内容可调整大小
        # self.verticalScrollBar().sliderPressed.connect(self.scrollContents) #滚动条滚动时，调用的函数
        #设置各个控件-----------------------------------------------------------------------------------------

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['Src', 'Dst']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序


        # songInfos = [
        #     ['かばん', 'aiko']
        # ]
        # for i, songInfo in enumerate(songInfos): #遍历数据
        #     for j in range(2): #遍历每一列
        #         self.tableView.setItem(i, j, QTableWidgetItem(songInfo[j])) #设置每个单元格的内容


        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        #设置“译前替换”标签
        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("原文替换")

        #设置“译前替换”显示
        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(翻译前，将根据字典内容对原文文本进行替换)")


        #设置“译前替换”开
        self.checkBox1 = CheckBox('启用功能')
        self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout2.addWidget(label1)
        layout2.addWidget(self.label2)
        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.checkBox1)
        box2.setLayout(layout2)




        # 把内容添加到容器中 
        container.addWidget(box2)   
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下 

    #滚动条滚动时，调用的函数
    def scrollContents(self, position):
        self.scrollWidget.move(0, position) 

    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    # 移除JSON内容中的行内注释
    def remove_inline_comments(self, json_content):
        # 正则表达式匹配行内注释并替换为空字符串
        return re.sub(r'//.*$', '', json_content, flags=re.MULTILINE)

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'JSON Files (*.json)')      #调用QFileDialog类里的函数来选择文件
        if Input_File:
            print(f'[INFO]  已选择字典导入文件: {Input_File}')
        else :
            print('[INFO]  未选择文件')
            return
        
        try:
            # 尝试读取文件内容
            with open(Input_File, 'r', encoding="utf-8") as f:
                content = f.read()
        except FileNotFoundError:
            print(f'[ERROR] 文件未找到: {Input_File}')
            return
        except Exception as e:
            print(f'[ERROR] 读取文件时发生未知错误: {str(e)}')
            return
        try:
            # 移除内容中的行内注释，并反序列化
            dictionary = json.loads(self.remove_inline_comments(content))
        except json.JSONDecodeError as e:
            print(f'[ERROR] JSON解析错误: {str(e)}')
            return
        except Exception as e:
            print(f'[ERROR] 反序列化时发生未知错误: {str(e)}')
            return       

        # 将字典中的数据从表格底部添加到表格中
        for key, value in dictionary.items():
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(key))
            self.tableView.setItem(row, 1, QTableWidgetItem(value))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

        user_interface_prompter.createSuccessInfoBar("导入成功")
        print(f'[INFO]  已导入字典文件')

        # 导入成功后删除空白行
        self.delete_blank_row()
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户译前替换字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        configurator.read_write_config("write") 
        user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')


    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启译前替换功能，将依据表格内容进行替换")
    

class Widget_after_dict(QFrame):# 译文修正字典界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # self.scrollWidget = QWidget() #创建滚动窗口
        # #self.scrollWidget.resize(500, 400)    #设置滚动窗口大小
        # #self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff) #设置水平滚动条不可见
        # self.setViewportMargins(0, 0, 0, 0)   #设置滚动窗口的边距      
        # self.setWidget(self.scrollWidget)  #设置滚动窗口的内容  
        # self.setWidgetResizable(True)   #设置滚动窗口的内容可调整大小
        # self.verticalScrollBar().sliderPressed.connect(self.scrollContents) #滚动条滚动时，调用的函数
        #设置各个控件-----------------------------------------------------------------------------------------

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['Src', 'Dst']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序


        # songInfos = [
        #     ['かばん', 'aiko']
        # ]
        # for i, songInfo in enumerate(songInfos): #遍历数据
        #     for j in range(2): #遍历每一列
        #         self.tableView.setItem(i, j, QTableWidgetItem(songInfo[j])) #设置每个单元格的内容


        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        #设置“译前替换”标签
        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("译文修正")

        #设置“译前替换”显示
        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(翻译完成后，根据字典内容对译文文本进行替换)")


        #设置“译前替换”开
        self.checkBox1 = CheckBox('启用功能')
        self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout2.addWidget(label1)
        layout2.addWidget(self.label2)
        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.checkBox1)
        box2.setLayout(layout2)




        # 把内容添加到容器中 
        container.addWidget(box2)   
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下    

    #滚动条滚动时，调用的函数
    def scrollContents(self, position):
        self.scrollWidget.move(0, position) 

    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    # 移除JSON内容中的行内注释
    def remove_inline_comments(self, json_content):
        # 正则表达式匹配行内注释并替换为空字符串
        return re.sub(r'//.*$', '', json_content, flags=re.MULTILINE)

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'JSON Files (*.json)')      #调用QFileDialog类里的函数来选择文件
        if Input_File:
            print(f'[INFO]  已选择字典导入文件: {Input_File}')
        else :
            print('[INFO]  未选择文件')
            return
        
        try:
            # 尝试读取文件内容
            with open(Input_File, 'r', encoding="utf-8") as f:
                content = f.read()
        except FileNotFoundError:
            print(f'[ERROR] 文件未找到: {Input_File}')
            return
        except Exception as e:
            print(f'[ERROR] 读取文件时发生未知错误: {str(e)}')
            return
        try:
            # 移除内容中的行内注释，并反序列化
            dictionary = json.loads(self.remove_inline_comments(content))
        except json.JSONDecodeError as e:
            print(f'[ERROR] JSON解析错误: {str(e)}')
            return
        except Exception as e:
            print(f'[ERROR] 反序列化时发生未知错误: {str(e)}')
            return
        
        # 将字典中的数据从表格底部添加到表格中
        for key, value in dictionary.items():
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(key))
            self.tableView.setItem(row, 1, QTableWidgetItem(value))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

        user_interface_prompter.createSuccessInfoBar("导入成功")
        print(f'[INFO]  已导入字典文件')

        # 导入成功后删除空白行
        self.delete_blank_row()
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户译后修正字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        configurator.read_write_config("write") 
        user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')


    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启译后修正功能，将依据表格内容进行修正")
    



class Widget_rulebook(QFrame):# 提示书界面

    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.hBoxLayout = QHBoxLayout(self)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))


class Widget_system_prompt(QFrame): # 基础提示页面


    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("修改系统提示词")


        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(全程修改)")


        self.checkBox1 = CheckBox('启用功能')
        #self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout1.addWidget(label1)
        layout1.addWidget(self.label2)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox1)
        box1.setLayout(layout1)


        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()


        Prompt = f'''None
'''


        self.TextEdit1 = TextEdit()
        #设置输入框最小高度
        self.TextEdit1.setMinimumHeight(180)
        #设置默认文本
        self.TextEdit1.setText(Prompt)


        layout2.addWidget(self.TextEdit1)
        box2.setLayout(layout2)

        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(20) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addWidget(box1)
        container.addWidget(box2)


class Widget_prompt_dict(QFrame): # 术语字典界面


    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(3) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['原文', '译文', '备注']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序

        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第三列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 2, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“译时提示”标签
        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("添加提示字典")

        #设置“译时提示”显示
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(原文触发，自动构建术语表)")


        #设置“译时提示”开
        self.checkBox2 = CheckBox('启用功能')
        self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox2)
        box3.setLayout(layout3)


        # 把内容添加到容器中
        container.addWidget(box3)    
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(20)     
        container.setContentsMargins(50, 70, 50, 30)      


    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    # 将条目添加到表格的辅助函数
    def add_to_table(self, srt, dst, info):
            row = self.tableView.rowCount() - 1 #获取表格的倒数行数
            self.tableView.insertRow(row)    # 在表格中插入一行
            self.tableView.setItem(row, 0, QTableWidgetItem(srt))
            self.tableView.setItem(row, 1, QTableWidgetItem(dst))
            if info:
                self.tableView.setItem(row, 2, QTableWidgetItem(info))
            #设置新行的高度与前一行相同
            self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'JSON Files (*.json)')      #调用QFileDialog类里的函数来选择文件
        if Input_File:
            print(f'[INFO]  已选择字典导入文件: {Input_File}')
        else :
            print('[INFO]  未选择文件')
            return
        
        # 读取文件
        with open(Input_File, 'r', encoding="utf-8") as f:
            dictionary = json.load(f)

        # 检查数据是列表还是字典
        if isinstance(dictionary, list):  # 如果是列表
            for item in dictionary:
                if item.get("srt", "") and item.get("dst", ""):
                    srt = item.get("srt", "")
                    dst = item.get("dst", "")
                    info = item.get("info", "")

                    self.add_to_table(srt, dst,info)
                    # 格式例
                    # [
                    #   {
                    #     "srt": "xxxx",
                    #     "dst": "xxxx",
                    #     "info": "xxx",
                    #   }
                    # ]
                else: # 代表是Paratranz的术语表，处理每一个字典项
                    key = item.get("term", "")
                    value = item.get("translation", "")
                    info = ""
                    self.add_to_table(key, value,info)
                    # 格式例
                    # [
                    #   {
                    #     "id": 359894,
                    #     "createdAt": "2024-04-06T18:43:56.075Z",
                    #     "updatedAt": "2024-04-06T18:43:56.075Z",
                    #     "updatedBy": null,
                    #     "pos": "noun",
                    #     "uid": 49900,
                    #     "term": "アイテム",
                    #     "translation": "道具",
                    #     "note": "",
                    #     "project": 9841,
                    #     "variants": []
                    #   }
                    # ]
        elif isinstance(dictionary, dict):  # 如果是字典，处理字典键值对
            for key, value in dictionary.items():
                info = ""
                self.add_to_table(key, value,info)
        else:
            print('[ERROR]  不支持的文件格式')
            return

        user_interface_prompter.createSuccessInfoBar("导入成功")
        print(f'[INFO]  已导入字典文件')
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        dictionary = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            info_item = self.tableView.item(row, 2)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                if info_item:
                    info = info_item.text()
                    dictionary.append({"srt":key,"dst":value,"info":info})
                else:
                    dictionary.append({"srt":key,"dst":value})


        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户提示字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第三列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 2, button)
        button.clicked.connect(self.delete_blank_row)

        user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        configurator.read_write_config("write") 
        user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')

    
    #消息提示函数
    def checkBoxChanged2(self, isChecked: bool):
        if isChecked :
            user_interface_prompter.createSuccessInfoBar("已开启译时提示功能,将根据发送文本自动添加翻译示例")


class Widget_characterization(QFrame): # 角色介绍页面

    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(7) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['原名', '译名', '性别', '年龄', '性格', '说话风格', '补充']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序

        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第三列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 6, button)
        button.clicked.connect(self.delete_blank_row)


        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“译时提示”标签
        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("添加角色设定")

        #设置“译时提示”显示
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(文本中出现了角色原名触发)")


        #设置“译时提示”开
        self.checkBox1 = CheckBox('启用功能')
        #self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox1)
        box3.setLayout(layout3)


        # 把内容添加到容器中
        container.addWidget(box3)    
        container.addWidget(self.tableView)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(20)     
        container.setContentsMargins(50, 70, 50, 30)      


    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        # 从倒数第二行开始反向遍历表格，因为最后一行是按钮行
        for i in range(self.tableView.rowCount() - 2, -1, -1):
            # 检查第一列和第二列是否为空
            item1 = self.tableView.item(i, 0)
            item2 = self.tableView.item(i, 1)
            if (item1 is None or item1.text() == '') and (item2 is None or item2.text() == ''):
                self.tableView.removeRow(i)


class Widget_world_building(QFrame): # 背景设定界面

    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("添加背景设定")


        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(全程加入)")


        self.checkBox1 = CheckBox('启用功能')
        #self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout1.addWidget(label1)
        layout1.addWidget(self.label2)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox1)
        box1.setLayout(layout1)


        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()


        Prompt = f'''None
'''


        self.TextEdit1 = TextEdit()
        #设置输入框最小高度
        self.TextEdit1.setMinimumHeight(180)
        #设置默认文本
        self.TextEdit1.setText(Prompt)


        layout2.addWidget(self.TextEdit1)
        box2.setLayout(layout2)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(20) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addWidget(box1)
        container.addWidget(box2)


class Widget_writing_style(QFrame): # 写作风格界面

    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("添加文风设定")


        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(全程加入)")


        self.checkBox1 = CheckBox('启用功能')
        #self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout1.addWidget(label1)
        layout1.addWidget(self.label2)
        layout1.addStretch(1)  # 添加伸缩项
        layout1.addWidget(self.checkBox1)
        box1.setLayout(layout1)


        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()


        Prompt = f'''None
'''


        self.TextEdit1 = TextEdit()
        #设置输入框最小高度
        self.TextEdit1.setMinimumHeight(180)
        #设置默认文本
        self.TextEdit1.setText(Prompt)


        layout2.addWidget(self.TextEdit1)
        box2.setLayout(layout2)



        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(20) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addWidget(box1)
        container.addWidget(box2)


class Widget_translation_example(QFrame): # 翻译示例界面


    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['原文', '译文']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序

        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第三列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)


        # -----创建第3个组，添加多个组件-----
        box3 = QGroupBox()
        box3.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout3 = QHBoxLayout()

        #设置“译时提示”标签
        label3 = QLabel( flags=Qt.WindowFlags())  
        label3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label3.setText("添加翻译示例")

        #设置“译时提示”显示
        self.label4 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label4.setText("(全程加入)")


        #设置“译时提示”开
        self.checkBox1 = CheckBox('启用功能')
        #self.checkBox2.stateChanged.connect(self.checkBoxChanged2)

        layout3.addWidget(label3)
        layout3.addWidget(self.label4)
        layout3.addStretch(1)  # 添加伸缩项
        layout3.addWidget(self.checkBox1)
        box3.setLayout(layout3)


        # 把内容添加到容器中
        container.addWidget(box3)    
        container.addWidget(self.tableView)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(20)     
        container.setContentsMargins(50, 70, 50, 30)      


    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break





class Widget_RPG(QFrame):  # RPG主界面

    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.hBoxLayout = QHBoxLayout(self)
        self.hBoxLayout.addWidget(self.label, 1, Qt.AlignCenter)
        self.setObjectName(text.replace(' ', '-'))


class Widget_export_source_text(QFrame):#  提取子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box = QGroupBox()
        box.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout = QHBoxLayout()


        self.labe1_3 = QLabel(flags=Qt.WindowFlags())  
        self.labe1_3.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        self.labe1_3.setText("RPG Maker MV/MZ 的文本提取注入工具")


        #设置官方文档说明链接按钮
        hyperlinkButton = HyperlinkButton(
            url='https://www.ai2moe.org/topic/10271-jt%EF%BC%8C%E7%9B%AE%E6%A0%87%E6%98%AF%E9%9B%B6%E9%97%A8%E6%A7%9B%E7%9A%84%EF%BC%8C%E5%86%85%E5%B5%8C%E4%BA%86%E5%A4%9A%E4%B8%AA%E8%84%9A%E6%9C%AC%E7%9A%84%E9%9D%92%E6%98%A5%E7%89%88t/',
            text='(作者页面)'
        )


        layout.addStretch(1)  # 添加伸缩项
        layout.addWidget(self.labe1_3)
        layout.addWidget(hyperlinkButton)
        layout.addStretch(1)  # 添加伸缩项
        box.setLayout(layout)




        # -----创建第1个组，添加多个组件-----
        box_switch = QGroupBox()
        box_switch.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_switch = QHBoxLayout()

        #设置“是否日语游戏”标签
        self.labe1_4 = QLabel(flags=Qt.WindowFlags())  
        self.labe1_4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        self.labe1_4.setText("是否日语游戏")



        # 设置“是否日语游戏”选择开关
        self.SwitchButton_ja = CheckBox('        ')
        self.SwitchButton_ja.setChecked(True)    
        # 绑定选择开关的点击事件
        self.SwitchButton_ja.clicked.connect(self.test)



        layout_switch.addWidget(self.labe1_4)
        layout_switch.addStretch(1)  # 添加伸缩项
        layout_switch.addWidget(self.SwitchButton_ja)
        box_switch.setLayout(layout_switch)



        # -----创建第2个组，添加多个组件-----
        box_input = QGroupBox()
        box_input.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input = QHBoxLayout()

        #设置“游戏文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("游戏文件夹")

        #设置“游戏文件夹”显示
        self.label_input_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_input_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_input_path.setText("(游戏根目录文件夹)")  

        #设置打开文件按钮
        self.pushButton_input = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_input.clicked.connect(self.Select_project_folder) #按钮绑定槽函数



        layout_input.addWidget(label4)
        layout_input.addWidget(self.label_input_path)
        layout_input.addStretch(1)  # 添加伸缩项
        layout_input.addWidget(self.pushButton_input)
        box_input.setLayout(layout_input)



        # -----创建第3个组，添加多个组件-----
        box_output = QGroupBox()
        box_output.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_output = QHBoxLayout()

        #设置“输出文件夹”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("原文存储文件夹")

        #设置“输出文件夹”显示
        self.label_output_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_output_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_output_path.setText("(游戏原文提取后存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_output = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_output.clicked.connect(self.Select_output_folder) #按钮绑定槽函数


        layout_output.addWidget(label6)
        layout_output.addWidget(self.label_output_path)
        layout_output.addStretch(1)  # 添加伸缩项
        layout_output.addWidget(self.pushButton_output)
        box_output.setLayout(layout_output)



        # -----创建第3个组，添加多个组件-----
        box_data = QGroupBox()
        box_data.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_data = QHBoxLayout()

        #设置“输出文件夹”标签
        label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        label6.setText("工程存储文件夹")

        #设置“输出文件夹”显示
        self.label_data_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_data_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_data_path.setText("(该游戏工程数据存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_data = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_data.clicked.connect(self.Select_data_folder) #按钮绑定槽函数


        layout_data.addWidget(label6)
        layout_data.addWidget(self.label_data_path)
        layout_data.addStretch(1)  # 添加伸缩项
        layout_data.addWidget(self.pushButton_data)
        box_data.setLayout(layout_data)





        # -----创建第x个组，添加多个组件-----
        box_start_export = QGroupBox()
        box_start_export.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_start_export = QHBoxLayout()


        #设置“开始翻译”的按钮
        self.primaryButton_start_export = PrimaryPushButton('开始提取原文', self, FIF.UPDATE)
        self.primaryButton_start_export.clicked.connect(self.Start_export) #按钮绑定槽函数


        layout_start_export.addStretch(1)  # 添加伸缩项
        layout_start_export.addWidget(self.primaryButton_start_export)
        layout_start_export.addStretch(1)  # 添加伸缩项
        box_start_export.setLayout(layout_start_export)



        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box)
        container.addWidget(box_switch)
        container.addWidget(box_input)
        container.addWidget(box_output)
        container.addWidget(box_data)
        container.addWidget(box_start_export)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

    #设置开关绑定函数
    def test(self, isChecked: bool):
        if isChecked== False:
            user_interface_prompter.createWarningInfoBar("不建议使用在非日语游戏上,容易出现问题")

    # 选择输入文件夹按钮绑定函数
    def Select_project_folder(self):
        Input_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            self.label_input_path.setText(Input_Folder)
            print('[INFO]  已选择游戏根目录文件夹: ',Input_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 选择原文文件夹按钮绑定函数
    def Select_output_folder(self):
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            self.label_output_path.setText(Output_Folder)
            print('[INFO]  已选择原文存储文件夹:' ,Output_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 选择工程文件夹按钮绑定函数
    def Select_data_folder(self):
        data_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if data_Folder:
            self.label_data_path.setText(data_Folder)
            print('[INFO]  已选择工程存储文件夹:' ,data_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作


    # 提取函数
    def Start_export(self):
        print('[INFO]  开始提取游戏原文,请耐心等待！！！')

        #读取配置文件
        config_path = os.path.join(script_dir, "StevExtraction", "config.yaml")
        with open(config_path, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)

        #修改输入输出路径及开关
        config['game_path'] = self.label_input_path.text()
        config['save_path'] = self.label_data_path.text()
        config['data_path'] = self.label_output_path.text()
        config['ja']=self.SwitchButton_ja.isChecked()
        #提取文本
        pj=jtpp.Jr_Tpp(config)
        pj.FromGame(config['game_path'],config['save_path'],config['data_path'])


class Widget_import_translated_text(QFrame):#  导入子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_input = QGroupBox()
        box_input.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("游戏文件夹")

        #设置“输入文件夹”显示
        self.label_input_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_input_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_input_path.setText("(原来的游戏根目录文件夹)")  

        #设置打开文件按钮
        self.pushButton_input = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_input.clicked.connect(self.Select_game_folder) #按钮绑定槽函数



        layout_input.addWidget(label4)
        layout_input.addWidget(self.label_input_path)
        layout_input.addStretch(1)  # 添加伸缩项
        layout_input.addWidget(self.pushButton_input)
        box_input.setLayout(layout_input)



        # -----创建第2个组，添加多个组件-----
        box_data = QGroupBox()
        box_data.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_data = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("工程文件夹")

        #设置“输入文件夹”显示
        self.label_data_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_data_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_data_path.setText("(原来导出的工程数据文件夹)")  

        #设置打开文件按钮
        self.pushButton_data = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_data.clicked.connect(self.Select_data_folder) #按钮绑定槽函数



        layout_data.addWidget(label4)
        layout_data.addWidget(self.label_data_path)
        layout_data.addStretch(1)  # 添加伸缩项
        layout_data.addWidget(self.pushButton_data)
        box_data.setLayout(layout_data)



        # -----创建第3个组，添加多个组件-----
        box_translation_folder = QGroupBox()
        box_translation_folder.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_folder = QHBoxLayout()

        #设置“输出文件夹”标签
        self.label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label6.setText("译文文件夹")

        #设置“输出文件夹”显示
        self.label_translation_folder = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_translation_folder.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_translation_folder.setText("(译文文件存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_translation_folder = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_translation_folder.clicked.connect(self.Select_translation_folder) #按钮绑定槽函数


        layout_translation_folder.addWidget(self.label6)
        layout_translation_folder.addWidget(self.label_translation_folder)
        layout_translation_folder.addStretch(1)  # 添加伸缩项
        layout_translation_folder.addWidget(self.pushButton_translation_folder)
        box_translation_folder.setLayout(layout_translation_folder)


        # -----创建第4个组，添加多个组件-----
        box_output_folder = QGroupBox()
        box_output_folder.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_putput_folder = QHBoxLayout()

        #设置“输出文件夹”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("存储文件夹")

        #设置“输出文件夹”显示
        self.label_output_folder = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_output_folder.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_output_folder.setText("(游戏文件注入译文后，存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_putput_folder = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_putput_folder.clicked.connect(self.Select_save_folder) #按钮绑定槽函数


        layout_putput_folder.addWidget(self.label7)
        layout_putput_folder.addWidget(self.label_output_folder)
        layout_putput_folder.addStretch(1)  # 添加伸缩项
        layout_putput_folder.addWidget(self.pushButton_putput_folder)
        box_output_folder.setLayout(layout_putput_folder)



        # -----创建第5个组，添加多个组件-----
        box_title_watermark1 = QGroupBox()
        box_title_watermark1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_title_watermark1 = QHBoxLayout()


        self.LineEdit_title_watermark = LineEdit()

        #设置微调距离用的空白标签
        self.labelB = QLabel()  
        self.labelB.setText("          ")

        # 设置“添加游戏标题水印”选择开关
        self.checkBox_title_watermark = CheckBox('添加标题水印', self)



        layout_title_watermark1.addWidget(self.LineEdit_title_watermark)
        layout_title_watermark1.addWidget(self.labelB)
        layout_title_watermark1.addWidget(self.checkBox_title_watermark)
        box_title_watermark1.setLayout(layout_title_watermark1)





        # -----创建第5个组，添加多个组件-----
        box_auto_wrap = QGroupBox()
        box_auto_wrap.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_auto_wrap = QHBoxLayout()

        #设置标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("换行字数")

        self.spinBox_auto_wrap = SpinBox(self)
        self.spinBox_auto_wrap.setRange(0, 1000)    
        self.spinBox_auto_wrap.setValue(0)


        # 设置“添加游戏标题水印”选择开关
        self.checkBox_auto_wrap = CheckBox('启用自动换行', self)


        layout_auto_wrap.addWidget(label4)
        layout_auto_wrap.addWidget(self.spinBox_auto_wrap)
        layout_auto_wrap.addStretch(1)
        layout_auto_wrap.addWidget(self.checkBox_auto_wrap)
        box_auto_wrap.setLayout(layout_auto_wrap)


        # -----创建第x个组，添加多个组件-----
        box_start_import = QGroupBox()
        box_start_import.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_start_import = QHBoxLayout()


        #设置“开始翻译”的按钮
        self.primaryButton_start_import = PrimaryPushButton('开始注入译文', self, FIF.UPDATE)
        self.primaryButton_start_import.clicked.connect(self.Start_import) #按钮绑定槽函数


        layout_start_import.addStretch(1)  # 添加伸缩项
        layout_start_import.addWidget(self.primaryButton_start_import)
        layout_start_import.addStretch(1)  # 添加伸缩项
        box_start_import.setLayout(layout_start_import)



        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_input)
        container.addWidget(box_data)
        container.addWidget(box_translation_folder)
        container.addWidget(box_output_folder)
        container.addWidget(box_title_watermark1)
        container.addWidget(box_auto_wrap)
        container.addWidget(box_start_import)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


    # 选择输入文件夹按钮绑定函数
    def Select_game_folder(self):
        Input_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            self.label_input_path.setText(Input_Folder)
            print('[INFO]  已选择原游戏文件夹: ',Input_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作
        
    # 选择工程文件夹按钮绑定函数
    def Select_data_folder(self):
        Data_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Data_Folder:
            self.label_data_path.setText(Data_Folder)
            print('[INFO]  已选择工程数据文件夹: ',Data_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

    # 选择译文文件夹按钮绑定函数
    def Select_translation_folder(self):
        translation_folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if translation_folder:
            self.label_translation_folder.setText(translation_folder)
            print('[INFO]  已选择译文文件夹:' ,translation_folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作
        
    # 选择存储文件夹按钮绑定函数
    def Select_save_folder(self):
        save_folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if save_folder:
            self.label_output_folder.setText(save_folder)
            print('[INFO]  已选择注入后存储文件夹:' ,save_folder)
        else :
            print('[INFO]  未选择文件夹')

    
    # 导入按钮绑定函数
    def Start_import(self):
        print('[INFO]  开始注入译文到游戏文件中,请耐心等待！！！')

        #读取配置文件
        config_path = os.path.join(script_dir, "StevExtraction", "config.yaml")
        with open(config_path, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)

        #修改配置信息
        config['game_path'] = self.label_input_path.text()
        config['save_path'] = self.label_data_path.text()
        config['translation_path'] = self.label_translation_folder.text()
        config['output_path'] = self.label_output_folder.text()

        if self.checkBox_title_watermark.isChecked():
            config['mark'] = self.LineEdit_title_watermark.text()
        else:
            config['mark'] = 0

        if self.checkBox_auto_wrap.isChecked():
            config['line_length'] = self.spinBox_auto_wrap.value()
        else:
            config['line_length'] = 0

        #导入文本
        pj=jtpp.Jr_Tpp(config,config['save_path'])
        pj.ToGame(config['game_path'],config['translation_path'],config['output_path'],config['mark'])
        

class Widget_update_text(QFrame):#  更新子界面
    def __init__(self, text: str, parent=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        #设置各个控件-----------------------------------------------------------------------------------------



        # -----创建第1个组，添加多个组件-----
        box_input = QGroupBox()
        box_input.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_input = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("游戏文件夹")

        #设置“输入文件夹”显示
        self.label_input_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_input_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_input_path.setText("(新版本的游戏根目录文件夹)")  

        #设置打开文件按钮
        self.pushButton_input = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_input.clicked.connect(self.Select_game_folder) #按钮绑定槽函数



        layout_input.addWidget(label4)
        layout_input.addWidget(self.label_input_path)
        layout_input.addStretch(1)  # 添加伸缩项
        layout_input.addWidget(self.pushButton_input)
        box_input.setLayout(layout_input)



        # -----创建第2个组，添加多个组件-----
        box_data = QGroupBox()
        box_data.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_data = QHBoxLayout()

        #设置“输入文件夹”标签
        label4 = QLabel(flags=Qt.WindowFlags())  
        label4.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px")
        label4.setText("工程文件夹")

        #设置“输入文件夹”显示
        self.label_data_path = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_data_path.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px")
        self.label_data_path.setText("(新版本游戏导出的工程数据文件夹)")  

        #设置打开文件按钮
        self.pushButton_data = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_data.clicked.connect(self.Select_data_folder) #按钮绑定槽函数



        layout_data.addWidget(label4)
        layout_data.addWidget(self.label_data_path)
        layout_data.addStretch(1)  # 添加伸缩项
        layout_data.addWidget(self.pushButton_data)
        box_data.setLayout(layout_data)



        # -----创建第3个组，添加多个组件-----
        box_translation_folder = QGroupBox()
        box_translation_folder.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_translation_folder = QHBoxLayout()

        #设置“输出文件夹”标签
        self.label6 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label6.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label6.setText("译文文件夹")

        #设置“输出文件夹”显示
        self.label_translation_folder = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_translation_folder.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_translation_folder.setText("(旧版本游戏的译文文件存放的文件夹)")

        #设置输出文件夹按钮
        self.pushButton_translation_folder = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_translation_folder.clicked.connect(self.Select_translation_folder) #按钮绑定槽函数


        layout_translation_folder.addWidget(self.label6)
        layout_translation_folder.addWidget(self.label_translation_folder)
        layout_translation_folder.addStretch(1)  # 添加伸缩项
        layout_translation_folder.addWidget(self.pushButton_translation_folder)
        box_translation_folder.setLayout(layout_translation_folder)


        # -----创建第4个组，添加多个组件-----
        box_output_folder = QGroupBox()
        box_output_folder.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_putput_folder = QHBoxLayout()

        #设置“输出文件夹”标签
        self.label7 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label7.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;  color: black")
        self.label7.setText("保存文件夹")

        #设置“输出文件夹”显示
        self.label_output_folder = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label_output_folder.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label_output_folder.setText("(新版游戏提取到的原文与旧版译文合并后，剩下的需要翻译的原文保存路径)")

        #设置输出文件夹按钮
        self.pushButton_putput_folder = PushButton('选择文件夹', self, FIF.FOLDER)
        self.pushButton_putput_folder.clicked.connect(self.Select_save_folder) #按钮绑定槽函数


        layout_putput_folder.addWidget(self.label7)
        layout_putput_folder.addWidget(self.label_output_folder)
        layout_putput_folder.addStretch(1)  # 添加伸缩项
        layout_putput_folder.addWidget(self.pushButton_putput_folder)
        box_output_folder.setLayout(layout_putput_folder)





        # -----创建第x个组，添加多个组件-----
        box_start_import = QGroupBox()
        box_start_import.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout_start_import = QHBoxLayout()


        #设置“开始翻译”的按钮
        self.primaryButton_start_import = PrimaryPushButton('开始提取原文', self, FIF.UPDATE)
        self.primaryButton_start_import.clicked.connect(self.Start_import) #按钮绑定槽函数


        layout_start_import.addStretch(1)  # 添加伸缩项
        layout_start_import.addWidget(self.primaryButton_start_import)
        layout_start_import.addStretch(1)  # 添加伸缩项
        box_start_import.setLayout(layout_start_import)



        # 最外层的垂直布局
        container = QVBoxLayout()

        # 把内容添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box_input)
        container.addWidget(box_data)
        container.addWidget(box_translation_folder)
        container.addWidget(box_output_folder)
        container.addWidget(box_start_import)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下


    # 选择输入文件夹按钮绑定函数
    def Select_game_folder(self):
        Input_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Input_Folder:
            self.label_input_path.setText(Input_Folder)
            print('[INFO]  已选择新版游戏文件夹: ',Input_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作
        
    # 选择工程文件夹按钮绑定函数
    def Select_data_folder(self):
        Data_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Data_Folder:
            self.label_data_path.setText(Data_Folder)
            print('[INFO]  已选择新版游戏工程数据文件夹: ',Data_Folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

    # 选择译文文件夹按钮绑定函数
    def Select_translation_folder(self):
        translation_folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if translation_folder:
            self.label_translation_folder.setText(translation_folder)
            print('[INFO]  已选择旧版译文文件夹:' ,translation_folder)
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作
        
    # 选择存储文件夹按钮绑定函数
    def Select_save_folder(self):
        save_folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if save_folder:
            self.label_output_folder.setText(save_folder)
            print('[INFO]  已选择保存文件夹:' ,save_folder)
        else :
            print('[INFO]  未选择文件夹')

    
    # 导入按钮绑定函数
    def Start_import(self):
        print('[INFO]  开始提取新版本游戏原文,请耐心等待！！！')

        #读取配置文件
        config_path = os.path.join(script_dir, "StevExtraction", "config.yaml")
        with open(config_path, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)

        #修改配置信息
        config['game_path'] = self.label_input_path.text()
        config['save_path'] = self.label_data_path.text()
        config['translation_path'] = self.label_translation_folder.text()
        config['data_path'] = self.label_output_folder.text()


        #导入文本
        pj=jtpp.Jr_Tpp(config)
        pj.Update(config['game_path'],config['translation_path'],config['save_path'],config['data_path'])




class Widget_sponsor(QFrame):# 赞助界面
    def __init__(self, text: str, parent=None):
        super().__init__(parent=parent)
        self.setObjectName(text.replace(' ', '-'))

        # -----创建第1个组，添加多个组件-----
        box1 = QGroupBox()
        box1.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1 = QHBoxLayout()


        # 创建 QLabel 用于显示图片
        self.image_label = QLabel(self)
        # 通过 QPixmap 加载图片
        pixmap = QPixmap(os.path.join(resource_dir,"sponsor","赞赏码.png"))
        # 设置 QLabel 的固定大小
        self.image_label.setFixedSize(350, 350)
        # 调整 QLabel 大小以适应图片
        self.image_label.setPixmap(pixmap)
        self.image_label.setScaledContents(True)


        layout1.addWidget(self.image_label)
        box1.setLayout(layout1)
        


        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 0px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()


        # 创建 QLabel 用于显示文字
        self.text_label = QLabel(self)
        self.text_label.setStyleSheet("font-family: 'SimSun'; font-size: 19px;")
        #self.text_label.setText("个人开发不易，如果这个项目帮助到了您，可以考虑请作者喝一杯奶茶。您的支持就是作者开发和维护项目的动力！🙌")
        self.text_label.setText("喜欢我的项目吗？如果这个项目帮助到了您，赞助一杯奶茶，让我能更有动力更新哦！💖")

        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.text_label)
        layout2.addStretch(1)  # 添加伸缩项
        box2.setLayout(layout2)



        
        # -----最外层容器设置垂直布局-----
        container = QVBoxLayout()

        # 设置窗口显示的内容是最外层容器
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(50, 70, 50, 30) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下

        # 把各个组添加到容器中
        container.addStretch(1)  # 添加伸缩项
        container.addWidget(box1)
        container.addWidget(box2)
        container.addStretch(1)  # 添加伸缩项


class AvatarWidget(NavigationWidget):#头像导航项
    """ Avatar widget """

    def __init__(self, parent=None):
        super().__init__(isSelectable=False, parent=parent)
        self.avatar = QImage(os.path.join(resource_dir, "Avatar.png")).scaled(
            24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)

    def paintEvent(self, e):
        painter = QPainter(self)
        painter.setRenderHints(
            QPainter.SmoothPixmapTransform | QPainter.Antialiasing)

        painter.setPen(Qt.NoPen)

        if self.isPressed:
            painter.setOpacity(0.7)

        # draw background
        if self.isEnter:
            c = 255 if isDarkTheme() else 0
            painter.setBrush(QColor(c, c, c, 10))
            painter.drawRoundedRect(self.rect(), 5, 5)

        # draw avatar
        painter.setBrush(QBrush(self.avatar))
        painter.translate(8, 6)
        painter.drawEllipse(0, 0, 24, 24)
        painter.translate(-8, -6)

        if not self.isCompacted:
            painter.setPen(Qt.white if isDarkTheme() else Qt.black)
            font = QFont('Segoe UI')
            font.setPixelSize(14)
            painter.setFont(font)
            painter.drawText(QRect(44, 0, 255, 36), Qt.AlignVCenter, 'NEKOparapa')



class window(FramelessWindow): #主窗口 v

    def __init__(self):
        super().__init__()
        # use dark theme mode
        self.setTitleBar(StandardTitleBar(self))

        self.hBoxLayout = QHBoxLayout(self)
        self.navigationInterface = NavigationInterface(self, showMenuButton=True)
        self.stackWidget = QStackedWidget(self)


        # 创建子界面控件，传入参数为对象名和parent
        self.Widget_AI = Widget_AI('Widget_AI', self)
        self.Widget_official_api = Widget_official_api('Widget_official_api', self)
        self.Widget_Openai = Widget_Openai('Widget_Openai', self)   
        self.Widget_Proxy = Widget_Proxy('Widget_Proxy', self)
        self.Widget_Anthropic = Widget_Anthropic('Widget_Anthropic', self)
        self.Widget_Google = Widget_Google('Widget_Google', self)
        self.Widget_Cohere = Widget_Cohere('Widget_Cohere', self)
        self.Widget_ZhiPu = Widget_ZhiPu('Widget_ZhiPu', self)
        self.Widget_Moonshot = Widget_Moonshot('Widget_Moonshot', self)
        self.Widget_Deepseek = Widget_Deepseek('Widget_Deepseek', self)
        self.Widget_Dashscope = Widget_Dashscope('Widget_Dashscope', self)
        self.Widget_Volcengine = Widget_Volcengine('Widget_Volcengine', self)
        self.Widget_SakuraLLM = Widget_SakuraLLM('Widget_SakuraLLM', self)

        self.Widget_translation_settings = Widget_translation_settings('Widget_translation_settings', self)
        self.Widget_translation_settings_A = Widget_translation_settings_A('Widget_translation_settings_A', self) 
        self.Widget_translation_settings_B1 = Widget_translation_settings_B1('Widget_translation_settings_B1', self) 
        self.Widget_translation_settings_B2 = Widget_translation_settings_B2('Widget_translation_settings_B2', self) 
        self.Widget_translation_settings_C = Widget_translation_settings_C('Widget_translation_settings_C', self)  
        self.Widget_start_translation = Widget_start_translation('Widget_start_translation', self) 

        self.Widget_RPG = Widget_RPG('Widget_RPG', self)  
        self.Widget_export_source_text = Widget_export_source_text('Widget_export_source_text', self)  
        self.Widget_import_translated_text = Widget_import_translated_text('Widget_import_translated_text', self)  
        self.Widget_update_text = Widget_update_text('Widget_update_text', self)    

        self.Widget_tune = Widget_tune('Widget_tune', self)
        self.Widget_tune_openai = Widget_tune_openai('Widget_tune_openai', self)
        self.Widget_tune_sakura = Widget_tune_sakura('Widget_tune_sakura', self)
        self.Widget_tune_anthropic = Widget_tune_anthropic('Widget_tune_anthropic', self)
        self.Widget_tune_google = Widget_tune_google('Widget_tune_google', self)

        self.Widget_sponsor = Widget_sponsor('Widget_sponsor', self)
        self.Widget_replace_dict = Widget_replace_dict('Widget_replace_dict', self)

        self.Widget_rulebook = Widget_rulebook('Widget_rulebook', self)
        self.Widget_system_prompt = Widget_system_prompt('Widget_system_prompt', self)  
        self.Widget_prompt_dict = Widget_prompt_dict('Widget_prompt_dict', self)
        self.Widget_translation_example = Widget_translation_example('Widget_translation_example', self)  
        self.Widget_characterization = Widget_characterization('Widget_characterization', self) 
        self.Widget_world_building = Widget_world_building('Widget_world_building', self) 
        self.Widget_writing_style = Widget_writing_style('Widget_writing_style', self) 

        self.initLayout() #调用初始化布局函数 

        self.initNavigation()   #调用初始化导航栏函数

        self.initWindow()  #调用初始化窗口函数


    # 初始化布局的函数
    def initLayout(self):   
        self.hBoxLayout.setSpacing(0)
        self.hBoxLayout.setContentsMargins(0, self.titleBar.height(), 0, 0)
        self.hBoxLayout.addWidget(self.navigationInterface)
        self.hBoxLayout.addWidget(self.stackWidget)
        self.hBoxLayout.setStretchFactor(self.stackWidget, 1)

    # 初始化导航栏的函数
    def initNavigation(self): # 详细介绍：https://pyqt-fluent-widgets.readthedocs.io/zh_CN/latest/navigation.html


 

        # 添加账号设置界面
        self.addSubInterface(self.Widget_AI, FIF.IOT, '账号设置',NavigationItemPosition.SCROLL) # NavigationItemPosition.SCROLL表示在可滚动伸缩区域
        # 添加官方接口界面
        self.addSubInterface(self.Widget_official_api, FIF.PEOPLE, '官方接口',parent=self.Widget_AI) # NavigationItemPosition.SCROLL表示在可滚动伸缩区域
        # 添加closeai官方账号界面
        self.addSubInterface(self.Widget_Openai, FIF.FEEDBACK, 'OpenAI官方',parent=self.Widget_official_api) 
        # 添加谷歌官方账号界面
        self.addSubInterface(self.Widget_Google, FIF.FEEDBACK, 'Google官方',parent=self.Widget_official_api)
        # 添加Cohere官方账号界面
        self.addSubInterface(self.Widget_Cohere, FIF.FEEDBACK, 'Cohere官方',parent=self.Widget_official_api)
        # 添加anthropic官方账号界面
        self.addSubInterface(self.Widget_Anthropic, FIF.FEEDBACK, 'Anthropic官方',parent=self.Widget_official_api)
        # 添加Moonshot官方账号界面
        self.addSubInterface(self.Widget_Moonshot, FIF.FEEDBACK, 'Moonshot官方',parent=self.Widget_official_api) 
        # 添加Deepseek官方账号界面
        self.addSubInterface(self.Widget_Deepseek, FIF.FEEDBACK, 'Deepseek官方',parent=self.Widget_official_api) 
        # 添加Dashscope官方账号界面
        self.addSubInterface(self.Widget_Dashscope, FIF.FEEDBACK, 'Dashscope官方',parent=self.Widget_official_api) 
        # 添加Volcengine官方账号界面
        self.addSubInterface(self.Widget_Volcengine, FIF.FEEDBACK, 'Volcengine官方',parent=self.Widget_official_api) 
        # 添加智谱官方账号界面
        self.addSubInterface(self.Widget_ZhiPu, FIF.FEEDBACK, '智谱官方',parent=self.Widget_official_api) 

        # 添加代理账号界面
        self.addSubInterface(self.Widget_Proxy, FIF.CLOUD, '代理平台',parent=self.Widget_AI) 
        # 添加sakura界面
        self.addSubInterface(self.Widget_SakuraLLM, FIF.CONNECT, 'SakuraLLM',parent=self.Widget_AI) 

        self.navigationInterface.addSeparator(NavigationItemPosition.SCROLL) # 添加分隔符

        # 添加翻译设置相关页面
        self.addSubInterface(self.Widget_translation_settings, FIF.APPLICATION, '翻译设置',NavigationItemPosition.SCROLL) 
        self.addSubInterface(self.Widget_translation_settings_A, FIF.REMOVE, '基础设置',parent=self.Widget_translation_settings) 
        self.addSubInterface(self.Widget_translation_settings_B1, FIF.ALIGNMENT, '发送设置',parent=self.Widget_translation_settings) 
        self.addSubInterface(self.Widget_translation_settings_B2, FIF.ALBUM, '专项设置',parent=self.Widget_translation_settings) 
        self.addSubInterface(self.Widget_translation_settings_C, FIF.EMOJI_TAB_SYMBOLS, '混合翻译设置',parent=self.Widget_translation_settings) 

        # 添加开始翻译页面
        self.addSubInterface(self.Widget_start_translation, FIF.ROBOT, '开始翻译',NavigationItemPosition.SCROLL)  

        self.navigationInterface.addSeparator(NavigationItemPosition.SCROLL) # 添加分隔符

        # 添加翻译设置相关页面
        self.addSubInterface(self.Widget_rulebook, FIF.BOOK_SHELF, '提示书',NavigationItemPosition.SCROLL) 
        self.addSubInterface(self.Widget_system_prompt, FIF.LABEL, '基础提示',parent=self.Widget_rulebook)
        self.addSubInterface(self.Widget_prompt_dict, FIF.DICTIONARY, '提示字典',parent=self.Widget_rulebook)   
        self.addSubInterface(self.Widget_characterization, FIF.EXPRESSIVE_INPUT_ENTRY, '角色介绍',parent=self.Widget_rulebook) 
        self.addSubInterface(self.Widget_world_building, FIF.QUICK_NOTE, '背景设定',parent=self.Widget_rulebook) 
        self.addSubInterface(self.Widget_writing_style, FIF.PENCIL_INK, '文风要求',parent=self.Widget_rulebook) 
        self.addSubInterface(self.Widget_translation_example, FIF.ZOOM, '翻译示例',parent=self.Widget_rulebook) 

        # 添加替换字典页面
        self.addSubInterface(self.Widget_replace_dict, FIF.DICTIONARY, '替换字典',NavigationItemPosition.SCROLL)  

        # 添加参数调整页面
        self.addSubInterface(self.Widget_tune, FIF.MIX_VOLUMES, '参数调整',NavigationItemPosition.SCROLL)  
        self.addSubInterface(self.Widget_tune_openai, FIF.SPEED_OFF, 'OpenAI',parent=self.Widget_tune)
        self.addSubInterface(self.Widget_tune_anthropic, FIF.SPEED_OFF, 'Anthropic',parent=self.Widget_tune)    
        self.addSubInterface(self.Widget_tune_sakura, FIF.SPEED_OFF, 'Sakura',parent=self.Widget_tune)  
        self.addSubInterface(self.Widget_tune_google, FIF.SPEED_OFF, 'Google',parent=self.Widget_tune)  

        self.navigationInterface.addSeparator(NavigationItemPosition.SCROLL)

        # 添加RPG界面
        self.addSubInterface(self.Widget_RPG, FIF.TILES, 'StevExtraction',NavigationItemPosition.SCROLL)
        self.addSubInterface(self.Widget_export_source_text, FIF.SHARE, '提取原文',parent=self.Widget_RPG)
        self.addSubInterface(self.Widget_import_translated_text, FIF.LABEL, '导入译文',parent=self.Widget_RPG)
        self.addSubInterface(self.Widget_update_text, FIF.PIE_SINGLE, '提取新版游戏原文',parent=self.Widget_RPG)

        self.navigationInterface.addSeparator(NavigationItemPosition.SCROLL) 

        # 添加赞助页面
        self.addSubInterface(self.Widget_sponsor, FIF.CAFE, '赞助一下', NavigationItemPosition.BOTTOM) 

       # 添加头像导航项
        self.navigationInterface.addWidget(
            routeKey='avatar',
            widget=AvatarWidget(),
            onClick=self.showMessageBox,
            position=NavigationItemPosition.BOTTOM
        )

        self.stackWidget.currentChanged.connect(self.onCurrentInterfaceChanged)
        self.stackWidget.setCurrentIndex(1)

    #初始化父窗口的函数
    def initWindow(self): 
        self.resize(1200 , 700)
        #self.setWindowIcon(QIcon('resource/logo.png'))
        self.setWindowTitle(Software_Version)
        self.titleBar.setAttribute(Qt.WA_StyledBackground)

        desktop = QApplication.desktop().availableGeometry()
        w, h = desktop.width(), desktop.height()
        self.move(w//2 - self.width()//2, h//2 - self.height()//2)

        dir1 = os.path.join(resource_dir, "light")
        dir2 = os.path.join(dir1, "demo.qss")
        with open(dir2, encoding='utf-8') as f:
            self.setStyleSheet(f.read())

    # 添加界面到导航栏布局函数
    def addSubInterface(self, interface, icon, text: str, position=NavigationItemPosition.TOP, parent=None):
        """ add sub interface """
        self.stackWidget.addWidget(interface)
        self.navigationInterface.addItem(
            routeKey=interface.objectName(),
            icon=icon,
            text=text,
            onClick=lambda: self.switchTo(interface),
            position=position,
            tooltip=text,
            parentRouteKey=parent.objectName() if parent else None
        )

    #切换到某个窗口的函数
    def switchTo(self, widget): 
        self.stackWidget.setCurrentWidget(widget) #设置堆栈窗口的当前窗口为widget

    #堆栈窗口的当前窗口改变时，调用的函数
    def onCurrentInterfaceChanged(self, index):    
        widget = self.stackWidget.widget(index)
        self.navigationInterface.setCurrentItem(widget.objectName())

    #头像导航项的函数调用的函数
    def showMessageBox(self):
        url = QUrl('https://github.com/NEKOparapa/AiNiee-chatgpt')
        QDesktopServices.openUrl(url)

    #窗口关闭函数，放在最后面，解决界面空白与窗口退出后子线程还在运行的问题
    def closeEvent(self, event):
        title = '确定是否退出程序?'
        content = """如果正在进行翻译任务，当前任务会停止。"""
        w = Dialog(title, content, self)

        if w.exec() :
            print("[INFO] 主窗口已经退出！")
            global Running_status
            Running_status = 10
            event.accept()
        else:
            event.ignore()





if __name__ == '__main__':

    #开启子进程支持
    multiprocessing.freeze_support() 

    # 启用了高 DPI 缩放
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)


    Software_Version = "AiNiee4.71"  #软件版本号
    cache_list = [] # 全局缓存数据
    Running_status = 0  # 存储程序工作的状态，0是空闲状态，1是接口测试状态
                        # 6是翻译任务进行状态，9是翻译任务暂停状态，10是强制终止任务状态


    # 定义线程锁
    lock1 = threading.Lock()  #这个用来锁缓存文件
    lock2 = threading.Lock()  #这个用来锁UI信号的
    lock3 = threading.Lock()  #这个用来锁自动备份缓存文件功能的


    # 工作目录改为python源代码所在的目录
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0])) # 获取当前工作目录
    print("[INFO] 当前工作目录是:",script_dir,'\n') 
    # 设置资源文件夹路径
    resource_dir = os.path.join(script_dir, "resource")


    # 创建全局UI通讯器
    user_interface_prompter = User_Interface_Prompter() 
    user_interface_prompter.signal.connect(user_interface_prompter.on_update_ui)  #创建信号与槽函数的绑定，使用方法为：user_interface_prompter.signal.emit("str","str"....)

    # 创建全局限制器
    request_limiter = Request_Limiter()

    # 创建全局配置器
    configurator = Configurator()


    #创建了一个 QApplication 对象
    app = QApplication(sys.argv)
    #创建全局窗口对象
    Window = window()
    
    #窗口对象显示
    Window.show()


    # 读取配置文件
    configurator.read_write_config("read")

    #进入事件循环，等待用户操作
    sys.exit(app.exec_())



