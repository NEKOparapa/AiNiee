import re
import os
import json
import pandas as pd
import traceback
import openpyxl
from chardet import detect
import csv
import logging

version = "v3.00"

# 设置 CSV 和 Pandas 的显示选项，防止文本被截断
csv.field_size_limit(2**30)
pd.options.display.max_colwidth = None
pd.options.display.max_rows = None
pd.options.display.max_columns = None
pd.options.display.width = None


class Jr_Tpp:
    def __init__(self, config: dict, path: str = False):
        """
        初始化 Jr_Tpp 对象。

        :param config: 配置字典，包含各种设置。
        :param path: 可选的项目路径，如果提供，将从该路径加载项目。
        """
        self.ProgramData = {}  # 存储项目数据的字典
        self.__tempdata = [  # 临时数据列表，用于存储提取的文本行
            "原文",
            "译文",
            "地址",
            "标签",
            "code",
        ]
        self.__sumlen = 0  # 记录合并文本的长度

        # 初始化日志记录器
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        ch = logging.StreamHandler()  # 创建一个控制台日志处理器
        ch.setFormatter(formatter)
        self.logger.addHandler(ch)

        self.ApplyConfig(config)  # 应用配置

        if path:
            self.load(path)  # 如果提供了路径，加载项目

    def ApplyConfig(self, config, clean=False):
        """
        应用配置设置。

        :param config: 配置字典。
        :param clean: 是否清除现有的项目数据（默认为 False）。
        """
        # 从配置中获取各项设置，如果配置中没有，则使用默认值
        self.BlackDir = config.get("BlackDir", [])  # 黑名单目录
        self.BlackFiles = config.get("BlackFiles", [])  # 黑名单文件
        self.BlackCode = config.get("BlackCode", [])  # 黑名单代码
        self.NameWithout = config.get("NameWithout", [])  # 提取人名时需要排除的文件
        self.codewithnames = config.get("codewithnames", [])  # 需要特殊处理的人名代码
        self.ReadCode = config.get("ReadCode", [])  # 需要读取的代码
        self.ja = config.get("ja", 1)  # 是否为日语模式（默认为 True）
        self.sumcode = config.get("sumcode", [])  # 需要合并的代码
        self.line_length = config.get("line_length", 40)  # 自动换行的行长度
        self.note_percent = config.get("note_percent", 0.2)  # note 文本的比例
        self.sptext = config.get("sptext", {})  # 特殊文本提取规则
        # 替换 sptext 中的句号为单引号，用于后续处理。
        for key in self.sptext:
            for mark in self.sptext[key]:
                self.sptext[key][mark] = self.sptext[key][mark].replace("。", "'")

        self.AutoLineFeed_jsdir = config.get(
            "auto_linefeed_js", "自动换行.js"
        )  # 自动换行插件的 JS 文件名
        self.need2check_filename = config.get(
            "need2check_filename", "need2check.json"
        )  # 可能需要人工检查的文件名
        self.project_data_dir = config.get("project_data_dir", "data")  # 项目数据目录名
        self.project_dir_name = config.get(
            "project_dir_name", "翻译工程文件"
        )  # 项目目录名

        # 从配置文件读取自定义表头和列号
        self.source_header = config.get("source_header", "Original Text")
        self.translation_header = config.get("translation_header", "Initial")
        self.source_column = config.get("source_column", 0)
        self.translation_column = config.get("translation_column", 1)

        if clean:
            self.ProgramData = {}  # 如果 clean 为 True，清空项目数据

    def __Readxlsx(self, name):
        """
        读取 xlsx 文件并返回 DataFrame。

        :param name: xlsx 文件名。
        :return: 包含 xlsx 文件数据的 DataFrame，如果读取失败则返回 None。
        """
        try:
            workbook = openpyxl.load_workbook(name)  # 加载 xlsx 文件
            sheet_names = workbook.sheetnames  # 获取所有 sheet 的名称
            worksheet = workbook[sheet_names[0]]  # 获取第一个 sheet
            # 获取第一行作为列名，并过滤掉空值
            column_names = [
                cell.value for cell in worksheet[1] if cell.value is not None
            ]
            if not column_names:
                # 如果没有列名，创建一个空的 DataFrame
                df = pd.DataFrame(columns=column_names)
            else:
                # 读取数据行
                data = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    data.append(row)
                # 创建 DataFrame
                df = pd.DataFrame(data, columns=column_names)
            workbook.close()  # 关闭 xlsx 文件
            return df
        except Exception as e:
            self.logger.error(f"读取xlsx文件失败: {name}")
            self.logger.error(traceback.format_exc())
            print(e)
            print("请关闭所有xlsx文件再试")
            return None

    def __Writexlsx(self, df, name, full=False, output_black=True):
        """
        将 DataFrame 写入 xlsx 文件。

        :param df: 要写入的 DataFrame。
        :param name: xlsx 文件名。
        :param full: 是否导出所有列（包括地址、标签和 code），默认为 False。
        :param output_black: 是否输出黑名单中的文本，默认为 True.
        """
        workbook = openpyxl.Workbook()  # 创建一个新的 xlsx 文件
        sheet = workbook.active  # 获取当前活动的 sheet

        # 根据 full 参数决定要导出的列
        columns_to_export = (
            ["原文", "译文", "地址", "标签", "code"] if full else ["原文", "译文"]
        )
        header_row = columns_to_export  # 设置表头
        sheet.append(header_row)  # 将表头添加到 sheet

        # 遍历 DataFrame 的每一行
        for index, row in df.iterrows():
            # 根据是否在黑名单中决定是否输出该行
            if (
                not self.__IfBlackDir(row["地址"]) and row["code"] not in self.BlackCode
            ) or output_black:
                data_row = [
                    row[column] for column in columns_to_export
                ]  # 获取要导出的数据
                sheet.append(data_row)  # 将数据行添加到 sheet

        # 将所有单元格的格式设置为文本格式
        for column_cells in sheet.columns:
            for cell in column_cells:
                cell.number_format = "@"

        # 确保所有字符串都以 UTF-8 编码
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    encoded_string = cell.value.encode("utf-8")
                    cell.value = encoded_string.decode("utf-8")
        try:
            workbook.save(name)  # 保存 xlsx 文件
        except Exception as e:
            self.logger.error(f"写入xlsx文件失败: {name}")
            self.logger.error(traceback.format_exc())
            print(e)
            input("导出失败，请关闭所有xlsx文件后再次尝试")

    def __GetSptext(self, data: str, Dir, code, rule) -> list:
        """
        根据特殊规则提取文本。

        :param data: 要提取文本的字符串。
        :param Dir: 文本的地址。
        :param code: 文本的 code。
        :param rule: 用于提取文本的正则表达式。
        :return: 包含提取的文本的列表。
        """
        res = []
        Dir += "\u200B" + "1"  # 在地址后添加一个零宽空格和一个数字 1，用于后续处理
        if self.ja:
            # 使用正则表达式查找所有匹配的文本
            data = re.findall(rule, data)
            for i in data:
                # 检查提取的文本中是否包含日语字符
                if re.search(
                    "[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5ー々〆〤]", i
                ):
                    res.append([i, "", Dir, "", code])  # 将提取的文本添加到结果列表
        return res

    def __ReadFile(self, data, FileName: str, code: int = False) -> list:
        """
        递归地读取文件内容，提取需要翻译的文本。

        :param data: 要读取的数据（可以是字典、列表或字符串）。
        :param FileName: 当前处理的文件名或路径。
        :param code: 当前文本的 code（可选，默认为 False）。
        :return: 包含提取的文本的列表，每个元素是一个包含原文、译文、地址、标签和 code 的列表。
        """
        res = []
        tp = type(data)

        if tp == dict:
            # 如果数据是字典，遍历键值对
            for key in data.keys():
                current_code = data.get("code", False)  # 获取当前的 code
                # 递归调用 __ReadFile 处理子元素
                res += self.__ReadFile(data[key], FileName + "\\" + key, current_code)

        elif tp == list:
            # 如果数据是列表，遍历元素
            for i in range(0, len(data)):
                # 递归调用 __ReadFile 处理子元素
                res += self.__ReadFile(data[i], FileName + "\\" + str(i), code)

        elif tp == str:
            # 如果数据是字符串，处理文本提取逻辑
            current_code = str(code) if code else "-1"  # 将 code 转换为字符串

            if current_code in self.sptext.keys():
                # 如果当前 code 在 sptext 字典中，应用特殊规则提取文本
                for mark in self.sptext[current_code].keys():
                    if mark in data or mark == "空":
                        rule = self.sptext[current_code][mark]
                        # 使用正则表达式提取文本
                        [
                            res.append(x)
                            for x in self.__GetSptext(
                                data, FileName, current_code, rule
                            )
                        ]
            else:
                # 检查是否包含日语字符或特殊情况下的处理
                if (
                    not self.ja
                    or re.search(
                        r"[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5ー々〆〤]",
                        data,
                    )
                    or r"System.json\gameTitle" in FileName
                ):
                    if r"System.json\gameTitle" in FileName and data == "":
                        data = " "  # 确保 gameTitle 不为空
                    textdata = [data, "", FileName, "", current_code]

                    # 处理 code 为 108 或 408 时的文本合并逻辑
                    if (
                        current_code == self.__tempdata[4]
                        or (self.__tempdata[4] == "108" and current_code == "408")
                    ) and current_code in self.sumcode:
                        self.__tempdata[0] += "\n" + textdata[0]
                        self.__sumlen += 1
                    else:
                        # 如果存在合并的文本，添加到结果列表
                        if self.__sumlen:
                            self.__tempdata[2] += "\u200B" + str(self.__sumlen)
                            if self.__tempdata[0] != "":
                                res.append(self.__tempdata)

                        # 根据 ReadCode 设置决定是否添加到结果列表
                        if current_code in self.ReadCode or not self.ReadCode:
                            self.__tempdata = textdata
                            self.__sumlen = 1
                        else:
                            self.__tempdata = ["原文", "译文", "地址", "标签", "code"]
                            self.__sumlen = 0
                else:
                    # 处理不包含日语字符的情况
                    if (
                        current_code == self.__tempdata[4]
                        or (self.__tempdata[4] == "108" and current_code == "408")
                    ) and current_code in self.sumcode:
                        self.__tempdata[0] += "\n" + data
                        self.__sumlen += 1
                    elif self.__sumlen:
                        self.__tempdata[2] += "\u200B" + str(self.__sumlen)
                        if self.__tempdata[0] != "":
                            res.append(self.__tempdata)
                        self.__tempdata = ["原文", "译文", "地址", "标签", "code"]
                        self.__sumlen = 0
                    else:
                        self.__tempdata = ["原文", "译文", "地址", "标签", "code"]
                        self.__sumlen = 0
        return res

    def __ReadFolder(self, dir: str) -> list:
        """
        读取指定目录下的所有文件。

        :param dir: 要读取的目录。
        :return: 包含所有文件路径的列表。
        """
        res = []
        # 遍历指定目录及其子目录
        for root, _, files in os.walk(dir):
            for file in files:
                res.append(os.path.join(root, file))  # 将文件路径添加到结果列表
        return res

    def __IfBlackDir(self, Dir: str) -> bool:
        """
        检查给定的目录是否在黑名单中。

        :param Dir: 要检查的目录。
        :return: 如果目录在黑名单中，则返回 True，否则返回 False。
        """
        for blackdir in self.BlackDir:
            pattern = re.compile(blackdir)  # 创建正则表达式对象
            # 如果目录与黑名单中的任何一个模式匹配，则返回 True
            if re.search(pattern, Dir):
                return True
        return False

    def __RemoveDuplicated(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        去除 DataFrame 中的重复行，并合并重复行的地址和 code。

        :param data: 要处理的 DataFrame。
        :return: 处理后的 DataFrame。
        """
        # 检查 DataFrame 是否包含 "地址" 和 "code" 列
        if "地址" in data.columns and "code" in data.columns:
            a = data[~data.index.duplicated()].copy()  # 获取不重复的行
            b = data[data.index.duplicated()].copy()  # 获取重复的行
            a_dict = {"a": a}
            b_dict = {"b": b}
            # 遍历不重复的行
            for index in a_dict["a"].index:
                if index in b_dict["b"].index:
                    # 获取重复行的地址和 code
                    Dirs = list(b_dict["b"][b_dict["b"].index == index]["地址"])
                    codes = list(b_dict["b"][b_dict["b"].index == index]["code"])
                    for i in range(len(Dirs)):
                        black = False
                        # 检查地址和 code 是否在黑名单中
                        if self.__IfBlackDir(Dirs[i]):
                            black = True
                        if codes[i] in self.BlackCode:
                            black = True
                        # 如果不在黑名单中，将地址和 code 合并到不重复的行中
                        if not black:
                            a_dict["a"].loc[index, "地址"] += "☆↑↓" + Dirs[i]
                            a_dict["a"].loc[index, "code"] += "," + codes[i]
            return a_dict["a"]  # 返回处理后的 DataFrame
        else:
            return data[
                ~data.index.duplicated()
            ]  # 如果没有 "地址" 和 "code" 列，只去除重复行

    def __toDataFrame(self, data: list) -> pd.DataFrame:
        """
        将数据列表转换为 DataFrame。

        :param data: 包含文本数据的列表。
        :return: 转换后的 DataFrame。
        """
        # 创建 DataFrame，并设置列名
        DataFrame = pd.DataFrame(data, columns=["原文", "译文", "地址", "标签", "code"])
        DataFrame.index = list(DataFrame["原文"])  # 将 "原文" 列设置为索引
        DataFrame = self.__RemoveDuplicated(DataFrame)  # 去除重复行并合并地址和 code
        return DataFrame

    def __nameswitch(self, name, csv: bool = False):
        """
        根据输入文件名和类型（xlsx 或 csv）返回相应的文件名。

        :param name: 输入文件名。
        :param csv: 是否为 csv 文件（默认为 False）。
        :return: 转换后的文件名。
        """
        base, ext = os.path.splitext(name)  # 分离文件名和扩展名
        # 根据输入文件类型返回对应的输出文件类型
        if not csv:
            return base + ".json" if ext.lower() == ".xlsx" else base + ".xlsx"
        else:
            return base + ".json" if ext.lower() == ".csv" else base + ".csv"

    def __WriteFile(
        self,
        data,
        untrs: str,
        trsed: str,
        Dir: list,
        length: int,
        code: str,
        key_is_list=False,
    ):
        """
        递归地将翻译后的文本写入文件数据结构。

        :param data: 要写入的文件数据（可以是字典、列表或字符串）。
        :param untrs: 未翻译的文本。
        :param trsed: 翻译后的文本。
        :param Dir: 文本在文件中的地址（列表形式）。
        :param length: 文本的长度（用于处理列表）。
        :param code: 文本的 code。
        :param key_is_list: 指示当前的key是否是list
        :return: 更新后的文件数据。
        """
        # 获取文本在文件内的地址
        if isinstance(data, list):
            i = int(Dir[0])  # 获取列表索引
            # 递归调用 __WriteFile 处理列表中的元素
            data[i] = self.__WriteFile(data[i], untrs, trsed, Dir[1:], length, code)
            if key_is_list:
                # 如果 key 是列表，并且文本长度大于 1，标记后续的元素为 "☆删除☆"
                for n in range(1, length):
                    data[i + n] = "☆删除☆"
        elif isinstance(data, dict):
            try:
                if Dir[0] == "list" and length > 1:
                    key_is_list = True  # 如果 key 是 "list" 并且长度大于 1，设置 key_is_list 为 True

                # 递归调用 __WriteFile 处理字典中的值
                data[Dir[0]] = self.__WriteFile(
                    data[Dir[0]],
                    untrs,
                    trsed,
                    Dir[1:],
                    length,
                    code,
                    key_is_list=key_is_list,
                )
            except Exception as e:
                self.logger.error(
                    f"写入文件错误，原文: {untrs}, 译文: {trsed}, 地址: {Dir}"
                )
                self.logger.error(traceback.format_exc())
                print(untrs)
                print(trsed)
                print(Dir)
                input(data)
        elif isinstance(data, str) and not Dir:
            # 如果数据是字符串并且地址列表为空，说明已经到达叶子节点，进行文本替换
            # 写code355,655
            if code in self.sptext.keys():
                # 如果 code 在 sptext 中，应用特殊规则替换文本
                for mark in self.sptext[code].keys():
                    if mark in data or mark == "空":
                        data = data.replace(untrs, trsed)
            else:
                # 否则直接替换文本
                data = trsed
        return data

    def __del_marked_list(self, data, scenario=False):
        """
        递归地删除标记为 "☆删除☆" 的列表元素。

        :param data: 要处理的数据（可以是字典、列表或字符串）。
        :param scenario: 是否为 Scenario.json 文件（默认为 False）。
        :return: 更新后的数据。
        """
        if isinstance(data, list):
            # 如果数据是列表，遍历元素
            for i in range(len(data)):
                # 递归调用 __del_marked_list 处理列表中的元素
                data[i] = self.__del_marked_list(data[i], scenario)
        elif isinstance(data, dict):
            # 如果数据是字典，遍历键值对
            for key in data.keys():
                if key == "list" or scenario:
                    # 如果 key 是 "list" 或者当前文件是 Scenario.json，删除标记为 "☆删除☆" 的元素
                    while "☆删除☆" in data[key]:
                        data[key].remove("☆删除☆")
                else:
                    # 递归调用 __del_marked_list 处理字典中的值
                    data[key] = self.__del_marked_list(data[key], scenario)
        return data

    def __CheckNAN(self):
        """
        检查译文中是否存在空值（NaN），如果存在，则用原文填充。
        """
        nanlist = []  # 存储空值所在的索引
        # 遍历所有 DataFrame
        for name in self.ProgramData.keys():
            DataFrame = self.ProgramData[name]
            # 获取译文为空的行
            NANFrame = DataFrame[DataFrame["译文"].isnull()]
            if not NANFrame.empty:
                # 将空值所在的索引添加到列表
                indexlist = list(NANFrame.index)
                nanlist += indexlist
                # 用原文填充空值
                for index in indexlist:
                    DataFrame.loc[index, "译文"] = index
            self.ProgramData[name] = DataFrame  # 更新 DataFrame
        if nanlist:
            print("以下原文没有对应译文，恢复为原文")
            for i in nanlist:
                print(i)

    def ReadGame(self, GameDir: str):
        """
        读取游戏目录中的 JSON 文件，提取需要翻译的文本。

        :param GameDir: 游戏目录的路径。
        """
        self.logger.info(f"开始读取游戏目录: {GameDir}")
        data_folder_name = "data"  # 数据文件夹的名称
        data_dir = os.path.join(GameDir, data_folder_name)  # 构造数据文件夹的完整路径
        Files = self.__ReadFolder(GameDir)  # 读取游戏目录下的所有文件
        for File in Files:
            # 计算文件相对于数据文件夹的路径
            relative_path = os.path.relpath(File, data_dir)
            # 如果文件在数据文件夹内，则使用相对路径作为名称，否则使用文件名
            name = (
                relative_path
                if os.sep + data_folder_name + os.sep in File
                else os.path.basename(File)
            )
            # 处理数据文件夹内的 JSON 文件
            if data_folder_name in File.lower() and name.endswith(".json"):
                if name not in self.BlackFiles:
                    print(f"正在读取{name}")
                    try:
                        # 尝试使用 UTF-8 编码读取 JSON 文件
                        with open(File, "r", encoding="utf8") as f:
                            data = json.load(f)
                    except Exception:
                        # 如果 UTF-8 编码失败，尝试自动检测编码并读取
                        try:
                            with open(File, "rb") as f:
                                encoding = detect(f.read())["encoding"]
                                encoding = (
                                    encoding if encoding else "ansi"
                                )  # 如果未检测到编码，则使用 ansi
                            with open(File, "r", encoding=encoding) as f:
                                data = json.load(f)
                        except Exception as e:
                            self.logger.error(f"读取文件 {name} 失败")
                            self.logger.error(traceback.format_exc())
                            print(
                                f"无法确定{name}文件编码, 且无法用ANSI编码打开，读取失败"
                            )
                            continue

                    # 调用 __ReadFile 提取文本数据
                    TextDatas = self.__ReadFile(data, name)
                    # 处理可能存在的文本合并
                    if self.__sumlen and self.__tempdata not in TextDatas:
                        if "\u200B" not in self.__tempdata[2]:
                            self.__tempdata[2] += "\u200B" + str(self.__sumlen)
                        if self.__tempdata[0] != "":
                            TextDatas.append(self.__tempdata)
                    # 重置临时数据和合并长度
                    self.__tempdata = [
                        "原文",
                        "译文",
                        "地址",
                        "标签",
                        "code",
                    ]
                    self.__sumlen = 0
                    # 将提取的文本数据转换为 DataFrame 并添加到 ProgramData 中
                    if TextDatas:
                        self.ProgramData.update({name: self.__toDataFrame(TextDatas)})
        self.logger.info("游戏读取完成")
        print("########################读取游戏完成########################")

    def InjectGame(
        self, GameDir: str, path: str, BlackLabel: list = None, BlackCode: list = None
    ):
        """
        将翻译后的文本注入回游戏目录中的 JSON 文件。

        :param GameDir: 游戏目录的路径。
        :param path: 翻译文件所在的路径。
        :param BlackLabel: 要排除的标签列表（默认为 ["Black"]）。
        :param BlackCode: 要排除的 code 列表（默认为 self.BlackCode）。
        """
        self.logger.info(f"开始注入翻译到游戏目录: {GameDir}, 输出路径: {path}")
        # 如果未提供 BlackLabel 和 BlackCode，则使用默认值
        if BlackLabel is None:
            BlackLabel = ["Black"]
        if BlackCode is None:
            BlackCode = self.BlackCode

        # 查找未翻译的行并导出
        untrsline = self.search("", 1)  # 查找译文为空的行
        if untrsline:
            output = pd.concat(list(untrsline.values()), axis=0)  # 合并搜索结果
            output_path = "未翻译行.xlsx"  # 设置输出文件名
            self.__Writexlsx(output, output_path, full=True)  # 导出未翻译行
            self.ApplyUntrs(untrsline)  # 将未翻译的行替换为原文
            self.logger.warning(
                f"存在未翻译行，已应用原文并导出为 {output_path} 以供确认"
            )
            print(f"存在未翻译行，已应用原文并导出为“{output_path}”以供确认")

        self.__CheckNAN()  # 检查并处理译文中的空值
        Files = self.__ReadFolder(GameDir)  # 读取游戏目录下的所有文件
        data_dir = os.path.join(path, "data")  # 构造输出数据目录的路径
        os.makedirs(data_dir, exist_ok=True)  # 确保输出目录存在

        for File in Files:
            # 提取文件名
            name_parts = File.split(os.path.join("data", ""))
            if len(name_parts) > 1 and name_parts[-1].endswith(".json"):
                name = name_parts[-1]
                Scenario = "Scenario.json" in File  # 检查是否为 Scenario.json 文件

                if os.sep + "data" + os.sep in File and name.endswith(".json"):
                    print(f"正在写入{name}")
                    try:
                        # 尝试使用 UTF-8 编码读取 JSON 文件
                        with open(File, "r", encoding="utf8") as f:
                            data = json.load(f)
                    except Exception as e:
                        self.logger.error(f"读取文件 {name} 失败")
                        self.logger.error(traceback.format_exc())
                        print(f"读取{name}失败")
                        continue

                    if name in self.ProgramData.keys():
                        # 如果存在对应的翻译数据，则进行文本替换
                        DataFrame = self.ProgramData[name]
                        for untrs in DataFrame.index:
                            trsed = DataFrame.loc[untrs, "译文"]  # 获取翻译后的文本
                            # 获取文本的地址、code 和标签
                            Dirlist = DataFrame.loc[untrs, "地址"].split("☆↑↓")
                            codelist = DataFrame.loc[untrs, "code"].split(",")
                            labellist = DataFrame.loc[untrs, "标签"].split(",")
                            # 检查是否存在需要排除的标签
                            black = any(label in BlackLabel for label in labellist)

                            for i in range(len(Dirlist)):
                                # 解析地址、长度和 code
                                Dir = Dirlist[i].split("\u200B")[0]
                                length = int(Dirlist[i].split("\u200B")[1])
                                code = codelist[i]
                                # 排除黑名单中的文本和目录
                                if (
                                    not black
                                    and code not in BlackCode
                                    and not self.__IfBlackDir(Dir)
                                ):
                                    # 解析地址为列表形式
                                    Dir = Dir.split("json\\")[1].split("\\")
                                    # 调用 __WriteFile 写入翻译后的文本
                                    data = self.__WriteFile(
                                        data,
                                        untrs,
                                        trsed,
                                        Dir,
                                        length,
                                        code,
                                        key_is_list=Scenario,
                                    )
                        # 删除标记为 "☆删除☆" 的列表元素
                        data = self.__del_marked_list(data, Scenario)

                    # 创建输出目录
                    self.__makedir(name, data_dir)
                    output_file = os.path.join(data_dir, name)  # 构造输出文件的完整路径
                    try:
                        # 将更新后的数据写回 JSON 文件，使用缩进和确保 ASCII 字符不转义
                        with open(output_file, "w", encoding="utf8") as f1:
                            json.dump(data, f1, ensure_ascii=False, indent=4)
                    except Exception as e:
                        self.logger.error(f"写入文件 {name} 失败")
                        self.logger.error(traceback.format_exc())
                        print(f"写入{name}失败")

        self.logger.info("游戏注入完成")
        print("########################写入游戏完成########################")

    def GetFileNames(self) -> list:
        """
        获取所有需要处理的文件名列表。

        :return: 文件名列表。
        """
        namelist = list(self.ProgramData.keys())  # 获取所有 DataFrame 的名称
        mapname = []  # 存储地图文件名的列表
        # 提取地图文件名并排序
        for name in namelist.copy():
            if "map" in name.lower() and "info" not in name.lower():
                namelist.remove(name)
                mapname.append(int(name.replace("Map", "").replace(".json", "")))
        mapname = sorted(mapname)  # 对地图文件名进行排序
        # 如果存在地图文件，则添加一个表示地图文件范围的名称
        if mapname:  # 检查 mapname 列表是否为空
            mapname[0] = f"{mapname[0]:03d}"
            mapname[-1] = f"{mapname[-1]:03d}"
            namelist.append(f"Map{mapname[0]}~{mapname[-1]}.json")
        return namelist

    def __makedir(self, name, path):
        """
        创建目录（如果目录不存在）。

        :param name: 文件名。
        :param path: 基础路径。
        :return: 创建的目录的完整路径。
        """
        dir_path = os.path.join(path, os.path.dirname(name))  # 构造目录的完整路径
        os.makedirs(dir_path, exist_ok=True)  # 创建目录，如果目录已存在，则不引发异常
        return dir_path

    def ToXlsx(self, name: str, path: str, output_black=True):
        """
        将指定名称的 DataFrame 导出为 xlsx 文件。

        :param name: DataFrame 的名称。
        :param path: 导出路径。
        :param output_black: 是否输出黑名单，默认为 True.
        """
        outputname = self.__nameswitch(name)  # 获取转换后的文件名
        output_dir = os.path.join(path, "data")  # 输出到data子目录
        os.makedirs(output_dir, exist_ok=True)  # 确保目录存在
        output_file_path = os.path.join(output_dir, outputname)  # 完整输出路径

        data = self.ProgramData[name]  # 获取指定名称的 DataFrame
        try:
            # 调用 __Writexlsx 导出 DataFrame
            self.__Writexlsx(data, output_file_path, output_black=output_black)
            print(f"正在导出{outputname} to {output_dir}")  # 打印导出路径
        except Exception as e:
            self.logger.error(f"导出xlsx文件 {outputname} 失败")
            self.logger.error(traceback.format_exc())
            print(e)
            input("导出失败，请关闭所有xlsx文件后再次尝试")

    def ToCsv(self, name: str, path: str):
        """
        将指定名称的 DataFrame 保存为 CSV 文件。

        :param name: DataFrame 的名称。
        :param path: 保存路径。
        """
        outputname = self.__nameswitch(name, True)  # 获取转换后的文件名（CSV 格式）
        output_dir = os.path.join(path, self.project_data_dir)  # 构造输出目录的路径
        os.makedirs(output_dir, exist_ok=True)  # 确保输出目录存在
        output_file_path = os.path.join(
            output_dir, outputname
        )  # 构造输出文件的完整路径

        data = self.ProgramData[name]  # 获取指定名称的 DataFrame
        try:
            # 将 DataFrame 保存为 CSV 文件，使用 UTF-8 编码，不包含索引
            data.to_csv(output_file_path, sep=",", encoding="utf8", index=False)
            print(f"正在保存{outputname} to {output_dir}")  # 打印保存路径
        except Exception as e:
            self.logger.error(f"保存csv文件 {outputname} 失败")
            self.logger.error(traceback.format_exc())
            print(e)
            input("保存失败，请关闭所有csv文件后再次尝试")

    def Output(self, path: str):
        """
        将所有 DataFrame 导出为 xlsx 文件。

        :param path: 导出路径。
        """
        output_data_dir = os.path.join(path, "data")  # 构造输出数据目录的路径
        os.makedirs(output_data_dir, exist_ok=True)  # 确保输出目录存在
        # 遍历所有 DataFrame，并调用 ToXlsx 导出
        for name in self.ProgramData.keys():
            self.ToXlsx(name, path, output_black=False)  # 不输出黑名单
        print("########################导出完成########################")

    def Save(self, path: str):
        """
        将所有 DataFrame 保存为 CSV 文件，用于创建翻译工程。

        :param path: 保存路径。
        """
        project_path = os.path.join(path, self.project_dir_name)  # 构造项目目录的路径
        project_data_path = os.path.join(
            project_path, self.project_data_dir
        )  # 构造项目数据目录的路径
        os.makedirs(project_path, exist_ok=True)  # 确保项目目录存在
        os.makedirs(project_data_path, exist_ok=True)  # 确保项目数据目录存在

        # 遍历所有 DataFrame，并调用 ToCsv 保存
        for name in self.ProgramData.keys():
            self.ToCsv(name, project_path)  # 使用 project_path 作为 toCsv 的路径
        print("########################保存工程完成########################")

    def InputFromJson(
        self, trsdata: dict = None, path: str = None, namelist: list = None
    ):
        """
        从 JSON 文件或字典导入翻译数据。

        :param trsdata: 包含翻译数据的字典（可选）。
        :param path: JSON 文件的路径（可选）。
        :param namelist: 要导入的文件名列表（可选，默认为所有文件）。
        """
        if path:
            # 如果提供了 JSON 文件路径，则尝试读取 JSON 文件
            try:
                with open(path, "r", encoding="utf8") as f:
                    trsdata = json.load(f)
            except Exception:
                # 如果 UTF-8 编码失败，尝试自动检测编码并读取
                try:
                    with open(path, "rb") as f:
                        encoding = detect(f.read())["encoding"]
                        encoding = encoding if encoding else "ansi"
                    with open(path, "r", encoding=encoding) as f:
                        trsdata = json.load(f)
                except Exception as e:
                    self.logger.error(f"从JSON文件导入翻译失败: {path}")
                    self.logger.error(traceback.format_exc())
                    print(
                        f"读取{path}失败,请确保json文件头格式正确\n"
                        f"若提示UnicodeDecodeError，请确保该文件编码可读"
                    )
                    return

        if not namelist:
            namelist = self.ProgramData.keys()  # 如果未提供文件名列表，则处理所有文件

        if trsdata:  # 确保 trsdata 不是 None
            # 遍历文件名列表
            for name in namelist:
                if name not in self.ProgramData:  # 检查 name 是否存在于 ProgramData 中
                    self.logger.warning(f"跳过文件 {name}，因为它不在工程数据中")
                    continue
                DataFrame = self.ProgramData[name]  # 获取对应的 DataFrame
                # 遍历翻译数据字典，并更新 DataFrame 中的译文
                for untrs in trsdata.keys():
                    if untrs in DataFrame.index:
                        DataFrame.loc[untrs, "译文"] = trsdata[untrs]
                self.ProgramData[name] = DataFrame  # 更新 DataFrame
        else:
            self.logger.warning("没有提供翻译数据 (trsdata 为 None)")
            print("没有提供翻译数据")

    def InputFromDataFrame(
        self, data: pd.DataFrame, namelist: list = None, samefile=False
    ):
        """
        从 DataFrame 导入翻译数据。

        :param data: 包含翻译数据的 DataFrame。
        :param namelist: 要导入的文件名列表（可选，默认为所有文件）。
        :param samefile: 是否为同一文件导入模式（默认为 False）。
        """
        source_col_name = None
        translation_col_name = None

        # 尝试使用自定义表头读取
        if (
            self.source_header in data.columns
            and self.translation_header in data.columns
        ):
            source_col_name = self.source_header
            translation_col_name = self.translation_header
        else:
            # 尝试使用自定义列号读取
            try:
                source_col_name = data.columns[self.source_column]
                translation_col_name = data.columns[self.translation_column]
            except IndexError:
                self.logger.error(
                    "无法找到原文或译文列，请检查配置文件中的 source_column 和 translation_column 设置。"
                )
                print("无法找到原文或译文列，请检查配置文件中的自定义表头和列号设置。")
                return

        if source_col_name is None or translation_col_name is None:
            return

        # 确保 DataFrame 至少有两列，并将前两列的名称设置为 "原文" 和 "译文"
        if len(data.columns) >= 2:
            data = data.rename(
                columns={
                    source_col_name: "原文",
                    translation_col_name: "译文",
                }
            )
            data.index = list(data["原文"])  # 将 "原文" 列设置为索引
            data["译文"] = data["译文"].fillna("")  # 将译文中的空值填充为空字符串
            data = self.__RemoveDuplicated(data)  # 去除重复行
            if not namelist:
                namelist = (
                    self.ProgramData.keys()
                )  # 如果未提供文件名列表，则处理所有文件
            # 遍历文件名列表
            for name in namelist:
                if name not in self.ProgramData:  # 检查 name 是否存在
                    self.logger.warning(f"跳过文件 {name}，因为它不在工程数据中")
                    continue
                DataFrame = self.ProgramData[name]  # 获取对应的 DataFrame
                if not samefile:
                    # 如果不是同一文件导入模式，则遍历 DataFrame 的索引并更新译文
                    for untrs in data.index:
                        if untrs in DataFrame.index:
                            DataFrame.loc[untrs, "译文"] = data.loc[untrs, "译文"]
                else:
                    # 如果是同一文件导入模式，则直接将 data 中的译文复制到 DataFrame
                    DataFrame["译文"] = data["译文"]
                self.ProgramData[name] = DataFrame  # 更新 DataFrame

    def InputFromeXlsx(self, path: str, namelist: list = None, samefile=False):
        """
        从 xlsx 文件导入翻译数据。

        :param path: xlsx 文件所在的路径。
        :param namelist: 要导入的文件名列表（可选，默认为所有文件）。
        :param samefile: 是否为同一文件导入模式（默认为 False）。
        """
        FileNames = self.__ReadFolder(path)  # 读取指定路径下的所有文件
        # 遍历所有文件
        for file in FileNames:
            if file.endswith(".xlsx"):
                name = os.path.basename(file)  # 获取文件名
                print(f"正在导入{name}")
                data = self.__Readxlsx(file)  # 读取xlsx文件
                if data is not None:  # 检查数据是否成功读取
                    if not samefile:
                        # 如果不是同一文件导入模式，则直接调用 InputFromDataFrame
                        self.InputFromDataFrame(data, namelist)
                    else:
                        # 如果是同一文件导入模式，则检查是否存在对应的 JSON 文件，并进行导入
                        jsonname = self.__nameswitch(file)
                        if jsonname in self.ProgramData.keys():
                            self.InputFromDataFrame(data, [jsonname], True)
                        else:
                            print(f"{name}没有对应的json文件，已跳过导入该文件")
                else:
                    print(f"导入{name}失败，请检查文件")

    def load(self, path: str):
        """
        从指定路径加载翻译工程。

        :param path: 翻译工程所在的路径。
        """
        self.ProgramData = {}  # 清空 ProgramData
        project_path = os.path.join(path, self.project_dir_name)  # 构造项目目录的路径
        project_data_path = os.path.join(
            project_path, self.project_data_dir
        )  # 构造项目数据目录的路径
        if not os.path.exists(project_data_path):  # 检查路径是否存在
            self.logger.warning(f"工程数据路径 {project_data_path} 不存在，加载失败")
            print(f"工程数据路径 {project_data_path} 不存在，加载失败")
            return

        FileNames = self.__ReadFolder(project_data_path)  # 读取 project_data_path

        for file in FileNames:
            name = file.split(os.sep + self.project_data_dir + os.sep)[
                -1
            ]  # 使用 self.project_data_dir
            if name.endswith(".csv"):
                csv_file_path = os.path.join(project_data_path, name)  # 完整CSV文件路径
                if os.path.exists(csv_file_path):  # 再次检查文件是否存在
                    print(f"正在加载 {name}")
                    try:
                        # 读取 CSV 文件，使用 UTF-8 编码，并指定引擎为 Python，数据类型为字符串
                        data = pd.read_csv(
                            csv_file_path,
                            sep=",",
                            encoding="utf8",
                            engine="python",
                            dtype=str,
                        )
                        # 检查 CSV 文件的列名是否正确
                        if not list(data.columns) == [
                            "原文",
                            "译文",
                            "地址",
                            "标签",
                            "code",
                        ]:
                            self.logger.warning(
                                f"{file} 文件列名不为 ['原文','译文','地址','标签','code']，读取失败"
                            )
                            print(
                                f"{file}文件列名不为['原文','译文','地址','标签','code']，读取失败"
                            )
                            continue
                        data.index = list(data["原文"])  # 将 "原文" 列设置为索引
                        data["译文"] = data["译文"].fillna(
                            ""
                        )  # 将译文中的空值填充为空字符串
                        data["标签"] = data["标签"].fillna(
                            ""
                        )  # 将标签中的空值填充为空字符串
                        data = self.__RemoveDuplicated(data)  # 去除重复行
                        jsonname = self.__nameswitch(
                            name, True
                        )  # 获取对应的 JSON 文件名
                        self.ProgramData.update(
                            {jsonname: data}
                        )  # 将 DataFrame 添加到 ProgramData 中
                    except Exception as e:
                        self.logger.error(f"加载CSV文件 {file} 失败")
                        self.logger.error(traceback.format_exc())
                        print(f"加载 {file} 失败")
                else:
                    self.logger.warning(f"CSV 文件 {csv_file_path} 不存在")
                    print(f"CSV 文件 {csv_file_path} 不存在")

        self.logger.info("工程加载完成")
        print("########################加载工程完成########################")

    def addlabel(self, target: dict, label: str):
        """
        给指定的文本行添加标签。

        :param target: 包含要添加标签的文本行信息的字典，格式为 {文件名: [索引列表]}。
        :param label: 要添加的标签。
        """
        for name, indices in target.items():  # 遍历字典的键值对
            if name not in self.ProgramData:  # 检查 name 是否存在
                self.logger.warning(f"跳过文件 {name}，因为它不在工程数据中")
                print(f'在工程中没有找到文件: "{name}"')
                continue
            for index in indices:  # 遍历索引列表
                if index in self.ProgramData[name].index:  # 检查索引是否存在
                    # 如果标签不在当前行的标签中，则添加标签
                    if label not in self.ProgramData[name].loc[index, "标签"]:
                        self.ProgramData[name].loc[index, "标签"] = (
                            self.ProgramData[name].loc[index, "标签"] + "," + label
                        ).lstrip(
                            ","
                        )  # 添加标签并去除开头的逗号
                else:
                    print(f'在{name}中没有找到原文为:"{index}"的行')
        print(f'添加标签"{label}"完成')

    def removelabel(self, target: dict, label: str):
        """
        从指定的文本行移除标签。

        :param target: 包含要移除标签的文本行信息的字典，格式为 {文件名: [索引列表]}。
        :param label: 要移除的标签。
        """
        for name, indices in target.items():  # 遍历字典的键值对
            if name not in self.ProgramData:  # 检查 name 是否存在
                self.logger.warning(f"跳过文件 {name}，因为它不在工程数据中")
                print(f'在工程中没有找到文件: "{name}"')
                continue
            for index in indices:  # 遍历索引列表
                if index in self.ProgramData[name].index:  # 检查索引是否存在
                    # 如果标签在当前行的标签中，则移除标签
                    if label in self.ProgramData[name].loc[index, "标签"]:
                        self.ProgramData[name].loc[index, "标签"] = (
                            ("," + self.ProgramData[name].loc[index, "标签"])
                            .replace(("," + label), "")  # 移除标签
                            .lstrip(",")  #  去除开头的逗号
                        )
                    else:
                        print(f'{name}:"{index}"行没有{label}标签')
                else:
                    print(f'在{name}中没有找到原文为:"{index}"的行')
        print(f'去除标签"{label}"完成')

    def Display(self, target: dict = None, namelist: list = None):
        """
        显示指定的文本行信息。

        :param target: 要显示的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要显示的文件名列表（可选，默认为所有文件）。
        """
        if target is None:
            target = self.ProgramData.copy()  # 如果未提供 target，则显示所有数据
        if namelist is None:
            namelist = target.keys()  # 如果未提供 namelist，则显示所有文件
        for key in namelist:  # 遍历文件名列表
            if key in target.keys():  # 检查文件名是否存在
                print(f"{key}:")
                # 打印 DataFrame，重置索引以避免显示原文索引
                print(target[key].reset_index(drop=True, inplace=False))
            else:
                print(f"{key}不在目标范围内")

    def search(
        self,
        string: str,
        col: int,
        target: dict = None,
        namelist: list = None,
        notin: bool = False,
        BigSmall=False,
        regex=True,
    ) -> dict:
        """
        在指定的列中搜索包含指定字符串的文本行。

        :param string: 要搜索的字符串。
        :param col: 要搜索的列的索引（0: 原文, 1: 译文, 2: 地址, 3: 标签, 4: code）。
        :param target: 要搜索的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要搜索的文件名列表（可选，默认为所有文件）。
        :param notin: 是否搜索不包含指定字符串的文本行（默认为 False）。
        :param BigSmall: 是否区分大小写 (默认 False, 即不区分大小写)。
        :param regex: 是否使用正则表达式搜索（默认为 True）。
        :return: 包含搜索结果的字典，格式为 {文件名: DataFrame}。
        """
        # 根据列索引确定列名
        if col == 0:
            col_name = "原文"
        elif col == 1:
            col_name = "译文"
        elif col == 2:
            col_name = "地址"
        elif col == 3:
            col_name = "标签"
        elif col == 4:
            col_name = "code"
        else:
            self.logger.error(f"无效的搜索列索引: {col}")
            print(f"无效的搜索列索引: {col}")
            return {}

        res = {}  # 存储搜索结果的字典
        if target is None:
            target = self.ProgramData.copy()  # 如果未提供 target，则搜索所有数据
        if namelist is None:
            namelist = target.keys()  # 如果未提供 namelist，则搜索所有文件

        for name in namelist:  # 遍历文件名
            if name not in target:  # 检查 name 是否存在
                self.logger.warning(f"跳过文件 {name}，因为它不在搜索目标中")
                print(f"{name}不在搜索目标范围内")
                continue

            DataFrame = target[name].copy()  # 获取对应的 DataFrame
            if string == "":
                # 如果搜索字符串为空，则查找指定列为空的行
                temp = DataFrame[DataFrame[col_name] == ""]
            else:
                # 根据 notin 参数决定是搜索包含还是不包含指定字符串的行
                if notin:
                    temp = DataFrame[
                        ~DataFrame[col_name].str.contains(
                            string, case=not BigSmall, regex=regex
                        )
                    ]
                else:
                    temp = DataFrame[
                        DataFrame[col_name].str.contains(
                            string, case=not BigSmall, regex=regex
                        )
                    ]
            if not temp.empty:
                res.update({name: temp})  # 将搜索结果添加到字典中
        return res

    def DoubleSearch(
        self,
        A: str,
        B: str,
        colA: int,
        colB: int = None,
        target: dict = None,
        namelist: list = None,
        BigSmall=False,
        regex=True,
    ):
        """
        进行双重搜索，先搜索包含字符串 A 的文本行，然后在结果中搜索不包含字符串 B 的文本行。

        :param A: 第一个要搜索的字符串。
        :param B: 第二个要搜索的字符串。
        :param colA: 第一个搜索的列的索引。
        :param colB: 第二个搜索的列的索引（可选，默认为与 colA 相同）。
        :param target: 要搜索的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要搜索的文件名列表（可选，默认为所有文件）。
        :param BigSmall: 是否区分大小写 (默认 False, 即不区分大小写)。
        :param regex: 是否使用正则表达式搜索（默认为 True）。
        :return: 包含搜索结果的字典，格式为 {文件名: DataFrame}。
        """
        # 进行第一次搜索
        res = self.search(
            A, colA, target, namelist, notin=False, BigSmall=BigSmall, regex=regex
        )
        if colB is None:
            colB = colA  # 如果未提供 colB，则使用与 colA 相同的列进行第二次搜索
        # 进行第二次搜索
        return self.search(B, colB, res, namelist, True, BigSmall, regex)

    def Replace(
        self, before: str, after: str, target: dict = None, namelist: list = None
    ):
        """
        在指定的文本行中进行文本替换。

        :param before: 要替换的字符串。
        :param after: 替换后的字符串。
        :param target: 要进行替换的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要进行替换的文件名列表（可选，默认为所有文件）。
        :return: 更新后的 target 字典。
        """
        if target is None:
            target = self.ProgramData  # 如果未提供 target，则替换所有数据
        if namelist is None:
            namelist = target.keys()  # 如果未提供 namelist，则替换所有文件
        for name in namelist:  # 遍历文件名列表
            if name not in target:  # 检查 name 是否存在
                self.logger.warning(f"跳过文件 {name}，因为它不在替换目标中")
                print(f"{name}不在替换目标范围内")
                continue
            for index in target[name].index:  # 遍历索引
                # 进行文本替换
                target[name].loc[index, "译文"] = (
                    target[name].loc[index, "译文"].replace(before, after)
                )
        return target

    def LabelBySearch(
        self,
        string: str,
        col: int,
        label: str,
        target: dict = None,
        namelist: list = None,
        notin: bool = False,
        BigSmall=False,
        add=True,
        regex=True,
    ):
        """
        根据搜索结果给文本行添加或移除标签。

        :param string: 要搜索的字符串。
        :param col: 要搜索的列的索引。
        :param label: 要添加或移除的标签。
        :param target:  要进行操作的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist:  要进行操作的文件名列表（可选，默认为所有文件）。
        :param notin: 是否搜索不包含指定字符串的文本行（默认为 False）。
        :param BigSmall: 是否区分大小写(默认为 False)。
        :param add: 是否添加标签（默认为 True），如果为 False，则移除标签。
        :param regex: 是否使用正则表达式（默认为 True）。
        :return: 包含搜索结果的字典，格式为 {文件名: DataFrame}。
        """
        # 进行搜索
        res = self.search(
            string,
            col,
            target=target,
            namelist=namelist,
            notin=notin,
            BigSmall=BigSmall,
            regex=regex,
        )
        if res:
            label_target = {}  # 存储要添加或移除标签的文本行的信息
            for name, df in res.items():  # 遍历搜索结果
                label_target[name] = list(
                    df.index
                )  # 获取要添加或移除标签的文本行的索引
            # 根据 add 参数决定是添加还是移除标签
            if add:
                self.addlabel(label_target, label)
            else:
                self.removelabel(label_target, label)
        else:
            print("搜索结果为空")
        return res

    def DisplayBySearch(
        self,
        string: str,
        col: int,
        target: dict = None,
        namelist: list = None,
        notin: bool = False,
        BigSmall=False,
        regex=True,
    ):
        """
        根据搜索结果显示文本行信息。

        :param string: 要搜索的字符串。
        :param col: 要搜索的列的索引。
        :param target: 要显示的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要显示的文件名列表（可选，默认为所有文件）。
        :param notin: 是否搜索不包含指定字符串的文本行（默认为 False）。
        :param BigSmall: 是否区分大小写(默认为 False)。
        :param regex: 是否使用正则表达式搜索（默认为 True）。
        :return: 包含搜索结果的字典，格式为 {文件名: DataFrame}。
        """
        # 进行搜索
        res = self.search(
            string,
            col,
            target=target,
            namelist=namelist,
            notin=notin,
            BigSmall=BigSmall,
            regex=regex,
        )
        if res:
            self.Display(res)  # 显示搜索结果
        else:
            print("搜索结果为空")
        return res

    def OutputBySearch(
        self,
        string: str,
        col: int,
        target: dict = None,
        namelist: list = None,
        notin: bool = False,
        BigSmall=False,
        OutputName: str = "SearchRes.xlsx",
        full=False,
        regex=True,
    ):
        """
        根据搜索结果将文本行导出为 xlsx 文件。

        :param string: 要搜索的字符串。
        :param col: 要搜索的列的索引。
        :param target: 要导出的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要导出的文件名列表（可选，默认为所有文件）。
        :param notin: 是否搜索不包含指定字符串的文本行（默认为 False）。
        :param BigSmall: 是否区分大小写(默认为 False)。
        :param OutputName: 输出文件名（默认为 "SearchRes.xlsx"）。
        :param full: 是否导出所有列（默认为 False）。
        :param regex: 是否使用正则表达式搜索（默认为 True）。
        :return: 包含搜索结果的字典，格式为 {文件名: DataFrame}。
        """
        # 进行搜索
        res = self.search(
            string,
            col,
            target=target,
            namelist=namelist,
            notin=notin,
            BigSmall=BigSmall,
            regex=regex,
        )
        if res:
            output = pd.concat(list(res.values()), axis=0)  # 合并搜索结果
            self.__Writexlsx(output, OutputName, full)  # 将搜索结果导出为 xlsx 文件
            print(f"已将搜索结果保存为{OutputName}")
        else:
            print("搜索结果为空")
        return res

    def JsonBySearch(
        self,
        string: str,
        col: int,
        target: dict = None,
        namelist: list = None,
        notin: bool = False,
        BigSmall=False,
        OutputName: str = "SearchRes.json",
        regex=True,
    ):
        """
        根据搜索结果将文本行导出为 JSON 文件。

        :param string: 要搜索的字符串。
        :param col: 要搜索的列的索引。
        :param target: 要导出的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要导出的文件名列表（可选，默认为所有文件）。
        :param notin: 是否搜索不包含指定字符串的文本行（默认为 False）。
        :param BigSmall: 是否区分大小写(默认为 False)。
        :param OutputName: 输出文件名（默认为 "SearchRes.json"）。
        :param regex: 是否使用正则表达式搜索（默认为 True）。
        :return: 包含搜索结果的字典，格式为 {原文: 译文}。
        """
        # 进行搜索
        res = self.search(
            string,
            col,
            target=target,
            namelist=namelist,
            notin=notin,
            BigSmall=BigSmall,
            regex=regex,
        )
        if res:
            res_concat = pd.concat(list(res.values()), axis=0)  # 合并搜索结果
            # 将搜索结果转换为字典，格式为 {原文: 译文}
            output = dict(zip(res_concat["原文"], res_concat["译文"].fillna("")))
            try:
                # 将字典导出为 JSON 文件，使用缩进和确保 ASCII 字符不转义
                with open(OutputName, "w", encoding="utf8") as f:
                    json.dump(output, f, indent=4, ensure_ascii=False)
                print(f"已将结果导出为{OutputName}")
            except Exception as e:
                self.logger.error(f"导出JSON文件 {OutputName} 失败")
                self.logger.error(traceback.format_exc())
                print(f"导出JSON文件 {OutputName} 失败")
            return output
        else:
            print("搜索结果为空")
            return {}

    def __IfAllBlack(self, string: str, code: bool) -> bool:
        """
        检查给定的字符串（地址或 code）是否全部在黑名单中。

        :param string: 要检查的字符串（地址或 code）。
        :param code: 是否为 code 模式（True 表示 code 模式，False 表示地址模式）。
        :return: 如果全部在黑名单中，则返回 True，否则返回 False。
        """
        blacklist = (
            self.BlackDir if not code else self.BlackCode
        )  # 根据模式选择黑名单列表
        qlist = (
            string.split("☆↑↓") if not code else string.split(",")
        )  # 根据模式分割字符串
        for black in blacklist:  # 遍历黑名单
            sig = True
            pattern = re.compile(black)  # 编译正则表达式
            for q in qlist:  # 遍历分割后的列表
                # 如果有任何一个元素不在黑名单中，则将 sig 设置为 False
                if not re.search(pattern, q):
                    sig = False
                    break
            if sig:
                return True  # 如果所有元素都在黑名单中，则返回 True
        return False

    def LabelBlack(self):
        """
        标记黑名单中的文本行，并应用原文。
        """
        target = {}  # 存储要标记的文本行的信息
        # 遍历所有 DataFrame
        for name, df in self.ProgramData.items():
            # 查找地址在黑名单中的文本行的索引
            black_dir_indices = df[
                df["地址"].apply(lambda x: self.__IfAllBlack(x, False))
            ].index.tolist()
            # 查找 code 在黑名单中的文本行的索引
            black_code_indices = df[
                df["code"].apply(lambda x: self.__IfAllBlack(x, True))
            ].index.tolist()
            # 合并两个索引列表，并去除重复项
            combined_indices = list(set(black_dir_indices + black_code_indices))
            if combined_indices:
                target[name] = combined_indices  # 将要标记的文本行的信息添加到字典中
        if target:
            self.addlabel(target, "Black")  # 添加 "Black" 标签
            self.ApplyUntrs_BySearch(
                "Black", 3
            )  # 将标记为 "Black" 的文本行的译文替换为原文
        print("########################已标记黑名单并应用原文########################")

    def LabelName(self, without: list = None):
        """
        标记包含 "name" 的文本行。

        :param without:  排除列表
        """
        target = self.ProgramData  # 获取所有数据
        # 如果提供了排除列表，则先排除包含指定字符串的文本行
        if without:
            for i in without:
                target = self.search(i, 2, target=target, notin=True, regex=False)
        # 标记包含 "name" 的文本行
        self.LabelBySearch("name", 2, "Name", target=target, BigSmall=True, regex=False)

    def GetName(self, data_path, without: list = None):
        """
        提取人名并保存到 JSON 文件。

        :param data_path: 数据路径，用于创建输出目录。
        :param without: 排除列表，用于排除不需要提取人名的文件。
        """
        path = os.path.join(data_path, "name")  # 构造输出目录的路径
        os.makedirs(path, exist_ok=True)  # 确保输出目录存在
        self.LabelName(without) # 标记地址中包含 "name" 的文本行，以便后续提取。
        # 搜索标签为 "Name" 的文本行，并导出为 JSON 文件
        namedict = self.JsonBySearch(
            "Name", 3, OutputName=os.path.join(path, "Name.json"), regex=False
        )
        if self.ja and namedict:  # 检查 namedict 是否为空
            splited_name = {}  # 存储分割后的人名
            # 遍历提取的人名
            for name in namedict.keys():
                # 使用正则表达式分割人名
                namelist = re.sub(
                    "[^\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5ー々〆〤]",
                    "↓☆←",
                    name,
                ).split("↓☆←")
                trsednamelist = re.sub(
                    "[^\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5ー々〆〤]",
                    "↓☆←",
                    name,
                ).split("↓☆←")
                # 如果分割后的原文和译文列表长度相同，则添加到字典中
                if len(namelist) == len(trsednamelist):
                    splited_name.update(dict(zip(namelist, trsednamelist)))
                else:
                    splited_name.update(
                        dict(zip(namelist, namelist))
                    )  # 如果长度不同，则使用原文
            if "" in splited_name:
                del splited_name[""]  # 删除空字符串
            try:
                # 将分割后的人名导出为 JSON 文件
                with open(os.path.join(path, "Name.json"), "w", encoding="utf8") as f:
                    json.dump(splited_name, f, indent=4, ensure_ascii=False)
            except Exception as e:
                self.logger.error(f"导出JSON文件 Name.json 失败")
                self.logger.error(traceback.format_exc())
                print(f"导出Name.json失败")

    def ApplyUntrs(self, target: dict = None):
        """
        将指定文本行的译文替换为原文。

        :param target: 要进行替换的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        """
        if target is None:
            target = self.ProgramData  # 如果未提供 target，则替换所有数据
        for name, df in target.items():  # 遍历字典
            if name not in self.ProgramData:  # 检查 name 是否存在
                self.logger.warning(f"跳过文件 {name}，因为它不在工程数据中")
                print(f'在工程中没有找到文件: "{name}"')
                continue
            for index in df.index:  # 遍历索引
                if index in self.ProgramData[name].index:  # 检查索引是否存在
                    # 将译文替换为原文
                    self.ProgramData[name].loc[index, "译文"] = index
        print("应用原文完成")

    def ApplyUntrs_BySearch(
        self,
        string: str,
        col: int,
        target: dict = None,
        namelist: list = None,
        notin: bool = False,
        BigSmall=False,
        regex=True,
    ):
        """
        根据搜索结果将文本行的译文替换为原文。

        :param string: 要搜索的字符串。
        :param col: 要搜索的列的索引。
        :param target: 要进行替换的文本行信息的字典，格式为 {文件名: DataFrame}（可选，默认为所有数据）。
        :param namelist: 要进行替换的文件名列表（可选，默认为所有文件）。
        :param notin: 是否搜索不包含指定字符串的文本行（默认为 False）。
        :param BigSmall: 是否区分大小写(默认为 False)。
        :param regex: 是否使用正则表达式搜索（默认为 True）。
        """
        # 进行搜索
        res = self.search(
            string,
            col,
            target=target,
            namelist=namelist,
            notin=notin,
            BigSmall=BigSmall,
            regex=regex,
        )
        if res:
            self.ApplyUntrs(res)  # 将搜索结果中的文本行的译文替换为原文
        else:
            print("搜索结果为空")

    def AddMark(self, mark: str):
        """
        给游戏标题添加水印。

        :param mark: 要添加的水印。
        """
        if "System.json" in self.ProgramData:  # 检查是否存在 System.json 文件
            try:
                data = self.ProgramData["System.json"]  # 获取 System.json 的 DataFrame
                # 获取游戏标题的索引
                index = list(
                    data[data["地址"] == r"System.json\gameTitle\u200B1"].index
                )[0]
                # 如果游戏标题的译文为空，则先将译文设置为原文
                if not self.ProgramData["System.json"].loc[index, "译文"]:
                    self.ProgramData["System.json"].loc[index, "译文"] = index
                # 在游戏标题的译文后添加水印
                self.ProgramData["System.json"].loc[index, "译文"] += mark
                print("########################已添加水印########################")
            except Exception as e:
                self.logger.error("添加水印失败，未找到游戏标题")
                self.logger.error(traceback.format_exc())
                print(f"没有找到游戏标题，添加水印失败")

    def __splitbychar(self, q, l):
        """
        根据给定的分隔符分割字符串。

        :param q: 要分割的字符串。
        :param l: 分隔符列表（包含两个分隔符）。
        :return: 分割后的字符串列表。
        """
        b = []  # 存储分割后的字符串的列表
        # 如果两个分隔符都在字符串中，则进行分割
        if l[0] in q and l[1] in q:
            a = q.split(l[0])  # 使用第一个分隔符分割字符串
            b.append(a[0])  # 将第一个分割后的字符串添加到列表
            # 遍历剩余的分割后的字符串
            for i in range(1, len(a)):
                a[i] = l[0] + a[i]  # 在每个分割后的字符串前添加第一个分隔符
                c = a[i].split(l[1])  # 使用第二个分隔符分割字符串
                # 遍历分割后的字符串
                for j in range(0, len(c) - 1):
                    b.append(
                        c[j] + l[1]
                    )  # 将分割后的字符串添加到列表，并在末尾添加第二个分隔符
                b.append(c[-1])  # 将最后一个分割后的字符串添加到列表
        return b

    def __dealin(self, untrs, trsed, filename):
        """
        处理包含文件名的特殊文本行。

        :param untrs: 未翻译的文本。
        :param trsed: 翻译后的文本。
        :param filename: 文件名。
        :return: 处理后的翻译后的文本。
        """
        check = False  # 标记是否需要人工检查
        for l in self.codewithnames:  # 遍历需要特殊处理的人名代码
            untrs_list = self.__splitbychar(untrs, l)  # 分割未翻译的文本
            trsed_list = self.__splitbychar(trsed, l)  # 分割翻译后的文本
            length = len(trsed_list)  # 获取分割后的列表的长度
            # 如果未翻译和翻译后的文本分割后的长度不同，则添加到 need2check 字典中
            if len(untrs_list) != length:
                self.need2check[untrs] = trsed
                return trsed
            elif length:
                if length > 3:
                    check = True  # 如果分割后的列表长度大于 3，则标记为需要人工检查
                # 根据分隔符的不同情况进行处理
                if l[0] != l[1]:
                    # 如果两个分隔符不同，则处理包含文件名的部分
                    for i in range(0, int(length / 2)):
                        if filename in untrs_list[2 * i + 1]:
                            trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                else:
                    # 如果两个分隔符相同，则处理包含文件名的部分
                    for i in range(0, int(length / 4)):
                        if filename in untrs_list[4 * i + 2]:
                            trsed_list[4 * i + 2] = untrs_list[4 * i + 2]
                trsed = "".join(trsed_list)  # 将分割后的列表重新组合成字符串
        if check:
            self.need2check[untrs] = (
                trsed  # 如果标记为需要人工检查，则添加到 need2check 字典中
            )
        return trsed

    def dnb(self, GameDir, files: list = None):
        """
        修正文件名相关的翻译。

        :param GameDir: 游戏目录。
        :param files: 文件名列表（可选）。
        """
        print("开始修正文件名")
        self.need2check = {}  # 存储可能需要人工修正的文本行
        if files is None:
            # 如果未提供文件名列表，则读取游戏目录下的所有文件
            temp = self.__ReadFolder(GameDir)
            files = []
            for filename in temp:
                filename = os.path.basename(filename)  # 获取文件名
                # 如果文件名包含日语字符或者不是日语模式，则添加到文件列表
                if (
                    re.search(
                        r"[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5ー々〆〤]",
                        filename,
                    )
                    or not self.ja
                ):
                    files.append(filename)
                    base, ext = os.path.splitext(filename)  # 分离文件名和扩展名
                    files.append(base)  # 将文件名（不包含扩展名）添加到文件列表
        # 遍历所有 DataFrame
        for name, df in self.ProgramData.items():
            for index in df.index:  # 遍历索引
                if isinstance(index, str):
                    for filename in files:  # 遍历文件名列表
                        # 如果索引与文件名相同，则将译文设置为文件名
                        if index == filename:
                            self.ProgramData[name].loc[index, "译文"] = index
                        # 如果文件名包含在索引中，并且地址中不包含 "note"，则调用 __dealin 处理
                        elif (
                            filename in index
                            and "note" not in self.ProgramData[name].loc[index, "地址"]
                        ):
                            self.ProgramData[name].loc[index, "译文"] = self.__dealin(
                                index,
                                self.ProgramData[name].loc[index, "译文"],
                                filename,
                            )
        print("########################修正文件名完成########################")
        if self.need2check:
            need2check_file = self.need2check_filename  # 从配置获取文件名
            try:
                # 将可能需要人工修正的文本行导出到 JSON 文件
                with open(need2check_file, "w", encoding="utf8") as f1:
                    json.dump(self.need2check, f1, indent=4, ensure_ascii=False)
                print(f"已将可能需要人工修正的文本行导出到 {need2check_file}")
            except Exception as e:
                self.logger.error(f"导出 need2check.json 文件失败")
                self.logger.error(traceback.format_exc())
                print(f"导出 {need2check_file} 失败")

    def AutoLineFeed(self, linelength: int):
        """
        自动换行。

        :param linelength: 每行的最大长度。
        """
        for name, DataFrame in self.ProgramData.items():  # 遍历所有 DataFrame
            for index in DataFrame.index:  # 遍历索引
                data = DataFrame.loc[index, "译文"]  # 获取译文
                lines = data.split("\n")  # 按换行符分割文本
                res = ""  # 存储处理后的文本
                for line in lines:  # 遍历每一行
                    q = line  # 当前行
                    # 如果当前行的长度大于最大长度，则进行换行
                    while len(q) > linelength:
                        # 查找下一个日语字符的位置
                        n = re.search(
                            r"[\u4e00-\u9fff\u3000-\u303f\uff01-\uffef\u2026《》【】]",
                            q[linelength:],
                        )
                        # 如果找到日语字符，则在日语字符前换行
                        if n:
                            n = n.span()[0] + linelength
                        else:
                            break  # 如果没有找到日语字符，则停止换行
                        res += q[:n] + "\n"  # 添加换行符
                        q = q[n:]  # 更新当前行
                    res += q + "\n"  # 添加剩余的文本和换行符
                DataFrame.loc[index, "译文"] = res.rstrip(
                    "\n"
                )  # 更新译文，并去除末尾的换行符
            self.ProgramData[name] = DataFrame  # 更新 DataFrame
        print(
            "########################已按每行{}字自动换行########################".format(
                linelength
            )
        )

    def AutoLineFeed_js(self, GameDir):
        """
        安装自动换行插件（JS 版）。

        :param GameDir: 游戏目录。
        """
        dirlist = self.__ReadFolder(GameDir)  # 读取游戏目录下的所有文件
        pgsdir = None  # 存储 plugins.js 文件的路径
        # 查找 plugins.js 文件
        for fdir in dirlist:
            if os.path.basename(fdir) == "plugins.js" and "js" in fdir:
                pgsdir = fdir
                pgsfolder = os.path.join(
                    os.path.dirname(pgsdir), "plugins"
                )  # 构造插件目录的路径
                break

        if pgsdir:  # 检查 pgsdir 是否为 None
            try:
                # 读取 plugins.js 文件的内容
                with open(pgsdir, "rb") as f:
                    encoding = detect(f.read())["encoding"]
                    encoding = encoding if encoding else "ansi"
                with open(pgsdir, "r", encoding=encoding) as f:
                    plugins_js = f.read()

                temp0 = "plugins =\n[\n"  # 定义 plugins 变量的开头
                body = plugins_js.split(temp0)  # 分割 plugins.js 文件的内容
                # 定义要添加的自动换行插件的配置
                addline = '{"name":"自动换行","status":true,"description":"auto linefeed","parameters":{}},\n'
                # 如果 plugins.js 文件中没有自动换行插件的配置，则添加
                if addline not in body[1]:
                    body[1] = addline + body[1]
                plugins_js = body[0] + temp0 + body[1]  # 重新组合 plugins.js 文件的内容

                # 将更新后的 plugins.js 文件写回
                with open(pgsdir, "w", encoding=encoding) as f:
                    f.write(plugins_js)

                js_file = self.AutoLineFeed_jsdir  # 从配置读取js文件名
                output_js_path = os.path.join(pgsfolder, js_file)  # 使用配置中的文件名

                # 读取自动换行插件的 JS 代码
                with open(js_file, "r", encoding=encoding) as f:  # 使用配置中的文件名
                    jscode = f.read()

                # 将自动换行插件的 JS 代码写入插件目录
                with open(output_js_path, "w", encoding=encoding) as f:
                    f.write(jscode)

                print(
                    "已将自动换行插件复制到游戏插件目录，如果出现bug，请自行修改plugins.js文件\n 将自动换行那一行删掉"
                )
            except Exception as e:
                self.logger.error("自动换行插件JS脚本安装失败")
                self.logger.error(traceback.format_exc())
                print("自动换行插件JS脚本安装失败，请检查错误日志")
        else:
            print("plugins.js 文件未找到，自动换行插件JS脚本安装跳过")

    def checknum(self):
        """
        检查文本中特定字符串的数量是否一致。
        """
        count = 0  # 记录不一致的数量
        checkdict_file = "checkdict.json"  # 直接使用文件名
        try:
            # 读取 checkdict.json 文件
            with open(checkdict_file, "r", encoding="utf8") as f:
                tempdict = json.load(f)
            # 将字典的键按长度降序排序
            sortedkey = sorted(tempdict.keys(), key=lambda x: len(x), reverse=True)
            checkdict = {i: tempdict[i] for i in sortedkey}  # 创建一个新的排序后的字典
        except Exception as e:
            self.logger.error(f"加载 {checkdict_file} 失败")
            self.logger.error(traceback.format_exc())
            print(
                f"没有找到格式正确的{checkdict_file}文件，请确认路径设置是否正确以及文件是否符合json格式"
            )
            return

        res = {}  # 存储检查结果的字典
        fixdict = {}  # 存储修正字典的字典
        print("处理中，请稍候,根据checkdict的长度，可能会花费较长时间")
        # 遍历 checkdict
        for CheckUntrs, CheckTrsed in checkdict.items():  # 直接迭代字典的 items()
            tempdict = {}  # 存储当前检查项的结果
            # 遍历所有 DataFrame
            for name, DataFrame in self.ProgramData.items():
                for untrs, row in DataFrame.iterrows():  # 使用 iterrows()
                    trsed = row["译文"]  # 获取译文
                    # 如果原文和译文中特定字符串的数量不一致，并且原文中包含该字符串，则添加到 tempdict 中
                    if (
                        untrs.count(CheckUntrs) != trsed.count(CheckTrsed)
                        and untrs.count(CheckUntrs) > 0
                    ):
                        tempdict[untrs] = trsed
            if tempdict:
                res[CheckUntrs] = tempdict  # 将当前检查项的结果添加到 res 字典中
                fixdict[CheckUntrs] = [CheckTrsed]  # 将当前检查项添加到 fixdict 字典中
                count += len(tempdict)  # 更新不一致的数量

        checkres_file = "checkres.json"  # 直接使用文件名
        try:
            # 将检查结果导出到 checkres.json 文件
            with open(checkres_file, "w", encoding="utf8") as f1:
                json.dump(res, f1, indent=4, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"导出 {checkres_file} 失败")
            self.logger.error(traceback.format_exc())
            print(f"导出 {checkres_file} 失败")
        print(f"已将检查结果保存到{checkres_file}")

        fixdict_file = "fixdict.json"  # 直接使用文件名
        overw = True  # 标记是否覆盖 fixdict.json 文件
        # 如果 fixdict.json 文件已存在，则询问用户是否覆盖
        if os.path.exists(fixdict_file):
            inp = ""  # 初始化 inp
            overw = False
            while inp not in ["y", "n"]:
                inp = input(f"{fixdict_file}已存在，请问要覆盖吗（y,n)\n").lower()
            if inp == "y":
                overw = True
        if overw:
            try:
                # 将修正字典导出到 fixdict.json 文件
                with open(fixdict_file, "w", encoding="utf8") as f:  # 直接使用文件名
                    json.dump(fixdict, f, indent=4, ensure_ascii=False)
                print(
                    "对应的修正词典正确翻译导出为fixdict.json\n"
                    "修正词典的格式为：\n"
                    "{\n"
                    '"原文1":["正确译文","错误译文1","错误译文2"],\n'
                    '"原文2":["正确译文","错误译文1","错误译文2"]\n'
                    "}"
                )
            except Exception as e:
                self.logger.error(f"导出 {fixdict_file} 失败")
                self.logger.error(traceback.format_exc())
                print(f"导出 {fixdict_file} 失败")

        print("########################核对完毕########################")

    def fixnum(self):
        """
        应用 checknum 的检查结果，修正文本中特定字符串的数量不一致的问题。
        """
        checkres_file = "checkres.json"  # 直接使用文件名
        fixdict_file = "fixdict.json"  # 直接使用文件名
        try:
            # 读取 checkres.json 文件
            with open(checkres_file, "r", encoding="utf8") as f:
                checkres = json.load(f)
        except Exception as e:
            self.logger.error(f"加载 {checkres_file} 失败")
            self.logger.error(traceback.format_exc())
            print(
                f"没有找到格式正确的{checkres_file}文件，请确认文件是否存在且符合json格式"
            )
            return

        try:
            # 读取 fixdict.json 文件
            with open(fixdict_file, "r", encoding="utf8") as f:
                tempdict = json.load(f)
            # 将字典的键按长度降序排序
            sortedkey = sorted(tempdict.keys(), key=lambda x: len(x), reverse=True)
            fixdict = {i: tempdict[i] for i in sortedkey}  # 创建一个新的排序后的字典
        except Exception as e:
            self.logger.error(f"加载 {fixdict_file} 失败")
            self.logger.error(traceback.format_exc())
            print(
                f"没有找到格式正确的{fixdict_file}文件，请确认路径设置是否正确以及文件是否符合json格式"
            )
            return

        print("处理中，请稍候")
        TrsData = {}  # 存储要修正的文本行
        # 遍历 fixdict
        for fixkey, fixlist in fixdict.items():  # 直接迭代字典的 items()
            if fixkey in checkres:  # 如果 fixkey 在 checkres 中
                righttrs = fixlist[0]  # 获取正确的译文
                for untrs, trsed in checkres[fixkey].items():  # 直接迭代字典的 items()
                    if fixkey in untrs:  # 如果 fixkey 在原文中
                        # 遍历错误译文列表，并进行替换
                        for i in fixlist[1:]:
                            if untrs in TrsData:
                                TrsData[untrs] = TrsData[untrs].replace(i, righttrs)
                            else:
                                checkres[fixkey][untrs] = checkres[fixkey][
                                    untrs
                                ].replace(i, righttrs)
                    if untrs not in TrsData:
                        TrsData[untrs] = checkres[fixkey][
                            untrs
                        ]  # 将修正后的译文添加到 TrsData 中
        self.InputFromJson(trsdata=TrsData)  # 应用修正结果
        print("已应用修正结果")

    def DNoteB(self):
        """
        处理可能存在的 note 问题。
        """
        print("正在处理可能存在的note问题")
        res = self.search("note", 2, regex=False)  # 搜索包含 "note" 的文本行
        for name, DataFrame in res.items():  # 遍历搜索结果
            for untrs, row in DataFrame.iterrows():  # 使用 iterrows()
                l = ["<", ">"]  # 定义分隔符
                trsed = row["译文"]  # 获取译文
                # 如果原文中不包含分隔符，则跳过
                if untrs.count(l[0]) == 0 and untrs.count(l[1]) == 0:
                    continue
                # 如果原文和译文中分隔符的数量一致，则进行处理
                elif (
                    untrs.count(l[0]) == untrs.count(l[1])
                    and untrs.count(l[0]) == trsed.count(l[0])
                    and untrs.count(l[1]) == trsed.count(l[1])
                ):
                    untrs_list = self.__splitbychar(untrs, l)  # 分割原文
                    trsed_list = self.__splitbychar(trsed, l)  # 分割译文
                    length = len(trsed_list)  # 获取分割后的列表的长度
                    # 如果分割后的列表长度不一致，则跳过
                    if length and len(untrs_list) == length:
                        for i in range(0, int(length / 2)):  # 遍历分割后的列表
                            untrs_code = untrs_list[2 * i + 1]  # 获取原文中的 code 部分
                            trsed_code = trsed_list[2 * i + 1]  # 获取译文中的 code 部分
                            # 计算原文中 code 部分的日语字符数量
                            jalen = len(
                                re.findall(
                                    "[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5ー々〆〤]",
                                    untrs_code,
                                )
                            )
                            if jalen == 0:
                                jalen = 1  # 如果日语字符数量为 0，则设置为 1
                            # 如果 code 部分的长度与日语字符数量的比例小于设定的阈值，则进行处理
                            if (
                                (untrs_code.find(":") - untrs_code.find("<")) / jalen
                            ) < self.note_percent:
                                l = ["<", ":"]  # 重新定义分隔符
                                # 如果 code 部分的分隔符数量一致，则进行处理
                                if untrs_code.count(l[0]) == trsed_code.count(
                                    l[0]
                                ) and untrs_code.count(l[1]) == trsed_code.count(l[1]):
                                    untrs_code_list = self.__splitbychar(
                                        untrs_code, l
                                    )  # 分割原文中的 code 部分
                                    trsed_code_list = self.__splitbychar(
                                        trsed_code, l
                                    )  # 分割译文中的 code 部分
                                    length = len(
                                        trsed_code_list
                                    )  # 获取分割后的列表的长度
                                    # 如果分割后的列表长度不一致，则跳过
                                    if length and len(untrs_code_list) == length:
                                        # 将译文中的 code 部分替换为原文中的 code 部分
                                        for j in range(0, int(length / 2)):
                                            trsed_code_list[2 * j + 1] = (
                                                untrs_code_list[2 * j + 1]
                                            )
                                        trsed_code = "".join(
                                            trsed_code_list
                                        )  # 将分割后的列表重新组合成字符串
                                        trsed_list[2 * i + 1] = (
                                            trsed_code  # 更新译文列表
                                        )
                                    else:
                                        # 如果分割后的列表长度不一致，则将译文中的 code 部分替换为原文
                                        trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                                else:
                                    # 如果 code 部分的分隔符数量不一致，则将译文中的 code 部分替换为原文
                                    trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                            else:
                                # 如果 code 部分的长度与日语字符数量的比例大于等于设定的阈值，则将译文中的 code 部分替换为原文
                                trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                        trsed = "".join(trsed_list)  # 将分割后的列表重新组合成字符串
                        DataFrame.loc[untrs, "译文"] = trsed  # 更新译文
                    else:
                        # 如果分割后的列表长度不一致，则将译文替换为原文
                        DataFrame.loc[untrs, "译文"] = untrs
                else:
                    # 如果原文和译文中分隔符的数量不一致，则将译文替换为原文
                    DataFrame.loc[untrs, "译文"] = untrs
            self.InputFromDataFrame(DataFrame, [name])  # 应用更改
        print("########################note处理完毕########################")

    def FromGame(self, GameDir, save_path, data_path):
        """
        从游戏目录提取文本，生成翻译工程。

        :param GameDir: 游戏目录。
        :param save_path: 翻译工程的保存路径。
        :param data_path: 数据文件的保存路径。
        """
        self.ReadGame(GameDir)  # 读取游戏目录
        self.LabelBlack()  # 标记黑名单
        self.GetName(data_path, self.NameWithout)  # 提取人名
        self.Output(data_path)  # 导出文本到 xlsx 文件
        self.Save(save_path)  # 保存翻译工程

    def ToGame(self, GameDir, path, OutputPath, mark: str = None):
        """
        将翻译后的文本注入回游戏目录。

        :param GameDir: 游戏目录。
        :param path: 翻译文件所在的路径。
        :param OutputPath: 输出目录。
        :param mark: 要添加的水印（可选）。
        """
        self.InputFromeXlsx(path)  # 从 xlsx 文件导入翻译数据
        jsonpath = os.path.join(path, "trans.json")  # 构造 trans.json 文件的路径
        # 如果存在 trans.json 文件，则导入翻译数据
        if os.path.exists(jsonpath):
            print("########################正在导入trans.json########################")
            self.InputFromJson(path=jsonpath)
        replacepath = os.path.join(path, "replace.json")  # 构造 replace.json 文件的路径
        # 如果存在 replace.json 文件，则应用替换字典
        if os.path.exists(replacepath):
            print("########################正在应用替换字典########################")
            try:
                with open(replacepath, "r", encoding="utf8") as f:
                    replacedict = json.load(f)
                for key, value in replacedict.items():  # 直接迭代字典的 items()
                    print(f"{key}--->{value}")
                    self.Replace(key, value)  # 进行文本替换
            except Exception as e:
                self.logger.error("加载替换字典失败")
                self.logger.error(traceback.format_exc())
                print("加载替换字典失败，请检查replace.json文件")

        self.dnb(GameDir)  # 修正文件名相关的翻译
        self.DNoteB()  # 处理 note 问题
        if (
            self.line_length and self.line_length != -1
        ):  # 如果设置了自动换行，则进行自动换行
            self.AutoLineFeed(self.line_length)
        if mark:  # 如果提供了水印，则添加水印
            self.AddMark(mark)
        self.InjectGame(GameDir, OutputPath)  # 将翻译后的文本注入回游戏目录
        if self.line_length == -1:  # 如果设置了使用 JS 插件自动换行，则安装插件
            self.AutoLineFeed_js(GameDir)

    def Update(self, GameDir, path, save_path, data_path):
        """
        更新翻译工程。

        :param GameDir: 游戏目录。
        :param path: 翻译文件所在的路径。
        :param save_path: 翻译工程的保存路径。
        :param data_path: 数据文件的保存路径。
        """
        self.ReadGame(GameDir)  # 读取游戏目录
        self.InputFromeXlsx(path)  # 从 xlsx 文件导入翻译数据
        self.GetName(data_path, self.NameWithout)  # 提取人名
        self.Output(data_path)  # 导出文本到 xlsx 文件
        self.Save(save_path)  # 保存翻译工程
