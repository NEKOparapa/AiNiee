
# coding:utf-8               
import copy
import datetime
import json
import random
import re
import os

import shutil
import zipfile


from openpyxl import Workbook  
import ebooklib #需要安装库pip install ebooklib
from ebooklib import epub
from bs4 import BeautifulSoup #需要安装库pip install beautifulsoup4
import openpyxl  #需安装库pip install openpyxl







# 文件读取器
class File_Reader():
    def __init__(self):
        pass


    # 生成项目ID
    def generate_project_id(self,prefix):
        # 获取当前时间，并将其格式化为数字字符串
        current_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

        # 生成5位随机数
        random_number = random.randint(10000, 99999)

        # 组合生成项目ID
        project_id = f"{current_time}{prefix}{random_number}"
        
        return project_id


    # 读取文件夹中树形结构Paratranz json 文件
    def read_paratranz_files(self, folder_path):
        # 待处理的json接口例
        # [
        #     {
        #         "key": "Activate",
        #         "original": "カードをプレイ",
        #         "translation": "出牌",
        #         "context": null
        #     }
        # ]
        # 缓存数据结构示例
        ex_cache_data = [
            {'project_type': 'Paratranz'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json', 'key': 'txtKey',
             'context': ''},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json', 'key': 'txtKey',
             'context': ''},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ',
             'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json',
             'file_name': 'Replace the original text.json', 'key': 'txtKey', 'context': ''},
        ]

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self, "Paratranz")
        json_data_list.append({
            "project_type": "Paratranz",
            "project_id": project_id,
        })

        # 文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file)  # 构建文件路径

                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for item in json_data:
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = item.get('original', '')  # 获取原文，如果没有则默认为空字符串
                            translated_text = item.get('translation', '')  # 获取翻译，如果没有则默认为空字符串
                            key = item.get('key', '')  # 获取键值，如果没有则默认为空字符串
                            context = item.get('context', '')  # 获取上下文信息，如果没有则默认为空字符串
                            storage_path = os.path.relpath(file_path, folder_path)
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "model": "none",
                                "storage_path": storage_path,
                                "file_name": file_name,
                                "key": key,
                                "context": context
                            })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    # 读取文件夹中树形结构Mtool文件
    def read_mtool_files (self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Mtool")
        json_data_list.append({
            "project_type": "Mtool",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for key, value in json_data.items():
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = key
                            translated_text = value
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "model": "none",
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    #读取文件夹中树形结构的xlsx文件， 存到列表变量中
    def read_xlsx_files(self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'T++'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 1},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.xlsx', 'file_name': 'TrsData.xlsx', "row_index": 2},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\text.xlsx', 'file_name': 'text.xlsx', "row_index": 3},
        ]

        # 创建列表
        cache_list = []
        # 添加文件头
        project_id = File_Reader.generate_project_id(self,"T++")
        cache_list.append({
            "project_type": "T++",
            "project_id": project_id,
        })
        #文本索引初始值
        i = 1

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    file_path = os.path.join(root, file) #构建文件路径

                    wb = openpyxl.load_workbook(file_path)
                    sheet = wb.active
                    for row in range(2, sheet.max_row + 1): # 从第二行开始读取，因为第一行是标识头，通常不用理会
                        cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                        cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值

                        source_text = cell_value1  # 获取原文
                        storage_path = os.path.relpath(file_path, folder_path) # 用文件的绝对路径和输入文件夹路径“相减”，获取相对的文件路径
                        file_name = file #获取文件名

                        #第1列的值不为空，和第2列的值为空，是未翻译内容
                        if cell_value1 and cell_value2 is  None:
                            
                            translated_text = "无"
                            cache_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "model": "none",
                                "storage_path": storage_path,
                                "file_name": file_name,
                                "row_index": row ,
                            })

                            i = i + 1 # 增加文本索引值

                        # 第1列的值不为空，和第2列的值不为空，是已经翻译内容
                        elif cell_value1 and cell_value2 :

                            translated_text = cell_value2
                            cache_list.append({
                                "text_index": i,
                                "translation_status": 1,
                                "source_text": source_text,
                                "translated_text": translated_text,
                                "storage_path": storage_path,
                                "model": "none",
                                "file_name": file_name,
                                "row_index": row ,
                            })

                            i = i + 1 # 增加文本索引值



        return cache_list
    

    # 读取文件夹中树形结构VNText导出文件
    def read_vnt_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Vnt")
        json_data_list.append({
            "project_type": "Vnt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".json"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    # 读取 JSON 文件内容
                    with open(file_path, 'r', encoding='utf-8') as json_file:
                        json_data = json.load(json_file)

                        # 提取键值对
                        for entry in json_data:
                            # 根据 JSON 文件内容的数据结构，获取相应字段值
                            source_text = entry["message"]
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file

                            name = entry.get("name")
                            if name:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "name": name,
                                    "model": "none",
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })
                            else:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "model": "none",
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })

                            # 增加文本索引值
                            i = i + 1

        return json_data_list


    #读取缓存文件
    def read_cache_files(self,folder_path):
        # 获取文件夹中的所有文件
        files = os.listdir(folder_path)

        # 查找以 "CacheData" 开头且以 ".json" 结尾的文件
        json_files = [file for file in files if file.startswith("AinieeCacheData") and file.endswith(".json")]

        if not json_files:
            print(f"Error: No 'CacheData' JSON files found in folder '{folder_path}'.")
            return None

        # 选择第一个符合条件的 JSON 文件
        json_file_path = os.path.join(folder_path, json_files[0])

        # 读取 JSON 文件内容
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
            return data


    # 读取文件夹中树形结构Srt字幕文件
    def read_srt_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Srt")
        json_data_list.append({
            "project_type": "Srt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1
        source_text = ''
        subtitle_number = ''
        subtitle_time = ''

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 srt 文件
                if file.endswith(".srt"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    # 将内容按行分割,并除去换行
                    lines = content.split('\n')
                    # 计数变量
                    j = 1

                    # 遍历每一行
                    for line in lines:

                        # 去除行首的BOM（如果存在）
                        line = line.lstrip('\ufeff')

                        # 如果行是数字，代表新的字幕开始
                        if line.isdigit() and (line == str(j)):
                            subtitle_number = line

                        # 时间码行
                        elif ' --> ' in line:
                            subtitle_time = line

                        # 空行代表字幕文本的结束
                        elif line == '':
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": source_text,
                                "model": "none",
                                "subtitle_number": subtitle_number,
                                "subtitle_time": subtitle_time,
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                            # 增加文本索引值
                            i = i + 1
                            j = j + 1
                            # 清空变量
                            source_text = ''
                            subtitle_number = ''
                            subtitle_time = ''

                        # 其他行是字幕文本，需要添加到文本中
                        else:
                            if  source_text:
                                source_text += '\n' + line
                            else:
                                source_text = line

        return json_data_list


    # 读取文件夹中树形结构vtt字幕文件
    def read_vtt_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Vtt")
        json_data_list.append({
            "project_type": "Vtt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1
        source_text = ''
        subtitle_number = ''
        subtitle_time = ''

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 存储开头注释文本
                top_text = ''

                # 判断文件是否为 Vtt 文件
                if file.endswith(".vtt"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    # 将内容按行分割,并除去换行
                    lines = content.split('\n')


                    # 先截取开头注释行
                    top_text = ''
                    for line in lines:
                        # 截断存储开头文本
                        if line.isdigit() and (line == "1"):
                            top_text +='\n'
                            break

                        elif line == '':
                            top_text += '\n'

                        elif line != '':
                            if  top_text:
                                top_text += '\n' + line
                            else:
                                top_text = line


                    # 去除开头注释行
                    cleaned_lines = []
                    in_header = True  # 标记是否在开头注释部分
                    for line in lines:
                        if in_header:
                            if line.isdigit() and line == "1":
                                in_header = False  # 遇到第一个数字行，退出开头注释部分
                        if not in_header:
                            cleaned_lines.append(line)  # 添加非注释部分的行

                    
                   
                    j = 1
                    # 提取字幕行
                    for line in cleaned_lines:

                        # 如果行是数字，代表新的字幕开始
                        if line.isdigit() and (line == str(j)):
                            subtitle_number = line

                        # 时间码行
                        elif '-->' in line:
                            subtitle_time = line

                        # 空行代表字幕文本的结束
                        elif line == '':
                            storage_path = os.path.relpath(file_path, folder_path) 
                            file_name = file
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": source_text,
                                "translated_text": source_text,
                                "model": "none",
                                "subtitle_number": subtitle_number,
                                "subtitle_time": subtitle_time,
                                "top_text": top_text,
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                            # 增加文本索引值
                            i = i + 1
                            j = j + 1
                            # 清空变量
                            source_text = ''
                            subtitle_number = ''
                            subtitle_time = ''

                        # 其他行是字幕文本行
                        else:
                            if  source_text:
                                source_text += '\n' + line
                            else:
                                source_text = line


        return json_data_list

    # 读取文件夹中树形结构Lrc音声文件
    def read_lrc_files (self,folder_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Lrc")
        json_data_list.append({
            "project_type": "Lrc",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1
        subtitle_title = ""

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".lrc"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    # 切行
                    lyrics = content.split('\n')
                    for line in lyrics:

                        # 使用正则表达式匹配标题标签行
                        title_pattern = re.compile(r'\[ti:(.*?)\]')
                        match = title_pattern.search(line)
                        if match:
                            subtitle_title =  match.group(1)  # 返回匹配到的标题全部内容


                        # 使用正则表达式匹配时间戳和歌词内容
                        pattern = re.compile(r'(\[([0-9:.]+)\])(.*)')
                        match = pattern.match(line)
                        if match:
                            timestamp = match.group(2)
                            source_text = match.group(3).strip()
                            if source_text == "":
                                continue
                            storage_path = os.path.relpath(file_path, folder_path)
                            file_name = file

                            if subtitle_title:                             
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "model": "none",
                                    "subtitle_time": timestamp,
                                    "subtitle_title":subtitle_title,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })
                                subtitle_title = ""

                            else:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": source_text,
                                    "translated_text": source_text,
                                    "model": "none",
                                    "subtitle_time": timestamp,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })

                            # 增加文本索引值
                            i += 1

        return json_data_list


    # 读取文件夹中树形结构Txt小说文件
    def read_txt_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Txt")
        json_data_list.append({
            "project_type": "Txt",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 JSON 文件
                if file.endswith(".txt"):
                    file_path = os.path.join(root, file) # 构建文件路径
                    
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()

                    storage_path = os.path.relpath(file_path, folder_path)
                    file_name = file

                    # 切行
                    lines = content.split('\n')


                    for j, line in enumerate(lines):
                        if line.strip() == '': # 跳过空行
                            continue
                        spaces = len(line) - len(line.lstrip()) # 获取行开头的空格数

                        if j < len(lines) - 1 and lines[j + 1].strip() == '': # 检查当前行是否是文本中的最后一行,并检测下一行是否为空行
                            if (j+1) < len(lines) - 1 and lines[j + 2].strip() == '': # 再检查下下行是否为空行，所以最多只会保留2行空行信息
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": line,
                                    "translated_text": line,
                                    "model": "none",
                                    "sentence_indent": spaces,
                                    "line_break":2,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })
                            else:
                                # 将数据存储在字典中
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": line,
                                    "translated_text": line,
                                    "model": "none",
                                    "sentence_indent": spaces,
                                    "line_break":1,
                                    "storage_path": storage_path,
                                    "file_name": file_name,
                                })

                        else:
                            # 将数据存储在字典中
                            json_data_list.append({
                                "text_index": i,
                                "translation_status": 0,
                                "source_text": line,
                                "translated_text": line,
                                "model": "none",
                                "sentence_indent": spaces,
                                "line_break":0,
                                "storage_path": storage_path,
                                "file_name": file_name,
                            })

                        i += 1


        return json_data_list


    # 读取文件夹中树形结构Epub文件
    def read_epub_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Epub")
        json_data_list.append({
            "project_type": "Epub",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 epub 文件
                if file.endswith(".epub"):

                    file_path = os.path.join(root, file)  # 构建文件路径

                    # 构建解压文件夹路径
                    parent_path = os.path.dirname(file_path)
                    extract_path = os.path.join(parent_path, 'EpubCache')

                    # 创建暂存文件夹
                    if not os.path.exists(extract_path):
                        os.makedirs(extract_path)

                    # 解压EPUB文件到暂存文件夹中
                    with zipfile.ZipFile(file_path, 'r') as epub_file:
                        # 提取所有文件
                        epub_file.extractall(extract_path)

                    # 加载EPUB文件
                    book = epub.read_epub(file_path)

                    # 获取文件路径和文件名
                    storage_path = os.path.relpath(file_path, folder_path)
                    book_name = file

                    # 遍历书籍中的所有内容
                    for item in book.get_items():
                        # 检查是否是文本内容
                        if item.get_type() == ebooklib.ITEM_DOCUMENT:


                            # 获取文件的唯一ID及文件名
                            item_id = item.get_id()
                            file_name = os.path.basename(item.get_name())

                            # 遍历文件夹中的所有文件,找到该文件，因为上面给的相对路径与epub解压后路径是不准的
                            for root_extract, dirs_extract, files_extract in os.walk(extract_path):
                                for filename in files_extract:
                                    # 如果文件名匹配
                                    if filename == file_name:
                                        # 构建完整的文件路径
                                        the_file_path = os.path.join(root_extract, filename)

                            # 打开对应HTML文件
                            with open(the_file_path, 'r', encoding='utf-8') as file:
                                # 读取文件内容
                                html_content = file.read()


                            # 获取文本内容并解码（为什么不用这个而进行解压操作呢，因为这个会自动将成对标签改为自适应标签）
                            #html_content = item.get_content().decode('utf-8')

                            # 正则表达式匹配<p>标签及其内容，包括自闭和的<p/>标签
                            p_pattern = r'<p[^>/]*>(.*?)</p>|<p[^>/]*/>'

                            # 使用findall函数找到所有匹配的内容
                            paragraphs = re.findall(p_pattern, html_content, re.DOTALL)

                            # 过滤掉空的内容
                            filtered_matches = [match for match in paragraphs if match.strip()]

                            # 遍历每个p标签，并提取文本内容
                            for p in filtered_matches:
                                # 保留原html内容文本
                                cleaned_text = p

                                # 提取纯文本
                                p_html = "<p>"+ p + "</p>"
                                soup = BeautifulSoup(p_html, 'html.parser')
                                text_content = soup.get_text()

                                # 去除前面的空格
                                text_content = text_content.lstrip()
                                cleaned_text = cleaned_text.lstrip() 

                                # 检查一下是否提取到空文本内容
                                if not text_content.strip():
                                    continue

                                # 获取项目的唯一ID
                                item_id = item.get_id()

                                # 录入缓存
                                json_data_list.append({
                                    "text_index": i,
                                    "translation_status": 0,
                                    "source_text": text_content,
                                    "translated_text": text_content,
                                    "html":cleaned_text,
                                    "model": "none",
                                    "item_id": item_id,
                                    "storage_path": storage_path,
                                    "file_name": book_name,
                                })                                    
                                # 增加文本索引值
                                i = i + 1

                    # 删除文件夹
                    shutil.rmtree(extract_path)

        return json_data_list


    # 读取文件夹中树形结构Docx文件
    def read_docx_files (self,folder_path):

        # 创建缓存数据，并生成文件头信息
        json_data_list = []
        project_id = File_Reader.generate_project_id(self,"Docx")
        json_data_list.append({
            "project_type": "Docx",
            "project_id": project_id,
        })

        #文本索引初始值
        i = 1

        # 遍历文件夹及其子文件夹
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 判断文件是否为 docx 文件
                if file.endswith(".docx"):

                    # 构建文件的路径
                    file_path = os.path.join(root, file)  

                    # 构建解压文件夹路径
                    parent_path = os.path.dirname(file_path)
                    extract_path = os.path.join(parent_path, 'DocxCache')

                    # 创建暂存文件夹
                    if not os.path.exists(extract_path):
                        os.makedirs(extract_path)

                    # 解压docx文件到暂存文件夹中
                    with zipfile.ZipFile(file_path, 'r') as docx_file:
                        # 提取所有文件
                        docx_file.extractall(extract_path)

                    # 获取文件路径和文件名
                    storage_path = os.path.relpath(file_path, folder_path)
                    file_name = file

                    # 构建存储文本的文件路径
                    the_file_path = os.path.join(extract_path,'word', 'document.xml')


                    # 打开对应xml文件
                    with open(the_file_path, 'r', encoding='utf-8') as file:
                        # 读取文件内容
                        xml_content = file.read()


                    # 正则表达式匹配运行标签及其内容
                    p_pattern = r'<w:t[^>/]*>(.*?)</w:t>'

                    # 使用findall函数找到所有匹配的内容
                    paragraphs = re.findall(p_pattern, xml_content, re.DOTALL)

                    # 过滤掉空的内容
                    filtered_matches = [match for match in paragraphs if match.strip()]

                    # 遍历每个标签，并提取文本内容
                    for text in filtered_matches:

                        # 检查一下是否提取到空文本内容和其他特殊内容
                        if text == "" or text == "\n" or text == " "or text == '\xa0':
                            continue

                        # 录入缓存
                        json_data_list.append({
                            "text_index": i,
                            "translation_status": 0,
                            "source_text": text,
                            "translated_text": text,
                            "model": "none",
                            "storage_path": storage_path,
                            "file_name": file_name,
                        })                                    
                        # 增加文本索引值
                        i = i + 1

            # 删除暂存文件夹
            shutil.rmtree(extract_path)

        return json_data_list

    # 根据文件类型读取文件
    def read_files (self,translation_project,label_input_path):

        if translation_project == "Mtool导出文件":
            cache_list = File_Reader.read_mtool_files(self,folder_path = label_input_path)
        elif translation_project == "T++导出文件":
            cache_list = File_Reader.read_xlsx_files (self,folder_path = label_input_path)
        elif translation_project == "VNText导出文件":
            cache_list = File_Reader.read_vnt_files(self,folder_path = label_input_path)
        elif translation_project == "Srt字幕文件":
            cache_list = File_Reader.read_srt_files(self,folder_path = label_input_path)
        elif translation_project == "Vtt字幕文件":
            cache_list = File_Reader.read_vtt_files(self,folder_path = label_input_path)
        elif translation_project == "Lrc音声文件":
            cache_list = File_Reader.read_lrc_files(self,folder_path = label_input_path)
        elif translation_project == "Txt小说文件":
            cache_list = File_Reader.read_txt_files(self,folder_path = label_input_path)
        elif translation_project == "Epub小说文件":
            cache_list = File_Reader.read_epub_files(self,folder_path = label_input_path)
        elif translation_project == "Docx文档文件":
            cache_list = File_Reader.read_docx_files(self,folder_path = label_input_path)
        elif translation_project == "Ainiee缓存文件":
            cache_list = File_Reader.read_cache_files(self,folder_path = label_input_path)
        if translation_project == "ParaTranz导出文件":
            cache_list = File_Reader.read_paratranz_files(self,folder_path = label_input_path)
        return cache_list
