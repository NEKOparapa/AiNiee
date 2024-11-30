import openpyxl
import rapidjson as json
from qfluentwidgets import TableWidget
from PyQt5.QtWidgets import QTableWidgetItem

class TableHelper():

    def __init__(self) -> None:
        super().__init__()

    # 从表格加载数据
    def load_from_table(table: TableWidget, keys: list[str]) -> list[dict]:
        result = []

        # 遍历每一行
        for row in range(table.rowCount()):
            # 获取当前行所有条目
            data: list[QTableWidgetItem] = [
                table.item(row, col)
                for col in range(table.columnCount())
            ]

            # 检查数据合法性
            if not isinstance(data[0], QTableWidgetItem) or len(data[0].text().strip()) == 0:
                continue

            # 添加数据
            result.append(
                {
                    keys[i]: (data[i].text().strip() if isinstance(data[i], QTableWidgetItem) else "")
                    for i in range(len(keys))
                }
            )

        return result

    # 向表格更新数据
    def update_to_table(table: TableWidget, data: list[dict], keys: list[str]) -> None:
        # 设置表格行数
        table.setRowCount(max(12, len(data)))

        # 去重
        data_unique = {v.get(keys[0], ""): v for v in data}
        data = [v for v in data_unique.values()]

        # 遍历表格
        for row, v in enumerate(data):
            for col in range(table.columnCount()):
                table.setItem(row, col, QTableWidgetItem(v.get(keys[col], "")))

    # 从文件加载数据
    def load_from_file(path: str, keys: list[str]) -> list[dict]:
        result = []

        # 从 json 文件加载数据
        if path.endswith(".json"):
            result = TableHelper.load_from_json_file(path, keys)

        # 从 xlsx 文件加载数据
        if path.endswith(".xlsx"):
            result = TableHelper.load_from_xlsx_file(path, keys)

        return result

    # 从 json 文件加载数据
    def load_from_json_file(path: str, keys: list[str]) -> list[dict]:
            result = []

            # 读取文件
            inputs = []
            with open(path, "r", encoding = "utf-8") as reader:
                inputs = json.load(reader)

            # 标准字典列表
            # [
            #     {
            #         "key": "value",
            #         "key": "value",
            #         "key": "value",
            #     }
            # ]
            if isinstance(inputs, list):
                for data in inputs:
                    # 数据校验
                    if not isinstance(data, dict) or str(data.get(keys[0], "")).strip() == "":
                        continue

                    # 添加数据
                    result.append({
                        keys[i]: str(data.get(keys[i], "")).strip()
                        for i in range(len(keys))
                    })

                # 兼容旧版，保留一段时间用以过度
                if len(result) == 0 and "src" in keys:
                    # 将 keys 中的 "src" 替换为 "srt" 然后重试
                    result = TableHelper.load_from_json_file(path, [("srt" if v == "src" else v) for v in keys])

                    # 将字段换回来
                    for v in result:
                        v["src"] = v.get("srt", "")

            # 标准 KV 字典
            # [
            #     "ダリヤ": "达莉雅"
            # ]
            if isinstance(inputs, dict):
                for k, v in inputs.items():
                    # 数据校验
                    if str(k).strip() == "":
                        continue

                    # 添加数据
                    item = {}
                    for i in range(len(keys)):
                        if i == 0:
                            item[keys[i]] = str(k).strip()
                        elif i == 1:
                            item[keys[i]] = str(v).strip() if v != None else ""
                        else:
                            item[keys[i]] = ""
                    result.append(item)

            return result

    # 从 xlsx 文件加载数据
    def load_from_xlsx_file(path: str, keys: list[str]) -> list[dict]:
        result = []

        sheet = openpyxl.load_workbook(path).active
        for row in range(1, sheet.max_row + 1):
            # 读取每一行的数据
            data: list[str] = [
                sheet.cell(row = row, column = col).value
                for col in range(1, len(keys) + 1)
            ]

            # 检查数据合法性
            if data[0] == None or str(data[0]).strip() == "":
                continue

            # 添加数据
            result.append(
                {
                    keys[i]: str(data[i]).strip() if data[i] != None else ""
                    for i in range(len(keys))
                }
            )

        return result