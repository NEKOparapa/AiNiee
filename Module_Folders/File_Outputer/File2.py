import re
import os
import copy
import json

import shutil
import zipfile

import ebooklib # 需要安装库pip install ebooklib
from ebooklib import epub
from openpyxl import Workbook

# 文件输出器
class File_Outputter():

    def __init__(self):
        pass

    # 输出json文件
    def output_json_file(self,cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
        {'project_type': 'Mtool'},
        {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ', 'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json'},
        {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222', 'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json', 'file_name': 'Replace the original text.json'},
        ]


        # 中间存储字典格式示例
        ex_path_dict = {
            "D:\\DEBUG Folder\\Replace the original text.json": {'translation_status': 1, 'Source Text': 'しこトラ！', 'Translated Text': 'しこトラ！'},
            "D:\\DEBUG Folder\\DEBUG Folder\\Replace the original text.json": {'translation_status': 0, 'Source Text': 'しこトラ！', 'Translated Text': 'しこトラ！'}
        }


        # 输出文件格式示例
        ex_output ={
        'しこトラ！': 'xxxx',
        '室内カメラ': 'yyyyy',
        '111111111': '无3',
        '222222222': '无4',
        }

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'



            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:

                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text']}
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text']}
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = {}

            #存储未翻译的文本
            output_file2 = {}

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 如果这个本已经翻译了，存放对应的文件中
                if content['translation_status'] == 1:
                    output_file[content['source_text']] = content['translated_text']
                # 如果这个文本没有翻译或者正在翻译
                elif content['translation_status'] == 0 or content['translation_status'] == 2:
                    output_file2[content['source_text']] = content['source_text']


            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出paratranz文件
    def output_paratranz_file(self, cache_data, output_path):
        # 缓存数据结构示例
        ex_cache_data = [
            {'project_type': 'Mtool'},
            {'text_index': 1, 'text_classification': 0, 'translation_status': 0, 'source_text': 'しこトラ！',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json',
             'key': 'txtKey', 'context': ''},
            {'text_index': 2, 'text_classification': 0, 'translation_status': 0, 'source_text': '室内カメラ',
             'translated_text': '无', 'storage_path': 'TrsData.json', 'file_name': 'TrsData.json',
             'key': 'txtKey', 'context': ''},
            {'text_index': 3, 'text_classification': 0, 'translation_status': 0, 'source_text': '111111111',
             'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json',
             'file_name': 'Replace the original text.json', 'key': 'txtKey', 'context': ''},
            {'text_index': 4, 'text_classification': 0, 'translation_status': 0, 'source_text': '222222222',
             'translated_text': '无', 'storage_path': 'DEBUG Folder\\Replace the original text.json',
             'file_name': 'Replace the original text.json', 'key': 'txtKey', 'context': ''},
        ]

        # 中间存储字典格式示例
        ex_path_dict = {
            "D:\\DEBUG Folder\\Replace the original text.json": {'translation_status': 1, 'Source Text': 'しこトラ！',
                                                                 'Translated Text': 'しこトラ！'},
            "D:\\DEBUG Folder\\DEBUG Folder\\Replace the original text.json": {'translation_status': 0,
                                                                               'Source Text': 'しこトラ！',
                                                                               'Translated Text': 'しこトラ！'}
        }

        # 输出文件格式示例
        ex_output = [
        {
            "key": "Activate",
            "original": "カードをプレイ",
            "translation": "出牌",
            "context": ""
        }]

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'

                # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:
                text = {'translation_status': item['translation_status'], 'source_text': item['source_text'],
                        'translated_text': item['translated_text'], 'context': item['context'], 'key': item['key']}
                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'], 'source_text': item['source_text'],
                        'translated_text': item['translated_text'], 'context': item['context'], 'key': item['key']}
                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)

            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)

            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = []

            # 存储未翻译的文本
            output_file2 = []

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                item = {
                    "key": content.get("key", ""),  # 假设每个 content 字典都有 'key' 字段
                    "original": content['source_text'],
                    "translation": content.get('translated_text', ""),
                    "context": content.get('context', "")  # 如果你有 'context' 字段，也包括它
                }
                # 根据翻译状态，选择存储到已翻译或未翻译的列表
                if content['translation_status'] == 1:
                    output_file.append(item)
                elif content['translation_status'] == 0 or content['translation_status'] == 2:
                    output_file2.append(item)

            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出vnt文件
    def output_vnt_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output =  [
            {
                "name": "玲",
                "message": "「……おはよう」"
            },
            {
                "message": "　心の内では、ムシャクシャした気持ちは未だに鎮まっていなかった。"
            }
            ]

        # 创建中间存储字典，这个存储已经翻译的内容
        path_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'



            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in path_dict:
                if'name' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text'],
                            'name': item['name']}

                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text']}

                path_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                if'name' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text'],
                            'name': item['name']}

                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text']}

                path_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in path_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_translated = old_filename.replace(".json", "") + "_translated.json"
            else:
                file_name_translated = old_filename + "_translated.json"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 创建未翻译文本的新文件路径
            if old_filename.endswith(".json"):
                file_name_untranslated = old_filename.replace(".json", "") + "_untranslated.json"
            else:
                file_name_untranslated = old_filename + "_untranslated.json"
            file_path_untranslated = os.path.join(folder_path, file_name_untranslated)

            # 存储已经翻译的文本
            output_file = []

            #存储未翻译的文本
            output_file2 = []

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 如果这个本已经翻译了，存放对应的文件中
                if'name' in content:
                    text = {'name': content['name'],
                            'message': content['translated_text'],}
                else:
                    text = {'message': content['translated_text'],}

                output_file.append(text)

                # 如果这个文本没有翻译或者正在翻译
                if content['translation_status'] == 0 or content['translation_status'] == 2:
                    if'name' in content:
                        text = {'name': content['name'],
                                'message': content['translated_text'],}
                    else:
                        text = {'message': content['translated_text'],}

                    output_file2.append(text)


            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                json.dump(output_file, file, ensure_ascii=False, indent=4)

            # 输出未翻译的内容
            if output_file2:
                with open(file_path_untranslated, 'w', encoding='utf-8') as file:
                    json.dump(output_file2, file, ensure_ascii=False, indent=4)

    # 输出表格文件
    def output_excel_file(self, cache_data, output_path):
        # 缓存数据结构示例
        # ex_cache_data = [
        #     {"project_type": "T++"},
        #     {"text_index": 1, "text_classification": 0, "translation_status": 1, "source_text": "しこトラ！", "translated_text": "无", "storage_path": "TrsData.xlsx", "file_name": "TrsData.xlsx", "row_index": 2},
        #     {"text_index": 2, "text_classification": 0, "translation_status": 0, "source_text": "室内カメラ", "translated_text": "无", "storage_path": "TrsData.xlsx", "file_name": "TrsData.xlsx", "row_index": 3},
        #     {"text_index": 3, "text_classification": 0, "translation_status": 0, "source_text": "草草草草", "translated_text": "11111", "storage_path": "DEBUG Folder\\text.xlsx", "file_name": "text.xlsx", "row_index": 3},
        #     {"text_index": 4, "text_classification": 0, "translation_status": 1, "source_text": "室内カメラ", "translated_text": "22222", "storage_path": "DEBUG Folder\\text.xlsx", "file_name": "text.xlsx", "row_index": 4},
        # ]

        # 创建一个字典，用于存储翻译数据
        translations_by_path = {}

        # 遍历缓存数据
        for item in cache_data:
            if "storage_path" in item:
                path = item["storage_path"]

                # 如果路径不存在，创建文件夹
                folder_path = os.path.join(output_path, os.path.dirname(path))
                os.makedirs(folder_path, exist_ok = True)

                # 提取信息
                source_text = item.get("source_text", "")
                translated_text = item.get("translated_text", "")
                row_index = item.get("row_index", -1)
                translation_status = item.get("translation_status", -1)

                # 构造字典
                translation_dict = {
                    "translation_status": translation_status,
                    "source_text": source_text,
                    "translated_text": translated_text,
                    "row_index": row_index
                }

                # 将字典添加到对应路径的列表中
                if path in translations_by_path:
                    translations_by_path[path].append(translation_dict)
                else:
                    translations_by_path[path] = [translation_dict]

        # 遍历字典，将数据写入 Excel 文件
        for path, translations_list in translations_by_path.items():
            file_path = os.path.join(output_path, path)

            # 创建一个工作簿
            wb = Workbook()

            # 选择默认的活动工作表
            ws = wb.active

            # 添加表头
            ws.append(
                [
                    "Original Text",
                    "Initial",
                    "Machine translation",
                    "Better translation",
                    "Best translation",
                ]
            )

            # 将数据写入工作表
            for item in translations_list:
                source_text = item.get("source_text", "")
                translated_text = item.get("translated_text", "")
                row_index = item.get("row_index", -1)
                translation_status = item.get("translation_status", -1)

                # 根据翻译状态写入原文及译文
                # 如果文本是以 = 开始，则加一个空格
                # 因为 = 开头会被识别成 Excel 公式导致 T++ 导入时 卡住
                # 加入空格后，虽然还是不能直接导入 T++ ，但是可以手动复制粘贴
                if translation_status != 1:
                    ws.cell(row = row_index, column = 1).value = re.sub(r"^=", " =", source_text)
                else:
                    ws.cell(row = row_index, column = 1).value = re.sub(r"^=", " =", source_text)
                    ws.cell(row = row_index, column = 2).value = re.sub(r"^=", " =", translated_text)

            # 保存工作簿
            wb.save(file_path)

    # 输出缓存文件
    def output_cache_file(self,cache_data,output_path):
        # 复制缓存数据到新变量
        try:
            modified_cache_data = copy.deepcopy(cache_data)
        except:
            print("[INFO]: 无法正常进行深层复制,改为浅复制")
            modified_cache_data = cache_data.copy()

        # 修改新变量的元素中的'translation_status'
        for item in modified_cache_data:
            if 'translation_status' in item and item['translation_status'] == 2:
                item['translation_status'] = 0

        # 输出为JSON文件
        with open(os.path.join(output_path, "AinieeCacheData.json"), "w", encoding="utf-8") as f:
            json.dump(modified_cache_data, f, ensure_ascii=False, indent=4)

    # 输出srt文件
    def output_srt_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output ="""
        1
        00:00:16,733 --> 00:00:19,733
        Does that feel good, Tetchan?

        2
        00:00:25,966 --> 00:00:32,500
        Just a little more... I'm really close too... Ahhh, I can't...!
        """

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:

                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text'],'subtitle_number': item['subtitle_number'],'subtitle_time': item['subtitle_time']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text'],'subtitle_number':  item['subtitle_number'],'subtitle_time': item['subtitle_time']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".srt"):
                file_name_translated = old_filename.replace(".srt", "") + "_translated.srt"
            else:
                file_name_translated = old_filename + "_translated.srt"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""
            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取字幕序号
                subtitle_number = content['subtitle_number']
                # 获取字幕时间轴
                subtitle_time = content['subtitle_time']
                # 获取字幕文本内容
                subtitle_text = content['translated_text']

                output_file += f'{subtitle_number}\n{subtitle_time}\n{subtitle_text}\n\n'



            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)

    # 输出vtt文件
    def output_vtt_file(self,cache_data, output_path):

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:

                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text'],'subtitle_number': item['subtitle_number'],'subtitle_time': item['subtitle_time'],'top_text': item['top_text']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],'source_text': item['source_text'], 'translated_text': item['translated_text'],'subtitle_number':  item['subtitle_number'],'subtitle_time': item['subtitle_time'],'top_text': item['top_text']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".vtt"):
                file_name_translated = old_filename.replace(".vtt", "") + "_translated.vtt"
            else:
                file_name_translated = old_filename + "_translated.vtt"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""

            # 恢复开头注释
            top_text = content_list[1]['top_text']
            output_file = output_file + top_text

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取字幕序号
                subtitle_number = content['subtitle_number']
                # 获取字幕时间轴
                subtitle_time = content['subtitle_time']
                # 获取字幕文本内容
                subtitle_text = content['translated_text']

                output_file += f'{subtitle_number}\n{subtitle_time}\n{subtitle_text}\n\n'



            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)


    # 输出lrc文件
    def output_lrc_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output ="""
        [ti:1.したっぱ童貞構成員へハニートラップ【手コキ】 (Transcribed on 15-May-2023 19-10-13)]
        [00:00.00]お疲れ様です大長 ただいま機会いたしました
        [00:06.78]法案特殊情報部隊一番対処得フィルレイやセルドツナイカーです 今回例の犯罪組織への潜入が成功しましたのでご報告させていただきます
        """

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                if 'subtitle_title' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time'],
                            'subtitle_title': item['subtitle_title']}
                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                if 'subtitle_title' in item:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time'],
                            'subtitle_title': item['subtitle_title']}
                else:
                    text = {'translation_status': item['translation_status'],
                            'source_text': item['source_text'],
                            'translated_text': item['translated_text'],
                            "subtitle_time": item['subtitle_time']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".lrc"):
                file_name_translated = old_filename.replace(".lrc", "") + "_translated.lrc"
            else:
                file_name_translated = old_filename + "_translated.lrc"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""
            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取字幕时间轴
                subtitle_time = content['subtitle_time']
                # 获取字幕文本内容
                subtitle_text = content['translated_text']

                if 'subtitle_title' in content:
                    subtitle_title = content['subtitle_title']
                    output_file += f'[{subtitle_title}]\n[{subtitle_time}]{subtitle_text}\n'
                else:
                    output_file += f'[{subtitle_time}]{subtitle_text}\n'



            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)

    # 输出txt文件
    def output_txt_file(self,cache_data, output_path):

        # 输出文件格式示例
        ex_output ="""
        　测试1
        　今ではダンジョンは、人々の営みの一部としてそれなりに定着していた。

        ***

        「正気なの？」
        测试2
        """

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'],
                        'translated_text': item['translated_text'],
                        "sentence_indent": item['sentence_indent'],
                        'line_break': item['line_break']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'],
                        'translated_text': item['translated_text'],
                        "sentence_indent": item['sentence_indent'],
                        'line_break': item['line_break']}
                text_dict[file_path] = [text]

        # 遍历 path_dict，并将内容写入文件
        for file_path, content_list in text_dict.items():

            # 提取文件路径的文件夹路径和文件名
            folder_path, old_filename = os.path.split(file_path)


            # 创建已翻译文本的新文件路径
            if old_filename.endswith(".txt"):
                file_name_translated = old_filename.replace(".txt", "") + "_translated.txt"
            else:
                file_name_translated = old_filename + "_translated.txt"
            file_path_translated = os.path.join(folder_path, file_name_translated)


            # 存储已经翻译的文本
            output_file = ""

            # 转换中间字典的格式为最终输出格式
            for content in content_list:
                # 获取记录的句首空格数
                expected_indent_count = content['sentence_indent']
                # 获取句尾换行符数
                line_break_count = content['line_break']

                # 删除句首的所有空格
                translated_text = content['translated_text'].lstrip()

                # 根据记录的空格数在句首补充空格
                sentence_indent = "　" * expected_indent_count

                line_break = "\n" * (line_break_count + 1)

                output_file += f'{sentence_indent}{translated_text}{line_break}'

            # 输出已经翻译的文件
            with open(file_path_translated, 'w', encoding='utf-8') as file:
                file.write(output_file)

    # 输出epub文件
    def output_epub_file(self,cache_data, output_path, input_path):

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'],
                        'translated_text': item['translated_text'],
                        'html': item['html'],
                        "item_id": item['item_id'],}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'],
                        'translated_text': item['translated_text'],
                        'html': item['html'],
                        "item_id": item['item_id'],}
                text_dict[file_path] = [text]




        # 创建输出
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        # 将输入路径里面的所有epub文件复制到输出路径
        for dirpath, dirnames, filenames in os.walk(input_path):
            for filename in filenames:
                # 检查文件扩展名是否为.epub
                if filename.endswith('.epub'):
                    # 构建源文件和目标文件的完整路径
                    src_file_path = os.path.join(dirpath, filename)
                    # 计算相对于输入路径的相对路径
                    relative_path = os.path.relpath(src_file_path, start=input_path)
                    # 构建目标文件的完整路径
                    dst_file_path = os.path.join(output_path, relative_path)

                    # 创建目标文件的目录（如果不存在）
                    os.makedirs(os.path.dirname(dst_file_path), exist_ok=True)
                    # 复制文件
                    shutil.copy2(src_file_path, dst_file_path)
                    #print(f'Copied: {src_file_path} -> {dst_file_path}')



        # 遍历 path_dict，并将内容写入对应文件中
        for file_path, content_list in text_dict.items():

            # 加载EPUB文件
            book = epub.read_epub(file_path)

            # 构建解压文件夹路径
            parent_path = os.path.dirname(file_path)
            extract_path = os.path.join(parent_path, 'EpubCache')

            # 创建解压文件夹
            if not os.path.exists(extract_path):
                os.makedirs(extract_path)

            # 使用zipfile模块打开并解压EPUB文件
            with zipfile.ZipFile(file_path, 'r') as epub_file:
                # 提取所有文件
                epub_file.extractall(extract_path)

            # 遍历书籍中的所有内容
            for item in book.get_items():
                # 检查是否是文本内容
                if item.get_type() == ebooklib.ITEM_DOCUMENT:

                    # 获取文件的唯一ID及文件名
                    item_id = item.get_id()
                    file_name = os.path.basename(item.get_name())

                    # 遍历文件夹中的所有文件,找到该文件，因为上面给的相对路径与epub解压后路径是不准的
                    for root, dirs, files in os.walk(extract_path):
                        for filename in files:
                            # 如果文件名匹配
                            if filename == file_name:
                                # 构建完整的文件路径
                                the_file_path = os.path.join(root, filename)

                    # 打开对应HTML文件
                    with open(the_file_path, 'r', encoding='utf-8') as file:
                        # 读取文件内容
                        content_html = file.read()

                    # 遍历缓存数据
                    for content in content_list:
                        # 如果找到匹配的文件id
                        if item_id == content['item_id']:
                            # 获取原文本
                            original = content['source_text']
                            # 获取翻译后的文本
                            replacement = content['translated_text']

                            # 获取html标签化的文本
                            html = content['html']
                            html = str(html)

                            # 删除 &#13;\n\t\t\t\t
                            html = html.replace("&#13;\n\t\t\t\t", "")

                            if"Others who have read our chapters and offered similar assistance have" in html:
                                print("ce")

                            # 有且只有一个a标签，则改变替换文本，以保留跳转功能
                            if (re.match( r'^(?:<a(?:\s[^>]*?)?>[^<]*?</a>)*$', html) is not None):
                                # 针对跳转标签的保留，使用正则表达式搜索<a>标签内的文本
                                a_tag_pattern = re.compile(r'<a[^>]*>(.*?)</a>')
                                matches = a_tag_pattern.findall(html)

                                if len(matches) == 1:
                                    html = matches[0]


                            # 如果原文与译文不为空，则替换原hrml文件中的文本
                            if (original and replacement):
                                # 替换第一个匹配项
                                content_html = content_html.replace(html, replacement, 1)
                                #content_html  = re.sub(original, replacement, content_html, count=1)


                    # 写入内容到HTML文件
                    with open(the_file_path, 'w', encoding='utf-8') as file:
                        file.write(content_html)

            # 构建修改后的EPUB文件路径
            modified_epub_file = file_path.rsplit('.', 1)[0] + '_translated.epub'

            # 创建ZipFile对象，准备写入压缩文件
            with zipfile.ZipFile(modified_epub_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 遍历文件夹中的所有文件和子文件夹
                for root, dirs, files in os.walk(extract_path):
                    for file in files:
                        # 获取文件的完整路径
                        full_file_path = os.path.join(root, file)
                        # 获取文件在压缩文件中的相对路径
                        relative_file_path = os.path.relpath(full_file_path, extract_path)
                        # 将文件添加到压缩文件中
                        zipf.write(full_file_path, relative_file_path)

            # 删除旧文件
            os.remove(file_path)
            # 删除文件夹
            shutil.rmtree(extract_path)


    # 输出docx文件
    def output_docx_file(self,cache_data, output_path, input_path):

        # 创建中间存储文本
        text_dict = {}

        # 遍历缓存数据
        for item in cache_data:
            # 忽略不包含 'storage_path' 的项
            if 'storage_path' not in item:
                continue

            # 获取相对文件路径
            storage_path = item['storage_path']
            # 获取文件名
            file_name = item['file_name']

            # 创建文件夹路径及文件路径
            if file_name != storage_path :
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'
                # 获取输出路径的上一级路径，使用os.path.dirname
                folder_path = os.path.dirname(file_path)
                # 如果路径不存在，则创建
                os.makedirs(folder_path, exist_ok=True)
            else:
                # 构建文件输出路径
                file_path = f'{output_path}/{storage_path}'


            # 如果文件路径已经在 path_dict 中，添加到对应的列表中
            if file_path in text_dict:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'],
                        'translated_text': item['translated_text']}
                text_dict[file_path].append(text)

            # 否则，创建一个新的列表
            else:
                text = {'translation_status': item['translation_status'],
                        'source_text': item['source_text'],
                        'translated_text': item['translated_text']}
                text_dict[file_path] = [text]


        # 创建输出文件夹
        if not os.path.exists(output_path):
            os.makedirs(output_path)

        # 将输入路径里面的所有docx文件复制到输出路径
        for dirpath, dirnames, filenames in os.walk(input_path):
            for filename in filenames:
                # 检查文件扩展名是否为.epub
                if filename.endswith('.docx'):
                    # 构建源文件和目标文件的完整路径
                    src_file_path = os.path.join(dirpath, filename)
                    # 计算相对于输入路径的相对路径
                    relative_path = os.path.relpath(src_file_path, start=input_path)
                    # 构建目标文件的完整路径
                    dst_file_path = os.path.join(output_path, relative_path)

                    # 创建目标文件的目录（如果不存在）
                    os.makedirs(os.path.dirname(dst_file_path), exist_ok=True)
                    # 复制文件
                    shutil.copy2(src_file_path, dst_file_path)



        # 遍历 path_dict，并将内容写入对应文件中
        for file_path, content_list in text_dict.items():

            # 构建解压文件夹路径
            parent_path = os.path.dirname(file_path)
            extract_path = os.path.join(parent_path, 'DocxCache')

            # 创建解压文件夹
            if not os.path.exists(extract_path):
                os.makedirs(extract_path)

            # 使用zipfile模块打开并解压docx文件
            with zipfile.ZipFile(file_path, 'r') as epub_file:
                # 提取所有文件
                epub_file.extractall(extract_path)

            # 构建存储文本的文件路径
            the_file_path = os.path.join(extract_path,'word', 'document.xml')

            # 打开对应xml文件
            with open(the_file_path, 'r', encoding='utf-8') as file:
                # 读取文件内容
                content_xml = file.read()

            # 遍历缓存数据
            for content in content_list:
                # 获取原文本
                original = content['source_text']
                # 获取翻译后的文本
                replacement = content['translated_text']

                # 替换第一个匹配项
                content_xml = content_xml.replace(original, replacement, 1)


            # 写入内容到xml文件
            with open(the_file_path, 'w', encoding='utf-8') as file:
                file.write(content_xml)

            # 构建修改后的docx文件路径
            modified_epub_file = file_path.rsplit('.', 1)[0] + '_translated.docx'

            # 创建ZipFile对象，准备写入压缩文件
            with zipfile.ZipFile(modified_epub_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 遍历文件夹中的所有文件和子文件夹
                for root, dirs, files in os.walk(extract_path):
                    for file in files:
                        # 获取文件的完整路径
                        full_file_path = os.path.join(root, file)
                        # 获取文件在压缩文件中的相对路径
                        relative_file_path = os.path.relpath(full_file_path, extract_path)
                        # 将文件添加到压缩文件中
                        zipf.write(full_file_path, relative_file_path)

            # 删除旧文件
            os.remove(file_path)
            # 删除文件夹
            shutil.rmtree(extract_path)

    # 输出已经翻译文件
    def output_translated_content(self, cache_data, output_path, input_path) -> None:
        if cache_data[0]["project_type"] == "Mtool":
            File_Outputter.output_json_file(self, cache_data, output_path)
        elif cache_data[0]["project_type"] == "Srt":
            File_Outputter.output_srt_file(self, cache_data, output_path)
        elif cache_data[0]["project_type"] == "Vtt":
            File_Outputter.output_vtt_file(self, cache_data, output_path)
        elif cache_data[0]["project_type"] == "Lrc":
            File_Outputter.output_lrc_file(self, cache_data, output_path)
        elif cache_data[0]["project_type"] == "Vnt":
            File_Outputter.output_vnt_file(self, cache_data, output_path)
        elif cache_data[0]["project_type"] == "Txt":
            File_Outputter.output_txt_file(self, cache_data, output_path)
        elif cache_data[0]["project_type"] == "Epub":
            File_Outputter.output_epub_file(self, cache_data, output_path, input_path)
        elif cache_data[0]["project_type"] == "Docx":
            File_Outputter.output_docx_file(self, cache_data, output_path, input_path)
        elif cache_data[0]["project_type"] == "Paratranz":
            File_Outputter.output_paratranz_file(self, cache_data, output_path)
        else:
            File_Outputter.output_excel_file(self, cache_data, output_path)