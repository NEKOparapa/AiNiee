
import json
import os
from openpyxl import Workbook  
import openpyxl  
import re
from PyQt5.QtGui import QBrush, QColor, QDesktopServices, QFont, QImage, QPainter, QPixmap#需要安装库 pip3 install PyQt5
from PyQt5.QtCore import  QObject,  QRect,  QUrl,  Qt, pyqtSignal 
from PyQt5.QtWidgets import QAbstractItemView,QHeaderView,QApplication, QTableWidgetItem, QFrame, QGridLayout, QGroupBox, QLabel,QFileDialog, QStackedWidget, QHBoxLayout, QVBoxLayout

from qfluentwidgets.components import Dialog  # 需要安装库 pip install "PyQt-Fluent-Widgets[full]" -i https://pypi.org/simple/
from qfluentwidgets import ProgressRing, SegmentedWidget, TableWidget,CheckBox, DoubleSpinBox, HyperlinkButton,InfoBar, InfoBarPosition, NavigationWidget, Slider, SpinBox, ComboBox, LineEdit, PrimaryPushButton, PushButton ,StateToolTip, SwitchButton, TextEdit, Theme,  setTheme ,isDarkTheme,qrouter,NavigationInterface,NavigationItemPosition, EditableComboBox
from qfluentwidgets import FluentIcon as FIF
from qframelesswindow import FramelessWindow, StandardTitleBar



class Widget_replace_dict(QFrame):  # 替换字典主界面
    def __init__(self, text: str, parent=None,configurator=None,user_interface_prompter=None):  # 构造函数，初始化实例时会自动调用
        super().__init__(parent=parent)  # 调用父类 QWidget 的构造函数
        self.setObjectName(text.replace(' ', '-'))  # 设置对象名，用于在 NavigationInterface 中的 addItem 方法中的 routeKey 参数中使用
        self.configurator = configurator
        self.user_interface_prompter = user_interface_prompter

        self.pivot = SegmentedWidget(self)  # 创建一个 SegmentedWidget 实例，分段式导航栏
        self.stackedWidget = QStackedWidget(self)  # 创建一个 QStackedWidget 实例，堆叠式窗口
        self.vBoxLayout = QVBoxLayout(self)  # 创建一个垂直布局管理器

        self.A_settings = Widget_before_dict('A_settings', self,configurator,user_interface_prompter)  # 创建实例，指向界面
        self.B_settings = Widget_after_dict('B_settings', self,configurator,user_interface_prompter)  # 创建实例，指向界面

        # 添加子界面到分段式导航栏
        self.addSubInterface(self.A_settings, 'A_settings', '译前替换')
        self.addSubInterface(self.B_settings, 'B_settings', '译后替换')

        # 将分段式导航栏和堆叠式窗口添加到垂直布局中
        self.vBoxLayout.addWidget(self.pivot)
        self.vBoxLayout.addWidget(self.stackedWidget)
        self.vBoxLayout.setContentsMargins(30, 50, 30, 30)  # 设置布局的外边距

        # 连接堆叠式窗口的 currentChanged 信号到槽函数 onCurrentIndexChanged
        self.stackedWidget.currentChanged.connect(self.onCurrentIndexChanged)
        self.stackedWidget.setCurrentWidget(self.A_settings)  # 设置默认显示的子界面为xxx界面
        self.pivot.setCurrentItem(self.A_settings.objectName())  # 设置分段式导航栏的当前项为xxx界面

    def addSubInterface(self, widget: QLabel, objectName, text):
        """
        添加子界面到堆叠式窗口和分段式导航栏
        """
        widget.setObjectName(objectName)
        #widget.setAlignment(Qt.AlignCenter) # 设置 widget 对象的文本（如果是文本控件）在控件中的水平对齐方式
        self.stackedWidget.addWidget(widget)
        self.pivot.addItem(
            routeKey=objectName,
            text=text,
            onClick=lambda: self.stackedWidget.setCurrentWidget(widget),
        )

    def onCurrentIndexChanged(self, index):
        """
        槽函数：堆叠式窗口的 currentChanged 信号的槽函数
        """
        widget = self.stackedWidget.widget(index)
        self.pivot.setCurrentItem(widget.objectName())


class Widget_before_dict(QFrame):# 原文替换字典界面
    def __init__(self, text: str, parent=None,configurator=None,user_interface_prompter=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        self.configurator = configurator
        self.user_interface_prompter = user_interface_prompter

        # self.scrollWidget = QWidget() #创建滚动窗口
        # #self.scrollWidget.resize(500, 400)    #设置滚动窗口大小
        # #self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff) #设置水平滚动条不可见
        # self.setViewportMargins(0, 0, 0, 0)   #设置滚动窗口的边距      
        # self.setWidget(self.scrollWidget)  #设置滚动窗口的内容  
        # self.setWidgetResizable(True)   #设置滚动窗口的内容可调整大小
        # self.verticalScrollBar().sliderPressed.connect(self.scrollContents) #滚动条滚动时，调用的函数
        #设置各个控件-----------------------------------------------------------------------------------------

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['Src', 'Dst']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序
        self.tableView.setBorderVisible(True) # 开启显示边框功能，从而可以修改表格角半径
        self.tableView.setBorderRadius(8) # 将表格组件的边角半径设置为x像素，从而实现圆角效果。

        # songInfos = [
        #     ['かばん', 'aiko']
        # ]
        # for i, songInfo in enumerate(songInfos): #遍历数据
        #     for j in range(2): #遍历每一列
        #         self.tableView.setItem(i, j, QTableWidgetItem(songInfo[j])) #设置每个单元格的内容


        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        #设置“译前替换”标签
        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("原文替换")

        #设置“译前替换”显示
        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(翻译前，将根据字典内容对原文文本进行替换)")


        #设置“译前替换”开
        self.checkBox1 = CheckBox('启用功能')
        self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout2.addWidget(label1)
        layout2.addWidget(self.label2)
        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.checkBox1)
        box2.setLayout(layout2)




        # 把内容添加到容器中 
        container.addWidget(box2)   
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下 

    #滚动条滚动时，调用的函数
    def scrollContents(self, position):
        self.scrollWidget.move(0, position) 

    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    # 移除JSON内容中的行内注释
    def remove_inline_comments(self, json_content):
        # 正则表达式匹配行内注释并替换为空字符串
        return re.sub(r'//.*$', '', json_content, flags=re.MULTILINE)

    #导入字典按钮
    def Importing_dictionaries(self):

        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'All Files (*)')

        if Input_File:
            print(f'[INFO] 已选择文件: {Input_File}')
            # 获取文件后缀
            file_suffix = Input_File.split('.')[-1].lower()
            
            # 根据文件后缀执行不同操作
            if file_suffix == 'json':

                try:
                    # 尝试读取文件内容
                    with open(Input_File, 'r', encoding="utf-8") as f:
                        content = f.read()
                except FileNotFoundError:
                    print(f'[ERROR] 文件未找到: {Input_File}')
                    return
                except Exception as e:
                    print(f'[ERROR] 读取文件时发生未知错误: {str(e)}')
                    return
                try:
                    # 移除内容中的行内注释，并反序列化
                    dictionary = json.loads(self.remove_inline_comments(content))
                except json.JSONDecodeError as e:
                    print(f'[ERROR] JSON解析错误: {str(e)}')
                    return
                except Exception as e:
                    print(f'[ERROR] 反序列化时发生未知错误: {str(e)}')
                    return       

                # 将字典中的数据从表格底部添加到表格中
                for key, value in dictionary.items():
                    row = self.tableView.rowCount() - 1 #获取表格的倒数行数
                    self.tableView.insertRow(row)    # 在表格中插入一行
                    self.tableView.setItem(row, 0, QTableWidgetItem(key))
                    self.tableView.setItem(row, 1, QTableWidgetItem(value))
                    #设置新行的高度与前一行相同
                    self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))


                # 导入成功后删除空白行
                self.delete_blank_row()

                # 输出日志
                self.user_interface_prompter.createSuccessInfoBar("导入成功")
                print(f'[INFO]  已导入字典文件')


            elif file_suffix == 'xlsx':
                # 执行XLSX文件的操作
                wb = openpyxl.load_workbook(Input_File)
                sheet = wb.active
                for row in range(2, sheet.max_row + 1): # 第一行是标识头，第二行才开始读取
                    cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                    cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值


                    row = self.tableView.rowCount() - 1 #获取表格的倒数行数
                    self.tableView.insertRow(row)    # 在表格中插入一行
                    self.tableView.setItem(row, 0, QTableWidgetItem(cell_value1))
                    self.tableView.setItem(row, 1, QTableWidgetItem(cell_value2))
                    #设置新行的高度与前一行相同
                    self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

                # 导入成功后删除空白行
                self.delete_blank_row()

                # 输出日志
                self.user_interface_prompter.createSuccessInfoBar("导入成功")
                print(f'[INFO]  已导入字典文件')
                    
            else:
                print(f'[INFO] 不支持的文件类型: .{file_suffix}')

        else:
            print('[INFO] 未选择文件')
            return        
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户译前替换字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        self.user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        self.user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        self.user_interface_prompter.read_write_config("write",self.configurator.resource_dir)
        self.user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')


    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            self.user_interface_prompter.createSuccessInfoBar("已开启译前替换功能，将依据表格内容进行替换")
    

class Widget_after_dict(QFrame):# 译文修正字典界面
    def __init__(self, text: str, parent=None,configurator=None,user_interface_prompter=None):#解释器会自动调用这个函数
        super().__init__(parent=parent)          #调用父类的构造函数
        self.setObjectName(text.replace(' ', '-'))#设置对象名，作用是在NavigationInterface中的addItem中的routeKey参数中使用
        self.configurator = configurator
        self.user_interface_prompter = user_interface_prompter

        # self.scrollWidget = QWidget() #创建滚动窗口
        # #self.scrollWidget.resize(500, 400)    #设置滚动窗口大小
        # #self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff) #设置水平滚动条不可见
        # self.setViewportMargins(0, 0, 0, 0)   #设置滚动窗口的边距      
        # self.setWidget(self.scrollWidget)  #设置滚动窗口的内容  
        # self.setWidgetResizable(True)   #设置滚动窗口的内容可调整大小
        # self.verticalScrollBar().sliderPressed.connect(self.scrollContents) #滚动条滚动时，调用的函数
        #设置各个控件-----------------------------------------------------------------------------------------

        # 最外层的垂直布局
        container = QVBoxLayout()

        # -----创建第1个组，添加放置表格-----
        self.tableView = TableWidget(self)
        self.tableView.setWordWrap(False) #设置表格内容不换行
        self.tableView.setRowCount(2) #设置表格行数
        self.tableView.setColumnCount(2) #设置表格列数
        #self.tableView.verticalHeader().hide() #隐藏垂直表头
        self.tableView.setHorizontalHeaderLabels(['Src', 'Dst']) #设置水平表头
        self.tableView.resizeColumnsToContents() #设置列宽度自适应内容
        self.tableView.resizeRowsToContents() #设置行高度自适应内容
        self.tableView.setEditTriggers(QAbstractItemView.AllEditTriggers)   # 设置所有单元格可编辑
        #self.tableView.setFixedSize(500, 300)         # 设置表格大小
        self.tableView.setMaximumHeight(400)          # 设置表格的最大高度
        self.tableView.setMinimumHeight(400)             # 设置表格的最小高度
        self.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  #作用是将表格填满窗口
        #self.tableView.setSortingEnabled(True)  #设置表格可排序
        self.tableView.setBorderVisible(True) # 开启显示边框功能，从而可以修改表格角半径
        self.tableView.setBorderRadius(8) # 将表格组件的边角半径设置为x像素，从而实现圆角效果。

        # songInfos = [
        #     ['かばん', 'aiko']
        # ]
        # for i, songInfo in enumerate(songInfos): #遍历数据
        #     for j in range(2): #遍历每一列
        #         self.tableView.setItem(i, j, QTableWidgetItem(songInfo[j])) #设置每个单元格的内容


        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)



        # -----创建第1_1个组，添加多个组件-----
        box1_1 = QGroupBox()
        box1_1.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout1_1 = QHBoxLayout()


        #设置导入字典按钮
        self.pushButton1 = PushButton('导入字典', self, FIF.DOWNLOAD)
        self.pushButton1.clicked.connect(self.Importing_dictionaries) #按钮绑定槽函数

        #设置导出字典按钮
        self.pushButton2 = PushButton('导出字典', self, FIF.SHARE)
        self.pushButton2.clicked.connect(self.Exporting_dictionaries) #按钮绑定槽函数

        #设置清空字典按钮
        self.pushButton3 = PushButton('清空字典', self, FIF.DELETE)
        self.pushButton3.clicked.connect(self.Empty_dictionary) #按钮绑定槽函数

        #设置保存字典按钮
        self.pushButton4 = PushButton('保存字典', self, FIF.SAVE)
        self.pushButton4.clicked.connect(self.Save_dictionary) #按钮绑定槽函数


        layout1_1.addWidget(self.pushButton1)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton2)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton3)
        layout1_1.addStretch(1)  # 添加伸缩项
        layout1_1.addWidget(self.pushButton4)
        box1_1.setLayout(layout1_1)




        # -----创建第2个组，添加多个组件-----
        box2 = QGroupBox()
        box2.setStyleSheet(""" QGroupBox {border: 1px solid lightgray; border-radius: 8px;}""")#分别设置了边框大小，边框颜色，边框圆角
        layout2 = QHBoxLayout()

        #设置“译前替换”标签
        label1 = QLabel( flags=Qt.WindowFlags())  
        label1.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 17px;")
        label1.setText("译文修正")

        #设置“译前替换”显示
        self.label2 = QLabel(parent=self, flags=Qt.WindowFlags())  
        self.label2.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 11px;  color: black")
        self.label2.setText("(翻译完成后，根据字典内容对译文文本进行替换)")


        #设置“译前替换”开
        self.checkBox1 = CheckBox('启用功能')
        self.checkBox1.stateChanged.connect(self.checkBoxChanged1)

        layout2.addWidget(label1)
        layout2.addWidget(self.label2)
        layout2.addStretch(1)  # 添加伸缩项
        layout2.addWidget(self.checkBox1)
        box2.setLayout(layout2)




        # 把内容添加到容器中 
        container.addWidget(box2)   
        container.addWidget(self.tableView)
        container.addWidget(box1_1)
        container.addStretch(1)  # 添加伸缩项

        # 设置窗口显示的内容是最外层容器
        #self.scrollWidget.setLayout(container)
        self.setLayout(container)
        container.setSpacing(28) # 设置布局内控件的间距为28
        container.setContentsMargins(20, 10, 20, 20) # 设置布局的边距, 也就是外边框距离，分别为左、上、右、下    

    #滚动条滚动时，调用的函数
    def scrollContents(self, position):
        self.scrollWidget.move(0, position) 

    #添加行按钮
    def add_row(self):
        # 添加新行在按钮所在行前面
        self.tableView.insertRow(self.tableView.rowCount()-1)
        #设置新行的高度与前一行相同
        self.tableView.setRowHeight(self.tableView.rowCount()-2, self.tableView.rowHeight(self.tableView.rowCount()-3))

    #删除空白行按钮
    def delete_blank_row(self):
        #表格行数大于2时，删除表格内第一列和第二列为空或者空字符串的行
        if self.tableView.rowCount() > 2:
            # 删除表格内第一列和第二列为空或者空字符串的行
            for i in range(self.tableView.rowCount()-1):
                if self.tableView.item(i, 0) is None or self.tableView.item(i, 0).text() == '':
                    self.tableView.removeRow(i)
                    break
                elif self.tableView.item(i, 1) is None or self.tableView.item(i, 1).text() == '':
                    self.tableView.removeRow(i)
                    break

    # 移除JSON内容中的行内注释
    def remove_inline_comments(self, json_content):
        # 正则表达式匹配行内注释并替换为空字符串
        return re.sub(r'//.*$', '', json_content, flags=re.MULTILINE)

    #导入字典按钮
    def Importing_dictionaries(self):
        # 选择文件
        Input_File, _ = QFileDialog.getOpenFileName(None, 'Select File', '', 'All Files (*)')

        if Input_File:
            print(f'[INFO] 已选择文件: {Input_File}')
            # 获取文件后缀
            file_suffix = Input_File.split('.')[-1].lower()
            
            # 根据文件后缀执行不同操作
            if file_suffix == 'json':

                try:
                    # 尝试读取文件内容
                    with open(Input_File, 'r', encoding="utf-8") as f:
                        content = f.read()
                except FileNotFoundError:
                    print(f'[ERROR] 文件未找到: {Input_File}')
                    return
                except Exception as e:
                    print(f'[ERROR] 读取文件时发生未知错误: {str(e)}')
                    return
                try:
                    # 移除内容中的行内注释，并反序列化
                    dictionary = json.loads(self.remove_inline_comments(content))
                except json.JSONDecodeError as e:
                    print(f'[ERROR] JSON解析错误: {str(e)}')
                    return
                except Exception as e:
                    print(f'[ERROR] 反序列化时发生未知错误: {str(e)}')
                    return       

                # 将字典中的数据从表格底部添加到表格中
                for key, value in dictionary.items():
                    row = self.tableView.rowCount() - 1 #获取表格的倒数行数
                    self.tableView.insertRow(row)    # 在表格中插入一行
                    self.tableView.setItem(row, 0, QTableWidgetItem(key))
                    self.tableView.setItem(row, 1, QTableWidgetItem(value))
                    #设置新行的高度与前一行相同
                    self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))


                # 导入成功后删除空白行
                self.delete_blank_row()

                # 输出日志
                self.user_interface_prompter.createSuccessInfoBar("导入成功")
                print(f'[INFO]  已导入字典文件')


            elif file_suffix == 'xlsx':
                # 执行XLSX文件的操作
                wb = openpyxl.load_workbook(Input_File)
                sheet = wb.active
                for row in range(2, sheet.max_row + 1): # 第一行是标识头，第二行才开始读取
                    cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                    cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值


                    row = self.tableView.rowCount() - 1 #获取表格的倒数行数
                    self.tableView.insertRow(row)    # 在表格中插入一行
                    self.tableView.setItem(row, 0, QTableWidgetItem(cell_value1))
                    self.tableView.setItem(row, 1, QTableWidgetItem(cell_value2))
                    #设置新行的高度与前一行相同
                    self.tableView.setRowHeight(row, self.tableView.rowHeight(row-1))

                # 导入成功后删除空白行
                self.delete_blank_row()

                # 输出日志
                self.user_interface_prompter.createSuccessInfoBar("导入成功")
                print(f'[INFO]  已导入字典文件')
                    
            else:
                print(f'[INFO] 不支持的文件类型: .{file_suffix}')

        else:
            print('[INFO] 未选择文件')
            return    
    
    #导出字典按钮
    def Exporting_dictionaries(self):
        #获取表格中从第一行到倒数第二行的数据，判断第一列或第二列是否为空，如果为空则不获取。如果不为空，则第一列作为key，第二列作为value，存储中间字典中
        data = []
        for row in range(self.tableView.rowCount() - 1):
            key_item = self.tableView.item(row, 0)
            value_item = self.tableView.item(row, 1)
            if key_item and value_item:
                key = key_item.text()
                value = value_item.text()
                data.append((key, value))

        # 将数据存储到中间字典中
        dictionary = {}
        for key, value in data:
            dictionary[key] = value

        # 选择文件保存路径
        Output_Folder = QFileDialog.getExistingDirectory(None, 'Select Directory', '')      #调用QFileDialog类里的函数来选择文件目录
        if Output_Folder:
            print(f'[INFO]  已选择字典导出文件夹: {Output_Folder}')
        else :
            print('[INFO]  未选择文件夹')
            return  # 直接返回，不执行后续操作

        # 将字典保存到文件中
        with open(os.path.join(Output_Folder, "用户译后修正字典.json"), 'w', encoding="utf-8") as f:
            json.dump(dictionary, f, ensure_ascii=False, indent=4)

        self.user_interface_prompter.createSuccessInfoBar("导出成功")
        print(f'[INFO]  已导出字典文件')

    #清空字典按钮
    def Empty_dictionary(self):
        #清空表格
        self.tableView.clearContents()
        #设置表格的行数为1
        self.tableView.setRowCount(2)
        
        # 在表格最后一行第一列添加"添加行"按钮
        button = PushButton('添新行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 0, button)
        button.clicked.connect(self.add_row)
        # 在表格最后一行第二列添加"删除空白行"按钮
        button = PushButton('删空行')
        self.tableView.setCellWidget(self.tableView.rowCount()-1, 1, button)
        button.clicked.connect(self.delete_blank_row)

        self.user_interface_prompter.createSuccessInfoBar("清空成功")
        print(f'[INFO]  已清空字典')

    #保存字典按钮
    def Save_dictionary(self):
        self.user_interface_prompter.read_write_config("write",self.configurator.resource_dir)
        self.user_interface_prompter.createSuccessInfoBar("保存成功")
        print(f'[INFO]  已保存字典')


    #提示函数
    def checkBoxChanged1(self, isChecked: bool):
        if isChecked :
            self.user_interface_prompter.createSuccessInfoBar("已开启译后修正功能，将依据表格内容进行修正")
    