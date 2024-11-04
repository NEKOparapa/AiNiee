import os
import json

import openpyxl
from PyQt5.Qt import Qt
from PyQt5.QtWidgets import QFrame
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QHeaderView
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QTableWidgetItem

from qfluentwidgets import Action
from qfluentwidgets import FluentIcon
from qfluentwidgets import MessageBox
from qfluentwidgets import TableWidget

from Base.Base import Base
from Widget.CommandBarCard import CommandBarCard
from Widget.SwitchButtonCard import SwitchButtonCard

class TextReplaceAPage(QFrame, Base):

    DEFAULT = {
        "pre_translation_switch": False,
        "pre_translation_content": {},
    }

    def __init__(self, text: str, window):
        super().__init__(window)
        self.setObjectName(text.replace(" ", "-"))

        # 载入配置文件
        config = self.load_config()

        # 设置主容器
        self.container = QVBoxLayout(self)
        self.container.setSpacing(8)
        self.container.setContentsMargins(24, 24, 24, 24) # 左、上、右、下

        # 添加控件
        self.add_widget_header(self.container, config)
        self.add_widget_body(self.container, config)
        self.add_widget_footer(self.container, config, window)

    # 头部
    def add_widget_header(self, parent, config):
        def widget_init(widget):
            widget.set_checked(config.get("pre_translation_switch"))

        def widget_callback(widget, checked: bool):
            config = self.load_config()
            config["pre_translation_switch"] = checked
            self.save_config(config)

        parent.addWidget(
            SwitchButtonCard(
                "译前替换",
                "在翻译开始前，将原文中匹配的部分替换为指定的文本，执行的顺序为从上到下依次替换",
                widget_init,
                widget_callback,
            )
        )

    # 主体
    def add_widget_body(self, parent, config):

        def item_changed(item):
            item.setTextAlignment(Qt.AlignCenter)

        self.table = TableWidget(self)
        parent.addWidget(self.table)

        # 设置表格属性
        self.table.setBorderRadius(4)
        self.table.setBorderVisible(True)
        self.table.setWordWrap(False)
        self.table.setColumnCount(2)
        self.table.resizeRowsToContents() # 设置行高度自适应内容
        self.table.resizeColumnsToContents() # 设置列宽度自适应内容
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 撑满宽度
        self.table.itemChanged.connect(item_changed)

        # 设置水平表头并隐藏垂直表头
        self.table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table.setHorizontalHeaderLabels(
            [
                "原文",
                "替换",
            ],
        )

        # 向表格更新数据
        self.update_to_table(self.table, config)

    # 底部
    def add_widget_footer(self, parent, config, window):
        self.command_bar_card = CommandBarCard()
        parent.addWidget(self.command_bar_card)

        # 添加命令
        self.add_command_bar_action_01(self.command_bar_card)
        self.add_command_bar_action_02(self.command_bar_card)
        self.command_bar_card.add_separator()
        self.add_command_bar_action_03(self.command_bar_card)
        self.add_command_bar_action_04(self.command_bar_card)
        self.command_bar_card.add_separator()
        self.add_command_bar_action_05(self.command_bar_card)
        self.add_command_bar_action_06(self.command_bar_card, window)

    # 向表格更新数据
    def update_to_table(self, table, config):
        datas = []
        dictionary = config.get("pre_translation_content", {})

        # 构建表格数据
        for k, v in dictionary.items():
            datas.append(
                [k.strip(), v.strip()]
            )

        # 向表格中填充数据
        table.setRowCount(max(12, len(dictionary)))
        for row in range(len(datas)):
            for col in range(table.columnCount()):
                table.setItem(row, col, QTableWidgetItem(datas[row][col]))

    # 从表格更新数据
    def update_from_table(self, table, config):
        config["pre_translation_content"] = {}

        for row in range(table.rowCount()):
            data_str = table.item(row, 0)
            data_dst = table.item(row, 1)

            # 判断是否有数据
            if data_str == None or data_dst == None:
                continue

            data_str = data_str.text().strip()
            data_dst = data_dst.text().strip()

            # 判断是否有数据
            if data_str == "" or data_dst == "":
                continue

            config["pre_translation_content"][data_str] = data_dst

        return config

    # 导入
    def add_command_bar_action_01(self, parent):

        def load_json_file(path):
            dictionary = {}

            inputs = []
            with open(path, "r", encoding = "utf-8") as reader:
                inputs = json.load(reader)

            if isinstance(inputs, list) and len(inputs) > 0:
                for v in inputs:
                    # 标准术语表
                    # [
                    #     {
                    #         "srt": "ダリヤ",
                    #         "dst": "达莉雅",
                    #         "info": "女性的名字"
                    #     }
                    # ]
                    if isinstance(v, dict) and v.get("srt", "") != "" and v.get("dst", "") != "":
                        dictionary[v.get("srt", "").strip()] = v.get("dst", "").strip()

                    # Paratranz的术语表
                    # [
                    #   {
                    #     "id": 359894,
                    #     "createdAt": "2024-04-06T18:43:56.075Z",
                    #     "updatedAt": "2024-04-06T18:43:56.075Z",
                    #     "updatedBy": null,
                    #     "pos": "noun",
                    #     "uid": 49900,
                    #     "term": "アイテム",
                    #     "translation": "道具",
                    #     "note": "",
                    #     "project": 9841,
                    #     "variants": []
                    #   }
                    # ]
                    if isinstance(v, dict) and v.get("term", "") != "" and v.get("translation", "") != "":
                        dictionary[v.get("term", "").strip()] = v.get("translation", "").strip()
            elif isinstance(inputs, dict):
                # 普通 KV 格式
                # [
                #     "ダリヤ": "达莉雅"
                # ]
                for k, v in inputs.items():
                    if isinstance(v, str) and k != "" and v != "":
                        dictionary[k.strip()] = v.strip()

            return dictionary

        def load_xlsx_file(path):
            dictionary = {}

            sheet = openpyxl.load_workbook(path).active
            for row in range(2, sheet.max_row + 1):                     # 第一行是标识头，第二行才开始读取
                cell_01 = sheet.cell(row = row, column = 1).value       # 第N行第一列的值
                cell_02 = sheet.cell(row = row, column = 2).value       # 第N行第二列的值
                cell_03 = sheet.cell(row = row, column = 3).value       # 第N行第三列的值

                if cell_01 != None and cell_02 != None and cell_01 != "" and cell_02 != "":
                    dictionary[str(cell_01).strip()] = str(cell_02).strip()

            return dictionary

        def callback():
            # 选择文件
            path, _ = QFileDialog.getOpenFileName(None, "选择文件", "", "json files (*.json);;xlsx files (*.xlsx)")
            if path == None or path == "":
                return

            # 获取文件后缀
            file_suffix = path.split(".")[-1].lower()

            datas = []
            if file_suffix == "json":
                datas = load_json_file(path)

            if file_suffix == "xlsx":
                datas = load_xlsx_file(path)

            # 读取配置文件
            config = self.load_config()
            config["pre_translation_content"].update(datas)

            # 保存配置文件
            config = self.save_config(config)

            # 向表格更新数据
            self.update_to_table(self.table, config)

            # 弹出提示
            self.success_toast("", "数据已导入 ...")

        parent.add_action(
            Action(FluentIcon.DOWNLOAD, "导入", parent, triggered = callback),
        )

    # 导出
    def add_command_bar_action_02(self, parent):
        def callback():
            # 读取配置文件
            config = self.load_config()

            # 从表格更新数据
            config = self.update_from_table(self.table, config)

            # 整理数据
            datas = []
            user_dictionary = config.get("pre_translation_content", {})
            for k, v in user_dictionary.items():
                datas.append(
                    {
                        "srt": k,
                        "dst": v,
                        "info": "",
                    }
                )

            # 选择文件导出路径
            path = QFileDialog.getExistingDirectory(None, "Select Directory", "")
            if path == None or path == "":
                return

            # 导出文件
            with open(os.path.join(path, "导出_译前替换.json"), "w", encoding = "utf-8") as writer:
                writer.write(json.dumps(datas, indent = 4, ensure_ascii = False))

            # 弹出提示
            self.success_toast("", "数据已导出为 \"导出_译前替换.json\" ...")

        parent.add_action(
            Action(FluentIcon.SHARE, "导出", parent, triggered = callback),
        )

    # 添加新行
    def add_command_bar_action_03(self, parent):
        def callback():
            # 添加新行
            self.table.setRowCount(self.table.rowCount() + 1)

            # 弹出提示
            self.success_toast("", "新行已添加 ...")

        parent.add_action(
            Action(FluentIcon.ADD_TO, "添加新行", parent, triggered = callback),
        )

    # 移除空行
    def add_command_bar_action_04(self, parent):
        def callback():
            # 从表格更新数据，生成一个临时的配置文件
            config = self.update_from_table(self.table, {})

            # 清空表格
            self.table.clearContents()

            # 向表格更新数据
            self.update_to_table(self.table, config)

            # 保存配置文件
            config = self.save_config(config)

            # 弹出提示
            self.success_toast("", "空行已移除 ...")

        parent.add_action(
            Action(FluentIcon.BROOM, "移除空行", parent, triggered = callback),
        )

    # 保存
    def add_command_bar_action_05(self, parent):
        def callback():
            # 读取配置文件
            config = self.load_config()

            # 从表格更新数据
            config = self.update_from_table(self.table, config)

            # 保存配置文件
            config = self.save_config(config)

            # 弹出提示
            self.success_toast("", "数据已保存 ...")

        parent.add_action(
            Action(FluentIcon.SAVE, "保存", parent, triggered = callback),
        )

    # 重置
    def add_command_bar_action_06(self, parent, window):
        def callback():
            message_box = MessageBox("警告", "是否确认重置为默认数据 ... ？", window)
            message_box.yesButton.setText("确认")
            message_box.cancelButton.setText("取消")

            if not message_box.exec():
                return

            # 清空表格
            self.table.clearContents()

            # 读取配置文件
            config = self.load_config()

            # 加载默认设置
            config["pre_translation_content"] = self.DEFAULT.get("pre_translation_content")

            # 保存配置文件
            config = self.save_config(config)

            # 向表格更新数据
            self.update_to_table(self.table, config)

            # 弹出提示
            self.success_toast("", "数据已重置 ...")

        parent.add_action(
            Action(FluentIcon.DELETE, "重置", parent, triggered = callback),
        )