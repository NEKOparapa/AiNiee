import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.escape import escape

from ModuleFolders.Cache.CacheFile import CacheFile
from ModuleFolders.Cache.CacheItem import TranslationStatus
from ModuleFolders.Cache.CacheProject import ProjectType
from ModuleFolders.FileOutputer.BaseWriter import (
    BaseTranslatedWriter,
    OutputConfig,
    PreWriteMetadata
)


class TPPWriter(BaseTranslatedWriter):
    def __init__(self, output_config: OutputConfig):
        super().__init__(output_config)

    WORKBOOK_HEADER = [
        "Original Text",
        "Initial",
        "Machine translation",
        "Better translation",
        "Best translation",
    ]

    def on_write_translated(
        self, translation_file_path: Path, cache_file: CacheFile,
        pre_write_metadata: PreWriteMetadata,
        source_file_path: Path = None,
    ):
        # 创建一个工作簿
        wb = Workbook()

        # 选择默认的活动工作表
        ws = wb.active

        # 添加表头
        ws.append(self.WORKBOOK_HEADER)

        # 将数据写入工作表
        for item in cache_file.items:
            source_text = item.source_text or ""
            translated_text = item.final_text or ""
            row_index = item.get_extra("row_index", -1)
            translation_status = item.translation_status

            # 根据翻译状态写入原文及译文
            # 如果文本是以 = 开始，则加一个空格
            # 因为 = 开头会被识别成 Excel 公式导致 T++ 导入时 卡住
            # 加入空格后，虽然还是不能直接导入 T++ ，但是可以手动复制粘贴
            if item.translation_status == TranslationStatus.TRANSLATED or item.translation_status == TranslationStatus.POLISHED:
                ws.cell(row=row_index, column=1).value = re.sub(r"^=", " =", source_text)

                # 防止含有特殊字符而不符合Excel公式时，导致的写入译文错误
                try:
                    ws.cell(row=row_index, column=2).value = re.sub(r"^=", " =", translated_text)
                except:
                    # 过滤非法控制字符并转义XML特殊字符
                    filtered_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', translated_text)
                    escaped_string = escape(filtered_text)
                    ws.cell(row=row_index, column=2).value = escaped_string

            else :
                ws.cell(row=row_index, column=1).value = re.sub(r"^=", " =", source_text)
                
        # 保存工作簿
        wb.save(translation_file_path)

    @classmethod
    def get_project_type(self):
        return ProjectType.TPP
