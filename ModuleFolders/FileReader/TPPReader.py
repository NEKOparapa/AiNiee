from pathlib import Path

import openpyxl  # 需安装库pip install openpyxl

from ModuleFolders.Cache.CacheFile import CacheFile
from ModuleFolders.Cache.CacheItem import CacheItem, TranslationStatus
from ModuleFolders.Cache.CacheProject import ProjectType
from ModuleFolders.FileReader.BaseReader import (
    BaseSourceReader,
    InputConfig,
    PreReadMetadata
)


class TPPReader(BaseSourceReader):
    def __init__(self, input_config: InputConfig):
        super().__init__(input_config)

    @classmethod
    def get_project_type(cls):
        return ProjectType.TPP

    @property
    def support_file(self):
        return "xlsx"

    def on_read_source(self, file_path: Path, pre_read_metadata: PreReadMetadata) -> CacheFile:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        items = []
        for row in range(2, sheet.max_row + 1):  # 从第二行开始读取，因为第一行是标识头，通常不用理会
            cell_value1 = sheet.cell(row=row, column=1).value  # 第N行第一列的值
            cell_value2 = sheet.cell(row=row, column=2).value  # 第N行第二列的值
            source_text = str(cell_value1) if cell_value1 is not None else ""  # 获取原文

            if cell_value1:
                # 第1列的值不为空，和第2列的值为空，是未翻译内容
                # 第1列的值不为空，和第2列的值不为空，是已经翻译内容
                if cell_value2 is not None:
                    translated_text = cell_value2
                    translation_status = TranslationStatus.TRANSLATED
                else:
                    translated_text = ''
                    translation_status = TranslationStatus.UNTRANSLATED

                item = CacheItem(
                    source_text=source_text,
                    translated_text=translated_text,
                    translation_status=translation_status,
                    extra={"row_index": row},
                )
                items.append(item)
        return CacheFile(items=items)
