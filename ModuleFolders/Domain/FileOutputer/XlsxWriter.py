from pathlib import Path

import openpyxl  # 需安装库pip install openpyxl
from openpyxl.utils.escape import escape
import re

from ModuleFolders.Infrastructure.Cache.CacheFile import CacheFile
from ModuleFolders.Infrastructure.Cache.CacheProject import ProjectType
from ModuleFolders.Domain.FileOutputer.BaseWriter import (
    BaseTranslatedWriter,
    OutputConfig,
    PreWriteMetadata
)


class XlsxWriter(BaseTranslatedWriter):
    def __init__(self, output_config: OutputConfig):
        super().__init__(output_config)

    @classmethod
    def get_project_type(cls):
        return ProjectType.XLSX

    def on_write_translated(
        self, translation_file_path: Path, cache_file: CacheFile,
        pre_write_metadata: PreWriteMetadata,
        source_file_path: Path = None,
    ):
        # 1. 从 extra 中获取表头
        header = cache_file.get_extra("header")
        if not header:
            print(f"Error: Header not found in cache for {translation_file_path.name}")
            return

        # 2. 构建数据映射以便快速查找: (row, col) -> final_text
        # 使用 final_text 确保获取的是 润色后 > 翻译后 > 原文
        data_map = {
            (item.get_extra("row"), item.get_extra("col")): item.final_text 
            for item in cache_file.items
        }

        # 3. 获取所有有数据的行号（去重并排序）
        rows_with_data = sorted(set(r for r, c in data_map.keys()))
        
        # 计算最大列数
        max_col = 0
        if data_map:
            max_col = max(c for r, c in data_map.keys())
        
        # 列数由表头决定，取表头长度和最大列数的最大值
        num_cols = max(len(header), max_col + 1)

        # 4. 创建工作簿并写入数据
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 写入表头
            for col_idx, header_text in enumerate(header, start=1):
                ws.cell(row=1, column=col_idx).value = header_text
            
            # 写入内容：只遍历有数据的行，每个单元格根据位置信息(row, col)写入
            # row 从0开始计数（表头不算），所以写入时需要 +2（+1因为从0开始，+1因为表头占一行）
            for r in rows_with_data:
                for c in range(num_cols):
                    cell_value = data_map.get((r, c), None)
                    row_index = r + 2  # +1 因为从0开始，+1 因为表头占一行
                    col_index = c + 1
                    
                    # 只写入有数据的单元格
                    if cell_value is not None:
                        # 如果文本是以 = 开始，则加一个空格
                        # 因为 = 开头会被识别成 Excel 公式
                        if isinstance(cell_value, str) and cell_value.startswith("="):
                            cell_value = " " + cell_value
                        
                        # 防止含有特殊字符而不符合Excel公式时，导致的写入译文错误
                        try:
                            ws.cell(row=row_index, column=col_index).value = cell_value
                        except:
                            # 过滤非法控制字符并转义XML特殊字符
                            if isinstance(cell_value, str):
                                filtered_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', cell_value)
                                escaped_string = escape(filtered_text)
                                ws.cell(row=row_index, column=col_index).value = escaped_string
                            else:
                                ws.cell(row=row_index, column=col_index).value = ""
            
            # 保存工作簿
            wb.save(translation_file_path)
            
        except Exception as e:
            print(f"Error writing translated XLSX: {e}")
