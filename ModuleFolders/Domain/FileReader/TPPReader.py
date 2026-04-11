from pathlib import Path

import openpyxl

from ModuleFolders.Service.Cache.CacheFile import CacheFile
from ModuleFolders.Service.Cache.CacheItem import CacheItem, TranslationStatus
from ModuleFolders.Service.Cache.CacheProject import ProjectType
from ModuleFolders.Domain.FileReader.BaseReader import (
    BaseSourceReader,
    InputConfig,
    PreReadMetadata,
)


class TPPReader(BaseSourceReader):
    HEADER_PAIRS = (
        ("Original Text", "Initial"),
        ("原文", "译文"),
    )

    def __init__(self, input_config: InputConfig):
        super().__init__(input_config)

    @classmethod
    def get_project_type(cls):
        return ProjectType.TPP

    @property
    def support_file(self):
        return "xlsx"

    @classmethod
    def is_tpp_header(cls, first_header: str, second_header: str) -> bool:
        first_header = (first_header or "").strip()
        second_header = (second_header or "").strip()
        return (first_header, second_header) in cls.HEADER_PAIRS

    @classmethod
    def is_tpp_workbook(cls, file_path: Path) -> bool:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            try:
                sheet = wb.active
                first_header = str(sheet.cell(row=1, column=1).value or "")
                second_header = str(sheet.cell(row=1, column=2).value or "")
                return cls.is_tpp_header(first_header, second_header)
            finally:
                wb.close()
        except Exception:
            return False

    def can_read_by_content(self, file_path: Path) -> bool:
        return self.is_tpp_workbook(file_path)

    def on_read_source(self, file_path: Path, pre_read_metadata: PreReadMetadata) -> CacheFile:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        items = []
        for row in range(2, sheet.max_row + 1):
            cell_value1 = sheet.cell(row=row, column=1).value
            cell_value2 = sheet.cell(row=row, column=2).value
            source_text = str(cell_value1) if cell_value1 is not None else ""

            if cell_value1:
                if cell_value2 is not None:
                    translated_text = cell_value2
                    translation_status = TranslationStatus.TRANSLATED
                else:
                    translated_text = ""
                    translation_status = TranslationStatus.UNTRANSLATED

                item = CacheItem(
                    source_text=source_text,
                    translated_text=translated_text,
                    translation_status=translation_status,
                    extra={"row_index": row},
                )
                items.append(item)
        return CacheFile(items=items)
