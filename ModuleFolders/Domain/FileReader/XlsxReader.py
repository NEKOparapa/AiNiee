from pathlib import Path

import openpyxl

from ModuleFolders.Service.Cache.CacheFile import CacheFile
from ModuleFolders.Service.Cache.CacheItem import CacheItem, TranslationStatus
from ModuleFolders.Service.Cache.CacheProject import ProjectType
from ModuleFolders.Domain.FileAccessor.WolfXlsxAccessor import WolfXlsxAccessor
from ModuleFolders.Domain.FileReader.BaseReader import (
    BaseSourceReader,
    InputConfig,
    PreReadMetadata,
)
from ModuleFolders.Domain.FileReader.TPPReader import TPPReader


class XlsxReader(BaseSourceReader):
    def __init__(self, input_config: InputConfig):
        super().__init__(input_config)

    @classmethod
    def get_project_type(cls):
        return ProjectType.XLSX

    @property
    def support_file(self):
        return "xlsx"

    def can_read_by_content(self, file_path: Path) -> bool:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            try:
                sheet = wb.active
                header = [
                    "" if value is None else str(value)
                    for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                ]
                first_header = header[0] if len(header) > 0 else ""
                second_header = header[1] if len(header) > 1 else ""
                return (
                    not TPPReader.is_tpp_header(first_header, second_header)
                    and not WolfXlsxAccessor.is_wolf_header(header)
                )
            finally:
                wb.close()
        except Exception:
            return False

    def on_read_source(self, file_path: Path, pre_read_metadata: PreReadMetadata) -> CacheFile:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        items = []
        header = []

        try:
            if sheet.max_row > 0:
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    header.append(str(cell_value) if cell_value is not None else "")

            if not header:
                return None

            for row in range(2, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        item = CacheItem(
                            source_text=str(cell_value),
                            translation_status=TranslationStatus.UNTRANSLATED,
                            extra={
                                "row": row - 2,
                                "col": col - 1,
                            },
                        )
                        items.append(item)

        except Exception as e:
            print(f"Error reading XLSX {file_path}: {e}")
            return None

        finally:
            wb.close()

        if not items:
            return None

        return CacheFile(items=items, extra={"header": header})
