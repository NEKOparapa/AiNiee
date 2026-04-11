import re
from pathlib import Path

import openpyxl
from openpyxl.utils.escape import escape


class WolfXlsxAccessor:
    SOURCE_COL = 6
    TARGET_COL = 7

    REQUIRED_HEADERS = (
        "Code (No Change)",
        "Flag (No Change)",
        "Type",
        "Info",
        "Your notes",
        "Original text (No Change)",
    )
    TRANSLATED_HEADER_PREFIX = "Translated text 1"
    HALF_WIDTH_FLAG = "<Half-Width Characters Only>"

    SEPARATOR_PATTERN = re.compile(r"^\s*[-=~_/\\\.]{3,}\s*$")
    VERSION_PATTERN = re.compile(r"^\s*v?\d+(?:\.\d+)+\s*$", re.IGNORECASE)
    ILLEGAL_CONTROL_CHARS_PATTERN = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
    FILE_SUFFIXES = (
        ".png", ".jpg", ".jpeg", ".gif", ".webp",
        ".wav", ".mp3", ".ogg", ".txt", ".json",
        ".bmp", ".aac", ".flac", ".avi", ".mov",
        ".mkv", ".flv",
    )
    COMPACT_FILE_REFERENCE_PATTERN = re.compile(
        r"(?i)\S+\.(png|jpg|jpeg|gif|webp|wav|mp3|ogg|txt|json|bmp|aac|flac|avi|mov|mkv|flv)\s*$"
    )
    AUDIO_SUFFIXES = (".wav", ".mp3", ".ogg")

    @classmethod
    def normalize_text(cls, value) -> str:
        return "" if value is None else str(value)

    @classmethod
    def read_header(cls, sheet) -> list[str]:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [cls.normalize_text(value) for value in header_row]

    @classmethod
    def is_wolf_header(cls, header: list[str]) -> bool:
        if len(header) < cls.TARGET_COL:
            return False

        if tuple(header[:len(cls.REQUIRED_HEADERS)]) != cls.REQUIRED_HEADERS:
            return False

        return header[cls.TARGET_COL - 1].startswith(cls.TRANSLATED_HEADER_PREFIX)

    @classmethod
    def is_wolf_workbook(cls, file_path: Path) -> bool:
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
            try:
                sheet = workbook[workbook.sheetnames[0]]
                return cls.is_wolf_header(cls.read_header(sheet))
            finally:
                workbook.close()
        except Exception:
            return False

    @classmethod
    def build_file_extra(cls, file_path: Path) -> dict:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header = cls.read_header(sheet)
            if not cls.is_wolf_header(header):
                raise ValueError(f"`{file_path}` is not a WOLF translation workbook")

            return {
                "sheet_name": sheet.title,
                "header": header,
                "target_col": cls.TARGET_COL,
            }
        finally:
            workbook.close()

    @classmethod
    def iter_rows(cls, file_path: Path):
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header = cls.read_header(sheet)
            if not cls.is_wolf_header(header):
                raise ValueError(f"`{file_path}` is not a WOLF translation workbook")

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_values = list(row)
                if len(row_values) < cls.TARGET_COL:
                    row_values.extend([None] * (cls.TARGET_COL - len(row_values)))

                yield {
                    "row_index": row_index,
                    "code": cls.normalize_text(row_values[0]) if len(row_values) > 0 else "",
                    "flag": cls.normalize_text(row_values[1]) if len(row_values) > 1 else "",
                    "type": cls.normalize_text(row_values[2]) if len(row_values) > 2 else "",
                    "info": cls.normalize_text(row_values[3]) if len(row_values) > 3 else "",
                    "notes": cls.normalize_text(row_values[4]) if len(row_values) > 4 else "",
                    "source_text": cls.normalize_text(row_values[cls.SOURCE_COL - 1]),
                    "translated_text": cls.normalize_text(row_values[cls.TARGET_COL - 1]),
                }
        finally:
            workbook.close()

    @classmethod
    def has_half_width_flag(cls, flag: str) -> bool:
        return cls.HALF_WIDTH_FLAG in flag

    @classmethod
    def should_translate(cls, source_text: str, info: str, item_type: str = "") -> bool:
        stripped_text = source_text.strip()
        if not stripped_text:
            return False

        if stripped_text.isdigit():
            return False

        if cls.SEPARATOR_PATTERN.fullmatch(source_text):
            return False

        if cls.VERSION_PATTERN.fullmatch(source_text):
            return False

        if "font" in info.lower():
            return False

        if cls.COMPACT_FILE_REFERENCE_PATTERN.fullmatch(stripped_text):
            return False

        if ("BGM" in item_type or "SE" in item_type) and stripped_text.lower().endswith(cls.AUDIO_SUFFIXES):
            return False

        return True

    @classmethod
    def sanitize_output_text(cls, text: str) -> str:
        text = cls.normalize_text(text)
        if text.startswith("="):
            text = " " + text

        return cls.ILLEGAL_CONTROL_CHARS_PATTERN.sub("", text)

    @classmethod
    def write_translations(cls, source_file_path: Path, translation_file_path: Path, translations: dict[int, str]):
        workbook = openpyxl.load_workbook(source_file_path)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header = cls.read_header(sheet)
            if not cls.is_wolf_header(header):
                raise ValueError(f"`{source_file_path}` is not a WOLF translation workbook")

            for row_index, translated_text in translations.items():
                sanitized_text = cls.sanitize_output_text(translated_text)
                try:
                    sheet.cell(row=row_index, column=cls.TARGET_COL).value = sanitized_text
                except Exception:
                    sheet.cell(row=row_index, column=cls.TARGET_COL).value = escape(sanitized_text)

            workbook.save(translation_file_path)
        finally:
            workbook.close()
